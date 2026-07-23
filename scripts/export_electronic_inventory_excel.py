#!/usr/bin/env python3
"""
Export the Electronic Inventory share into an Excel workbook saved to the
user's Downloads folder — one sheet per document category listing which files
each counterparty has on disk.

Source layout (same as apps/pages/routes.py — Electronic Inventory page):
    ELECTRONIC_INVENTORY_ROOT\\{Counterparty}\\Transactional      (flat)
    ELECTRONIC_INVENTORY_ROOT\\{Counterparty}\\SSI                (flat)
    ELECTRONIC_INVENTORY_ROOT\\{Counterparty}\\Confirmations\\<yyyy>\\<mm>. <Month>\\<dd>\\<Product>

Output:  ~/Downloads/electronic_inventory_export_<YYYYMMDD_HHMMSS>.xlsx

Sheets:
  - Transacionais : Contraparte | Arquivo
  - SSI           : Contraparte | Arquivo
  - Confirmações  : Contraparte | Arquivo | Subpasta (yyyy/mm/dd/produto)

The counterparty name is the on-disk folder name (the sanitized COUNTERPARTY
from RefData that the app itself creates). Category folders are matched
case-insensitively; every category is walked recursively so files filed in an
unexpected subfolder still show up (the subfolder goes in Subpasta). Junk
files (Thumbs.db, desktop.ini, .DS_Store, ~$ locks, dotfiles) are skipped.

Usage:
    python scripts/export_electronic_inventory_excel.py
    python scripts/export_electronic_inventory_excel.py --root "D:\\OTC\\Electronic Inventory"
"""

import argparse
import os
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

DEFAULT_ROOT = os.getenv('ELECTRONIC_INVENTORY_ROOT',
                         r'I:\Confirmation\Derivativos\OTC Tracker\Electronic Inventory')

# (sheet title, on-disk category folder). Matched case-insensitively.
SHEETS = [
    ('Transacionais', 'Transactional'),
    ('SSI', 'SSI'),
    ('Confirmações', 'Confirmations'),
]

JUNK = {'thumbs.db', 'desktop.ini', '.ds_store'}


def long_path(path):
    r"""Windows extended-length form (\\?\...) — the Confirmations tree repeats
    long counterparty names and can pass MAX_PATH (260), where every os call
    fails as a plain "not found". No-op off Windows / on short paths."""
    if os.name != 'nt' or len(path) < 250 or path.startswith('\\\\?\\'):
        return path
    if path.startswith('\\\\'):
        return '\\\\?\\UNC' + path[1:]
    return '\\\\?\\' + path


def is_junk(name):
    return (name.lower() in JUNK or name.startswith('~$') or name.startswith('.'))


def find_category_dir(client_dir, category):
    """Case-insensitive match of the category folder inside a counterparty dir."""
    try:
        for entry in os.listdir(long_path(client_dir)):
            if entry.lower() == category.lower():
                p = os.path.join(client_dir, entry)
                if os.path.isdir(long_path(p)):
                    return p
    except OSError:
        pass
    return None


def walk_files(base):
    """[(relative subfolder or '', filename)] under `base`, junk skipped."""
    out = []
    for cur, _dirs, files in os.walk(long_path(base)):
        rel = os.path.relpath(cur, long_path(base))
        rel = '' if rel == '.' else rel.replace('\\', '/')
        for f in files:
            if not is_junk(f):
                out.append((rel, f))
    return out


def collect(root):
    """{sheet title: [row, ...]} for the whole share."""
    rows = {title: [] for title, _cat in SHEETS}
    try:
        clients = sorted((e for e in os.listdir(long_path(root))
                          if os.path.isdir(long_path(os.path.join(root, e)))),
                         key=str.upper)
    except OSError as exc:
        sys.exit('ERRO: não foi possível ler a raiz %r (%s) — o share I: está acessível?'
                 % (root, exc))
    for client in clients:
        client_dir = os.path.join(root, client)
        for title, cat in SHEETS:
            cat_dir = find_category_dir(client_dir, cat)
            if not cat_dir:
                continue
            for rel, fname in sorted(walk_files(cat_dir)):
                if title == 'Confirmações':
                    rows[title].append([client, fname, rel])
                else:
                    # Flat categories: a stray subfolder is appended to the file
                    # name so the information is never silently dropped.
                    rows[title].append([client, ('%s/%s' % (rel, fname)) if rel else fname])
    return rows


def build_workbook(rows):
    wb = Workbook()
    wb.remove(wb.active)
    head_font = Font(bold=True, color='FFFFFF')
    head_fill = PatternFill('solid', fgColor='0066CC')
    for title, _cat in SHEETS:
        ws = wb.create_sheet(title)
        headers = ['Contraparte', 'Arquivo'] + (['Subpasta'] if title == 'Confirmações' else [])
        ws.append(headers)
        for cell in ws[1]:
            cell.font = head_font
            cell.fill = head_fill
        for r in rows[title]:
            ws.append(r)
        widths = [len(h) for h in headers]
        for r in rows[title]:
            for i, v in enumerate(r):
                widths[i] = max(widths[i], len(str(v)))
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = min(w + 2, 80)
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions
    return wb


def main():
    ap = argparse.ArgumentParser(description='Exporta o Electronic Inventory para Excel.')
    ap.add_argument('--root', default=DEFAULT_ROOT,
                    help='raiz do Electronic Inventory (default: %(default)s)')
    ap.add_argument('--out', default=None,
                    help='pasta de saída (default: ~/Downloads)')
    args = ap.parse_args()

    rows = collect(args.root)
    wb = build_workbook(rows)

    out_dir = args.out or os.path.join(os.path.expanduser('~'), 'Downloads')
    os.makedirs(out_dir, exist_ok=True)
    out = os.path.join(out_dir, 'electronic_inventory_export_%s.xlsx'
                       % datetime.now().strftime('%Y%m%d_%H%M%S'))
    wb.save(out)

    print('Gerado: %s' % out)
    for title, _cat in SHEETS:
        print('  %-14s %d arquivo(s)' % (title, len(rows[title])))


if __name__ == '__main__':
    main()
