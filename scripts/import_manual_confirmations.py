"""
import_manual_confirmations.py
------------------------------
Cria os dois bancos de Manual Confirmations e carrega o MANUAIS.xlsx neles.

  apps/static/data/db/manual_confirmations_pending.db  — a esteira ainda não fechou
  apps/static/data/db/manual_confirmations_ok.db       — fechou

Os dois têm a mesma tabela `manual_confirmations`, com as colunas da página. Em
qual banco cada linha entra NÃO vem da planilha: sai do estado da própria linha
(quem já validou, e quem ainda precisa validar segundo o cadastro
`manual-conf-validation`). Uma coluna "Pending" digitada na planilha discordaria
das datas ao lado dela na primeira mudança, e a tela mostraria uma etapa que já
passou.

O cabeçalho da planilha é casado por SEMELHANÇA (sem acento, sem caixa, sem
pontuação), porque o arquivo veio de uma planilha de trabalho: 'VALIDADO p/ MO '
com espaço no fim, 'Data Operação' com e sem acento, e três colunas chamadas
'Time Stamp'. As três são resolvidas pela POSIÇÃO — a que vem depois de
'Conferido OTC' é a do OTC, a de depois de 'VALIDADO p/ MO' é a do MO, e assim
por diante —, que é a única informação que distingue as três no arquivo.

Uso
    python scripts/import_manual_confirmations.py
    python scripts/import_manual_confirmations.py --xlsx ~/Downloads/MANUAIS.xlsx
    python scripts/import_manual_confirmations.py --sheet "Manuais"
    python scripts/import_manual_confirmations.py --dry-run     # só relata
    python scripts/import_manual_confirmations.py --schema-only # só cria os bancos

É idempotente: o import REESCREVE os dois bancos a partir da planilha. Rodar
duas vezes não duplica linha.
"""
import argparse
import os
import re
import sys
import unicodedata
from datetime import datetime

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT = os.path.dirname(SCRIPT_DIR)
sys.path.insert(0, REPO_ROOT)

from apps.pages import manual_conf as MC                          # noqa: E402

DEFAULT_XLSX_CANDIDATES = [
    os.path.join(SCRIPT_DIR, 'MANUAIS.xlsx'),
    os.path.expanduser('~/Downloads/MANUAIS.xlsx'),
    os.path.expanduser('~/Desktop/MANUAIS.xlsx'),
]


def norm_header(s):
    """Cabeçalho comparável: sem acento, sem pontuação, sem caixa, sem espaço."""
    s = unicodedata.normalize('NFKD', str(s or '')).encode('ascii', 'ignore').decode('ascii')
    return re.sub(r'[^a-z0-9]', '', s.lower())


# Nome na planilha → coluna do banco, para os casos em que a semelhança não
# basta. As três 'Time Stamp' NÃO entram aqui: elas são resolvidas por posição.
ALIASES = {
    norm_header('Aging Confirmacao'): 'Aging Confirmação',
    norm_header('Data Operacao'): 'Data Operação',
    norm_header('Data de envio validacao Registro'): 'Data de envio validação Registro',
    norm_header('Data validacao Registro'): 'Data validação Registro',
    norm_header('Data envio validacao OTC'): 'Data envio validação OTC',
    norm_header('Data envio validacao MO/FO'): 'Data envio validação MO/FO',
    norm_header('Enviado p/ cliente (desbloqueado no fep)'): 'Enviado p/ cliente (desbloqueado no fep)',
    norm_header('Cliente'): 'Cliente',
    norm_header('Legal Entity'): 'Legal Entity',
}

# A coluna que vem ANTES de cada 'Time Stamp' no arquivo original. É por ela que
# se sabe de quem é o carimbo.
TIMESTAMP_AFTER = {
    norm_header('Conferido OTC'): 'Time Stamp OTC',
    norm_header('VALIDADO p/ MO'): 'Time Stamp MO',
    norm_header('VALIDADO p/ FO'): 'Time Stamp FO',
}


def map_headers(header_row):
    """[(índice, coluna do banco)] a partir da linha de cabeçalho da planilha.

    Devolve também a lista de cabeçalhos que não casaram, para o relatório: uma
    coluna que a planilha tem e o banco não é informação perdida, e o script tem
    de dizer isso em voz alta em vez de descartá-la em silêncio.
    """
    by_norm = {norm_header(c): c for c in MC.COLUMNS}
    out, unmatched, prev = [], [], None
    for i, raw in enumerate(header_row):
        n = norm_header(raw)
        if not n:
            prev = None
            continue
        if n == norm_header('Time Stamp'):
            col = TIMESTAMP_AFTER.get(prev)
            if col:
                out.append((i, col))
            else:
                unmatched.append((i, raw, 'Time Stamp sem coluna de validação antes dele'))
            prev = n
            continue
        col = by_norm.get(n) or ALIASES.get(n)
        if col:
            # 'Trade ID' aparecia duas vezes no arquivo; a segunda é cópia.
            if col in [c for _i, c in out]:
                unmatched.append((i, raw, 'coluna repetida — a primeira ocorrência vale'))
            else:
                out.append((i, col))
        else:
            unmatched.append((i, raw, 'sem coluna correspondente no banco'))
        prev = n
    return out, unmatched


def cell(v):
    """Valor da célula como TEXTO, no formato que a página lê.

    Datas viram dd/mm/aaaa: o openpyxl devolve datetime para as células
    formatadas como data, e um str() nelas escreveria '2026-08-04 00:00:00' na
    tabela.
    """
    if v is None:
        return ''
    if isinstance(v, datetime):
        return v.strftime('%d/%m/%Y')
    if hasattr(v, 'strftime') and hasattr(v, 'year'):
        return v.strftime('%d/%m/%Y')
    if isinstance(v, float) and float(v).is_integer():
        return str(int(v))
    return str(v).strip()


def read_xlsx(path, sheet=None):
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        raise SystemExit('A planilha "%s" está vazia.' % (sheet or ws.title))
    return ws.title, rows


def build_rows(raw_rows):
    header, data = raw_rows[0], raw_rows[1:]
    mapping, unmatched = map_headers(header)
    if not mapping:
        raise SystemExit('Nenhuma coluna da planilha casou com o banco — confira a aba.')
    rules = MC.validation_rules()
    out = []
    for r in data:
        row = {c: '' for c in MC.DB_COLUMNS}
        for i, col in mapping:
            row[col] = cell(r[i]) if i < len(r) else ''
        if not str(row.get(MC.KEY_COLUMN, '')).strip():
            continue          # linha sem Trade ID não tem chave — provavelmente vazia
        MC.refresh_derived(row, rules)
        out.append(row)
    return out, mapping, unmatched


def rewrite(category, rows):
    """Reescreve um banco inteiro. Só o import faz isso — a aplicação grava linha
    a linha; aqui a planilha é a verdade e o banco tem de espelhá-la."""
    import duckdb
    path = MC.db_path(category)
    MC.ensure_db(path)
    con = duckdb.connect(path)
    try:
        cols = ', '.join('"{}"'.format(c) for c in MC.DB_COLUMNS)
        ph = ', '.join('?' for _ in MC.DB_COLUMNS)
        con.execute('DELETE FROM {}'.format(MC.TABLE))
        con.executemany(
            'INSERT INTO {} ({}) VALUES ({})'.format(MC.TABLE, cols, ph),
            [[str(r.get(c, '') or '') for c in MC.DB_COLUMNS] for r in rows])
    finally:
        con.close()


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('--xlsx', help='caminho do MANUAIS.xlsx')
    ap.add_argument('--sheet', help='nome da aba (padrão: a primeira)')
    ap.add_argument('--dry-run', action='store_true', help='lê e relata, sem gravar')
    ap.add_argument('--schema-only', action='store_true',
                    help='só cria os dois bancos vazios, sem planilha')
    args = ap.parse_args()

    for cat in ('pending', 'ok'):
        MC.ensure_db(MC.db_path(cat))
        print('[db] %s' % MC.db_path(cat))
    if args.schema_only:
        print('\nBancos criados (%d colunas). Nada importado (--schema-only).'
              % len(MC.DB_COLUMNS))
        return 0

    path = args.xlsx
    if not path:
        path = next((p for p in DEFAULT_XLSX_CANDIDATES if os.path.isfile(p)), None)
    if not path or not os.path.isfile(path):
        print('\nMANUAIS.xlsx não encontrado. Procurei em:')
        for p in DEFAULT_XLSX_CANDIDATES:
            print('   ', p)
        print('\nOs bancos já estão criados e a página abre vazia. Rode de novo\n'
              'apontando o arquivo:  --xlsx "caminho/MANUAIS.xlsx"')
        return 1

    sheet, raw = read_xlsx(path, args.sheet)
    rows, mapping, unmatched = build_rows(raw)

    print('\n[planilha] %s · aba "%s" · %d linha(s) com Trade ID' % (path, sheet, len(rows)))
    print('[colunas ] %d casadas de %d no banco' % (len(mapping), len(MC.COLUMNS)))
    faltando = [c for c in MC.COLUMNS if c not in [x for _i, x in mapping]]
    if faltando:
        print('           sem correspondente na planilha (ficam vazias): %s'
              % ', '.join(faltando))
    for i, raw_h, why in unmatched:
        print('           ! coluna %d %r: %s' % (i + 1, raw_h, why))

    por_etapa = {}
    for r in rows:
        por_etapa[r['Pending']] = por_etapa.get(r['Pending'], 0) + 1
    print('[esteira ] ' + (', '.join('%s: %d' % kv for kv in sorted(por_etapa.items()))
                           or 'nenhuma linha'))

    pend = [r for r in rows if MC.target_category(r) == 'pending']
    ok = [r for r in rows if MC.target_category(r) == 'ok']
    print('[destino ] pending: %d · ok: %d' % (len(pend), len(ok)))

    if args.dry_run:
        print('\n--dry-run: nada gravado.')
        return 0

    rewrite('pending', pend)
    rewrite('ok', ok)
    print('\nImportado.')
    return 0


if __name__ == '__main__':
    sys.exit(main())
