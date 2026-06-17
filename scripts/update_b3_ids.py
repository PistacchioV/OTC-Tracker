"""
update_b3_ids.py
----------------
Reads mapping.xlsx and updates B3_ID (and TradeDate) in the NDF and
Option Commodities JSON cache files.

Excel layout
    Sheet 1 — NDF Commodities : col A = Deal ID | col B = B3 ID | col C = Trade Date
    Sheet 2 — Option Commodities : col A = Deal ID | col B = B3 ID | col C = Trade Date

Usage
    python scripts\\update_b3_ids.py
    python scripts\\update_b3_ids.py "C:\\path\\to\\mapping.xlsx"
"""

import json
import os
import sys
import tempfile
from datetime import datetime, date

try:
    import openpyxl
except ImportError:
    sys.exit("ERROR: openpyxl not installed — run: pip install openpyxl")

# ── paths ────────────────────────────────────────────────────────────────────

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT  = os.path.dirname(SCRIPT_DIR)

DEFAULT_XLSX = r"C:\Users\e930179\Downloads\mapping.xlsx"

NDF_CACHE = os.path.normpath(os.path.join(
    REPO_ROOT, "apps", "static", "data", "cache", "new deals", "NDF", "Commodities"
))
OPT_CACHE = os.path.normpath(os.path.join(
    REPO_ROOT, "apps", "static", "data", "cache", "new deals", "Option", "Commodities"
))


# ── helpers ──────────────────────────────────────────────────────────────────

def _normalize_deal(v):
    return str(v).strip() if v is not None else ""


def _normalize_date(v):
    """Return 'YYYY-MM-DD' string or empty string. Primary format: YYYYMMDD."""
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip().replace("-", "").replace("/", "")
    # Primary: YYYYMMDD (8 digits)
    if len(s) == 8 and s.isdigit():
        try:
            return datetime.strptime(s, "%Y%m%d").strftime("%Y-%m-%d")
        except ValueError:
            pass
    return ""


def _atomic_write_json(file_path, data):
    dir_name = os.path.dirname(file_path)
    fd, tmp_path = tempfile.mkstemp(dir=dir_name, suffix=".tmp")
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as fh:
            json.dump(data, fh, ensure_ascii=False, indent=2)
        try:
            os.replace(tmp_path, file_path)
            return
        except PermissionError:
            pass
        with open(file_path, "w", encoding="utf-8") as fh:
            json.dump(data, fh, ensure_ascii=False, indent=2)
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
    except Exception:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
        raise


def _read_sheet(ws):
    """Yield (deal_id, b3_id, trade_date) tuples from a worksheet, skipping header."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return
    # Skip header row (first row with non-numeric first cell)
    start = 0
    if rows and not str(rows[0][0] or "").strip().lstrip("-").isdigit():
        start = 1

    for row in rows[start:]:
        if not row or row[0] is None:
            continue
        deal_id    = _normalize_deal(row[0])
        b3_id      = _normalize_deal(row[1]) if len(row) > 1 else ""
        trade_date = _normalize_date(row[2]) if len(row) > 2 else ""
        if not deal_id:
            continue
        yield deal_id, b3_id, trade_date


# ── core update logic ────────────────────────────────────────────────────────

FIXED_UPDATES = {
    "Status":  "Success",
    "Maker":   "E930179",
    "Checker": "E964956",
}


def _update_cache(cache_dir, suffix, mapping, has_b3_field):
    """
    Walk cache_dir, find *suffix* files, update matching deals.

    mapping      : dict  deal_id (upper) → (b3_id, trade_date)
    has_b3_field : bool  True = B3_ID already in schema (NDF); False = must insert (OPT)

    For every matched deal also sets: Status, Maker, Checker (FIXED_UPDATES).
    """
    stats = {"files_touched": 0, "deals_updated": 0, "not_found": []}
    remaining = dict(mapping)  # track unmatched deal IDs

    if not os.path.isdir(cache_dir):
        print(f"  WARNING: cache directory not found: {cache_dir}")
        return stats

    for root, _, files in os.walk(cache_dir):
        for fname in sorted(files):
            if not fname.endswith(suffix):
                continue
            fp = os.path.join(root, fname)
            try:
                with open(fp, "r", encoding="utf-8") as fh:
                    deals = json.load(fh)
            except Exception as exc:
                print(f"  SKIP (parse error): {fp} — {exc}")
                continue

            changed = False
            for deal in deals:
                if not isinstance(deal, dict):
                    continue
                deal_id_raw = (deal.get("Deal") or "").strip()
                if not deal_id_raw:
                    continue
                key = deal_id_raw.upper()
                if key not in remaining:
                    continue

                b3_id, trade_date = remaining[key]

                # ── B3_ID ───────────────────────────────────────────────────
                if has_b3_field:
                    if b3_id and deal.get("B3_ID") != b3_id:
                        deal["B3_ID"] = b3_id
                        changed = True
                else:
                    # Insert B3_ID right after "Deal" key (preserve field order)
                    if b3_id and deal.get("B3_ID") != b3_id:
                        rebuilt, inserted = {}, False
                        for k, v in deal.items():
                            rebuilt[k] = v
                            if k == "Deal" and not inserted:
                                rebuilt["B3_ID"] = b3_id
                                inserted = True
                        if not inserted:
                            rebuilt["B3_ID"] = b3_id
                        deal.clear()
                        deal.update(rebuilt)
                        changed = True

                # ── TradeDate (only if currently empty) ─────────────────────
                if trade_date and not deal.get("TradeDate", "").strip():
                    deal["TradeDate"] = trade_date
                    changed = True

                # ── Fixed fields: Status, Maker, Checker ────────────────────
                for field, value in FIXED_UPDATES.items():
                    if deal.get(field) != value:
                        deal[field] = value
                        changed = True

                stats["deals_updated"] += 1
                del remaining[key]

            if changed:
                _atomic_write_json(fp, deals)
                stats["files_touched"] += 1
                print(f"  Updated: {os.path.relpath(fp, cache_dir)}")

    stats["not_found"] = list(remaining.keys())
    return stats


# ── main ─────────────────────────────────────────────────────────────────────

def main():
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX

    if not os.path.isfile(xlsx_path):
        sys.exit(f"ERROR: Excel file not found: {xlsx_path}")

    print(f"Reading: {xlsx_path}")
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as exc:
        sys.exit(f"ERROR opening workbook: {exc}")

    sheets = wb.sheetnames
    if len(sheets) < 2:
        sys.exit(f"ERROR: Expected at least 2 sheets, found {len(sheets)}: {sheets}")

    # --- NDF sheet ---
    ndf_ws = wb[sheets[0]]
    ndf_mapping = {d.upper(): (b, td) for d, b, td in _read_sheet(ndf_ws) if d}
    print(f"\nSheet 1 (NDF Commodities — '{sheets[0]}'): {len(ndf_mapping)} deal(s) to update")

    # --- OPT sheet ---
    opt_ws = wb[sheets[1]]
    opt_mapping = {d.upper(): (b, td) for d, b, td in _read_sheet(opt_ws) if d}
    print(f"Sheet 2 (Option Commodities — '{sheets[1]}'): {len(opt_mapping)} deal(s) to update")

    # --- Apply to NDF cache ---
    print(f"\n── NDF Commodities cache ({NDF_CACHE}) ──")
    ndf_stats = _update_cache(NDF_CACHE, "_ndfcomm.json", ndf_mapping, has_b3_field=True)

    # --- Apply to OPT cache ---
    print(f"\n── Option Commodities cache ({OPT_CACHE}) ──")
    opt_stats = _update_cache(OPT_CACHE, "_optcomm.json", opt_mapping, has_b3_field=False)

    # --- Summary ---
    print("\n══ Summary ══════════════════════════════")
    print(f"NDF:    {ndf_stats['deals_updated']} deal(s) updated across {ndf_stats['files_touched']} file(s)")
    print(f"Option: {opt_stats['deals_updated']} deal(s) updated across {opt_stats['files_touched']} file(s)")

    all_not_found = ndf_stats["not_found"] + opt_stats["not_found"]
    if all_not_found:
        print(f"\nWARNING — {len(all_not_found)} deal(s) not found in cache:")
        for d in all_not_found:
            print(f"  {d}")
    else:
        print("\nAll deals matched successfully.")


if __name__ == "__main__":
    main()
