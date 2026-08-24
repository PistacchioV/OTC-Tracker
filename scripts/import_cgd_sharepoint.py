#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""import_cgd_sharepoint.py — a lista de CGDs do SharePoint vira o DuckDB do app.

A mesa exporta a lista do SharePoint como `Sharepoint-CGD.xlsx` (o padrão é o
Downloads) e este script a carrega no banco que a tela **Onboarding › Tracking
Docs** lê: `cgd_sharepoint.db`, na MESMA pasta dos demais bancos
(`Config.DATABASE_DIR` — §307), nunca num caminho montado aqui.

    python scripts/import_cgd_sharepoint.py
    python scripts/import_cgd_sharepoint.py --xlsx ~/Downloads/Sharepoint-CGD.xlsx
    python scripts/import_cgd_sharepoint.py --sheet "Lista"
    python scripts/import_cgd_sharepoint.py --dry-run      # só relata
    python scripts/import_cgd_sharepoint.py --schema-only  # só cria o banco

É idempotente: o import REESCREVE a tabela a partir da planilha. Rodar duas
vezes dá o mesmo resultado, e a linha que alguém apagou do SharePoint some daqui
— um upsert deixaria no banco o que já não existe na fonte.

Três coisas que não são óbvias:

- **As colunas casam por NOME, não por posição.** O SharePoint reordena colunas
  quando alguém mexe na vista, e por posição a importação escreveria o CNPJ na
  coluna do SPN sem errar nada visível. A comparação é cega a caixa, acento,
  espaço e pontuação, então `Data Solicitação`, `DATA SOLICITACAO` e
  `Data  Solicitacao` são a mesma coluna.
- **O cabeçalho é procurado, não presumido na linha 1.** A exportação costuma
  vir com uma ou duas linhas de título antes; o cabeçalho é a primeira linha que
  casa com pelo menos um terço das colunas conhecidas.
- **O `Aging` da planilha é IGNORADO.** Ele é do dia da exportação e envelhece
  parado no banco; quem o calcula é o `cgd_docs.aging_of`, em dias úteis ANBIMA,
  a cada leitura da tela. A coluna continua sendo gravada como veio, para se
  poder comparar o que o SharePoint dizia com o que o app conta.
"""

import argparse
import os
import re
import sys
import unicodedata

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT = os.path.dirname(SCRIPT_DIR)
if REPO_ROOT not in sys.path:
    sys.path.insert(0, REPO_ROOT)

# Fora do Windows o `Config` exige o share absoluto (§8), e este script não
# encosta no share — o default deixa o import do config passar em qualquer
# máquina, sem mudar nada para quem já tem a variável definida.
os.environ.setdefault('OTC_SHARED_DRIVE_ROOT', os.path.join(REPO_ROOT, '.import-share'))

from apps.pages import cgd_docs as CGD                              # noqa: E402

DEFAULT_XLSX_CANDIDATES = [
    os.path.expanduser('~/Downloads/Sharepoint-CGD.xlsx'),
    os.path.expanduser('~/Desktop/Sharepoint-CGD.xlsx'),
    os.path.join(SCRIPT_DIR, 'Sharepoint-CGD.xlsx'),
]


def norm_header(s):
    """Cabeçalho comparável: sem acento, sem pontuação, sem caixa, sem espaço.

    `CSA?` e `CSA`, `OTC - STAMP` e `OTC STAMP` são a mesma coluna — a
    pontuação do nome é decoração de planilha, não identidade.
    """
    s = unicodedata.normalize('NFKD', str(s or ''))
    s = ''.join(c for c in s if not unicodedata.combining(c))
    return re.sub(r'[^a-z0-9]', '', s.lower())


ALVO = {norm_header(c): c for c in CGD.COLUMNS}


def achar_cabecalho(linhas):
    """`(índice, {coluna do arquivo → coluna do banco})` da linha de cabeçalho.

    Procurar em vez de presumir a primeira linha: a exportação do SharePoint vem
    com título e filtros em cima com frequência, e um cabeçalho lido uma linha
    acima importaria a planilha inteira deslocada.
    """
    melhor = (None, {}, 0)
    for i, linha in enumerate(linhas[:25]):
        mapa = {}
        for j, cel in enumerate(linha):
            alvo = ALVO.get(norm_header(cel))
            if alvo and alvo not in mapa.values():
                mapa[j] = alvo
        if len(mapa) > melhor[2]:
            melhor = (i, mapa, len(mapa))
    idx, mapa, n = melhor
    # Um terço das colunas: com menos que isso o que casou foi coincidência
    # (uma linha de dados com a palavra "Status" dentro, por exemplo).
    if n < max(3, len(CGD.COLUMNS) // 3):
        return None, {}
    return idx, mapa


def ler_xlsx(path, sheet=None):
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True, read_only=True)
    nome = sheet or wb.sheetnames[0]
    if nome not in wb.sheetnames:
        raise SystemExit('aba "%s" não existe. Abas: %s' % (nome, ', '.join(wb.sheetnames)))
    ws = wb[nome]
    linhas = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return nome, linhas


def montar_linhas(linhas):
    idx, mapa = achar_cabecalho(linhas)
    if idx is None:
        raise SystemExit('não achei o cabeçalho: nenhuma das 25 primeiras linhas '
                         'casa com as colunas conhecidas.\nColunas esperadas:\n  '
                         + '\n  '.join(CGD.COLUMNS))
    out = []
    for linha in linhas[idx + 1:]:
        if not any(str(c).strip() for c in linha if c is not None):
            continue                                    # linha em branco no fim
        rec = {c: '' for c in CGD.COLUMNS}
        for j, alvo in mapa.items():
            v = linha[j] if j < len(linha) else None
            if alvo in CGD.DATE_COLUMNS:
                # Normaliza aqui: a planilha traz `datetime`, texto e serial do
                # Excel na MESMA coluna, e o banco é todo VARCHAR.
                rec[alvo] = CGD.fmt_date(v)
            else:
                rec[alvo] = '' if v is None else str(v).strip()
        # Linha sem nada que identifique o documento não é linha — é resto de
        # formatação da planilha.
        if not any((rec.get(c) or '').strip() for c in ('CNPJ', 'Razão Social',
                                                        'Grupo Economico', 'Doc Type')):
            continue
        out.append(rec)
    return idx, mapa, out


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('--xlsx', help='caminho da planilha (padrão: ~/Downloads/Sharepoint-CGD.xlsx)')
    ap.add_argument('--sheet', help='aba (padrão: a primeira)')
    ap.add_argument('--dry-run', action='store_true', help='relata sem gravar')
    ap.add_argument('--schema-only', action='store_true', help='só cria o banco vazio')
    args = ap.parse_args()

    print('[banco   ] %s' % CGD.DB_PATH)

    if args.schema_only:
        CGD.ensure_db()
        print('[db] criado/atualizado com %d colunas. Nada importado (--schema-only).'
              % len(CGD.DB_COLUMNS))
        return 0

    path = args.xlsx
    if not path:
        path = next((p for p in DEFAULT_XLSX_CANDIDATES if os.path.isfile(p)), None)
    if not path or not os.path.isfile(path):
        print('planilha não encontrada. Procurei em:\n  ' +
              '\n  '.join(DEFAULT_XLSX_CANDIDATES) +
              '\nUse --xlsx <caminho>.')
        return 1

    aba, linhas = ler_xlsx(path, args.sheet)
    idx, mapa, rows = montar_linhas(linhas)

    print('[planilha] %s · aba "%s"' % (path, aba))
    print('[cabeçalho] linha %d · %d de %d colunas casadas'
          % (idx + 1, len(mapa), len(CGD.COLUMNS)))
    faltando = [c for c in CGD.COLUMNS if c not in mapa.values()]
    if faltando:
        # Aviso, não erro: a lista do SharePoint ganha e perde colunas, e a
        # coluna ausente entra vazia. Parar por causa dela deixaria a mesa sem
        # tela nenhuma por causa de um campo.
        print('[aviso   ] sem correspondência na planilha (entram vazias): %s'
              % ', '.join(faltando))
    print('[linhas  ] %d' % len(rows))

    if rows:
        print('\nPrimeira linha lida:')
        for c in CGD.COLUMNS[:8]:
            print('   %-22s %s' % (c, rows[0].get(c, '')))
        idade = CGD.aging_of(rows[0])
        print('   %-22s %s (dias úteis, calculado — o da planilha é ignorado)'
              % ('Aging', idade))

    if args.dry_run:
        print('\n--dry-run: nada foi gravado.')
        return 0

    n = CGD.replace_all(rows)
    print('\n[gravado ] %d linha(s) em %s' % (n, CGD.TABLE))
    return 0


if __name__ == '__main__':
    sys.exit(main())
