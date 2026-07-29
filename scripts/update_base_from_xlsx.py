#!/usr/bin/env python3
r"""Atualiza RefData.json e CounterpartyDetails.json a partir de "Atualizar Base.xlsx".

Planilha de origem (aba "Base"), uma linha por contraparte:

    A Counterparty   B SPN    C CASID   D UCN   E ECI
    F Client Role Status      G Contatos        H Banker   I Grupo Economico
    J Tax ID (só dígitos)     K Signature Type  L B3 Account

O que é atualizado, casando pelo SPN:
  * RefData.json          → BANKER (col. H), ECONOMIC GROUP (col. I),
                            TAX ID (col. J), SIGNATURE TYPE (col. K) e
                            B3 ACCOUNT (col. L)
                            — coluna vazia na planilha = não mexe no que já está na base
  * CounterpartyDetails   → CONTACTS a partir da col. G (e-mails separados por ';')

SPN sem match no RefData
------------------------
Vira um registro NOVO, com o que a planilha traz (COUNTERPARTY, SPN, CASID, UCN,
ECI, STATUS, BANKER, ECONOMIC GROUP, TAX ID, SIGNATURE TYPE, B3 ACCOUNT). Os
campos que a planilha não tem (os dois acrônimos, MAKER, CHECKER) e as colunas
vazias ficam em branco — o cadastro é completado na tela de Reference Data.

Regras aplicadas aos contatos
-----------------------------
A planilha traz só o endereço, sem dizer para que serve cada um, então todo
contato importado recebe TODAS as regras (_ALL_RULES abaixo) — foi o combinado.
Ajuste essa lista aqui se o conjunto mudar.

E-mails placeholder ('xxx', 'x-x', 'xx@xx.com') são descartados usando o MESMO
predicado de apps/pages/routes.py, para os dois não divergirem.

Merge, não substituição
-----------------------
Por padrão os contatos são mesclados por e-mail: os que já existem têm as regras
atualizadas (preservando nome e telefone já cadastrados), os novos são
acrescentados e os que não estão na planilha são mantidos. Use --replace para
trocar a lista inteira pelo que veio da planilha.

Uso
---
    python scripts/update_base_from_xlsx.py                    # dry-run
    python scripts/update_base_from_xlsx.py --apply
    python scripts/update_base_from_xlsx.py --xlsx "C:\...\Atualizar Base.xlsx"
    python scripts/update_base_from_xlsx.py --apply --replace

Grava um .bak com timestamp de cada JSON antes de qualquer alteração.
"""
import argparse
import json
import os
import re
import shutil
import sys
from datetime import datetime

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(ROOT, 'apps', 'static', 'data')
REFDATA_PATH = os.path.join(DATA, 'RefData.json')
CPD_PATH = os.path.join(DATA, 'CounterpartyDetails.json')
ROUTES_PATH = os.path.join(ROOT, 'apps', 'pages', 'routes.py')

DEFAULT_XLSX = r'C:\Users\e930179\Downloads\Atualizar Base.xlsx'
SHEET = 'Base'

# Colunas (1-based, como na planilha).
COL_COUNTERPARTY, COL_SPN, COL_CASID, COL_UCN, COL_ECI = 1, 2, 3, 4, 5
COL_STATUS, COL_CONTATOS, COL_BANKER, COL_GRUPO = 6, 7, 8, 9
COL_TAXID = 10       # J — Tax ID (só dígitos na planilha)
COL_SIGNATURE = 11   # K — Signature Type
COL_B3 = 12          # L — B3 Account

# Ordem/conjunto de chaves de um registro do RefData — registros criados por este
# script nascem com TODAS elas (a tela lê r['CAMPO'] direto e quebra se faltar).
REF_KEYS = ['STATUS', 'COUNTERPARTY', 'ECONOMIC GROUP', 'BANKER', 'SIGNATURE TYPE',
            'TAX ID', 'ECI', 'SPN', 'CASID', 'UCN', 'B3 ACCOUNT',
            'COMMODITIES ACCRONYM', 'FX CASH ACCRONYM', 'MAKER', 'CHECKER']

# Todas as regras que um contato importado recebe. Espelha o que a Reference
# Data exibe hoje; manter em sincronia com CP_RULES em reference-data.html.
_ALL_RULES = [
    'Contact Confirmation',
    'IOF',
    'NEGOTIATION LETTER',
    'Only for Confirmation',
    'Only for Repurchase',
    'Settlement',
    'Settlement Advice',
]


def norm_spn(v):
    """Dígitos sem zeros à esquerda — regra global do projeto (HANDOFF §21.0).
    '0009530701', '9530701' e 9530701.0 são o mesmo SPN."""
    if v is None:
        return ''
    s = str(v).strip()
    if s.endswith('.0'):
        s = s[:-2]
    d = ''.join(ch for ch in s if ch.isdigit())
    return d.lstrip('0')


def norm_b3(v):
    """Conta B3 na máscara da base ('73760.10-2'). O consumidor
    (_opb3_refdata_by_account em routes.py) casa a STRING exata, então uma conta
    que venha só em dígitos ('73760102') precisa entrar mascarada. Qualquer outro
    formato passa como veio — melhor manter o texto do que inventar máscara."""
    s = str(v or '').strip()
    d = ''.join(ch for ch in s if ch.isdigit())
    if len(d) == 8 and not any(ch in s for ch in '.-'):
        return '%s.%s-%s' % (d[:5], d[5:7], d[7])
    return s


def norm_taxid(v):
    """Col. J vem só com dígitos; a base guarda com máscara ('45.985.371/0001-08')
    — os 438 registros preenchidos hoje são todos CNPJ mascarado.

    Excel entrega a coluna como número quando ela não é texto, o que corta os
    zeros à esquerda ('01.234.567/…' vira 13 dígitos) e pode trazer o '.0' do
    float; por isso 12 ou 13 dígitos são completados com zeros à esquerda antes
    de mascarar. 11 dígitos = CPF. Qualquer outro tamanho passa como veio — sem
    palpite sobre onde vai cada separador."""
    s = str(v or '').strip()
    if s.endswith('.0'):
        s = s[:-2]
    d = ''.join(ch for ch in s if ch.isdigit())
    if 12 <= len(d) <= 13:
        d = d.zfill(14)
    if len(d) == 14:
        return '%s.%s.%s/%s-%s' % (d[:2], d[2:5], d[5:8], d[8:12], d[12:])
    if len(d) == 11:
        return '%s.%s.%s-%s' % (d[:3], d[3:6], d[6:9], d[9:])
    return s


def norm_status(v):
    """Col. F ('Client Role Status') → STATUS do RefData, que só conhece
    ACTIVE/PENDING. Rótulo diferente disso vira PENDING (registro novo entra
    para ser revisado na tela, não já valendo como ativo)."""
    s = str(v or '').strip().upper()
    return s if s in ('ACTIVE', 'PENDING') else 'PENDING'


def load_email_predicate():
    """Extrai _cc_email_is_usable de routes.py.

    Importar o módulo puxaria Flask, DuckDB e o app inteiro; o bloco do filtro é
    autocontido, então dar exec só nele mantém o script rodável em qualquer lugar
    e garante comportamento idêntico ao do import da tela."""
    src = open(ROUTES_PATH, encoding='utf-8').read()
    try:
        start = src.index('_CC_EMAIL_RE = re.compile')
        end = src.index('def _cc_parse_rules')
    except ValueError:
        sys.exit('Bloco do filtro de placeholder não encontrado em routes.py — '
                 'provavelmente foi renomeado. Corrija este script antes de rodar.')
    ns = {'re': re}
    exec(src[start:end], ns)
    return ns['_cc_email_is_usable']


def name_from_email(email):
    """'felipe.garcia@co.abb.com' -> 'Felipe Garcia'. Só um rótulo inicial para
    a linha não ficar sem nome; qualquer nome já cadastrado tem precedência."""
    local = email.split('@')[0]
    parts = [p for p in re.split(r'[._\-+]+', local) if p]
    if not parts or not any(p.isalpha() for p in parts):
        return ''
    return ' '.join(p.capitalize() for p in parts)


def read_sheet(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        sys.exit('Aba %r não encontrada em %s. Abas: %s' % (SHEET, path, wb.sheetnames))
    ws = wb[SHEET]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 1:
            continue                      # cabeçalho
        def cell(idx):
            return '' if idx > len(row) or row[idx - 1] is None else str(row[idx - 1]).strip()
        spn = norm_spn(cell(COL_SPN))
        if not spn:
            continue
        rows.append({
            'spn': spn,
            'counterparty': cell(COL_COUNTERPARTY),
            'contatos': cell(COL_CONTATOS),
            'banker': cell(COL_BANKER),
            'grupo': cell(COL_GRUPO),
            'signature': cell(COL_SIGNATURE),
            'b3': norm_b3(cell(COL_B3)),
            'taxid': norm_taxid(cell(COL_TAXID)),
            'casid': cell(COL_CASID),
            'ucn': cell(COL_UCN),
            'eci': cell(COL_ECI),
            'status': cell(COL_STATUS),
        })
    wb.close()
    return rows


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('--xlsx', default=DEFAULT_XLSX, help='planilha de origem')
    ap.add_argument('--apply', action='store_true', help='grava (padrão é dry-run)')
    ap.add_argument('--replace', action='store_true',
                    help='substitui a lista de contatos em vez de mesclar por e-mail')
    ap.add_argument('--refdata', default=REFDATA_PATH)
    ap.add_argument('--cpd', default=CPD_PATH)
    args = ap.parse_args()

    if not os.path.exists(args.xlsx):
        sys.exit('Planilha não encontrada: %s\n(use --xlsx para apontar o caminho)' % args.xlsx)

    is_usable = load_email_predicate()
    rows = read_sheet(args.xlsx)
    print('Linhas com SPN na aba %r: %d\n' % (SHEET, len(rows)))

    refdata = json.load(open(args.refdata, encoding='utf-8'))
    cpd = json.load(open(args.cpd, encoding='utf-8'))
    ref_by_spn = {}
    for r in refdata:
        ref_by_spn.setdefault(norm_spn(r.get('SPN')), r)
    cpd_by_spn = {}
    for r in cpd:
        cpd_by_spn.setdefault(norm_spn(r.get('SPN')), r)

    n_banker = n_grupo = n_signature = n_b3 = n_taxid = 0
    n_contacts_new = n_contacts_upd = 0
    dropped, created_ref, created_cpd = [], [], []

    for row in rows:
        spn = row['spn']

        # ── RefData: banker, grupo econômico, assinatura e conta B3 ──────
        ref = ref_by_spn.get(spn)
        if ref is None:
            # SPN novo: entra na base com o que a planilha tem; o resto em branco.
            ref = dict.fromkeys(REF_KEYS, '')
            ref.update({
                'STATUS': norm_status(row['status']),
                'COUNTERPARTY': row['counterparty'],
                'ECONOMIC GROUP': row['grupo'],
                'BANKER': row['banker'],
                'SIGNATURE TYPE': row['signature'],
                'TAX ID': row['taxid'],
                'ECI': row['eci'],
                'SPN': row['spn'],
                'CASID': row['casid'],
                'UCN': row['ucn'],
                'B3 ACCOUNT': row['b3'],
            })
            refdata.append(ref)
            ref_by_spn[spn] = ref
            created_ref.append('%s · %s' % (spn, row['counterparty']))
        else:
            if row['banker'] and str(ref.get('BANKER') or '').strip() != row['banker']:
                ref['BANKER'] = row['banker']
                n_banker += 1
            if row['grupo'] and str(ref.get('ECONOMIC GROUP') or '').strip() != row['grupo']:
                ref['ECONOMIC GROUP'] = row['grupo']
                n_grupo += 1
            if row['signature'] and str(ref.get('SIGNATURE TYPE') or '').strip() != row['signature']:
                ref['SIGNATURE TYPE'] = row['signature']
                n_signature += 1
            if row['b3'] and str(ref.get('B3 ACCOUNT') or '').strip() != row['b3']:
                ref['B3 ACCOUNT'] = row['b3']
                n_b3 += 1
            if row['taxid'] and str(ref.get('TAX ID') or '').strip() != row['taxid']:
                ref['TAX ID'] = row['taxid']
                n_taxid += 1

        # ── CounterpartyDetails: contatos ────────────────────────────────
        emails, seen = [], set()
        for raw in re.split(r'[;\n]+', row['contatos']):
            e = raw.strip()
            if not e:
                continue
            if not is_usable(e):
                dropped.append('%s · %s' % (spn, e))
                continue
            key = e.lower()
            if key not in seen:
                seen.add(key)
                emails.append(e)
        if not emails:
            continue

        rec = cpd_by_spn.get(spn)
        if rec is None:
            rec = {'SPN': str(row['spn']), 'COUNTERPARTY': row['counterparty'], 'CGD': [],
                   'BANKING': {'PAY': [], 'RECEIVE': []}, 'CONTACTS': []}
            cpd.append(rec)
            cpd_by_spn[spn] = rec
            created_cpd.append('%s · %s' % (spn, row['counterparty']))
        existing = rec.get('CONTACTS') or []

        if args.replace:
            rec['CONTACTS'] = [{'name': name_from_email(e), 'phone': '', 'email': e,
                                'rules': list(_ALL_RULES), 'status': 'Active'} for e in emails]
            n_contacts_new += len(emails)
            continue

        by_email = {}
        for c in existing:
            by_email.setdefault(str((c or {}).get('email', '') or '').strip().lower(), c)
        for e in emails:
            hit = by_email.get(e.lower())
            if hit is None:
                existing.append({'name': name_from_email(e), 'phone': '', 'email': e,
                                 'rules': list(_ALL_RULES), 'status': 'Active'})
                n_contacts_new += 1
            elif sorted(hit.get('rules') or []) != sorted(_ALL_RULES):
                hit['rules'] = list(_ALL_RULES)   # nome/telefone existentes preservados
                n_contacts_upd += 1
        rec['CONTACTS'] = existing

    # ── Relatório ────────────────────────────────────────────────────────
    print('RefData    : BANKER atualizado em %d · ECONOMIC GROUP em %d · SIGNATURE TYPE em %d'
          ' · B3 ACCOUNT em %d · TAX ID em %d'
          % (n_banker, n_grupo, n_signature, n_b3, n_taxid))
    print('RefData    : %d registros criados (SPN sem match)' % len(created_ref))
    print('Contatos   : %d novos · %d com regras atualizadas' % (n_contacts_new, n_contacts_upd))
    print('Placeholder: %d e-mails descartados' % len(dropped))
    if dropped:
        for d in dropped[:30]:
            print('     ', d)
        if len(dropped) > 30:
            print('      … e mais %d' % (len(dropped) - 30))
    if created_ref:
        print('\nRegistros criados no RefData: %d' % len(created_ref))
        for s in created_ref[:20]:
            print('     ', s)
        if len(created_ref) > 20:
            print('      … e mais %d' % (len(created_ref) - 20))
    if created_cpd:
        print('\nRegistros criados no CounterpartyDetails: %d' % len(created_cpd))
        for s in created_cpd[:20]:
            print('     ', s)

    if not args.apply:
        print('\nDry-run — nada foi gravado. Rode com --apply para aplicar.')
        return

    stamp = datetime.now().strftime('%Y%m%d-%H%M%S')
    for path, data in ((args.refdata, refdata), (args.cpd, cpd)):
        bak = '%s.%s.bak' % (path, stamp)
        shutil.copy2(path, bak)
        with open(path, 'w', encoding='utf-8') as fh:
            json.dump(data, fh, ensure_ascii=False, indent=2)
        print('Gravado %s (backup: %s)' % (os.path.basename(path), os.path.basename(bak)))


if __name__ == '__main__':
    main()
