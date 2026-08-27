#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""import_cgd_auxiliar.py — as abas do `Auxiliar.xlsx` viram os cadastros do
/mapping que a Recon de CGD lê.

O batimento de CGD usava três abas de uma planilha mantida à mão numa pasta de
rede; hoje elas são cadastros do /mapping, e este script traz o conteúdo da
planilha para os JSONs — a carga inicial, ou uma recarga quando alguém atualizar
o Auxiliar por fora.

    python scripts/import_cgd_auxiliar.py
    python scripts/import_cgd_auxiliar.py --xlsx ~/Downloads/Auxiliar.xlsx
    python scripts/import_cgd_auxiliar.py --dry-run      # só relata

As três abas e os cadastros em que cada uma desemboca:

| Aba                | Cadastro              | Colunas                            |
|--------------------|-----------------------|------------------------------------|
| `Mapping`          | `cgd-b3-participante` | Razão Social · Nome Simplificado · CNPJ · Conta |
| `Garantidores`     | `cgd-garantidor`      | CNPJ / CPF · Empresa · Cliente     |
| `Contas encerradas`| `cgd-conta-encerrada` | CNPJ / CPF · Nome                  |

É idempotente: cada aba REESCREVE o JSON correspondente. Rodar duas vezes dá o
mesmo resultado, e a linha apagada da planilha some do cadastro — um upsert
deixaria valendo o que já não existe na fonte.

Duas coisas que não são óbvias:

- **As colunas casam por NOME, não por posição** (cego a caixa, acento e
  espaço): a ordem das colunas da planilha muda quando alguém a reorganiza, e
  por posição o CNPJ entraria na coluna do nome sem errar nada visível. Quando
  o cabeçalho não casa, a POSIÇÃO da planilha original (documentada no mapa de
  cada aba) é a segunda tentativa — avisada, nunca silenciosa.
- **Os valores NÃO são trimados além das pontas** e a pontuação do CNPJ fica
  como está: quem compara é o motor da recon, por dígito, e "limpar" aqui
  criaria uma grafia que a planilha original não tem.
"""

import argparse
import json
import os
import sys
import unicodedata

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT = os.path.dirname(SCRIPT_DIR)
if REPO_ROOT not in sys.path:
    sys.path.insert(0, REPO_ROOT)

# Fora do Windows o `Config` exige o share absoluto (§8), e este script não
# encosta no share.
os.environ.setdefault('OTC_SHARED_DRIVE_ROOT', os.path.join(REPO_ROOT, '.import-share'))

from apps.pages.data_paths import mapping_write                     # noqa: E402

DEFAULT_XLSX_CANDIDATES = [
    os.path.expanduser('~/Downloads/Auxiliar.xlsx'),
    os.path.expanduser('~/Desktop/Auxiliar.xlsx'),
    os.path.join(SCRIPT_DIR, 'Auxiliar.xlsx'),
]

# Cada aba: o cadastro de destino e as colunas dele — para cada CHAVE do JSON,
# os nomes de cabeçalho aceitos (normalizados) e a COLUNA da planilha original
# (0-based), que é a segunda tentativa quando o cabeçalho não casa.
SHEETS = (
    {
        'sheet': 'Mapping',
        'key': 'cgd-b3-participante',
        'cols': (
            ('RAZAO SOCIAL',     ('participante razao social', 'razao social'), 0),
            ('NOME CONTRAPARTE', ('participante nome simplificado', 'nome simplificado',
                                  'nome contraparte'), 1),
            ('CNPJ',             ('participante cnpj', 'cnpj', 'cnpj do participante'), 2),
            ('CONTA',            ('participante conta', 'conta'), 3),
        ),
    },
    {
        'sheet': 'Garantidores',
        'key': 'cgd-garantidor',
        'cols': (
            ('CNPJ / CPF', ('cnpj cpf', 'cnpj / cpf', 'cpf cnpj'), 0),
            ('EMPRESA',    ('empresa', 'nome'), 1),
            ('CLIENTE',    ('cliente',), 2),
        ),
    },
    {
        'sheet': 'Contas encerradas',
        'key': 'cgd-conta-encerrada',
        'cols': (
            ('CNPJ / CPF', ('cnpj cpf', 'cnpj / cpf', 'cpf cnpj'), 0),
            ('NOME',       ('nome', 'empresa', 'razao social', 'cliente'), 1),
        ),
    },
)


def _norm(s):
    """Caixa, acento, espaço e pontuação fora — o casamento de cabeçalho é cego
    aos quatro."""
    s = unicodedata.normalize('NFKD', str(s or ''))
    s = ''.join(c for c in s if not unicodedata.combining(c))
    s = ''.join(c if c.isalnum() else ' ' for c in s)
    return ' '.join(s.lower().split())


def _txt(v):
    if v is None:
        return ''
    return str(v).strip()


def carrega_aba(wb, spec):
    """As linhas de uma aba como dicts do cadastro, mais os avisos."""
    avisos = []
    nomes = {_norm(n): n for n in wb.sheetnames}
    real = nomes.get(_norm(spec['sheet']))
    if not real:
        return None, ['aba "%s" não encontrada (existem: %s)'
                      % (spec['sheet'], ', '.join(wb.sheetnames))]
    ws = wb[real]
    linhas = [[_txt(c) for c in row] for row in ws.iter_rows(values_only=True)]
    linhas = [l for l in linhas if any(l)]
    if not linhas:
        return [], []

    # O cabeçalho é PROCURADO: a primeira linha em que ao menos uma coluna
    # conhecida casa por nome. Sem cabeçalho nenhum, vale a posição documentada
    # — com aviso, porque importar deslocado é a falha que não aparece.
    idx = {}
    corpo = linhas
    for i, l in enumerate(linhas[:5]):
        achou = {}
        celulas = {_norm(c): j for j, c in enumerate(l) if _txt(c)}
        for chave, aliases, _pos in spec['cols']:
            for a in aliases:
                if a in celulas:
                    achou[chave] = celulas[a]
                    break
        if achou:
            idx = achou
            corpo = linhas[i + 1:]
            break
    for chave, _aliases, pos in spec['cols']:
        if chave not in idx:
            idx[chave] = pos
            avisos.append('coluna "%s" não casou pelo cabeçalho; usando a '
                          'posição %s da planilha original' % (chave, chr(65 + pos)))

    out = []
    for l in corpo:
        r = {chave: (l[j] if j < len(l) else '') for chave, j in idx.items()}
        if any(r.values()):
            out.append(r)
    return out, avisos


def main():
    ap = argparse.ArgumentParser(description=__doc__.split('\n')[0])
    ap.add_argument('--xlsx', help='caminho do Auxiliar.xlsx (padrão: Downloads)')
    ap.add_argument('--dry-run', action='store_true',
                    help='só relata; não grava os cadastros')
    args = ap.parse_args()

    xlsx = args.xlsx
    if not xlsx:
        xlsx = next((p for p in DEFAULT_XLSX_CANDIDATES if os.path.isfile(p)), None)
    if not xlsx or not os.path.isfile(xlsx):
        print('ERRO: não achei o Auxiliar.xlsx. Passe o caminho com --xlsx.')
        print('Procurei em:')
        for p in DEFAULT_XLSX_CANDIDATES:
            print('  ' + p)
        return 1

    try:
        from openpyxl import load_workbook
    except ImportError:
        print('ERRO: openpyxl não está instalado (pip install openpyxl).')
        return 1

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    falhou = False
    for spec in SHEETS:
        rows, avisos = carrega_aba(wb, spec)
        for a in avisos:
            print('AVISO [%s]: %s' % (spec['sheet'], a))
        if rows is None:
            falhou = True
            continue
        destino = mapping_write(spec['key'])
        print('%s: %d linha(s) -> %s%s'
              % (spec['sheet'], len(rows), destino,
                 '  (dry-run, nada gravado)' if args.dry_run else ''))
        if args.dry_run:
            continue
        os.makedirs(os.path.dirname(destino), exist_ok=True)
        # Escrita atômica, como a do /mapping: meia lista em disco é pior que a
        # lista de ontem.
        tmp = destino + '.tmp'
        with open(tmp, 'w', encoding='utf-8') as fh:
            json.dump(rows, fh, ensure_ascii=False, indent=2)
        os.replace(tmp, destino)
    return 1 if falhou else 0


if __name__ == '__main__':
    sys.exit(main())
