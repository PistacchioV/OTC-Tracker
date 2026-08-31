from gettext import install
import os
import io
import re
import random
import secrets
import string
import smtplib
import json
import threading
import traceback
import unicodedata
import uuid
import shutil
import base64
import logging
import time
from collections import Counter
from contextlib import contextmanager
from datetime import datetime, timedelta, timezone
from urllib.parse import quote
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

import awmpy
from flask import (
    render_template, request, redirect, send_from_directory,
    url_for, session, flash, jsonify, make_response, has_app_context
)
from jinja2 import TemplateNotFound
from werkzeug.exceptions import NotFound

# Caminhos de infraestrutura (banco de usuários e raiz do share) saem do
# `Config`, que os resolve para ABSOLUTOS e recusa um valor relativo. Era aqui
# que eles nasciam: o `I:\...` cru é um caminho relativo em qualquer coisa que
# não seja Windows, e o `os.makedirs` do dia a dia criava a árvore inteira
# dentro do diretório de trabalho — foi assim que apareceram as pastas
# `I:\Confirmation\...` na raiz do repositório.
from apps.config import Config
from apps.pages.data_paths import (
    data_dir, data_path, data_write,
    PACKAGED_DIR as PACKAGED_DATA_DIR,
)
# Cache de leitura por request + TTL curto entre requests para os JSONs do
# dia (Operations B3, OTM Settlements...) — ver apps/pages/request_cache.py.
from apps.pages.request_cache import (
    req_cached as _req_cached, once_per_request as _once_per_request,
    bump_cache_gen as _bump_cache_gen,
)
from apps.pages import blueprint
# Porte Python do parser de booking recap (o mesmo que otc-fileupload.js faz no
# navegador) — usado pela varredura agendada do box. Sem dependência externa.
from apps.pages import otc_boxparse
# Locks e transações dos bancos de arquivo (DuckDB/SQLite avulsos). Import no
# TOPO, e não preguiçoso: as exceções são capturadas em `except` de rotas, e um
# nome resolvido só na primeira chamada faria o `except` referenciar algo que
# ainda não existe.
# As exceções de transação (DatabaseLockTimeout etc.) foram com a vertical da
# Recon de Comitentes — o entrypoint dela as importa da mesma fonte.
from apps.pages.database_access import (
    duckdb_read_unlocked,
    duckdb_read,
    duckdb_write,
)
# Os tipos de confirmação são UMA lista só, definida no módulo da esteira: ela
# alimenta o Confirmation Type do upload do Electronic Inventory, o cadastro
# Produto × LOB de `manual-conf-validation` e o dropdown de Produto do Track
# Confirmations. Import no topo (e não preguiçoso como os demais usos de
# `manual_conf`) porque `_MAPPING_DEFS` precisa dela em tempo de importação.
from apps.pages import manual_conf as _mc_mod
# O banco da lista de CGDs do SharePoint (Onboarding).
from apps.pages import cgd_docs as _cgd_mod
# O batimento de CGD (FEP × B3) — tradução do workflow Alteryx.
_CONFIRMATION_TYPES = _mc_mod.CONFIRMATION_TYPES


# ==============================================================================
# APPLICATION CONTEXT PARA O QUE RODA FORA DE UM REQUEST
# ==============================================================================
# Os schedulers vivem em threads próprias, e lá NÃO existe application context —
# `render_template` (o corpo dos e-mails) exige um e estoura com "Working outside
# of application context". O sintoma é traiçoeiro: o botão Run do Control Panel
# funciona, porque aquele roda dentro de um request, e só o envio automático
# falha. Foi assim que o aviso das 19:00 do Deals Monitor morreu em silêncio.
#
# O app é capturado no REGISTRO do blueprint (`record_once`), que é o único
# momento em que ele existe e este módulo é alcançável — a fábrica `create_app`
# importa as rotas, então guardar a referência ao contrário seria circular.
_FLASK_APP = None


@blueprint.record_once
def _capture_flask_app(state):
    global _FLASK_APP
    _FLASK_APP = state.app
    # O banco de notificações é criado (e migrado do antigo) AQUI, uma vez,
    # antes de existir tráfego. Antes ele nascia na primeira chamada de
    # `get_notif_connection`, e num processo recém-subido essa chamada é quase
    # sempre o poll do sino: a migração — 9,4 segundos de lock exclusivo no
    # share — acontecia dentro do request mais frequente e mais barato do app.
    #
    # A falha aqui NÃO derruba a subida. Sem o banco de notificações o sino
    # fica vazio e todo o resto funciona; recusar subir por causa do sino
    # trocaria um aviso que não aparece por um app que não abre.
    try:
        _ensure_notif_db()
    except Exception:                                       # noqa: BLE001
        log.error('[notif-db] a subida não conseguiu preparar %s — o sino fica '
                  'vazio até a próxima gravação conseguir:\n%s',
                  NOTIF_DB_PATH, traceback.format_exc())
    _start_schedulers()


@contextmanager
def _app_context():
    """Garante application context. No-op dentro de um request (e quando o app
    ainda não foi registrado, para o erro continuar aparecendo em vez de virar
    um envio silenciosamente vazio)."""
    if has_app_context() or _FLASK_APP is None:
        yield
    else:
        with _FLASK_APP.app_context():
            yield


# ── Os schedulers sobem com o APP, nunca com o import ────────────────────────
# Os dez laços agendados (as duas APIs da Athena, a varredura do box e os sete
# e-mails do Control Panel) eram disparados no corpo do módulo, então bastava
# `from apps.pages import routes` para as threads começarem a rodar. Quem faz
# isso não é só a fábrica: os 67 scripts de `scripts/tests/` importam este
# módulo, e o `backfill`, e qualquer `python -c` de depuração.
#
# O efeito não é ruído. O laço do Deals Monitor faz um catch-up na primeira
# volta — os slots de HOJE que já passaram e ninguém reivindicou —, então rodar
# um teste depois das 19h fazia o processo do TESTE mandar o e-mail de
# pendências de verdade. E, com o claim agora gravado em disco e enxergado por
# todos os processos (`_claim_daily_slot`), ele ainda RESERVAVA o slot: o app
# real chegava ao horário, via o dia já reivindicado e não mandava nada. Um
# e-mail que some sem erro nenhum, por causa de um teste.
#
# Agora cada laço se REGISTRA aqui e quem os sobe é o `record_once` do
# blueprint, que é o registro do app — o mesmo momento em que o `_FLASK_APP` é
# capturado. Importar o módulo deixou de ter efeito; `create_app` continua
# tendo o mesmo. As funções `_x_start_scheduler` seguem idempotentes (cada uma
# tem o seu flag), então um segundo app no mesmo processo não duplica thread.
_SCHEDULERS = []


def _schedule_on_start(label, start):
    """Registra um laço para subir junto com o app. NÃO o inicia."""
    _SCHEDULERS.append((label, start))


def _start_schedulers():
    """Sobe os laços registrados. Um que falhe não pode levar os outros.

    `OTC_DISABLE_SCHEDULERS=1` é o kill-switch dos testes: os scripts que sobem
    o app várias vezes (o guard do config purga e reimporta `apps.*`) corriam
    com os laços vivos ao lado — e um catch-up de 16h/17h/19h30 num processo de
    TESTE tenta reivindicar o slot REAL do dia e mandar o e-mail de verdade.
    """
    if os.getenv('OTC_DISABLE_SCHEDULERS', '').strip() == '1':
        log.info('[schedulers] desligados por OTC_DISABLE_SCHEDULERS=1')
        return
    for label, start in _SCHEDULERS:
        try:
            start()
        except Exception:                                   # noqa: BLE001
            log.warning('[%s] scheduler não iniciou:\n%s', label,
                        traceback.format_exc())

# ==============================================================================
# LOGGING CONFIG
# ==============================================================================

logging.basicConfig(
    level=logging.DEBUG,
    format='[%(asctime)s] %(levelname)s [%(funcName)s:%(lineno)d] %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
log = logging.getLogger('otc_tracker')


# ==============================================================================
# SESSION EXPIRY — server-side check (independente do browser restaurar cookies)
# ==============================================================================

@blueprint.before_request
def enforce_session_expiry():
    if not session.get('authenticated'):
        return
    expires_at = session.get('session_expires_at')
    if not expires_at:
        session.clear()
        return
    try:
        expiry = datetime.fromisoformat(expires_at)
        now = datetime.now(tz=timezone.utc)
        # Garante comparação timezone-aware
        if expiry.tzinfo is None:
            expiry = expiry.replace(tzinfo=timezone.utc)
        if now > expiry:
            session.clear()
            log.info("[enforce_session_expiry] Session expired — cleared")
    except (ValueError, TypeError):
        session.clear()


# Endpoints reachable while the screen is locked. Everything else is bounced to
# the lock screen so the user cannot navigate (or "go back") into the app
# without re-entering their SID and passing IP verification / 2FA.
_LOCK_ALLOWED_ENDPOINTS = {
    'static',
    'pages_blueprint.lock_screen_page',   # the lock screen itself
    'pages_blueprint.unlock',             # SID submit to unlock
    'pages_blueprint.two_factor_page',    # 2FA page (IP-mismatch unlock path)
    'pages_blueprint.verify_2fa',         # 2FA code submit
    'pages_blueprint.resend_code',        # resend 2FA code
    'pages_blueprint.sign_in_page',       # "Not you? Sign in"
    'pages_blueprint.login',              # sign in as a different user
    'pages_blueprint.logout',             # allow logging out while locked
    # O /dev-login do DEV BYPASS. O bloco que o define NÃO vai para o
    # repositório, então em produção este endpoint simplesmente não existe e a
    # entrada é inerte — `request.endpoint` nunca vai valer isto. Ela está aqui
    # porque, sem ela, a tela travada devolvia o /dev-login para o próprio lock
    # ANTES de ele rodar o `session.clear()` que desbloquearia: fora da rede JPM
    # não há phonebook para o Unlock consultar, e o desenvolvimento ficava preso
    # num laço sem saída visível.
    'pages_blueprint.dev_login',
}


@blueprint.before_request
def enforce_screen_lock():
    """While session['locked'] is set, only the unlock flow is reachable."""
    if not session.get('locked') or not session.get('authenticated'):
        return
    if request.endpoint in _LOCK_ALLOWED_ENDPOINTS:
        return
    return redirect(url_for('pages_blueprint.lock_screen_page'))


@blueprint.after_request
def add_no_store_on_authed_pages(response):
    """Prevent the browser back/forward cache from showing protected HTML after
    the screen is locked or the session ends. Limited to HTML so static assets
    keep caching normally."""
    try:
        if session.get('authenticated') and response.mimetype == 'text/html':
            response.headers['Cache-Control'] = 'no-store, max-age=0'
            response.headers['Pragma'] = 'no-cache'
    except Exception:
        pass
    return response


# Content-Security-Policy in REPORT-ONLY mode: the browser blocks nothing, it just
# posts a report to /csp-report for anything this policy would forbid. The app
# still relies on inline scripts/handlers, so 'unsafe-inline' stays for now; the
# external hosts are the CDNs/fonts/embeds the pages actually load. Watch the
# csp-report logs, tighten the allowlist, then flip the header name to
# 'Content-Security-Policy' to start enforcing.
_CSP_REPORT_ONLY = (
    "default-src 'self'; "
    "base-uri 'self'; "
    "object-src 'none'; "
    "frame-ancestors 'self'; "
    "img-src 'self' data:; "
    "font-src 'self' data: https://fonts.gstatic.com https://cdn.jsdelivr.net; "
    "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com https://cdn.jsdelivr.net; "
    "script-src 'self' 'unsafe-inline' "
    "https://cdn.jsdelivr.net https://cdnjs.cloudflare.com https://apexcharts.com; "
    "connect-src 'self'; "
    "frame-src 'self' https://www.youtube.com https://player.vimeo.com; "
    "report-uri /csp-report"
)


@blueprint.after_request
def add_security_headers(response):
    """Baseline security headers on every response. CSP ships in report-only mode
    (see _CSP_REPORT_ONLY) so nothing breaks while violations are collected."""
    try:
        response.headers.setdefault('X-Content-Type-Options', 'nosniff')
        response.headers.setdefault('X-Frame-Options', 'SAMEORIGIN')
        response.headers.setdefault('Referrer-Policy', 'strict-origin-when-cross-origin')
        response.headers.setdefault('Content-Security-Policy-Report-Only', _CSP_REPORT_ONLY)
        # HSTS is only meaningful over HTTPS; request.is_secure reflects the
        # proxy's X-Forwarded-Proto via ProxyFix.
        if request.is_secure:
            response.headers.setdefault(
                'Strict-Transport-Security', 'max-age=31536000; includeSubDomains')
    except Exception:
        pass
    return response


@blueprint.route('/csp-report', methods=['POST'])
def csp_report():
    """Collector for Content-Security-Policy-Report-Only violations. Report-only
    means nothing is blocked yet — these logs show what an enforcing policy would
    break, so the allowlist can be tuned before the switch. Unauthenticated by
    design: browsers post these reports without credentials."""
    try:
        raw = request.get_data(as_text=True) or ''
        if raw:
            log.warning('[csp-report] %s', raw[:2000])
    except Exception:
        pass
    return ('', 204)


# ==============================================================================
# PER-USER PAGE ACCESS
# An admin grants each user a subset of the sidebar pages. Users with no
# configuration keep full access (backward compatible); admins always have full
# access. Enforced here (before_request), in the sidebar (JS), and in the
# notification feed (only pages the user can reach).
# ==============================================================================

# Pages everyone can always reach regardless of configuration. The dashboards are
# NOT here — they are grantable like any other page (Main › Dashboards). '/users-profile'
# stays open so a fully-restricted user always has a safe landing (no redirect loop).
# `_ALWAYS_ALLOWED_PATHS` mora em `platform/authz.py`; alias logo abaixo,
# junto dos demais (depois do import da platform).

# As notificações moram em `apps/pages/platform/notifications.py` (segunda
# fatia da fase platform/ — CLAUDE.md §10). Os nomes ficam aqui como ALIAS:
# as features alcançam por `routes.<nome>` e os 22 testes que trocam
# `R._create_notification` por espião seguem interceptando todo mundo. O
# ESTADO da subida (`_notif_db_done`, `_notif_db_retry_at`) mora LÁ.
from apps.pages.platform import notifications as _pf_notif  # noqa: E402

_NOTIF_DS_OTHERPUB = _pf_notif._NOTIF_DS_OTHERPUB

_NOTIF_PAGE_URL = _pf_notif._NOTIF_PAGE_URL
_notif_page_url = _pf_notif._notif_page_url

# A autorização por página/card mora em `platform/authz.py` — aliases; os
# dois `before_request` que a APLICAM ficam aqui (registro em blueprint é
# casca). O estado (`_page_access_cache`) mora lá e é mutado in place.
from apps.pages.platform import authz as _pf_authz  # noqa: E402

_ALWAYS_ALLOWED_PATHS = _pf_authz._ALWAYS_ALLOWED_PATHS
_load_nav_urls = _pf_authz._load_nav_urls
_NAV_URLS = _pf_authz._NAV_URLS
_CONTROL_PANEL_CARDS = _pf_authz._CONTROL_PANEL_CARDS
_CP_CARD_TOKENS = _pf_authz._CP_CARD_TOKENS
_CP_ENDPOINT_CARD = _pf_authz._CP_ENDPOINT_CARD
_cp_page_allowed = _pf_authz._cp_page_allowed
_cp_card_allowed = _pf_authz._cp_card_allowed
_page_access_forget = _pf_authz._page_access_forget
_get_page_access = _pf_authz._get_page_access
_read_page_access = _pf_authz._read_page_access
_get_user_authz = _pf_authz._get_user_authz
_read_user_authz = _pf_authz._read_user_authz
_get_user_role = _pf_authz._get_user_role
_set_page_access = _pf_authz._set_page_access
_MASTER_SIDS = _pf_authz._MASTER_SIDS
_session_is_master = _pf_authz._session_is_master
_session_is_admin = _pf_authz._session_is_admin
_safe_landing = _pf_authz._safe_landing
_user_can_access_page = _pf_authz._user_can_access_page


@blueprint.before_request
def refresh_session_role():
    """O papel do cadastro alcança quem já está logado (ver `platform/authz.py`).

    É um `before_request` PRÓPRIO, e não uma carona no `enforce_page_access`:
    aquele desiste cedo para master, `/api/*`, `/static*` e todo path fora do
    menu — e é exatamente em `/api/*` que a mesa valida a confirmação. Preso
    ali, o papel só se atualizaria ao navegar por uma página do menu.
    """
    _pf_authz.refresh_session_role()


@blueprint.before_request
def enforce_page_access():
    """Block direct navigation to a page the user isn't allowed to see. Master and
    unconfigured users pass; anyone with a configured allowlist (including admins
    restricted by the master) is enforced. API and static requests are never
    blocked here (they keep their own auth)."""
    if not session.get('authenticated') or _session_is_master():
        return
    path = request.path or ''
    if path.startswith('/api/') or path.startswith('/static') or path not in _NAV_URLS:
        return
    configured, allowed = _get_page_access(session.get('user_sid', ''))
    if not configured:
        return
    # Control Panel is card-gated: the page opens if at least one card is granted.
    if path == '/control-panel':
        if not _cp_page_allowed(allowed):
            return redirect(_safe_landing(allowed))
        return
    if path not in allowed:
        return redirect(_safe_landing(allowed))


@blueprint.before_request
def enforce_control_panel_cards():
    """Block a Control Panel routine's API call when the user isn't granted that
    specific card. Master and unconfigured users pass."""
    if not session.get('authenticated') or _session_is_master():
        return
    card = _CP_ENDPOINT_CARD.get(request.path or '')
    if not card:
        return
    configured, allowed = _get_page_access(session.get('user_sid', ''))
    if configured and not _cp_card_allowed(allowed, card):
        return jsonify({'success': False, 'message': 'Access denied for this routine.'}), 403


# ==============================================================================
# CONFIGURAÇÕES
# ==============================================================================

DB_PATH = Config.DATABASE_PATH
CACHE_BASE_DIR = os.path.normpath(os.path.join(
    data_dir(), "cache", "new deals", "Option", "Commodities"
))
# FXO has its own cache dir so the dashboard labels it "Option FXO" (not Commodities)
OPT_FXO_CACHE_DIR = os.path.normpath(os.path.join(
    data_dir(), "cache", "new deals", "Option", "FXO"
))
# A infraestrutura de e-mail mora em `platform/mail.py` — aliases.
from apps.pages.platform import mail as _pf_mail  # noqa: E402

SHARED_MAILBOX = _pf_mail.SHARED_MAILBOX

# A porta em que a instância roda. UMA constante porque ela aparece em três
# lugares que se leem de fora do código — o endereço dos botões de e-mail
# (`_otc_app_url`), o link do e-mail de versão nova e o `run.py` — e as três
# diziam 8050 enquanto a instância subia na 8051. Um botão de e-mail com a porta
# errada não dá erro: leva a pessoa a uma página que não abre.
#
# Os `.bat` não conseguem ler daqui (são do cmd), então a porta deles é conferida
# pelo `check_bat_files.py` contra esta constante.
APP_PORT = int(os.getenv('OTC_TRACKER_PORT', '8051'))
RETURN_PATH = os.getenv('RETURN_PATH', os.path.join(
    Config.SHARED_DRIVE_ROOT, 'Confirmation', 'Derivativos', 'OTC Tracker', 'Batch Conecta', 'Return'))
CONECTA_NEW_PATH = os.getenv('CONECTA_NEW_PATH', os.path.join(
    Config.SHARED_DRIVE_ROOT, 'Confirmation', 'Derivativos', 'OTC Tracker', 'Batch Conecta', 'New'))
# Electronic Inventory: one folder per counterparty with Confirmations /
# Transactional / SSI subfolders. Created here on Reference Data checker approval
# and in bulk by scripts/create_counterparty_folders.py (kept in sync).
ELECTRONIC_INVENTORY_ROOT = os.getenv(
    'ELECTRONIC_INVENTORY_ROOT',
    os.path.join(Config.SHARED_DRIVE_ROOT, 'Confirmation', 'Derivativos', 'OTC Tracker',
                 'Electronic Inventory'))
# ── Electronic Inventory: movido para platform/electronic_inventory.py (§317) ─
# Os nomes ficam como ALIAS; o ELECTRONIC_INVENTORY_ROOT fica AQUI porque e a
# superficie de patch do check_ei_api (a platform o le por routes.<nome>).
from apps.pages.platform import electronic_inventory as _pf_ei  # noqa: E402
EI_SUBFOLDERS = _pf_ei.EI_SUBFOLDERS
_EI_ILLEGAL = _pf_ei._EI_ILLEGAL
_ei_sanitize = _pf_ei._ei_sanitize
_ei_match_key = _pf_ei._ei_match_key
_ei_actual_dir_name = _pf_ei._ei_actual_dir_name
_ei_client_dir_names = _pf_ei._ei_client_dir_names
_EI_TRANSACTIONAL_TYPES = _pf_ei._EI_TRANSACTIONAL_TYPES
_EI_CONFIRMATION_TYPES = _pf_ei._EI_CONFIRMATION_TYPES
_EI_PREVIEWABLE = _pf_ei._EI_PREVIEWABLE
_EI_ALLOWED_UPLOAD = _pf_ei._EI_ALLOWED_UPLOAD
_ei_refdata_clients = _pf_ei._ei_refdata_clients
_ei_resolve_client_dir = _pf_ei._ei_resolve_client_dir
_EI_MONTH_NAMES = _pf_ei._EI_MONTH_NAMES
_EI_MONTH_DIR_RE = _pf_ei._EI_MONTH_DIR_RE
_ei_month_folder = _pf_ei._ei_month_folder
_ei_ordinal = _pf_ei._ei_ordinal
_ei_version_prefix = _pf_ei._ei_version_prefix
_ei_next_ordinal = _pf_ei._ei_next_ordinal
_ei_human_size = _pf_ei._ei_human_size
_ei_iter_files = _pf_ei._ei_iter_files
_EI_ROOT_CACHE = _pf_ei._EI_ROOT_CACHE
_EI_ROOT_CACHE_TTL = _pf_ei._EI_ROOT_CACHE_TTL
_EI_ROOT_CACHE_LOCK = _pf_ei._EI_ROOT_CACHE_LOCK
_ei_scan_root_worker = _pf_ei._ei_scan_root_worker
_ei_scan_root = _pf_ei._ei_scan_root
_ei_long_path = _pf_ei._ei_long_path
_ei_locate_file = _pf_ei._ei_locate_file


def _ensure_counterparty_folders(company):
    """Create ELECTRONIC_INVENTORY_ROOT\\<company>\\{Confirmations,Transactional,SSI}
    if missing. Tolerant existence match (case/whitespace/illegal-char insensitive)
    so a folder created earlier under a slightly different name is reused, not
    duplicated. Best-effort: never raises (the share may be offline in dev)."""
    folder = _ei_sanitize(company)
    if not folder:
        return
    try:
        parent = os.path.join(ELECTRONIC_INVENTORY_ROOT, _ei_actual_dir_name(folder))
        for sub in EI_SUBFOLDERS:
            os.makedirs(os.path.join(parent, sub), exist_ok=True)
    except Exception as exc:
        log.warning('Electronic Inventory folder creation failed for %r: %s', company, exc)
# O armazém JSON (escrita atômica, _cache_lock, claims diários, arquivos-dia)
# mora em `platform/json_cache.py` — aliases. Os objetos de estado (o RLock, o
# memo do daycache) são mutados in place e nunca rebindados, então o alias
# continua vivo.
from apps.pages.platform import json_cache as _pf_jcache  # noqa: E402

_cache_lock = _pf_jcache._cache_lock
_claim_daily_slot = _pf_jcache._claim_daily_slot
_release_daily_slot = _pf_jcache._release_daily_slot
_atomic_write_json = _pf_jcache._atomic_write_json
_unique_filepath = _pf_jcache._unique_filepath


SMTP_HOST = _pf_mail.SMTP_HOST
SMTP_PORT = _pf_mail.SMTP_PORT
CODE_EXPIRY_MINUTES = 10

# 2FA hardening. A 6-digit code has a 10^6 space; without a cap on wrong tries it
# is brute-forceable inside the 10-minute window. MAX_2FA_ATTEMPTS burns a code
# after that many wrong guesses. The send limits stop an attacker who knows a SID
# from bombing the victim's inbox (or minting endless fresh codes to keep guessing).
MAX_2FA_ATTEMPTS = 5              # wrong guesses per code before it is invalidated
CODE_RESEND_COOLDOWN_SECONDS = 30  # minimum gap between two code emails for a SID
CODE_WINDOW_MINUTES = 15         # rolling window for the code-issue cap
MAX_CODES_PER_WINDOW = 5         # max codes emailed to a SID within that window

ROLE_META = {
    'MASTER':       {'display': 'Master',        'icon': 'ti-crown',                'description': 'Top-level authority — manages page access for everyone, including admins.', 'responsibilities': ['Control All Access', 'Manage Admins', 'Manage Users', 'Configure System']},
    'ADMIN':        {'display': 'Admin',         'icon': 'ti-shield-lock',          'description': 'Full platform administration and user management.',         'responsibilities': ['Manage Users', 'Configure System', 'View All Data', 'Assign Roles']},
    'BO':           {'display': 'Back Office',   'icon': 'ti-briefcase',            'description': 'Back office operations and settlement processing.',         'responsibilities': ['Settlement Processing', 'Position Reconciliation', 'Trade Confirmation', 'Reporting']},
    'MO':           {'display': 'Middle Office', 'icon': 'ti-calculator',           'description': 'Risk management and trade operations oversight.',           'responsibilities': ['Risk Monitoring', 'P&L Attribution', 'Trade Validation', 'Limit Monitoring']},
    'FO':           {'display': 'Front Office',  'icon': 'ti-chart-arrows-vertical','description': 'Trading and client-facing OTC operations.',                'responsibilities': ['OTC Trading', 'Client Management', 'Trade Execution', 'Market Analysis']},
    'INSTITUTIONAL':{'display': 'Institutional', 'icon': 'ti-building-bank',        'description': 'Institutional client operations and portfolio management.', 'responsibilities': ['Portfolio Management', 'Client Reporting', 'Compliance Review', 'Investment Analysis']},
    'HUB':          {'display': 'Hub',           'icon': 'ti-topology-star-3',      'description': 'Hub operations coordinating cross-desk activity.',          'responsibilities': ['Cross-Desk Coordination', 'Deal Routing', 'Workflow Management', 'Escalation Handling']},
}


# ==============================================================================
# FUNÇÕES AUXILIARES — BANCO DE DADOS (DuckDB)
# ==============================================================================
# Quem abre os bancos de arquivo é o `database_access`: contexto (`with`) com
# lock no ARQUIVO, que vale ENTRE PROCESSOS — o `threading.Lock` de módulo que
# ficava aqui só protegia um. A conexão singleton, o lock de thread, o retry
# manual e a reabertura com quarentena de WAL saíram junto: retry e backoff são
# do contexto agora.

# Lazy one-time schema init (see _ensure_db_initialized). Deferred so the
# Werkzeug auto-reloader's supervisor process never opens the single-writer
# DuckDB file — only the worker that actually serves requests does.
_db_init_done = False
_db_init_lock = threading.RLock()     # re-entrant: init_db() re-enters via get_db_connection()
_db_init_tls  = threading.local()     # per-thread "currently initializing" flag


# O handle e a abertura do banco de usuários moram em `platform/db.py` —
# aliases. O caminho (`DB_PATH`), as primitivas e o `_ensure_db_initialized`
# ficam AQUI e a platform os alcança por busca atrasada: são a superfície
# que os testes trocam.
from apps.pages.platform import db as _pf_db  # noqa: E402

_DuckDBHandle = _pf_db._DuckDBHandle
get_db_connection = _pf_db.get_db_connection


# ── O banco das NOTIFICAÇÕES, separado do de usuários ───────────────────────
# O lock da camada de acesso é por ARQUIVO. Com `notifications` e
# `push_subscriptions` morando no mesmo DuckDB de `users`, cada gravação de
# notificação — e elas acontecem a cada ação de qualquer pessoa da mesa —
# segurava o arquivo inteiro em modo EXCLUSIVO, e com ele o login, a allowlist
# do `Page_Access` e a gestão de usuários. Some a isso o sino, que consulta por
# aba aberta: o banco vivia travado, e o que travava não era o dado que
# importa, era o aviso.
#
# Separados, os dois tráfegos deixam de se ver. A gravação de notificação não
# encosta em quem está entrando no app, e a consulta do sino (que é a mais
# frequente do sistema) disputa apenas com outras notificações.
NOTIF_DB_PATH = Config.NOTIFICATIONS_DATABASE_PATH


# O banco de notificações mora em `platform/notifications.py` — aliases; o
# ESTADO da subida (`_notif_db_done`, `_notif_db_retry_at`, o lock) mora lá.
get_notif_connection = _pf_notif.get_notif_connection
_notif_init_schema = _pf_notif._notif_init_schema
_notif_migrar_do_antigo = _pf_notif._notif_migrar_do_antigo
_notif_maior_id_antigo = _pf_notif._notif_maior_id_antigo
_notif_avanca_sequencia = _pf_notif._notif_avanca_sequencia
_notif_schema_pronto = _pf_notif._notif_schema_pronto
_ensure_notif_db = _pf_notif._ensure_notif_db


def init_db():
    log.info("[init_db] Initializing database schema…")
    conn = get_db_connection()
    try:
        log.debug("[init_db] Creating sequence seq_vc_id")
        conn.execute("CREATE SEQUENCE IF NOT EXISTS seq_vc_id START 1")
        log.debug("[init_db] Creating table users")
        conn.execute("""
            CREATE TABLE IF NOT EXISTS users (
                SID              VARCHAR PRIMARY KEY,
                Name             VARCHAR NOT NULL,
                Email            VARCHAR NOT NULL,
                Role_Description VARCHAR DEFAULT '',
                Position         VARCHAR DEFAULT '',
                Role             VARCHAR DEFAULT '',
                Status           VARCHAR DEFAULT 'Pending',
                IP_Address       VARCHAR,
                created_at       TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        log.debug("[init_db] Creating table verification_codes")
        conn.execute("""
            CREATE TABLE IF NOT EXISTS verification_codes (
                id         INTEGER  DEFAULT nextval('seq_vc_id') PRIMARY KEY,
                SID        VARCHAR  NOT NULL,
                code       VARCHAR(6) NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                expires_at TIMESTAMP NOT NULL,
                used       BOOLEAN  DEFAULT FALSE,
                attempts   INTEGER  DEFAULT 0
            )
        """)
        conn.execute("""
            CREATE INDEX IF NOT EXISTS idx_vc_lookup
            ON verification_codes (SID, used, expires_at)
        """)
        # `notifications` e `push_subscriptions` NÃO nascem mais aqui: elas
        # moram no banco de notificações (`_notif_init_schema`). Criá-las nos
        # dois lugares seria manter dois schemas para a mesma tabela, e eles
        # divergiriam na primeira coluna nova. Na instância elas continuam
        # existindo neste arquivo — é de lá que a migração as lê —, mas ninguém
        # mais escreve nelas.
        conn.commit()
        log.info("[init_db] Schema ready")
    except Exception:
        log.error("[init_db] FAILED:\n%s", traceback.format_exc())
        raise
    finally:
        conn.close()


def _migrate_schema():
    """Migra schemas antigos: sequence nula e colunas Role/Status ausentes."""
    log.info("[migrate] Checking schema migrations…")
    conn = get_db_connection()
    try:
        # Fix verification_codes: id NULL (schema sem sequence)
        try:
            row = conn.execute("SELECT COUNT(*) FROM verification_codes WHERE id IS NULL").fetchone()
            null_count = row[0] if row else 0
            log.debug("[migrate] verification_codes rows with NULL id: %d", null_count)
            if null_count > 0:
                log.warning("[migrate] Dropping verification_codes to fix NULL ids")
                conn.execute("DROP TABLE verification_codes")
                conn.execute("DROP SEQUENCE IF EXISTS seq_vc_id")
                conn.commit()
        except Exception:
            log.debug("[migrate] verification_codes check skipped: %s", traceback.format_exc())

        # Add verification_codes.attempts (per-code wrong-guess counter) if missing.
        try:
            vc_cols = [c[0] for c in conn.execute(
                "SELECT column_name FROM information_schema.columns WHERE table_name='verification_codes'"
            ).fetchall()]
            if vc_cols and 'attempts' not in vc_cols:
                log.warning("[migrate] Adding missing column verification_codes.attempts")
                conn.execute("ALTER TABLE verification_codes ADD COLUMN attempts INTEGER DEFAULT 0")
                conn.commit()
        except Exception:
            log.debug("[migrate] verification_codes.attempts check skipped: %s", traceback.format_exc())

        # Fix users: Role -> Role_Description + add Role + add Status
        try:
            cols = [c[0] for c in conn.execute(
                "SELECT column_name FROM information_schema.columns WHERE table_name='users'"
            ).fetchall()]
            log.debug("[migrate] users columns: %s", cols)

            if 'Role' in cols and 'Role_Description' not in cols:
                log.warning("[migrate] Renaming Role → Role_Description")
                conn.execute("ALTER TABLE users RENAME COLUMN Role TO Role_Description")
                conn.commit()
                cols = [c[0] for c in conn.execute(
                    "SELECT column_name FROM information_schema.columns WHERE table_name='users'"
                ).fetchall()]
                log.debug("[migrate] users columns after rename: %s", cols)

            if 'Role' not in cols:
                log.warning("[migrate] Adding missing column Role")
                conn.execute("ALTER TABLE users ADD COLUMN Role VARCHAR DEFAULT ''")
                conn.commit()

            if 'Status' not in cols:
                log.warning("[migrate] Adding missing column Status")
                conn.execute("ALTER TABLE users ADD COLUMN Status VARCHAR")
                conn.execute("UPDATE users SET Status = 'Pending' WHERE Status IS NULL")
                conn.commit()

            if 'Position' not in cols:
                log.warning("[migrate] Adding missing column Position")
                conn.execute("ALTER TABLE users ADD COLUMN Position VARCHAR DEFAULT ''")
                conn.commit()

            if 'Page_Access' not in cols:
                # JSON array of accessible page URLs. NULL/'' = not configured yet →
                # full access (backward compatible); a stored array = enforced allowlist.
                log.warning("[migrate] Adding missing column Page_Access")
                conn.execute("ALTER TABLE users ADD COLUMN Page_Access VARCHAR DEFAULT ''")
                conn.commit()

            log.info("[migrate] Schema migration complete")
        except Exception:
            log.error("[migrate] users schema migration FAILED:\n%s", traceback.format_exc())

        # As duas tabelas de notificação NÃO são mais migradas aqui. Elas
        # moram no banco de notificações, e é o `_ensure_notif_db` que cria o
        # schema e traz o que está neste arquivo. O `target_sid` inclusive: a
        # migração copia as colunas que EXISTEM no antigo, então uma base sem
        # ele nasce com a coluna do schema novo e as linhas antigas entram com
        # o default.
    finally:
        conn.close()


# O disparo do sino/push mora em `platform/notifications.py` — aliases.
_notif_roles = _pf_notif._notif_roles
_create_notification = _pf_notif._create_notification
_push_notify = _pf_notif._push_notify


def _nd_token(value):
    """Return a ' [ND:YYYY-MM-DD]' suffix for a notification detail so the bell can
    deep-link to that date (Accrual → ?date=, New Deals → ?tradedate=). Accepts a
    date, YYYYMMDD, YYYY-MM-DD or dd/mm/yyyy string; returns '' when unparseable.
    The topbar strips the token before displaying the detail."""
    if not value:
        return ''
    s = str(value).strip()
    d = None
    if re.match(r'^\d{8}$', s):
        try:
            d = datetime.strptime(s, '%Y%m%d')
        except Exception:
            d = None
    if d is None:
        try:
            d = _parse_date_any(s)
        except Exception:
            d = None
    return ' [ND:{}]'.format(d.strftime('%Y-%m-%d')) if d else ''


def get_user_by_sid(sid):
    log.debug("[get_user_by_sid] Looking up SID=%s", sid)
    conn = get_db_connection(readonly=True)
    try:
        result = conn.execute(
            "SELECT SID, Name, Email, Role_Description, Position, Role, Status, IP_Address FROM users WHERE SID = ?",
            [sid]
        ).fetchone()
        if result:
            user = {
                "SID": result[0],
                "Name": result[1],
                "Email": result[2],
                "Role_Description": result[3],
                "Position": result[4] or "",
                "Role": result[5],
                "Status": result[6] or "Pending",
                "IP_Address": result[7]
            }
            log.debug("[get_user_by_sid] Found: Name=%s Role=%s Status=%s IP=%s",
                      user["Name"], user["Role"], user["Status"], user["IP_Address"])
            return user
        log.debug("[get_user_by_sid] SID=%s not found in DB", sid)
        return None
    except Exception:
        log.error("[get_user_by_sid] Query error:\n%s", traceback.format_exc())
        raise
    finally:
        conn.close()


def get_all_users():
    conn = get_db_connection(readonly=True)
    try:
        rows = conn.execute("""
            SELECT SID, Name, Email, Role_Description, Position, Role, Status, IP_Address, created_at
            FROM users
            ORDER BY LOWER(COALESCE(Name, SID)) ASC
        """).fetchall()
        users = []
        for r in rows:
            users.append({
                "SID": r[0],
                "Name": r[1],
                "Email": r[2],
                "Role_Description": r[3] or "",
                "Position": r[4] or "",
                "Role": r[5] or "",
                "Status": r[6] or "Pending",
                "IP_Address": r[7],
                "created_at": r[8].strftime("%d %b, %Y") if r[8] else ""
            })
        return users
    finally:
        conn.close()


def get_role_groups():
    conn = get_db_connection(readonly=True)
    try:
        rows = conn.execute("""
            SELECT Role, COUNT(*) AS cnt,
                   LIST(SID   ORDER BY created_at DESC) AS sids,
                   LIST(Name  ORDER BY created_at DESC) AS names
            FROM users
            WHERE Role IS NOT NULL AND Role != ''
            GROUP BY Role
        """).fetchall()

        groups = {}
        for role, cnt, sids, names in rows:
            meta = ROLE_META.get(role, {
                'display': role, 'icon': 'ti-user',
                'description': '', 'responsibilities': [],
            })
            preview = [{'SID': s, 'Name': n}
                       for s, n in zip((sids or [])[:4], (names or [])[:4])]
            groups[role] = {
                'role': role,
                'display': meta['display'],
                'icon': meta['icon'],
                'description': meta['description'],
                'responsibilities': meta['responsibilities'],
                'count': cnt,
                'users': preview,
            }

        result = []
        for key in ['ADMIN', 'FO', 'MO', 'BO', 'INSTITUTIONAL', 'HUB']:
            if key in groups:
                result.append(groups[key])
        for key, val in groups.items():
            if key not in ['ADMIN', 'FO', 'MO', 'BO', 'INSTITUTIONAL', 'HUB']:
                result.append(val)
        return result
    finally:
        conn.close()


def insert_new_user(sid, name, email, role_description, ip_address):
    log.info("[insert_new_user] Inserting SID=%s Name=%s Email=%s IP=%s", sid, name, email, ip_address)
    conn = get_db_connection()
    try:
        conn.execute("""
            INSERT INTO users (SID, Name, Email, Role_Description, Position, Role, Status, IP_Address)
            VALUES (?, ?, ?, ?, '', '', 'Pending', ?)
        """, [sid, name, email, role_description or "", ip_address])
        conn.commit()
        log.info("[insert_new_user] Inserted SID=%s OK", sid)
    except Exception:
        log.error("[insert_new_user] FAILED for SID=%s:\n%s", sid, traceback.format_exc())
        raise
    finally:
        conn.close()


def update_user_ip(sid, ip_address):
    conn = get_db_connection()
    try:
        conn.execute("UPDATE users SET IP_Address = ? WHERE SID = ?", [ip_address, sid])
        conn.commit()
    finally:
        conn.close()


def update_user_role_status(sid, role, status, position=None):
    conn = get_db_connection()
    try:
        if position is not None:
            conn.execute(
                "UPDATE users SET Role = ?, Status = ?, Position = ? WHERE SID = ?",
                [role, status, position, sid]
            )
        else:
            conn.execute(
                "UPDATE users SET Role = ?, Status = ? WHERE SID = ?",
                [role, status, sid]
            )
        conn.commit()
    finally:
        conn.close()


def save_verification_code(sid, code):
    log.info("[save_verification_code] Saving code for SID=%s (expiry=%d min)", sid, CODE_EXPIRY_MINUTES)
    conn = get_db_connection()
    try:
        invalidated = conn.execute(
            "UPDATE verification_codes SET used = TRUE WHERE SID = ? AND used = FALSE",
            [sid]
        ).rowcount
        log.debug("[save_verification_code] Invalidated %s old codes for SID=%s", invalidated, sid)
        conn.execute(
            f"INSERT INTO verification_codes (SID, code, expires_at) "
            f"VALUES (?, ?, CURRENT_TIMESTAMP + INTERVAL '{CODE_EXPIRY_MINUTES}' MINUTE)",
            [sid, code]
        )
        conn.commit()
        log.info("[save_verification_code] Code saved for SID=%s", sid)
    except Exception:
        log.error("[save_verification_code] FAILED for SID=%s:\n%s", sid, traceback.format_exc())
        raise
    finally:
        conn.close()


def verify_code(sid, code):
    log.info("[verify_code] Verifying code for SID=%s", sid)
    conn = get_db_connection()
    try:
        # A match only counts while the code still has attempts left — once the
        # wrong-guess cap is hit the code is burned and can never validate again.
        result = conn.execute("""
            SELECT id FROM verification_codes
            WHERE SID = ? AND code = ? AND used = FALSE
              AND expires_at > CURRENT_TIMESTAMP
              AND attempts < ?
            ORDER BY created_at DESC
            LIMIT 1
        """, [sid, code, MAX_2FA_ATTEMPTS]).fetchone()

        if result:
            conn.execute("UPDATE verification_codes SET used = TRUE WHERE id = ?", [result[0]])
            conn.commit()
            log.info("[verify_code] Code verified OK for SID=%s (row id=%s)", sid, result[0])
            return True, "Code verified successfully."

        # No match: spend one attempt on the live code so the 6-digit space can't
        # be brute-forced. At the cap the code is invalidated (used = TRUE).
        active = conn.execute("""
            SELECT id, attempts, (expires_at > CURRENT_TIMESTAMP) AS live
            FROM verification_codes
            WHERE SID = ? AND used = FALSE
            ORDER BY created_at DESC
            LIMIT 1
        """, [sid]).fetchone()

        if active and active[2]:
            new_attempts = (active[1] or 0) + 1
            if new_attempts >= MAX_2FA_ATTEMPTS:
                conn.execute("UPDATE verification_codes SET used = TRUE, attempts = ? WHERE id = ?",
                             [new_attempts, active[0]])
                conn.commit()
                log.warning("[verify_code] SID=%s hit the attempt cap — code invalidated", sid)
                return False, "Too many incorrect attempts. Please request a new code."
            conn.execute("UPDATE verification_codes SET attempts = ? WHERE id = ?",
                         [new_attempts, active[0]])
            conn.commit()
            log.warning("[verify_code] Invalid code attempt %d/%d for SID=%s",
                        new_attempts, MAX_2FA_ATTEMPTS, sid)
            return False, "Invalid verification code."

        # No live code at all → it expired (or was already burned / never issued).
        expired = conn.execute(
            "SELECT 1 FROM verification_codes WHERE SID = ? AND code = ? AND used = FALSE",
            [sid, code]
        ).fetchone()
        if expired:
            log.warning("[verify_code] Code for SID=%s is EXPIRED", sid)
            return False, "Verification code has expired. Please request a new one."
        log.warning("[verify_code] Invalid code attempt for SID=%s (no live code)", sid)
        return False, "Invalid verification code."
    except Exception:
        log.error("[verify_code] Error for SID=%s:\n%s", sid, traceback.format_exc())
        raise
    finally:
        conn.close()


def _code_send_allowed(sid):
    """Throttle verification-code emails for a SID. Returns (allowed, message).

    Enforces a short cooldown between sends and a cap per rolling window so a
    known SID can't be used to flood a mailbox or to mint an endless stream of
    fresh codes for brute-forcing. Fails open on any DB error — 2FA email must
    never be bricked by the throttle itself."""
    try:
        conn = get_db_connection(readonly=True)
        try:
            row = conn.execute(
                "SELECT "
                "  SUM(CASE WHEN created_at > CURRENT_TIMESTAMP - INTERVAL '%d' SECOND "
                "           THEN 1 ELSE 0 END), "
                "  SUM(CASE WHEN created_at > CURRENT_TIMESTAMP - INTERVAL '%d' MINUTE "
                "           THEN 1 ELSE 0 END) "
                "FROM verification_codes WHERE SID = ?"
                % (CODE_RESEND_COOLDOWN_SECONDS, CODE_WINDOW_MINUTES),
                [sid]
            ).fetchone()
        finally:
            conn.close()
    except Exception:
        return True, ''
    just_now = (row[0] if row else 0) or 0
    recent = (row[1] if row else 0) or 0
    if just_now > 0:
        return False, "A verification code was just sent. Please wait a moment before requesting another."
    if recent >= MAX_CODES_PER_WINDOW:
        return False, "Too many code requests. Please wait a few minutes and try again."
    return True, ''


def cleanup_expired_codes():
    conn = get_db_connection()
    try:
        conn.execute("""
            DELETE FROM verification_codes
            WHERE used = TRUE OR expires_at < CURRENT_TIMESTAMP
        """)
        conn.commit()
    finally:
        conn.close()


# ==============================================================================
# FUNÇÕES AUXILIARES — UTILITÁRIOS
# ==============================================================================

def get_client_ip():
    # ProxyFix(x_for=1) (see create_app) already rewrites request.remote_addr to
    # the real client IP from the single trusted proxy hop. Do NOT read
    # X-Forwarded-For directly: its leftmost value is fully client-controlled and
    # could be spoofed to match a user's stored IP, skipping 2FA entirely.
    return request.remote_addr


def generate_verification_code():
    # 2FA secret: use a cryptographically secure RNG (secrets), never random —
    # the module-level Mersenne Twister is predictable from observed outputs.
    return ''.join(secrets.choice(string.digits) for _ in range(6))


def get_user_data_from_phonebook(sid):
    log.info("[phonebook] Fetching data for SID=%s", sid)
    try:
        data = awmpy.get_phonebook_data(sid)
        log.debug("[phonebook] Raw response keys for SID=%s: %s", sid, list(data.keys()) if data else None)
        result = {
            "nameFull": data.get("nameFull", ""),
            "email": data.get("email", ""),
            "positionName": data.get("positionName", "")
        }
        log.info("[phonebook] SID=%s → name=%s email=%s position=%s",
                 sid, result["nameFull"], result["email"], result["positionName"])
        return result
    except Exception:
        log.error("[phonebook] FAILED for SID=%s:\n%s", sid, traceback.format_exc())
        return None


def get_masked_email(email):
    if not email or '@' not in email:
        return "*******"
    local, domain = email.split('@', 1)
    if len(local) <= 2:
        masked_local = local[0] + '*****'
    else:
        masked_local = local[0] + '*' * (len(local) - 2) + local[-1]
    return f"{masked_local}@{domain}"


def get_masked_phone():
    return "******6789"


# ==============================================================================
# FUNÇÕES AUXILIARES — EMAIL
# ==============================================================================

def send_verification_email(to_email, code, recipient_name):
    from email.mime.image import MIMEImage
    from flask import current_app

    html_body = render_email_template(code, recipient_name)

    msg = MIMEMultipart('mixed')
    msg['Subject'] = "OTC Tracker - Verification Code"
    msg['From'] = SHARED_MAILBOX
    msg['To'] = to_email

    msg_related = MIMEMultipart('related')
    msg_alternative = MIMEMultipart('alternative')
    msg_alternative.attach(MIMEText('Please use an HTML email client to view this message.', 'plain'))
    msg_alternative.attach(MIMEText(html_body, 'html'))
    msg_related.attach(msg_alternative)

    logo_path = _get_logo_path()
    if logo_path:
        try:
            with open(logo_path, 'rb') as f:
                logo_data = f.read()
            logo_mime = MIMEImage(logo_data)
            logo_mime.add_header('Content-ID', '<otc_logo>')
            logo_mime.add_header('Content-Disposition', 'inline', filename='logo.png')
            msg_related.attach(logo_mime)
        except Exception as e:
            print(f"Warning: Could not attach logo: {e}")
    _attach_email_gradient(msg_related)

    msg.attach(msg_related)

    log.info("[send_email] Connecting to SMTP %s:%d", SMTP_HOST, SMTP_PORT)
    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            server.sendmail(SHARED_MAILBOX, to_email, msg.as_string())
        log.info("[send_email] Verification email sent to %s", to_email)
        return True
    except Exception:
        log.error("[send_email] FAILED sending to %s:\n%s", to_email, traceback.format_exc())
        return False


def send_account_activated_email(to_email, first_name):
    from email.mime.image import MIMEImage
    from datetime import datetime

    html_body = render_template(
        'pages/email-template-account-activated.html',
        first_name=first_name,
        sign_in_url=url_for('pages_blueprint.sign_in_page', _external=True),
        current_year=datetime.now().year
    )

    msg = MIMEMultipart('mixed')
    msg['Subject'] = "OTC Tracker - Your Account Has Been Activated"
    msg['From'] = SHARED_MAILBOX
    msg['To'] = to_email

    msg_related = MIMEMultipart('related')
    msg_alternative = MIMEMultipart('alternative')
    msg_alternative.attach(MIMEText('Your OTC Tracker account has been activated. Please sign in.', 'plain'))
    msg_alternative.attach(MIMEText(html_body, 'html'))
    msg_related.attach(msg_alternative)

    logo_path = _get_logo_path()
    if logo_path:
        try:
            with open(logo_path, 'rb') as f:
                logo_data = f.read()
            logo_mime = MIMEImage(logo_data)
            logo_mime.add_header('Content-ID', '<otc_logo>')
            logo_mime.add_header('Content-Disposition', 'inline', filename='logo.png')
            msg_related.attach(logo_mime)
        except Exception as e:
            print(f"Warning: Could not attach logo to activation email: {e}")
    else:
        print("Warning: logo not found, activation email will have no logo.")
    _attach_email_gradient(msg_related)

    msg.attach(msg_related)

    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            server.sendmail(SHARED_MAILBOX, to_email, msg.as_string())
        return True
    except Exception as e:
        print(f"Error sending activation email to {to_email}: {e}")
        return False


_get_logo_path = _pf_mail._get_logo_path
_get_email_asset = _pf_mail._get_email_asset
_attach_email_gradient = _pf_mail._attach_email_gradient


@blueprint.app_context_processor
def _inject_email_grad_url():
    """Expose `grad_url` (the header gradient image) to every template, so the
    shared e-mail header partial renders the Outlook gradient without each route
    having to pass it.

    It is ALWAYS the inline `cid:` attachment, never an absolute http URL: every
    sender attaches the image via `_attach_email_gradient()`, and a remote URL is
    a download Outlook may block, delay or complete only in part — which is what
    printed a header half gradient / half flat #4f8ae2 (see the CETIP Files Saved
    e-mail of 03/08/2026). The cid: reference cannot fail halfway."""
    return {'grad_url': 'cid:otc_gradient'}


_STATIC_DIR = os.path.normpath(os.path.join(os.path.dirname(__file__), '..', 'static'))


def _asset_v(rel_path):
    """Token de cache de um estático: o mtime do arquivo.

    Substitui o `?v=20260728b` escrito à mão, que já falhou em produção: o §164
    mudou o `otc-fileupload.js` e ninguém trocou a string, então o navegador
    seguiu servindo o JS ANTERIOR — que lê o cadastro NOVO e trata o padrão
    `HO"MY"` como prefixo literal. Resultado na tela: Underlying Asset
    `HO"MY"U6` (§170). Com o mtime, publicar o arquivo já invalida o cache; não
    há string para alguém esquecer de bumpar.

    Falha (arquivo ausente) devolve '0': pior caso o cache não é invalidado,
    que é o comportamento de hoje — nunca uma página quebrada.
    """
    try:
        return str(int(os.path.getmtime(os.path.join(_STATIC_DIR, rel_path))))
    except Exception:
        return '0'


@blueprint.app_context_processor
def _inject_asset_v():
    return {'asset_v': _asset_v}


def render_email_template(code, recipient_name):
    return render_template(
        'pages/email-verification.html',
        recipient_name=recipient_name,
        digits=list(code),
        expiry_minutes=CODE_EXPIRY_MINUTES,
        current_year=datetime.now().year
    )


# ==============================================================================
# INICIALIZAÇÃO DO BANCO DE DADOS
# ==============================================================================

def _ensure_duckdb_file():
    """Remove o arquivo se for SQLite (criado pelo SQLAlchemy antigo) para o DuckDB recriar."""
    db_path = os.path.abspath(DB_PATH)
    os.makedirs(os.path.dirname(db_path), exist_ok=True)
    if os.path.exists(db_path):
        try:
            with open(db_path, 'rb') as f:
                header = f.read(16)
            if header.startswith(b'SQLite format 3\x00'):
                print(f"[DB] Arquivo SQLite detectado em {db_path} — removendo para recriar como DuckDB.")
                os.remove(db_path)
        except Exception as e:
            print(f"[DB] Erro ao verificar formato do arquivo: {e}")


_ensure_duckdb_file()


def _ensure_db_initialized():
    """Run schema creation + migrations exactly once, lazily, in the process
    that first needs the DB. Deferred (NOT run at import) so the Werkzeug
    auto-reloader's supervisor process — which imports this module but never
    serves requests — does not open the single-writer DuckDB file and lock it
    out from the worker. In production (gunicorn) the first request triggers it.

    Re-entrant safe: init_db()/_migrate_schema() themselves call
    get_db_connection(), which calls back here. The per-thread `running` flag lets
    those nested calls fall straight through to open the connection without
    recursing into init again, while concurrent first-callers from OTHER threads
    block on the lock until init has fully completed (done set only on success)."""
    global _db_init_done
    if _db_init_done:
        return
    # Nested call on the SAME thread (from init_db's own get_db_connection):
    # don't recurse — let it proceed to open the connection.
    if getattr(_db_init_tls, 'running', False):
        return
    with _db_init_lock:
        if _db_init_done:
            return
        _db_init_tls.running = True
        try:
            init_db()
            _migrate_schema()
            cleanup_expired_codes()
            _db_init_done = True        # only mark done AFTER schema/migrations succeed
            log.info("[startup] Database initialized successfully at %s", os.path.abspath(DB_PATH))
        except Exception:
            log.error("[startup] Could not initialize database:\n%s", traceback.format_exc())
            raise
        finally:
            _db_init_tls.running = False


# ==============================================================================
# ROTAS — PÁGINAS DE AUTENTICAÇÃO
# ==============================================================================

@blueprint.route('/')
def index():
    if session.get('authenticated'):
        return redirect(url_for('pages_blueprint.dashboard'))
    return render_template('pages/auth-2-sign-in.html', segment='auth-2-sign-in')


@blueprint.route('/auth-2-sign-in')
def sign_in_page():
    return render_template('pages/auth-2-sign-in.html', segment='auth-2-sign-in')


@blueprint.route('/auth-2-sign-up')
def sign_up_page():
    return render_template('pages/auth-2-sign-up.html', segment='auth-2-sign-up')


@blueprint.route('/auth-2-two-factor')
def two_factor_page():
    return render_template(
        'pages/auth-2-two-factor.html',
        segment='auth-2-two-factor',
        masked_email=session.get('masked_email', '******'),
        masked_phone=session.get('masked_phone', '******6789')
    )


# ==============================================================================
# ROTAS — LÓGICA DE AUTENTICAÇÃO (POST)
# ==============================================================================

def _validate_sid(sid):
    return sid and re.match(r'^[A-Z][0-9]{6}$', sid)


@blueprint.route('/register', methods=['POST'])
def register():
    sid = request.form.get('sid', '').strip().upper()
    log.info("[register] Attempt SID=%s IP=%s", sid, get_client_ip())

    if not _validate_sid(sid):
        log.warning("[register] Invalid SID format: %r", sid)
        flash("Invalid SID format. Must be 1 letter + 6 numbers.", "error")
        return redirect(url_for('pages_blueprint.sign_up_page'))

    client_ip = get_client_ip()
    existing_user = get_user_by_sid(sid)

    if existing_user:
        log.info("[register] SID=%s already exists — delegating to _handle_existing_user", sid)
        return _handle_existing_user(existing_user, sid, client_ip,
                                     redirect_page='pages_blueprint.sign_up_page')
    else:
        log.info("[register] SID=%s is new — delegating to _handle_new_user", sid)
        return _handle_new_user(sid, client_ip,
                                redirect_page='pages_blueprint.sign_up_page')


@blueprint.route('/login', methods=['POST'])
def login():
    sid = request.form.get('sid', '').strip().upper()
    remember_me = request.form.get('remember_me') == 'on'
    client_ip = get_client_ip()
    log.info("[login] Attempt SID=%s IP=%s remember=%s", sid, client_ip, remember_me)

    if not _validate_sid(sid):
        log.warning("[login] Invalid SID format: %r", sid)
        flash("Invalid SID format. Must be 1 letter + 6 numbers.", "error")
        return redirect(url_for('pages_blueprint.sign_in_page'))

    existing_user = get_user_by_sid(sid)

    if existing_user:
        log.info("[login] SID=%s found in DB (Status=%s) — delegating to _handle_existing_user",
                 sid, existing_user.get("Status"))
        return _handle_existing_user(existing_user, sid, client_ip,
                                     redirect_page='pages_blueprint.sign_in_page',
                                     remember_me=remember_me)
    else:
        log.info("[login] SID=%s not in DB — delegating to _handle_new_user", sid)
        return _handle_new_user(sid, client_ip,
                                redirect_page='pages_blueprint.sign_in_page')


def _handle_existing_user(user, sid, client_ip, redirect_page, remember_me=False):
    status = user.get("Status", "Pending")
    stored_ip = user.get("IP_Address")
    log.info("[_handle_existing_user] SID=%s Status=%s StoredIP=%s ClientIP=%s remember=%s",
             sid, status, stored_ip, client_ip, remember_me)

    if status == 'Inactive':
        log.warning("[_handle_existing_user] SID=%s is Inactive — blocking login", sid)
        flash("Your account is inactive. Please contact the OTC Tracker administrator.", "error")
        return redirect(url_for(redirect_page))

    if status == 'Pending':
        log.warning("[_handle_existing_user] SID=%s is Pending — blocking login", sid)
        flash("Your account is pending approval. You will receive an email once it is activated.", "warning")
        return redirect(url_for(redirect_page))

    # Active
    if stored_ip == client_ip:
        log.info("[_handle_existing_user] SID=%s IP match — granting session directly", sid)
        _set_session(user, remember_me=remember_me)
        return redirect(url_for('pages_blueprint.dashboard'))
    else:
        log.info("[_handle_existing_user] SID=%s IP mismatch (stored=%s vs current=%s) — triggering 2FA",
                 sid, stored_ip, client_ip)
        # Do NOT persist the new IP yet: storing it before the code is verified
        # would let anyone holding the SID overwrite the trusted IP and then log
        # in directly (IP-match shortcut) without ever passing 2FA. Stash it in
        # the session and only commit it after verify_2fa succeeds.
        session['pending_remember_me'] = remember_me
        session['pending_ip'] = client_ip
        return _initiate_2fa(sid, user["Email"], user["Name"])


def _handle_new_user(sid, client_ip, redirect_page):
    log.info("[_handle_new_user] SID=%s — querying phonebook", sid)
    user_data = get_user_data_from_phonebook(sid)
    if not user_data:
        log.error("[_handle_new_user] Phonebook returned None for SID=%s", sid)
        flash("Could not retrieve user data. Please verify your SID.", "error")
        return redirect(url_for(redirect_page))

    log.info("[_handle_new_user] Phonebook OK for SID=%s — inserting into DB", sid)
    insert_new_user(sid, user_data["nameFull"], user_data["email"],
                    user_data["positionName"], client_ip)

    first_name = user_data["nameFull"].split()[0] if user_data["nameFull"] else sid
    log.info("[_handle_new_user] SID=%s registered successfully, showing success page", sid)
    _create_notification(sid, user_data.get("nameFull", sid), 'Access Request', 'Users',
                         sid + ' — ' + user_data.get("positionName", ''), target_role='ADMIN')
    return render_template('pages/auth-2-success-mail.html',
                           segment='auth-2-success-mail',
                           first_name=first_name)


def _set_session(user, remember_me=False):
    session.permanent = remember_me
    session['authenticated'] = True
    session['user_sid'] = user["SID"]
    session['user_name'] = user["Name"]
    session['user_email'] = user["Email"]
    # Master is pinned by SID and outranks every stored role.
    session['user_role'] = 'MASTER' if (user["SID"] or '').strip().upper() in _MASTER_SIDS else user["Role"]
    session['remember_me'] = remember_me
    # A freshly established session is never locked.
    session.pop('locked', None)
    # Without "Keep me signed in": hard cap of 5 hours (absolute, even with
    # activity). With it: 30 days + IP re-verification on a new IP.
    lifetime = timedelta(days=30) if remember_me else timedelta(hours=5)
    session['session_expires_at'] = (datetime.now(tz=timezone.utc) + lifetime).isoformat()


@blueprint.route('/verify-2fa', methods=['POST'])
def verify_2fa():
    sid = session.get('pending_sid')
    log.info("[verify_2fa] Request from IP=%s session_sid=%s is_json=%s",
             get_client_ip(), sid, request.is_json)

    if not sid:
        log.warning("[verify_2fa] No pending_sid in session — session keys: %s", list(session.keys()))
        flash("Session expired. Please try again.", "error")
        return redirect(url_for('pages_blueprint.sign_in_page'))

    if request.is_json:
        body = request.get_json()
        log.debug("[verify_2fa] JSON body keys: %s", list(body.keys()) if body else None)
        code = (body or {}).get('code', '').strip()
    else:
        code = request.form.get('code', '').strip()

    log.debug("[verify_2fa] Code length=%d for SID=%s", len(code), sid)

    if not code or len(code) != 6:
        log.warning("[verify_2fa] Invalid code format for SID=%s: %r", sid, code)
        if request.is_json:
            return jsonify({"success": False, "message": "Please enter a valid 6-digit code."}), 400
        flash("Please enter a valid 6-digit code.", "error")
        return redirect(url_for('pages_blueprint.two_factor_page'))

    is_valid, message = verify_code(sid, code)
    log.info("[verify_2fa] verify_code result for SID=%s: valid=%s msg=%s", sid, is_valid, message)

    if is_valid:
        user = get_user_by_sid(sid)
        remember_me = session.pop('pending_remember_me', False)
        # Now that the code is verified, trust this IP for future direct logins.
        pending_ip = session.pop('pending_ip', None)
        if pending_ip:
            update_user_ip(sid, pending_ip)
        session.pop('pending_sid', None)
        session.pop('masked_email', None)
        session.pop('masked_phone', None)
        _set_session(user, remember_me=remember_me)
        log.info("[verify_2fa] 2FA SUCCESS for SID=%s remember=%s — session set", sid, remember_me)

        if request.is_json:
            return jsonify({"success": True, "redirect": url_for('pages_blueprint.dashboard')})
        return redirect(url_for('pages_blueprint.dashboard'))
    else:
        log.warning("[verify_2fa] 2FA FAILED for SID=%s: %s", sid, message)
        if request.is_json:
            return jsonify({"success": False, "message": message}), 400
        flash(message, "error")
        return redirect(url_for('pages_blueprint.two_factor_page'))


@blueprint.route('/resend-code', methods=['POST'])
def resend_code():
    sid = session.get('pending_sid')

    if not sid:
        if request.is_json:
            return jsonify({"success": False, "message": "Session expired."}), 400
        flash("Session expired. Please try again.", "error")
        return redirect(url_for('pages_blueprint.sign_in_page'))

    user = get_user_by_sid(sid)
    if not user:
        if request.is_json:
            return jsonify({"success": False, "message": "User not found."}), 404
        flash("User not found.", "error")
        return redirect(url_for('pages_blueprint.sign_in_page'))

    # Cooldown + per-window cap: stop the resend button from bombing the mailbox
    # or minting endless fresh codes.
    allowed, wait_msg = _code_send_allowed(sid)
    if not allowed:
        if request.is_json:
            return jsonify({"success": False, "message": wait_msg}), 429
        flash(wait_msg, "warning")
        return redirect(url_for('pages_blueprint.two_factor_page'))

    code = generate_verification_code()
    save_verification_code(sid, code)
    email_sent = send_verification_email(user["Email"], code, user["Name"])

    if request.is_json:
        if email_sent:
            return jsonify({"success": True, "message": "New code sent successfully."})
        return jsonify({"success": False, "message": "Failed to send email."}), 500

    if email_sent:
        flash("A new verification code has been sent to your email.", "success")
    else:
        flash("Failed to send verification email. Please try again.", "error")
    return redirect(url_for('pages_blueprint.two_factor_page'))


# ==============================================================================
# ROTAS — LOCK SCREEN
# ==============================================================================

@blueprint.route('/lock')
def lock():
    """Lock the current session and send the user to the lock screen.
    Used by the topbar 'Lock Screen' item and the 3h idle auto-lock."""
    if not session.get('authenticated'):
        return redirect(url_for('pages_blueprint.sign_in_page'))
    session['locked'] = True
    log.info("[lock] Screen locked for SID=%s", session.get('user_sid'))
    return redirect(url_for('pages_blueprint.lock_screen_page'))


@blueprint.route('/auth-2-lock-screen')
def lock_screen_page():
    if not session.get('authenticated'):
        return redirect(url_for('pages_blueprint.sign_in_page'))
    return render_template(
        'pages/auth-2-lock-screen.html',
        segment='auth-2-lock-screen',
        user_name=session.get('user_name', ''),
        user_sid=session.get('user_sid', ''),
    )


@blueprint.route('/unlock', methods=['POST'])
def unlock():
    """Unlock the screen: the SID must match the locked account and pass the
    same IP verification as login (direct on IP match, otherwise 2FA)."""
    locked_sid = session.get('user_sid')
    if not session.get('authenticated') or not locked_sid:
        return redirect(url_for('pages_blueprint.sign_in_page'))

    sid = request.form.get('sid', '').strip().upper()
    log.info("[unlock] Attempt SID=%s lockedSID=%s IP=%s", sid, locked_sid, get_client_ip())

    if not _validate_sid(sid):
        flash("Invalid SID format. Must be 1 letter + 6 numbers.", "error")
        return redirect(url_for('pages_blueprint.lock_screen_page'))

    if sid != locked_sid:
        log.warning("[unlock] SID mismatch (entered=%s locked=%s)", sid, locked_sid)
        flash("This SID does not match the locked account.", "error")
        return redirect(url_for('pages_blueprint.lock_screen_page'))

    user = get_user_by_sid(sid)
    if not user:
        session.clear()
        return redirect(url_for('pages_blueprint.sign_in_page'))

    # Reuse the login IP-verification flow. On IP match _set_session clears the
    # 'locked' flag; on mismatch it routes to 2FA, after which verify_2fa does.
    remember_me = session.get('remember_me', False)
    return _handle_existing_user(user, sid, get_client_ip(),
                                 redirect_page='pages_blueprint.lock_screen_page',
                                 remember_me=remember_me)


# ==============================================================================
# ROTAS — APLICAÇÃO (PÓS-AUTENTICAÇÃO)
# ==============================================================================

@blueprint.route('/dashboard')
def dashboard():
    if not session.get('authenticated'):
        return redirect(url_for('pages_blueprint.sign_in_page'))
    return render_template('pages/index.html', segment='index')


# O painel percorre a árvore INTEIRA dos arquivos-dia do New Deals e ABRE cada
# um deles. Em disco local isso custa milissegundos; no share da instância cada
# `open()` é uma ida à rede, e um ano de pregão em dez pastas de produto são
# milhares delas — o request leva minutos e o navegador fica com o traço do
# placeholder o tempo todo. O sintoma é o pior possível: o servidor ESTÁ lendo, e
# a tela parece um dia sem operação.
#
# O painel é só leitura e agregação, então a resposta serve para todo mundo por
# alguns minutos. O TTL é curto de propósito: o dashboard é o que a mesa abre
# depois de importar, e uma janela grande faria a importação "não aparecer".
_DASH_TTL = float(os.getenv('OTC_DASHBOARD_TTL', '120'))
_dash_cache = {}
_dash_cache_lock = threading.Lock()


def _dash_stats_cached(period):
    """A resposta ainda fresca daquele período, ou `None`."""
    if _DASH_TTL <= 0:
        return None
    with _dash_cache_lock:
        item = _dash_cache.get(period)
    if not item:
        return None
    quando, dados = item
    return dados if (time.time() - quando) < _DASH_TTL else None


def _dash_stats_store(period, dados):
    """Guarda e DEVOLVE o payload — para o `return` do endpoint ser um só."""
    if _DASH_TTL > 0:
        with _dash_cache_lock:
            _dash_cache[period] = (time.time(), dados)
    return dados


# O leitor memoizado dos arquivos-dia mora em `platform/json_cache.py`.
_daycache_memo = _pf_jcache._daycache_memo
_daycache_lock = _pf_jcache._daycache_lock
_DAYCACHE_MAX = _pf_jcache._DAYCACHE_MAX
_daycache_dir_ok = _pf_jcache._daycache_dir_ok
_day_files = _pf_jcache._day_files
_day_json = _pf_jcache._day_json
_daycache_forget = _pf_jcache._daycache_forget

# ── A varredura do painel: podar, e ler a árvore UMA vez ─────────────────────
# O painel varria a árvore INTEIRA do cache de New Deals e abria todo JSON do
# período. Na dev é um SSD com dezenas de arquivos e ninguém nota; na instância
# do JPM a árvore está num share, cada operação é ida e volta de rede, e dois
# anos de histórico são milhares de arquivos — a tela fica em "Carregando os
# dados do painel…" por minutos, sem erro nenhum, porque nada falhou: o servidor
# está lendo.
#
# Duas correções, e a segunda quase se pagou sozinha na primeira tentativa:
#
#   • PODAR. A árvore termina em `{Produto}/{Sub}/{AAAA}/{MM}/arquivo.json`, e o
#     período pedido descarta ano e mês INTEIROS antes de entrar neles.
#   • LEMBRAR. Arquivo-dia já lido não precisa ser reaberto se não mudou.
#
# O memo só compensa se a verificação for de GRAÇA, e é aqui que o `os.scandir`
# entra no lugar do `os.walk`. A listagem de um diretório no SMB já devolve nome,
# tamanho e mtime de cada entrada, e o `DirEntry` os guarda: no Windows,
# `entry.stat()` não custa chamada nenhuma. Com `os.walk` essa informação é
# jogada fora e cada verificação vira um `os.stat` — uma ida a mais por arquivo,
# que é justamente o que a medição mostrou: a primeira abertura do painel ficava
# MAIS lenta do que antes, e só a segunda ganhava. Pelo scandir, a primeira
# abertura custa o mesmo de antes e a segunda deixa de abrir o que não mudou.
#
# O memo guarda os deals PROJETADOS nos campos que o endpoint usa, e não o
# registro inteiro: são 11 campos de umas 40, e guardar tudo seria trocar minutos
# de rede por centenas de MB no processo único que serve a mesa. Campo novo lido
# do deal no endpoint tem de entrar nesta tupla — `check_dashboard_walk.py` varre
# a função por AST e recusa o que ficar de fora, porque a projeção silenciosa
# devolveria `None` sem erro nenhum.
_DASH_DEAL_FIELDS = ('Client', 'Commodities', 'Commodity', 'Deal', 'LE',
                     'Status', 'TradeDate', 'UnderlyingAsset')

_dash_file_memo = {}                    # caminho → (mtime, tamanho, [deals])
_dash_memo_lock = threading.Lock()
# Teto do memo. Estourado, ele é ESVAZIADO inteiro em vez de despejar por idade:
# o custo é uma varredura completa a mais, uma vez, e a alternativa (LRU) é
# estado a mais para manter no caminho mais quente da tela.
_DASH_MEMO_MAX = 8000


def _dash_dir_matters(nome, pai, period, now):
    """Este diretório pode conter arquivo do período pedido?

    Nome de 4 dígitos é ano; de 2 dígitos, mês — e o mês só é descartado dentro
    do ano corrente, porque em outro ano ele já não é alcançado. Nome que não
    seja número é pasta de produto e nunca é descartado: quem decide o que é ano
    ou mês é o FORMATO do nome, não a profundidade — a árvore tem produto com um
    nível de subpasta e produto com dois.
    """
    if period == 'all':
        return True
    if len(nome) == 4 and nome.isdigit():
        return int(nome) == now.year
    if period == 'month' and len(nome) == 2 and nome.isdigit():
        if len(pai) == 4 and pai.isdigit():
            return int(nome) == now.month
    return True


def _dash_scan_files(raiz, period, now):
    """Gera (caminho, nome, mtime, tamanho) dos arquivos da árvore, podando.

    `os.scandir` no lugar de `os.walk` de propósito — ver o comentário acima: é
    o que faz a checagem do memo não custar uma ida a mais por arquivo.

    Um diretório que não abre é PULADO com aviso, não derruba a varredura: o
    share fica indisponível de vez em quando, e meia tela é melhor do que um 500
    no painel inteiro.
    """
    pilha = [(raiz, '')]
    while pilha:
        atual, pai = pilha.pop()
        subdirs, arquivos = [], []
        try:
            with os.scandir(atual) as entradas:
                for e in entradas:
                    try:
                        if e.is_dir():
                            if _dash_dir_matters(e.name, pai, period, now):
                                subdirs.append((e.name, e.path))
                            continue
                        if not e.name.endswith('.json'):
                            continue
                        st = e.stat()
                        arquivos.append((e.name, e.path, st.st_mtime, st.st_size))
                    except OSError:
                        continue
        except OSError:
            log.warning('[dashboard] não consegui listar %s', atual)
            continue
        # ORDENADO, e por nome — nos dois níveis. A ordem de leitura decide o
        # desempate da lista de "Recent deals" (deals do MESMO dia saem na ordem
        # em que entraram), e a ordem crua do `scandir` é a do sistema de
        # arquivos: o mesmo dado renderia listas diferentes no share do JPM e no
        # disco da dev. A pilha é LIFO, então os subdiretórios entram ao
        # contrário para sair em ordem.
        for nome, caminho in sorted(subdirs, reverse=True):
            pilha.append((caminho, nome))
        for nome, caminho, mtime, size in sorted(arquivos):
            yield (caminho, nome, mtime, size)


def _product_from_path(file_path):
    """Derive product label from directory path relative to new deals/ root.
    e.g. .../Option/Commodities/2026/06/file.json  → 'Option Commodities'
         .../NDF/FWD Start/2026/06/file.json       → 'NDF FWD Start'

    De MÓDULO, e não aninhada no endpoint: quem grava o memo (`_dash_file_deals`)
    guarda o `_product` e o `_type` junto de cada deal, então o aquecimento
    precisa derivá-los do MESMO jeito que o endpoint — aninhada, ele não a
    alcança, e a thread de aquecimento morria com `NameError` sem nada aparecer
    na tela (o painel seguia certo, só voltava a pagar a leitura inteira). É
    pura: só lê o `NEW_DEALS_CACHE_ROOT`, que é global.
    """
    rel = os.path.relpath(file_path, NEW_DEALS_CACHE_ROOT).replace('\\', '/')
    parts = rel.split('/')
    pretty = {'FwdStart': 'FWD Start', 'OtherPublisher': 'Other Publisher'}
    label_parts = [pretty.get(p, p) for p in parts[:-1] if not p.isdigit()][:2]
    return ' '.join(label_parts) if label_parts else 'Other'


def _type_from_product(product):
    p = product.lower()
    if p.startswith('option'):
        return 'OPT'
    if p.startswith('swap'):
        return 'SWAP'
    return 'NDF'


def _dash_file_deals(fp, fname, mtime, size, fdate, product, deal_type):
    """Os deals de UM arquivo-dia, já projetados, filtrados e anotados.

    A chave do memo é (mtime, tamanho) e não só o caminho: o arquivo-dia de HOJE
    é reescrito a cada importação, e um amend entra no arquivo do dia da operação
    — que pode ser antigo. Pelo caminho sozinho o painel mostraria o dia
    congelado na primeira leitura do processo.
    """
    with _dash_memo_lock:
        item = _dash_file_memo.get(fp)
    if item and item[0] == mtime and item[1] == size:
        return item[2]
    try:
        with open(fp, 'r', encoding='utf-8') as fh:
            data = json.load(fh)
    except Exception:                                       # noqa: BLE001
        return []
    saida = []
    fdate_txt = fdate.strftime('%Y-%m-%d')
    for d in (data if isinstance(data, list) else []):
        if not isinstance(d, dict) or not (d.get('Deal') or '').strip():
            continue
        # Cancelado via API não conta em nenhuma métrica.
        if str(d.get('Status') or '').strip() == 'Canceled':
            continue
        projetado = {k: d.get(k) for k in _DASH_DEAL_FIELDS}
        projetado['_fdate'] = fdate_txt
        projetado['_product'] = product
        projetado['_type'] = deal_type
        saida.append(projetado)
    with _dash_memo_lock:
        if len(_dash_file_memo) >= _DASH_MEMO_MAX:
            _dash_file_memo.clear()
        _dash_file_memo[fp] = (mtime, size, saida)
    return saida


# ── O memo do painel nasce QUENTE ───────────────────────────────────────────
# A varredura ficou barata da SEGUNDA abertura em diante — o memo evita reabrir
# o que não mudou. Só que a instância reinicia várias vezes ao dia (o reloader
# está desligado, então todo deploy pede restart), e a cada restart o memo volta
# a zero: quem abrir o painel primeiro paga a leitura inteira da árvore, e é
# quase sempre alguém.
#
# Esta thread faz essa leitura FORA do request, no start. Ela não calcula nada e
# não muda resposta nenhuma: só enche o `_dash_file_memo`, que é exatamente o
# trabalho que o primeiro visitante fazia. Quem abrir o painel no meio da
# varredura não espera por ela — cada arquivo que ela já leu é um a menos para o
# request, e os que faltam ele lê como antes.
#
# `year` primeiro porque é o período com que o painel ABRE (`loadDashboard('year')`
# no dashboard.js); `all` depois, para quem troca o filtro. A ordem importa: com
# `all` primeiro, o período que a tela usa só ficaria pronto no fim.
def _dash_warm_memo():
    if not os.path.isdir(NEW_DEALS_CACHE_ROOT):
        return
    agora = datetime.now()
    for periodo in ('year', 'all'):
        t0 = time.time()
        lidos = 0
        try:
            for fp, fname, mtime, size in _dash_scan_files(NEW_DEALS_CACHE_ROOT, periodo, agora):
                if fname.endswith('.tmp') or fname.endswith('.bak'):
                    continue
                try:
                    fdate = datetime.strptime(fname[:8], '%Y%m%d')
                except ValueError:
                    continue
                produto = _product_from_path(fp)
                _dash_file_deals(fp, fname, mtime, size, fdate,
                                 produto, _type_from_product(produto))
                lidos += 1
        except Exception:                                   # noqa: BLE001
            # Aquecer é otimização: falhar aqui não pode derrubar nada. O
            # painel continua funcionando, só volta a pagar a leitura no
            # primeiro acesso.
            log.warning('[dashboard] aquecimento de %s parou:\n%s',
                        periodo, traceback.format_exc())
            return
        log.info('[dashboard] memo aquecido: %s — %d arquivo(s) em %.1fs',
                 periodo, lidos, time.time() - t0)


def _dash_warm_start():
    threading.Thread(target=_dash_warm_memo, name='dashboard-warm',
                     daemon=True).start()


# Sobe com o APP, como os laços agendados: no import ele rodaria em todo script
# que importa o módulo, e a varredura do share não tem o que fazer ali.
_schedule_on_start('dashboard-warm', _dash_warm_start)

@blueprint.route('/api/data-files/status')
def api_data_files_status():
    """De ONDE o app está lendo cada JSON, e o que ele achou lá.

    Existe porque a falha desta família não parece falha: a tela abre, a API
    responde 200 e não há dado (CLAUDE.md §4). Da máquina de quem desenvolve não
    dá para ver o `DATA_DIR` da instância, e "não está carregando" sozinho não
    distingue as quatro causas — arquivo ausente, arquivo vazio, arquivo no
    lugar errado e share fora do ar.

    Ele responde as quatro de uma vez, por arquivo: o caminho RESOLVIDO (que é o
    que o `data_path` escolheu, e não onde ele deveria estar), se veio do
    `DATA_DIR` ou da cópia empacotada, o tamanho, quantos registros e a data.
    Sem esta rota, a única forma de saber isso é alguém abrir o share e olhar.
    """
    if not session.get('authenticated'):
        return jsonify({'success': False, 'error': 'Not authenticated'}), 401

    # Os arquivos que as telas leem por URL estática e que, vazios, deixam a
    # tela em branco sem erro. Não é a lista COMPLETA do `DATA_DIR` — é a dos
    # que já causaram esse sintoma.
    ARQUIVOS = ('Subjacente.json', 'VCP.json', 'Dominio.json', 'SwapIndex.json',
                'RefData.json', 'CounterpartyDetails.json', 'anbima.json')

    def _olha(nome):
        alvo = data_path(nome)
        item = {'file': nome, 'resolved': alvo,
                'from': 'DATA_DIR' if os.path.dirname(alvo) == os.path.normpath(Config.DATA_DIR)
                        else 'packaged'}
        try:
            st = os.stat(alvo)
        except OSError as e:
            item.update(exists=False, error='{}: {}'.format(type(e).__name__, e))
            return item
        item.update(exists=True, bytes=st.st_size,
                    modified=datetime.fromtimestamp(st.st_mtime).strftime('%d/%m/%Y %H:%M:%S'))
        # A contagem é o que separa "o arquivo está lá" de "o arquivo serve":
        # um `[]` de 2 bytes existe, tem data e não enche select nenhum.
        try:
            with open(alvo, encoding='utf-8') as fh:
                dados = json.load(fh)
            item['records'] = len(dados) if isinstance(dados, (list, dict)) else None
            item['is_list'] = isinstance(dados, list)
        except Exception as e:                              # noqa: BLE001
            item['error'] = 'não é JSON válido: {}'.format(e)
        return item

    arquivos = [_olha(n) for n in ARQUIVOS]
    return jsonify({
        'success': True,
        # As três raízes, para a resposta dizer sozinha em que ambiente ela foi
        # tirada — o mesmo JSON vindo da dev e da instância se parece.
        'data_dir': Config.DATA_DIR,
        'packaged_dir': PACKAGED_DATA_DIR,
        'same_folder': os.path.normpath(Config.DATA_DIR) == os.path.normpath(PACKAGED_DATA_DIR),
        'database_dir': Config.DATABASE_DIR,
        'shared_drive_root': Config.SHARED_DRIVE_ROOT,
        'files': arquivos,
        # O mapa que o gerador de confirmação usa. Ele é lido do MESMO
        # `Subjacente.json`, e ficar em zero é o que faz toda operação sair com
        # "sem cadastro no Subjacente" mesmo estando cadastrada.
        'subjacente_map_size': len(_conf_subjacente_map()),
    })

@blueprint.route('/api/dashboard-stats')
def api_dashboard_stats():
    if not session.get('authenticated'):
        return jsonify({'error': 'unauthorized'}), 401

    period = request.args.get('period', 'all')  # month | year | all
    cache = _dash_stats_cached(period)
    if cache is not None:
        return jsonify(cache)
    now = datetime.now()
    cur_month, cur_year = now.month, now.year

    def _file_in_period(fdate):
        if period == 'month':
            return fdate.month == cur_month and fdate.year == cur_year
        if period == 'year':
            return fdate.year == cur_year
        return True

    def _is_lawton(d):
        return 'lawton' in (d.get('Client') or '').lower()

    def _is_bank(d):
        cl = (d.get('Client') or '').lower()
        return 'banco' in cl or 'j.p morgan' in cl or 'jp morgan' in cl or 'jpmorgan' in cl

    _jpm_re = re.compile(r'J\.?P\.?\s*MORGAN', re.IGNORECASE)

    def _gen_ndf_counted(d):
        """NDF Vanilla / Other Publisher / FWD Start: nos gráficos de
        distribuição e deal flow só contam os pares JPM×cliente, JPM×MGT,
        JPM×Lawton e MGT×cliente. Os espelhos (MGT×JPM) e demais combinações
        (MGT×Lawton, LE fora de JPM/MGT) ficam de fora para o mesmo deal
        intragrupo não ser contado duas vezes."""
        le = (d.get('LE') or '').strip().upper()
        if le == 'JPM':
            return True
        if le == 'MGT':
            cl = (d.get('Client') or '')
            return 'LAWTON' not in cl.upper() and not _jpm_re.search(cl)
        return False

    # Generic scan of all new deals cache directories
    all_deals = []
    if os.path.isdir(NEW_DEALS_CACHE_ROOT):
        for fp, fname, mtime, size in _dash_scan_files(NEW_DEALS_CACHE_ROOT, period, now):
            if fname.endswith('.tmp') or fname.endswith('.bak'):
                continue
            try:
                fdate = datetime.strptime(fname[:8], '%Y%m%d')
            except ValueError:
                continue
            # A poda por diretório é grossa (ano e mês); quem decide de fato é a
            # data no NOME do arquivo, como sempre foi.
            if not _file_in_period(fdate):
                continue
            product = _product_from_path(fp)
            all_deals.extend(_dash_file_deals(
                fp, fname, mtime, size, fdate, product, _type_from_product(product)))

    def _is_fxo(d):
        return 'fxo' in (d.get('_product') or '').lower()
    # Deal counting rule (ALL products): count every leg that carries a deal,
    # EXCEPT the Banco J.P. Morgan counterparty leg. Each intragroup deal is
    # recorded from two views — "Banco x Lawton" and "Lawton x Banco" — which are
    # the SAME deal; the mirror whose Client is the Banco leg is dropped, so it is
    # counted only once. The client-facing leg (Client = external client) and the
    # single remaining intragroup leg are both kept.
    counted_deals = [d for d in all_deals if not _is_bank(d)]
    client_deals  = [d for d in all_deals if not _is_lawton(d) and not _is_bank(d)]

    def _fam(d):
        # FXO is split out of the OPT bucket so the dashboard can show it apart
        return 'FXO' if _is_fxo(d) else d['_type']

    def _ndf_bucket(product):
        """NDF sub-bucket by product label: vanilla/otherpub/fwdstart, '' = rest
        (Commodities, Intrag, …). Same split the distribution/flow charts show."""
        q = (product or '').lower().replace(' ', '')
        if 'vanilla' in q:
            return 'vanilla'
        if 'otherpublisher' in q:
            return 'otherpub'
        if 'fwdstart' in q:
            return 'fwdstart'
        return ''
    ndf_deals     = [d for d in counted_deals if _fam(d) == 'NDF']
    optcomm_deals = [d for d in counted_deals if _fam(d) == 'OPT']
    fxo_deals     = [d for d in counted_deals if _fam(d) == 'FXO']
    swap_deals    = [d for d in counted_deals if _fam(d) == 'SWAP']
    # Os três NDF genéricos usam a regra própria de pares (LE × contraparte) em
    # vez do filtro _is_bank — por isso partem de all_deals, não de counted_deals.
    _gen_ndf_all = [d for d in all_deals if d.get('_type') == 'NDF']
    ndf_vanilla_deals  = [d for d in _gen_ndf_all
                          if _ndf_bucket(d.get('_product')) == 'vanilla' and _gen_ndf_counted(d)]
    ndf_otherpub_deals = [d for d in _gen_ndf_all
                          if _ndf_bucket(d.get('_product')) == 'otherpub' and _gen_ndf_counted(d)]
    ndf_fwdstart_deals = [d for d in _gen_ndf_all
                          if _ndf_bucket(d.get('_product')) == 'fwdstart' and _gen_ndf_counted(d)]
    ndf_comm_deals     = [d for d in ndf_deals if not _ndf_bucket(d.get('_product'))]
    opt_deals     = optcomm_deals + fxo_deals  # all options (stat card)
    pending_statuses = {'Pending', 'New', 'pending', 'new'}
    pending_total = sum(1 for d in counted_deals if (d.get('Status') or '').strip() in pending_statuses)

    swap_total = len(swap_deals)

    client_counts = Counter(
        (d.get('Client') or '').strip()
        for d in client_deals
        if (d.get('Client') or '').strip()
    )
    # `most_common` desempata pela ordem de INSERÇÃO, que aqui é a ordem em que a
    # árvore foi lida — a do sistema de arquivos. Dois clientes com a MESMA
    # contagem trocavam de lugar entre o share do JPM e o disco da dev, e entre
    # duas leituras no mesmo lugar. O desempate passa a ser o nome, que é
    # arbitrário do mesmo jeito mas é sempre o mesmo. Vale para os três Top 5.
    def _top5(contador):
        return sorted(contador.items(), key=lambda kv: (-kv[1], kv[0]))[:5]

    top5_clients = []
    for c, n in _top5(client_counts):
        by_product = Counter(
            d['_product'] for d in client_deals if (d.get('Client') or '').strip() == c
        )
        top5_clients.append({'label': c, 'count': n, 'by_product': dict(by_product)})

    product_counts = Counter(d['_product'] for d in counted_deals)
    top5_products  = [{'label': p, 'count': n} for p, n in _top5(product_counts)]

    # Top 5 Underlying Assets — commodities show the Commodity name; FXO (no
    # Commodity) falls back to UnderlyingAsset (the currency).
    def _underlying_label(d):
        return (d.get('Commodities') or d.get('Commodity') or d.get('UnderlyingAsset') or '').strip()
    underlying_counts = Counter(
        _underlying_label(d) for d in counted_deals if _underlying_label(d)
    )
    top5_underlying = [{'label': c, 'count': n} for c, n in _top5(underlying_counts)]

    # Monthly counts for current year (always full year, ignores period filter)
    monthly_opt = [0] * 12
    monthly_ndf = [0] * 12
    monthly_fxo = [0] * 12
    monthly_swap = [0] * 12
    monthly_ndf_vanilla  = [0] * 12
    monthly_ndf_otherpub = [0] * 12
    monthly_ndf_fwdstart = [0] * 12
    if os.path.isdir(NEW_DEALS_CACHE_ROOT):
        # Segunda passada pela MESMA árvore — os contadores por mês são sempre do
        # ano inteiro e ignoram o período pedido. Ela usa o mesmo `_dash_scan_files`
        # e o mesmo memo da primeira: com `period='year'` os arquivos já foram
        # lidos e esta passada não abre nenhum, e com 'month' ou 'all' ela lê o
        # ano uma vez e a primeira passada aproveita. Antes eram dois `os.walk`
        # independentes, cada um abrindo os arquivos do próprio critério — o mesmo
        # arquivo lido DUAS vezes do share na mesma tela.
        for fp, fname, mtime, size in _dash_scan_files(NEW_DEALS_CACHE_ROOT, 'year', now):
            if fname.endswith('.tmp') or fname.endswith('.bak'):
                continue
            try:
                fdate = datetime.strptime(fname[:8], '%Y%m%d')
            except ValueError:
                continue
            if fdate.year != cur_year:
                continue
            product = _product_from_path(fp)
            is_fxo_file = 'fxo' in product.lower()
            ptype = _type_from_product(product)
            gen_bucket = ''
            if is_fxo_file:
                target = monthly_fxo
            elif ptype == 'OPT':
                target = monthly_opt
            elif ptype == 'SWAP':
                target = monthly_swap
            else:
                gen_bucket = _ndf_bucket(product)
                target = {'vanilla': monthly_ndf_vanilla,
                          'otherpub': monthly_ndf_otherpub,
                          'fwdstart': monthly_ndf_fwdstart}.get(
                              gen_bucket, monthly_ndf)
            # Same rule as the totals above: count every leg that carries a
            # deal except the Banco J.P. Morgan leg — this keeps the client
            # leg and one intragroup leg while dropping the mirror view, so
            # a deal seen from two sides is counted only once. Os NDF
            # genéricos (vanilla/other pub/fwd start) usam a regra de
            # pares LE × contraparte (_gen_ndf_counted).
            # O `Deal` vazio e o `Canceled` já foram descartados na leitura.
            target[fdate.month - 1] += sum(
                1 for d in _dash_file_deals(fp, fname, mtime, size, fdate, product, ptype)
                if (_gen_ndf_counted(d) if gen_bucket else not _is_bank(d))
            )

    # Recent deals: last 50 client rows sorted desc — frontend filters by product
    # A chave leva o Deal junto: `sorted` é estável, então sem ele o desempate
    # entre deals do MESMO dia era a ordem de leitura da árvore — que é a do
    # sistema de arquivos, e portanto diferente no share do JPM e no disco da
    # dev. A mesma base rendia listas diferentes, e ninguém tinha como notar.
    recent_sorted = sorted(client_deals,
                           key=lambda d: (d.get('_fdate', ''), d.get('Deal', '')),
                           reverse=True)[:50]
    recent_deals = [
        {
            'deal':    d.get('Deal', ''),
            'client':  d.get('Client', ''),
            'date':    d.get('TradeDate', '') or d.get('_fdate', ''),
            'status':  d.get('Status', ''),
            'product': d['_product'],
            'type':    d['_type'],
        }
        for d in recent_sorted
    ]

    return jsonify(_dash_stats_store(period, {
        'ndf_total':     len(ndf_deals),
        'opt_total':     len(opt_deals),
        'pending_total': pending_total,
        'swap_total':    swap_total,
        'total_deals':   len(counted_deals),
        'top5_clients':  top5_clients,
        'top5_products': top5_products,
        'top5_underlying': top5_underlying,
        'dist_ndf':      len(ndf_comm_deals),
        'dist_ndf_vanilla':  len(ndf_vanilla_deals),
        'dist_ndf_otherpub': len(ndf_otherpub_deals),
        'dist_ndf_fwdstart': len(ndf_fwdstart_deals),
        'dist_opt':      len(optcomm_deals),
        'dist_fxo':      len(fxo_deals),
        'dist_swap':     len(swap_deals),
        'monthly_opt':   monthly_opt,
        'monthly_ndf':   monthly_ndf,
        'monthly_ndf_vanilla':  monthly_ndf_vanilla,
        'monthly_ndf_otherpub': monthly_ndf_otherpub,
        'monthly_ndf_fwdstart': monthly_ndf_fwdstart,
        'monthly_fxo':   monthly_fxo,
        'monthly_swap':  monthly_swap,
        'recent_deals':  recent_deals,
    }))


# Live Position entity breakdown. The Banco (holder 73760) is a party to EVERY
# intragroup trade, so its bucket AGGREGATES all operations it faces against the
# four intragroup counterparty accounts below (its own 73760.10-2 book + Lawton +
# MGT + Atacama). Lawton/MGT/Atacama remain their own counterparty-specific tallies.
# Order fixed as Banco → Lawton → MGT → Atacama.
_LIVE_ENTITY_MAP = {
    '73760009': 'BANCO',    # holder book 73760.00-9 (mock data)
    '73760102': 'BANCO',    # Banco counterparty book 73760.10-2
    '00041007': 'LAWTON',   # 00041.00-7
    '04880006': 'MGT',      # 04880.00-6
    '85398005': 'ATACAMA',  # 85398.00-5
}
_LIVE_ENTITY_ORDER = ['BANCO', 'LAWTON', 'MGT', 'ATACAMA']
# Counterparties whose trades the BANCO bucket aggregates (all intragroup).
_LIVE_BANCO_COUNTERPARTIES = {'BANCO', 'LAWTON', 'MGT', 'ATACAMA'}

# Every standard product is listed even at 0 (mirrors the Settlement Forecast
# card), so the bar set is stable and never "loses" a product — e.g. Swap CEMHYB —
# just because the current snapshot happens to have none. COE is tracked but not
# yet counted (no logic wired) — shows 0 until the counting rule arrives.
_LIVE_PLACEHOLDER_PRODUCTS = ['NDF Moeda', 'NDF Commodities', 'Option FXO',
                             'Option Commodities', 'Option EDG',
                             'SWAP CEM', 'SWAP EDG', 'SWAP CEMHYB', 'COE']
# Fixed display order for the Live Position product bar (unknown products last).
_LIVE_PRODUCT_ORDER = {p: i for i, p in enumerate(_LIVE_PLACEHOLDER_PRODUCTS)}


def _live_map_entity(raw):
    """Like _fcst_map_entity but keeps BANCO (holder account) in the breakdown.
    Accepts the account dotted / plain / embedded (digit-key substring match)."""
    s = (raw or '').strip()
    if not s:
        return None
    digits = ''.join(ch for ch in s if ch.isdigit())
    for code, name in _LIVE_ENTITY_MAP.items():
        if code in digits:
            return name
    up = s.upper()
    for nm in _LIVE_ENTITY_ORDER:
        if nm in up:
            return nm
    return None


# One entry per B3 position (DPOSICAO*) snapshot file. Each row = one live
# operation still in custody on the reference date. Product/entity resolved by
# name token (reusing the forecast classifiers) so it survives header drift.
_LIVE_POSITION_SOURCES = [
    {'key': 'ndf', 'label': 'NDF', 'category': 'NDF',
     'file': lambda r: '73760_{}_DPOSICAO-TER.json'.format(r),
     'entity': ['titular', 'contraparte', 'parte', 'conta'],
     'product': ('ndfclass', ['classe do ativo', 'ativo subjacente', 'mercadoria', 'classe'])},
    {'key': 'opc', 'label': 'Options', 'category': 'Option',
     'file': lambda r: '73760_{}_DPOSICAO.json'.format(r),
     'entity': ['titular', 'contraparte', 'conta'],
     'product': ('optclass', ['classe do ativo subjacente', 'classe do ativo', 'classe'])},
    {'key': 'swap', 'label': 'Swap', 'category': 'Swap',
     'file': lambda r: '73760_{}_DPOSICAO-SWAP.json'.format(r),
     'entity': ['contraparte', 'titular', 'parte'],
     'product': ('lob', ['código identificador', 'codigo identificador', 'identificador'])},
]


@blueprint.route('/api/dashboard-live-position')
def api_dashboard_live_position():
    """Snapshot of open operations still in custody on a reference date, read from
    the B3 position (DPOSICAO*) JSONs. Independent of the trade-date period filter:
    it's a photo of current inventory. `date` (YYYY-MM-DD) defaults to D-1 ANBIMA."""
    if not session.get('authenticated'):
        return jsonify({'error': 'unauthorized'}), 401

    date_str = request.args.get('date')
    ref = None
    if date_str:
        try:
            ref = datetime.strptime(date_str, '%Y-%m-%d')
        except ValueError:
            ref = None
    if ref is None:
        ref = _prev_anbima_bizday(datetime.now())

    # Fallback: if the requested date has no snapshot yet, walk back (up to ~10 ANBIMA
    # business days) to the latest date that DOES have position files, so the dashboard
    # stays populated instead of showing zeros.
    def _live_has_files(dr):
        for src in _LIVE_POSITION_SOURCES:
            p = os.path.join(B3_JSON_ROOT, src['category'], _b3_date_subpath(dr), src['file'](dr))
            if os.path.isfile(p):
                return True
        return False

    probe = ref
    for _ in range(11):
        if _live_has_files(probe.strftime('%y%m%d')):
            ref = probe
            break
        probe = _prev_anbima_bizday(probe)
    dref = ref.strftime('%y%m%d')

    by_product, by_entity = {}, {}
    sources = []
    for src in _LIVE_POSITION_SOURCES:
        path = os.path.join(B3_JSON_ROOT, src['category'], _b3_date_subpath(dref), src['file'](dref))
        st = {'label': src['label'], 'file': os.path.basename(path), 'found': False, 'count': 0}
        if not os.path.isfile(path):
            sources.append(st)
            continue
        try:
            with open(path, 'r', encoding='utf-8') as fh:
                rows = json.load(fh)
        except Exception:
            sources.append(st)
            continue
        st['found'] = True
        if not rows:
            sources.append(st)
            continue
        keys = list(rows[0].keys())
        ent_key = _fcst_resolve_key(keys, src['entity'])
        pmode, pspec = src['product']
        prod_key = _fcst_resolve_key(keys, pspec)
        cnt = 0
        for row in rows:
            if pmode == 'ndfclass':
                product = _fcst_ndf_product(row.get(prod_key, '') if prod_key else '')
            elif pmode == 'optclass':
                product = _fcst_opt_class_product(row.get(prod_key, '') if prod_key else '')
            elif pmode == 'lob':
                lob = _fcst_lob(row.get(prod_key, '') if prod_key else '')
                if lob is None:
                    continue      # unclassified swap: leave uncounted, not mislabeled
                product = 'SWAP ' + lob
            else:
                product = src['label']
            by_product[product] = by_product.get(product, 0) + 1
            ent = _live_map_entity(row.get(ent_key, '')) if ent_key else None
            if ent:
                # Counterparty-specific bucket (LAWTON / MGT / ATACAMA / Banco own book).
                by_entity[ent] = by_entity.get(ent, 0) + 1
                # BANCO aggregates EVERY intragroup trade it is a party to. A row
                # already resolving to BANCO (its own 73760.10-2 book) is counted
                # once above; Lawton/MGT/Atacama rows add to BANCO on top of their
                # own tally.
                if ent in _LIVE_BANCO_COUNTERPARTIES and ent != 'BANCO':
                    by_entity['BANCO'] = by_entity.get('BANCO', 0) + 1
            cnt += 1
        st['count'] = cnt
        sources.append(st)

    # Only surface the product bar when there is real position data. COE (and any
    # other placeholder) is always shown at 0 alongside the real products.
    if by_product:
        for p in _LIVE_PLACEHOLDER_PRODUCTS:
            by_product.setdefault(p, 0)
    product_rows = [{'label': k, 'count': by_product[k]}
                    for k in sorted(by_product, key=lambda k: (_LIVE_PRODUCT_ORDER.get(k, 999), k))]
    entity_rows = [{'label': k, 'count': by_entity[k]}
                   for k in _LIVE_ENTITY_ORDER if k in by_entity]
    return jsonify({
        'ref_date':     ref.strftime('%Y-%m-%d'),
        'ref_date_fmt': ref.strftime('%d/%m/%Y'),
        'total':        sum(by_product.values()),
        'by_product':   product_rows,
        'by_entity':    entity_rows,
        'sources':      sources,
    })


@blueprint.route('/about')
def about():
    if not session.get('authenticated'):
        return redirect(url_for('pages_blueprint.sign_in_page'))
    return render_template('pages/about.html', segment='about')


@blueprint.route('/control-panel')
def control_panel():
    if not session.get('authenticated'):
        return redirect(url_for('pages_blueprint.sign_in_page'))
    cetip_default_date = _prev_anbima_bizday(datetime.now()).strftime('%Y-%m-%d')
    today = datetime.now()
    return render_template('pages/control-panel.html', segment='control-panel',
                           cetip_default_date=cetip_default_date,
                           today_date=today.strftime('%Y-%m-%d'),
                           today_fmt=today.strftime('%d/%m/%Y'))


# ============================================================================
#  CONTROL PANEL — Daily Settlement routines
# ============================================================================
#
#  Routine: "Salvar Arquivos CETIP" — Python translation of the Alteryx flow
#  (Directory → Filter → DynamicInput → Formula → Select → DbFileOutput → Email).
#
#  It reads the raw CETIP files B3 drops in the daily download folder, filters
#  them by type, renames each to the standard `73760_{YYMMDD}_{TYPE}` convention
#  and saves them into a single per-day destination folder the KPI process reads.
#  Two HTML e-mails are then sent from the OTC Tracker mailbox (best-effort):
#  one to Brazil OTC Ops and one to Brazil Sales Support MO (cc Ops).
#
#  Source folder (per run):  CETIP_SOURCE_ROOT\{YYYY}\{mm. Month}\{DD}
#  Destination folder:       CETIP_DEST_ROOT\{YYYY}\{mm. Month}\{DD}
#  (both keyed on the reference date; the file date in the rename still comes
#   from the source filename via Substring)
# ----------------------------------------------------------------------------
CETIP_SOURCE_ROOT = os.getenv('CETIP_SOURCE_ROOT', os.path.join(
    Config.SHARED_DRIVE_ROOT, 'Confirmation', 'Derivativos', 'OTC Tracker', 'Alteryx',
    'Posição B3', 'ARQUIVOS CETIP'))
CETIP_DEST_ROOT   = os.getenv('CETIP_DEST_ROOT', os.path.join(
    Config.SHARED_DRIVE_ROOT, 'Confirmation', 'Derivativos', 'OTC Tracker', 'CETIP Files',
    'Position Files'))
# Use the PRIMARY SMTP address (jpmorgan.com) — the jpmchase.com one is a
# secondary alias and the relay was not delivering to it.
CETIP_OTC_OPS_EMAIL       = os.getenv('CETIP_OTC_OPS_EMAIL',       'brazil.otc.ops@jpmorgan.com')
CETIP_SALES_SUPPORT_EMAIL = os.getenv('CETIP_SALES_SUPPORT_EMAIL', 'brazil_sales_support_mo@jpmchase.com')
# CEM Latam BA (Buenos Aires CIB Ops) — receive the Option Position .OPC file, cc OTC Ops.
# Editable TO lists for the four distribution e-mails (Sales Support / CEM Latam /
# BACC / BACC HUB EQT MO), persisted from the Save CETIP Files card. Empty/absent →
# the hardcoded defaults below; BACC e BACC HUB NÃO têm default, então uma lista
# vazia ali quer dizer "não envia", e o card diz isso. The CC (OTC Ops) stays
# hardcoded.
#
# BACC e BACC HUB são destinos DIFERENTES do mesmo time, e é por isso que são duas
# listas e dois e-mails: o BACC recebe o intragrupo RECORTADO (features/cetip),
# o HUB recebe os arquivos de posição **inteiros**, para reconciliar posição. Um
# e-mail só com os dois conjuntos entregaria a cada lado um arquivo que ele não
# pediu — e o recorte e o arquivo cheio têm o mesmo nome de origem.


CETIP_CEM_LATAM_EMAILS    = [e.strip() for e in os.getenv(
    'CETIP_CEM_LATAM_EMAILS',
    'lautaro.larriera@jpmchase.com,sacha.yebrin@jpmchase.com,candela.ferreiro@jpmorgan.com,'
    'martina.rambert@jpmchase.com,mercedes.e.mino@jpmchase.com').split(',') if e.strip()]


# Network shares for the secondary (flat) copies of two types, mirroring the
# Alteryx second outputs (commented date subfolder → flat folder).
CETIP_OPTIONS_SHARE = os.getenv('CETIP_OPTIONS_SHARE',
                                os.path.join(Config.SHARED_DRIVE_ROOT, 'CETIP_OPTIONS'))
CETIP_NDF_SHARE     = os.getenv('CETIP_NDF_SHARE',
                                os.path.join(Config.SHARED_DRIVE_ROOT, 'CETIP_NDF'))

# ── Quais arquivos a rotina Save CETIP Files considera ────────────────────────
#
# A LISTA (que arquivo entra, com que nome ele chega e com que nome é salvo) é
# cadastrada na tela /mapping → "CETIP Files"; o COMPORTAMENTO de cada tipo
# (como o arquivo vira JSON, se atualiza o VCP, a que e-mail ele é anexado) mora
# aqui, porque é lógica de parsing e não de-para. As duas metades se juntam em
# `_cetip_rules()` pela chave TYPE.
#
# Padrão dos nomes no cadastro: o literal `YYMMDD` marca ONDE a data está.
#   CETIP21_YYMMDD_DPOSICAO-SWAP   →  CETIP21_260731_DPOSICAO-SWAP.TXT
# A data que entra no nome de destino é a que veio no nome do arquivo de origem;
# a data do card do Control Panel é o que escolhe a PASTA do dia (origem e
# destino são `…\YYYY\mm. Month\dd`), e é ela que monta o nome esperado quando o
# arquivo não aparece (a linha "Not found" do e-mail).

# Seed do cadastro: exatamente os 15 tipos que estavam no código, com o padrão
# de origem reconstruído a partir do par (token do `match`, `date_start`) antigo
# — o `date_start` dizia o TAMANHO do prefixo e o token dizia o final do nome.
_CETIP_FILES_SEED = [
    {'TYPE': 'NDF Position (DPOSICAO C21)',
     'SOURCE': 'CETIP21_YYMMDD_DPOSICAO_C21',
     'DEST': '73760_YYMMDD_DPOSICAO.CETIP21', 'EXTRA DEST': ''},
    {'TYPE': 'SWAP Position (DPOSICAO-SWAP)',
     'SOURCE': 'CETIP21_YYMMDD_DPOSICAO-SWAP',
     'DEST': '73760_YYMMDD_DPOSICAO-SWAP.CETIP21', 'EXTRA DEST': ''},
    {'TYPE': 'Option Position (OPC DPOSICAO)',
     'SOURCE': 'OPC_YYMMDD_DPOSICAO',
     'DEST': '73760_YYMMDD_DPOSICAO.OPC', 'EXTRA DEST': CETIP_OPTIONS_SHARE},
    {'TYPE': 'Option Movement (OPC DMOVIMENTO)',
     'SOURCE': 'OPC_YYMMDD_DMOVIMENTO',
     'DEST': '73760_YYMMDD_DMOVIMENTO_3.OPC', 'EXTRA DEST': ''},
    {'TYPE': 'NDF Movement (DMOVIMENTO C21)',
     'SOURCE': 'CETIP21_YYMMDD_DMOVIMENTO_C21',
     'DEST': '73760_YYMMDD_DMOVIMENTO.CETIP21', 'EXTRA DEST': ''},
    {'TYPE': 'SWAP Movement (DMOVIMENTO-SWAP)',
     'SOURCE': 'CETIP21_YYMMDD_DMOVIMENTO-SWAP',
     'DEST': '73760_YYMMDD_DMOVIMENTO-SWAP.CETIP21', 'EXTRA DEST': ''},
    {'TYPE': 'SWAP Flow (DFLUXO_SWAP)',
     'SOURCE': 'CETIP21_YYMMDD_DFLUXO_SWAP',
     'DEST': '73760_YYMMDD_DFLUXO.CETIP21', 'EXTRA DEST': ''},
    {'TYPE': 'SWAP Premium Agenda (DAGENDAPREMIOS)',
     'SOURCE': 'CETIP21_YYMMDD_DAGENDAPREMIOS',
     'DEST': '73760_YYMMDD_DAGENDAPREMIOS.CETIP21', 'EXTRA DEST': ''},
    # Estratégia (MID): o quarto arquivo do BACC HUB EQT MO. TYPE, SOURCE e DEST
    # são os que já estavam no cadastro do time — o DEST dele **já termina em
    # `.txt`**, que é por que `_cetip_txt_name` não acrescenta um segundo.
    {'TYPE': 'SWAP (Strategy)',
     'SOURCE': 'CETIP21_YYMMDD_DPOSICAOESTRATEGIA_MID',
     'DEST': 'CETIP21_YYMMDD_DPOSICAOESTRATEGIA_MID.txt', 'EXTRA DEST': ''},
    {'TYPE': 'SWAP Indexers (INDEXADORESSWAP_VCP)',
     'SOURCE': 'CETIP21_YYMMDD_INDEXADORESSWAP_VCP',
     'DEST': 'CETIP21_YYMMDD_INDEXADORESSWAP_VCP.TXT', 'EXTRA DEST': ''},
    {'TYPE': 'Operations (DOPERACOES)',
     'SOURCE': 'CETIP21_YYMMDD_DOPERACOES',
     'DEST': '73760_YYMMDD_DOPERACOES.CETIP21', 'EXTRA DEST': ''},
    {'TYPE': 'COE (DRESUMOEMISSOR-COE)',
     'SOURCE': 'CETIP21_YYMMDD_DRESUMOEMISSOR-COE',
     'DEST': 'CETIP21_YYMMDD_SP_DRESUMOEMISSOR-COE.TXT', 'EXTRA DEST': ''},
    {'TYPE': 'Accelerator Agent (MID DAGENTEACELERADOR)',
     'SOURCE': 'CETIP21_YYMMDD_MID_DAGENTEACELERADOR',
     'DEST': '73760_YYMMDD_MID_DAGENTEACELERADOR.CETIP21', 'EXTRA DEST': ''},
    {'TYPE': 'NDF Position (DPOSICAO-TER)',
     'SOURCE': 'TER_YYMMDD_DPOSICAO-TER',
     'DEST': '73760_YYMMDD_DPOSICAO-TER.TER', 'EXTRA DEST': CETIP_NDF_SHARE},
    {'TYPE': 'SIC Contract Position (DPOSCONTRATOSIC)',
     'SOURCE': 'SIC_YYMMDD_DPOSCONTRATOSIC',
     'DEST': '73760_YYMMDD_DPOSCONTRATOSIC.txt', 'EXTRA DEST': ''},
    {'TYPE': 'Comitente Registry (DCADCOMITENTES)',
     'SOURCE': 'SIC_YYMMDD_DCADCOMITENTES',
     # Keep the original SIC name so the Comitente reconciliation finds it unchanged.
     'DEST': 'SIC_YYMMDD_DCADCOMITENTES.txt', 'EXTRA DEST': ''},
    # CGD: salvo na rotina e nada mais — não vira JSON e não vai para área nenhuma.
    {'TYPE': 'CGD (NET)',
     'SOURCE': 'CETIP21_YYMMDD_DPOSICAO-NET',
     'DEST': 'CETIP21_YYMMDD_DPOSICAO-NET.txt', 'EXTRA DEST': ''},
]


# ── B3 JSON export (feeds the Settlement Forecast) ────────────────────────────
# While saving the CETIP files, the relevant position files are ALSO parsed into
# tidy JSON under static/data/B3 Files/<category>/, so downstream routines (the
# Settlement Forecast) read named fields instead of guessing column positions.
#   NDF    → TER files          (DPOSICAO-TER)         — file has its own header
#   Option → OPC files          (DPOSICAO.OPC)         — file has its own header
#   Swap   → SWAP position/flow/premium agenda          — HEADERLESS: column names
#            come from _B3_SWAP_HEADERS (stored standard, keyed per file type)
B3_JSON_ROOT = data_write('cache', 'b3 files')
# Network folder the VCP / CEM / EDG / HYB source files are dropped into, per run.
# Layout: ACCRUAL_SOURCE_ROOT\{YYYY}\{mm. Month}\{DD} (run = last ANBIMA bizday of the
# month). Only reachable on the JPM environment; override with the env var off-site.

# Standard column headers for the HEADERLESS SWAP-family files (';'-delimited),
# in file order. These are the authoritative field names (the SWAP files ship with
# no header row). Stored as raw ';' strings and split on load. NOTE: the SWAP
# position layout repeats several column names (e.g. "Percentual", "Data de
# Cotação"); _b3_export_json de-duplicates repeats by appending _2, _3, …
_B3_SWAP_HEADERS_RAW = {
    # 73760_*_DPOSICAO-SWAP.CETIP21
    'swap_position': (
        "Tipo de Contrato;Data;Contrato;Participante;CPF/CNPJ Cliente Parte;Cesta Garantias Parte;"
        "Comissão Parte;Contraparte;CPF/CNPJ Cliente Contraparte;Cesta Garantias Contraparte;"
        "Comissão Contraparte;Data início;Data vencimento;Tipo de Adesão;Valor base;"
        "Valor Base Remanescente;Valor Antecipado;Saldo;Sinal Saldo;Data do Saldo;Funcionalidade;"
        "Agenda de Prêmio;Reset;Observação;Valor base inicial;Data operação termo;Índice Termo;"
        "Percentual Termo;PU Inicial;Tipo/Classe;Nome Tipo/Classe;Denominação;Juros a cada;"
        "Expresso em;Data inicio pagamento juros;Amortização a cada;Expresso em;"
        "Data inicio pagamento amortização;Tipo de amortização;Percentual;Código índice;TR Escolhida;"
        "Sinal Taxa;Taxa;Lim. Inferior (Floor);Lim. Superior (Cap);Valor Curva Atualizado;"
        "Data Correção;Fator Original de Juros;Percentual;Código índice;TR Escolhida;Sinal Taxa;Taxa;"
        "Lim. Inferior (Floor);Lim. Superior (Cap);Valor Curva Atualizado;Data Correção;"
        "Fator Original de Juros;Parte/Contraparte;Cupom Limpo;Percentual;Curva;Sinal Taxa;"
        "Taxa de Juros;Limitador;Pu inicial;Pu atual;Tipo/Classe;Nome Tipo/Classe;Denominação;"
        "Pu inicial;Pu atual;Tipo/Classe;Nome Tipo/Classe;Denominação;Cupom Limpo;Data de Cotação;"
        "Cupom Limpo;Data de Cotação;Tipo Libor - moeda;Tipo Libor - período;Data de Cotação;"
        "Variação Cambial;Tipo Classe;Nome Tipo/Classe;Outros - Cotação;Alíquota - IR;"
        "Limite inferior (FLOOR) - Perc.;Limite superior (CAP) - Perc.;Tipo Libor - moeda;"
        "Tipo Libor - período;Data de Cotação;Variação Cambial;Tipo Classe;Nome Tipo/Classe;"
        "Outros - Cotação;Alíquota - IR;Limite inferior (FLOOR) - Perc.;Limite superior (CAP) - Perc.;"
        "Taxa Juros;Troca de Fluxo;Variação Cambial;Tipo Classe;Nome Tipo/Classe;Outros - Cotação;"
        "Alíquota - IR;Limite inferior (FLOOR) - Perc.;Limite superior (CAP) - Perc.;Taxa Juros;"
        "Troca de Fluxo;Variação Cambial;Tipo Classe;Nome Tipo/Classe;Outros - Cotação;Alíquota - IR;"
        "Limite inferior (FLOOR) - Perc.;Limite superior (CAP) - Perc.;Parte/Contraparte;"
        "Fator/Valor/Taxa;Verificação;Data Disparo;Parte/Contraparte;Fator/Valor/Taxa;Verificação;"
        "Data Disparo;Titular;Prêmio 1;Rebate;Liquidação do Rebate;Dias Úteis após o Trigger Out;"
        "Prêmio 2;Data Exercício Prêmio 2;Estratégia;Amortiza sem Troca de Diferencial;"
        "Data da Cotação - Variação Cambial;Data da Cotação - Variação Cambial;Cotação Inicial;"
        "Código Commodity;Media Asiática Verificação;Data Cotação para Ajuste;Cotação Inicial;"
        "Código Commodity;Media Asiática Verificação;Data Cotação para Ajuste;Código Identificador;"
        "Data de Cotação Final – Termo;Tipo de Cotação (Parte);Tipo de Cotação (Contraparte);"
        "Data Liquidação;Cotação Inicial Moeda Parte;Metodologia de composição da taxa Parte;"
        "Deslocamento da taxa Parte;Expressão Juros Parte;Alíquota IR (em %) Parte;"
        "Cotação Inicial Moeda Contraparte;Metodologia de composição da taxa Contraparte;"
        "Deslocamento da taxa Contraparte;Expressão Juros Contraparte;Alíquota IR (em %) Contraparte;"
        "Data de Fixing IPCA (Parte);Data de Fixing IPCA (Contraparte);Sinal Spread (Parte);"
        "Spread (Parte);Sinal Spread (Contraparte);Spread (Contraparte);Variação Cambial;"
        "Cotação Inicial Moeda;Variação Cambial;Cotação Inicial Moeda"
    ),
    # 73760_*_DFLUXO.CETIP21
    'swap_fluxo': (
        "Código do contrato;Tipo Sistema;Código Conta Cetip Parte;Nome Simplificado Parte;"
        "Papel Parte;Código Conta Cetip Contraparte;Nome Simplificado Contraparte;Papel Contraparte;"
        "Tipo Amortização;Data Pagamento de Juros;Código Identificador;Data de ocorrência do Evento;"
        "Sinal Juros Parte;Taxa de Juros Parte;Limite Inferior Parte;Limite Superior Parte;"
        "Taxa Amortização;Sinal Juros Contraparte;Taxa de Juros Contraparte;Limite Inferior Contraparte;"
        "Limite Superior Contraparte;Taxa Amortização;Data Início Composição da Taxa Parte;"
        "Data Final Composição da Taxa Parte;Data Fixing Moeda Parte;"
        "Data Início Composição da Taxa Contraparte"
    ),
    # 73760_*_DAGENDAPREMIOS.CETIP21
    'swap_premio': (
        "Codigo do Contrato;Data;ID do Sistema;Parte;Nome Simplificado;Data do Evento;"
        "Operacao;Valor;Titular;Estado"
    ),
    # 73760_*_DOPERACOES.CETIP21
    'operations': (
        "Participante (Nome Simpl.);Conta;Liquidante;Cod.Operacao;Tipo Operacao;C/V;"
        "Tipo Compra/Venda;Titulo;Codigo IF Anterior;Tipo Titulo;Data Emissao;Data Vencimento;"
        "Quantidade;PU;Valor;Tx Colocacao;Sistema;Modalidade Liquidacao;Status;Numero Operacao;"
        "Numero Associacao;Data Liquidacao;Data Origem;Instituicao Confirmadora(Conta);"
        "Instituicao Confirmadora(Papel);Contraparte (Nome Simpl.);Conta Contraparte;"
        "Data Compromisso;PU/Ida Compromisso;Numero Operacao Original;"
        "Data da Operacao Original/Data Operacao Original da Antecipacao;PU Op Original;"
        "Qtd Op Original;ISPB Liq. Contraparte;Nu Op Msg;Num Ctrl Operacao;Programa de Emissao"
    ),
}
_B3_SWAP_HEADERS = {k: [h.strip() for h in v.split(';')]
                    for k, v in _B3_SWAP_HEADERS_RAW.items()}


def _b3_date_subpath(dref):
    """YYMMDD ref → 'YYYY/MM/DD' subfolders so the per-day JSON files are split by
    year/month/day inside each product folder. '' if the ref can't be parsed."""
    try:
        d = datetime.strptime(dref, '%y%m%d')
    except (ValueError, TypeError):
        return ''
    return os.path.join(d.strftime('%Y'), d.strftime('%m'), d.strftime('%d'))


def _b3_dref_to_iso(dref):
    """YYMMDD → 'YYYY-MM-DD' ('' quando não há ref). É a data do ARQUIVO que foi
    lido, e ela vai no payload das telas de posição para quem monta uma série
    saber que dia está olhando de verdade."""
    try:
        return datetime.strptime(dref, '%y%m%d').strftime('%Y-%m-%d')
    except (ValueError, TypeError):
        return ''


def _b3_filter_rows(rows, filt):
    """Keep only rows whose value in one column is in `filt['allowed']`.
    filt = {'column': [name-substrings], 'index': N, 'allowed': [...],
            'match': 'digits' (default) | 'text'}. The column is resolved by the
    first header whose lower-case name contains one of the tokens, falling back to
    the positional `index`. 'digits' compares digit-only (account numbers);
    'text' compares trimmed, upper-cased (e.g. Tipo Titulo ∈ TER/SWAP/OPC).
    Returns rows unchanged if the column can't be resolved (fail-open)."""
    if not filt or not rows:
        return rows
    keys = list(rows[0].keys())
    col_key = None
    for tok in filt.get('column', []):
        for k in keys:
            if tok in k.lower():
                col_key = k
                break
        if col_key:
            break
    if col_key is None and 'index' in filt:
        ix = filt['index']
        col_key = keys[ix] if 0 <= ix < len(keys) else None
    if filt.get('match') == 'text':
        _norm = lambda s: str(s).strip().upper()
    else:
        _norm = lambda s: ''.join(ch for ch in str(s) if ch.isdigit())
    allowed = set(_norm(a) for a in filt.get('allowed', []))
    if col_key and allowed:
        before = len(rows)
        rows = [r for r in rows if _norm(r.get(col_key, '')) in allowed]
        log.info("[b3-json] filter on %r (%s) kept %d/%d rows",
                 col_key, filt.get('match', 'digits'), len(rows), before)
    elif not col_key:
        log.warning("[b3-json] filter column not found (tokens=%s, index=%s)",
                    filt.get('column'), filt.get('index'))
    return rows


def _b3_export_json(src_path, json_cfg, dest_name, dref, skip_existing=False):
    """Parse a saved CETIP file into a list-of-dicts JSON under
    B3 Files/<category>/YYYY/MM/DD/<dest_name>.json. Header files use their own
    first line; headerless files use the stored standard header
    (_B3_SWAP_HEADERS) or positional Field_N names. Best-effort — returns the
    JSON path on success or None on failure. When skip_existing is True an
    already-created JSON for that day is left untouched (used by the backfill)."""
    try:
        out_dir = os.path.join(B3_JSON_ROOT, json_cfg['category'], _b3_date_subpath(dref))
        json_name = os.path.splitext(dest_name)[0] + '.json'
        json_path = os.path.join(out_dir, json_name)
        if skip_existing and os.path.exists(json_path):
            return json_path

        with open(src_path, 'r', encoding='latin-1', newline='') as fh:
            lines = [ln for ln in fh.read().splitlines() if ln.strip()]
        if not lines:
            return None

        if json_cfg.get('has_header'):
            header = [h.strip() for h in lines[0].split(';')]
            data_lines = lines[1:]
        else:
            header = list(_B3_SWAP_HEADERS.get(json_cfg.get('header_key', ''), []) or [])
            data_lines = lines

        # De-duplicate repeated header names (SWAP position repeats 38 of its
        # 170) so no field is silently overwritten: 1st keeps its name, repeats
        # get _2, _3… A regra é a MESMA do `nomes_unicos` do conversor, com a
        # igualdade trocada: aqui é a CHAVE do JSON, onde `PU Inicial` e
        # `Pu inicial` são dois campos e têm de continuar sendo — o desempate
        # por CAIXA é do nome de coluna do DuckDB, e acontece lá.
        #
        # O que a função traz de novo é conferir o candidato contra TODOS os
        # nomes do header: um layout com `X` duas vezes E uma coluna chamada
        # `X_2` produzia dois `X_2`, e no dicionário o segundo apagava o
        # primeiro — uma coluna perdida sem erro nenhum.
        from apps.pages.json_to_duckdb import nomes_unicos
        uniq_header = nomes_unicos(header, chave=lambda s: s) if header else []

        rows = []
        for ln in data_lines:
            fields = ln.split(';')
            if uniq_header:
                row = {}
                for i, val in enumerate(fields):
                    key = uniq_header[i] if i < len(uniq_header) else 'Field_{}'.format(i + 1)
                    row[key] = val.strip()
            else:                              # no stored header → positional names
                row = {'Field_{}'.format(i + 1): v.strip() for i, v in enumerate(fields)}
            rows.append(row)

        # Optional row filters: a single `filter` (back-compat) plus a `filters`
        # list, all applied as AND (keep only rows allowed by every filter). Each
        # filter keeps rows whose value in one column is in its `allowed` set —
        # e.g. DOPERACOES keeps Conta ∈ {73760009, 04880006} AND Tipo Titulo ∈
        # {TER, SWAP, OPC}.
        for filt in ([json_cfg['filter']] if json_cfg.get('filter') else []) + list(json_cfg.get('filters') or []):
            rows = _b3_filter_rows(rows, filt)

        os.makedirs(out_dir, exist_ok=True)
        _atomic_write_json(json_path, rows)     # funil: atômico + espelho (§335)
        return json_path
    except Exception:
        log.warning("[b3-json] export failed for %s:\n%s", src_path, traceback.format_exc())
        return None


# Existing VCP qualification table (Descrição/Classificação/STATUS per Qualification
# ID) — the Save CETIP Files routine refreshes it in place from the
# INDEXADORESSWAP_VCP file. Also read by the Swap Characteristics page and index-b3.
VCP_JSON = data_path('VCP.json')


# ── BACC: os quatro arquivos, recortados para o INTRAGRUPO ───────────────────
#  O BACC recebe DFLUXO swap, posição swap, posição OPC e posição TER — mas não o
#  arquivo cheio: só as linhas em que PARTE e CONTRAPARTE são as três contas de
#  casa. Uma linha com só um dos lados intragrupo é operação com cliente e não
#  entra (o `and` é o pedido; um `or` mandaria a carteira inteira).
#
#  Onde está cada conta em cada arquivo mora no `_CETIP_BEHAVIOUR['…']['bacc']`,
#  ao lado do resto do comportamento do tipo.


# ============================================================================
#  SETTLEMENT FORECAST  (Alteryx "Settlement Forecast v2" → Python)
# ----------------------------------------------------------------------------
#  Reads the tidy JSON emitted by the File-Saving routine (B3 Files/<category>/),
#  projects the upcoming settlements per business day broken down by product and
#  by entity, and returns the data to the page. The page renders dashboard-style
#  ApexCharts and (on Run) exports them to PNG, which the e-mail endpoint embeds
#  into the report sent to Brazil OTC Ops.
# ============================================================================

# ── motor do Forecast: movido para platform/forecast.py (§317) ───────────────
# Os nomes ficam como ALIAS — features, a família de liquidação e os testes
# seguem alcançando por `routes.<nome>`, resolvido em tempo de chamada.
from apps.pages.platform import forecast as _pf_fcst  # noqa: E402
FORECAST_BIZDAYS = _pf_fcst.FORECAST_BIZDAYS
FORECAST_RANGE_CHOICES = _pf_fcst.FORECAST_RANGE_CHOICES
_FCST_ENTITY_MAP = _pf_fcst._FCST_ENTITY_MAP
_FCST_ENTITY_ORDER = _pf_fcst._FCST_ENTITY_ORDER
_FCST_PRODUCT_ORDER = _pf_fcst._FCST_PRODUCT_ORDER
_FORECAST_SOURCES = _pf_fcst._FORECAST_SOURCES
_FCST_MONTH_ABBR = _pf_fcst._FCST_MONTH_ABBR
_fcst_parse_date = _pf_fcst._fcst_parse_date
_fcst_map_entity = _pf_fcst._fcst_map_entity
_fcst_option_product = _pf_fcst._fcst_option_product
_fcst_ndf_product = _pf_fcst._fcst_ndf_product
_fcst_opt_class_product = _pf_fcst._fcst_opt_class_product
_fcst_lob = _pf_fcst._fcst_lob
_forecast_spine = _pf_fcst._forecast_spine
_fcst_norm = _pf_fcst._fcst_norm
_fcst_resolve_key = _pf_fcst._fcst_resolve_key
_fcst_norm_contract = _pf_fcst._fcst_norm_contract
_swap_contract_ident_map = _pf_fcst._swap_contract_ident_map
_swap_contract_cpty_map = _pf_fcst._swap_contract_cpty_map
_forecast_collect = _pf_fcst._forecast_collect
_forecast_matrix = _pf_fcst._forecast_matrix
_forecast_payload = _pf_fcst._forecast_payload
_forecast_has_files = _pf_fcst._forecast_has_files
_forecast_latest_ref = _pf_fcst._forecast_latest_ref


def _decode_data_uri(d):
    """Decode a 'data:image/png;base64,...' URI into raw bytes (or None)."""
    if not d:
        return None
    try:
        if ',' in d:
            d = d.split(',', 1)[1]
        return base64.b64decode(d)
    except Exception:
        return None


# ──────────────────────────────────────────────────────────────────────────
# Control Panel — Daily Metric: Outstanding Confirmation Brazil OTC
# Emails a daily metric to a SAVED recipient list (TO/CC/BCC persisted on disk,
# so they don't have to be retyped each run). The reference date is always today.
# NOTE: the metric body is a placeholder for now — only the delivery plumbing and
# the persisted-recipients flow are wired; the actual metric content (source +
# format) is still to be defined.
# ──────────────────────────────────────────────────────────────────────────
_DAILY_METRIC_DIR = os.path.normpath(os.path.join(
    data_dir(), 'control-panel'))
# Settlement Forecast keeps its own saved TO/CC (same folder, no BCC) so the card
# recipients replace the previously hardcoded OTC Ops / accrual-cc addresses.


_parse_emails = _pf_mail._parse_emails


# ── Pending Confirmation: motor movido para platform/pending_confirmation.py (§318)
# Os nomes ficam como ALIAS; _PC_DB_DIR e _B3_DATA_DIR ficam AQUI (superficie
# de patch dos testes — a platform le por routes.<nome>).
from apps.pages.platform import pending_confirmation as _pf_pc  # noqa: E402
_pc_refdata_lookup = _pf_pc._pc_refdata_lookup
_PC_DBS = _pf_pc._PC_DBS
_PC_TABLE = _pf_pc._PC_TABLE
_PC_COLUMNS = _pf_pc._PC_COLUMNS
_PC_SNAPSHOT_DIR = _pf_pc._PC_SNAPSHOT_DIR
_pc_norm = _pf_pc._pc_norm
_pc_category_from_filters = _pf_pc._pc_category_from_filters
_pc_ensure_db = _pf_pc._pc_ensure_db
_pc_load_rows = _pf_pc._pc_load_rows
_PC_DERIVED_COLUMNS = _pf_pc._PC_DERIVED_COLUMNS
_pc_derive_row = _pf_pc._pc_derive_row
_PC_UPDATE_HEADERS = _pf_pc._PC_UPDATE_HEADERS
_pc_signature_pending_status = _pf_pc._pc_signature_pending_status
_pc_signature_status = _pf_pc._pc_signature_status
_pc_import_update = _pf_pc._pc_import_update
_pc_is_intragroup = _pf_pc._pc_is_intragroup
_pc_refdata_by_name = _pf_pc._pc_refdata_by_name
_pc_is_internal_counterparty = _pf_pc._pc_is_internal_counterparty
_pc_aging_band_label = _pf_pc._pc_aging_band_label
_pc_banker_for_spn = _pf_pc._pc_banker_for_spn
_PC_OK_STATUSES = _pf_pc._PC_OK_STATUSES
_PC_PASTDUE_STATUS = _pf_pc._PC_PASTDUE_STATUS
_PC_TENOR_EXCEPTION = _pf_pc._PC_TENOR_EXCEPTION
_PC_INTERNAL_EXCEPTION = _pf_pc._PC_INTERNAL_EXCEPTION
_PC_ESTEIRA_STATUSES = _pf_pc._PC_ESTEIRA_STATUSES
_pc_is_esteira_status = _pf_pc._pc_is_esteira_status
_pc_is_ok_status = _pf_pc._pc_is_ok_status
_pc_cutoff_date = _pf_pc._pc_cutoff_date
_pc_apply_auto_rules = _pf_pc._pc_apply_auto_rules
_pc_refresh_aging_status = _pf_pc._pc_refresh_aging_status
_pc_target_category = _pf_pc._pc_target_category
_pc_write_exec = _pf_pc._pc_write_exec
_pc_delete_tn = _pf_pc._pc_delete_tn
_pc_insert_into = _pf_pc._pc_insert_into
_pc_upsert_row = _pf_pc._pc_upsert_row
_pc_rewrite_db = _pf_pc._pc_rewrite_db
_pc_snapshot_pending = _pf_pc._pc_snapshot_pending
_pc_run_daily_maintenance = _pf_pc._pc_run_daily_maintenance
_PC_DAILY_TIME = _pf_pc._PC_DAILY_TIME
_pc_scheduler_started = _pf_pc._pc_scheduler_started
_pc_scheduler_lock = _pf_pc._pc_scheduler_lock
_pc_scheduler_loop = _pf_pc._pc_scheduler_loop
_pc_start_scheduler = _pf_pc._pc_start_scheduler
_pc_refdata_enrich = _pf_pc._pc_refdata_enrich
_pc_save_from_deal = _pf_pc._pc_save_from_deal
_PC_METRICS_AGING_THRESHOLD = _pf_pc._PC_METRICS_AGING_THRESHOLD
_PC_METRICS_HISTORY_FILE = _pf_pc._PC_METRICS_HISTORY_FILE
_pc_metrics_int = _pf_pc._pc_metrics_int
_pc_latest_snapshot_rows = _pf_pc._pc_latest_snapshot_rows
_pc_metrics_offenders = _pf_pc._pc_metrics_offenders
_pc_metrics_history = _pf_pc._pc_metrics_history


# ══════════════════════════════════════════════════════════════════════════
# Control Panel — Pending Signature Confirmations (Collection / "Cobrança")
# ══════════════════════════════════════════════════════════════════════════
# Segregates the pending-signature confirmations (base: Pending Confirmation page)
# by counterparty and builds one editable .eml draft per counterparty (zipped when
# many). Mirrors the legacy Excel "MassEmail" macro but generates review drafts
# instead of auto-sending. To = counterparty confirmation contacts (Counterparty
# Details); Cc = that counterparty's bankers (registry `bankers-email`, i.e.
# Mapping > Bankers E-mails, matched to the RefData BANKER group) + Brazil OTC
# Ops + IS Trade Doc.


# ── Counterparty Details: movido para platform/counterparty.py (§316) ────────
# Os nomes ficam como ALIAS — features e testes seguem alcançando por
# `routes.<nome>`, e todo chamador resolve o atributo em tempo de chamada.
from apps.pages.platform import counterparty as _pf_cpd  # noqa: E402
_CONTACTS_DATA_START_ROW = _pf_cpd._CONTACTS_DATA_START_ROW
_CC_SPN = _pf_cpd._CC_SPN
_CC_NAME = _pf_cpd._CC_NAME
_CC_ACTIVE = _pf_cpd._CC_ACTIVE
_CC_CONTACT = _pf_cpd._CC_CONTACT
_CC_PHONE = _pf_cpd._CC_PHONE
_CC_EMAIL = _pf_cpd._CC_EMAIL
_CC_RULE = _pf_cpd._CC_RULE
_CONTACT_RULE_MAP = _pf_cpd._CONTACT_RULE_MAP
_cc_cell = _pf_cpd._cc_cell
_CC_EMAIL_RE = _pf_cpd._CC_EMAIL_RE
_CC_PLACEHOLDER_TOKENS = _pf_cpd._CC_PLACEHOLDER_TOKENS
_CC_PLACEHOLDER_DOMAINS = _pf_cpd._CC_PLACEHOLDER_DOMAINS
_cc_is_placeholder_token = _pf_cpd._cc_is_placeholder_token
_cc_email_is_usable = _pf_cpd._cc_email_is_usable
_cc_drop_placeholder_contacts = _pf_cpd._cc_drop_placeholder_contacts
_cc_parse_rules = _pf_cpd._cc_parse_rules
_cc_read_rows = _pf_cpd._cc_read_rows
_cpd_path = _pf_cpd._cpd_path
_cpd_load = _pf_cpd._cpd_load
_norm_spn = _pf_cpd._norm_spn
_cpd_find = _pf_cpd._cpd_find
_cpd_save_list = _pf_cpd._cpd_save_list
_contacts_norm = _pf_cpd._contacts_norm
_net_norm = _pf_cpd._net_norm
_default_slot = _pf_cpd._default_slot
_bank_norm = _pf_cpd._bank_norm
_cgd_norm = _pf_cpd._cgd_norm
_CP_NET_TYPES = _pf_cpd._CP_NET_TYPES


# Network folder scanned for Daily Settlement source files when the dropzone is
# left empty (see api_cp_daily_settlement_save).
SETTLEMENTS_ROOT = os.getenv('SETTLEMENTS_ROOT', os.path.join(
    Config.SHARED_DRIVE_ROOT, 'Confirmation', 'Derivativos', 'OTC Tracker', 'Settlements'))

# Text-import specs translated from the VBA "ImportarTexto" (OpenText, TAB
# delimited) — one per source file, EXCLUDING Settlement OTM (done on its own
# page). Each cashflows/blotter file is a tab-delimited export; we read it, keep
# the header row, filter the data rows and write a per-type JSON. The Excel
# XLOOKUP enrichment (Tipo/Contraparte columns) is NOT part of the text import
# and is left out.
#   header  : 1-based row that holds the column names
#   filters : list of (kind, col, allowed) applied to each data row (ALL must pass)
#             kind 'digits' → compare digits-only cell; 'set' → compare UPPER cell
#   json    : output base name (…_YYYYMMDD.json under the daily-settlement folder)
# Kapital Hybrids (Other Products › Swap): BANCO_UPCOMING_PAYMENTS.csv is comma-
# delimited and needs a today()-Settlement-Date filter + per-trade aggregation, so
# it is handled by its own extractor (_swaphyb_*) rather than the generic tab path.
_SWAPHYB_JSON = 'swap-kapital-hybrids'

_DS_IMPORTS = [
    # OTM cashflows — same output as the OTM Settlements page (18 columns, its own
    # cleaning); processed here too so the card handles every file. `otm` flag →
    # _ds_handle uses _otm_extract + the OTM json path.
    {'key': 'otm', 'label': 'OTM Settlements', 'json': 'otm-settlement', 'header': 1,
     'match': lambda n: n.startswith('cashflows'), 'otm': True, 'filters': []},
    {'key': 'ndfc', 'label': 'NDF Cockpit', 'json': 'ndf-cockpit', 'header': 4,
     'match': lambda n: n.startswith('settlement') and n.endswith('.xlsx'), 'ndfc': True, 'filters': []},
    # As contas do Banco J.P. Morgan que entram (coluna Conta): a PRÓPRIA
    # (73760.00-9) e a de CLIENTE 2 (73760.20-5) — as duas estão no cadastro
    # `b3-accounts`, que é onde se lê o que cada uma é. A de CLIENTE 1
    # (73760.10-2) fica de fora, como sempre esteve.
    #
    # O que entra aqui alcança mais do que o card: este é o JSON que a página
    # Operations B3 lê (`opb3`), e é dele que saem a mensageria, os avisos de
    # liquidação e os cards de reconciliação. Uma conta que não passa por este
    # filtro não existe para nenhum deles — some sem erro nenhum.
    {'key': 'operacoes-jpm', 'label': 'Operações JPM', 'json': 'operacoes-jpm', 'header': 5,
     'match': lambda n: n.startswith('operacoes'),
     'filters': [('digits', 2, {'73760009', '73760205'}),
                 ('set', 10, {'OPC', 'OFVC', 'OFCC', 'SWAP', 'TER', 'COE'})],
     'opb3': True},                                 # also feeds the Operations B3 page json
    {'key': 'operacoes-mgt', 'label': 'Operações MGT', 'json': 'operacoes-mgt', 'header': 5,
     'match': lambda n: n.startswith('mgt.'),
     'filters': [('digits', 2, {'04880006'}),
                 ('set', 10, {'OPC', 'OFVC', 'OFCC', 'SWAP', 'TER', 'COE'})],
     'opb3': True},                                 # MGT operations ALSO feed the Operations B3 page json
    {'key': 'eventos-swap-jpm', 'label': 'Eventos Swap', 'json': 'eventos-swap-jpm', 'header': 7,
     'match': lambda n: n.startswith('swap-instrumentofinanceiro-consultacontrato'),
     'filters': [('set', 2, {'CONFIRMADO'}), ('digits', 23, {'73760009'})],
     'strip_hash': [1]},                            # remove "#" dos IDs da coluna A
    {'key': 'eventos-swap-mgt', 'label': 'Eventos Swap MGT', 'json': 'eventos-swap-mgt', 'header': 7,
     'match': lambda n: n.startswith('swapmgt.'),
     'filters': [('set', 2, {'CONFIRMADO'}), ('digits', 23, {'04880006'})],
     'strip_hash': [1]},                            # remove "#" dos IDs da coluna A
    {'key': 'cognos', 'label': 'Cognos (FXO Detail)', 'json': 'cognos', 'header': 1,
     'match': lambda n: n.startswith('fxo detail'),
     'cog': True, 'filters': [], 'skip_no_data': True},   # feeds the Cognos page
    # Opened via Workbooks.Open in the VBA (real workbook / .txt) — header on row 1.
    {'key': 'br-onshore', 'label': 'BR Onshore Settlements', 'json': 'br-onshore-settlements',
     # Header on ROW 2 (values from row 3). "Owner Legal Entity" (col 3) rows that
     # are LAWTON MULTIMERCADO EXCLUSIVO* are excluded from the JSON.
     'header': 2, 'match': lambda n: n.startswith('brazilonshoresettlementswarningfile'),
     'filters': [('not_startswith', 3, {'LAWTON MULTIMERCADO EXCLUSIVO'})]},
    # FbiRptLatamDeskPostion-NY-* → página Latam Desk Position (o nome do arquivo tem
    # o typo "Postion"; a página é "Position"). `latam` → _ds_handle usa o extractor
    # da página (_latam_extract: filtro das colunas 62/63 da macro, colunas do
    # relatório e datas dd/mm/yyyy), então card e página gravam o MESMO JSON.
    {'key': 'latam-desk', 'label': 'Latam Desk Position', 'json': 'latam-desk-position',
     'header': 1, 'match': lambda n: n.startswith('fbirptlatamdeskpo'),
     'latam': True, 'filters': []},
    # Kapital Hybrids — BANCO_UPCOMING_PAYMENTS.csv (comma-delimited). Own extractor
    # (_swaphyb_extract) filters Settlement Date = today; the page aggregates per trade.
    {'key': 'swap-kapital-hybrids', 'label': 'Swap Kapital Hybrids', 'json': _SWAPHYB_JSON,
     'match': lambda n: n.startswith('banco_upcoming_payments') and n.endswith('.csv'),
     'swaphyb': True, 'filters': []},
]


def _ds_read_rows(raw):
    """Rows from a Daily Settlement source file — tab-delimited text (as the VBA
    OpenText Tab:=True treats them) with a real-.xlsx (zip) fallback."""
    if raw[:2] == b'PK':
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        return [list(r) for r in ws.iter_rows(values_only=True)]
    return [ln.split('\t') for ln in raw.decode('latin-1').splitlines()]


def _ds_cell(row, i):
    return '' if (i < 0 or i >= len(row) or row[i] is None) else str(row[i]).strip()


def _ds_match_spec(name):
    n = (name or '').lower()
    for spec in _DS_IMPORTS:
        if spec['match'](n):
            return spec
    return None


def _ds_process(raw, spec):
    """Apply one spec: header row + row filters → list of dicts (kept rows). Returns
    (records, total_data_rows)."""
    rows = _ds_read_rows(raw)
    hidx = spec['header'] - 1
    if len(rows) <= hidx:
        return [], 0
    if spec.get('skip_no_data') and _ds_cell(rows[hidx], 0).lower().startswith('no data available'):
        return [], 0
    raw_header = rows[hidx]
    seen, header = {}, []
    for i in range(len(raw_header)):
        h = _ds_cell(raw_header, i) or 'Field_{}'.format(i + 1)
        seen[h] = seen.get(h, 0) + 1
        header.append(h if seen[h] == 1 else '{}_{}'.format(h, seen[h]))
    out, total = [], 0
    for row in rows[hidx + 1:]:
        if not any(_ds_cell(row, i) for i in range(len(row))):
            continue                                   # skip fully-blank lines
        total += 1
        keep = True
        for f in spec['filters']:
            if f[0] == 'nonempty_any':           # keep if ANY of the listed cols is non-empty
                if not any(_ds_cell(row, c - 1) for c in f[1]):
                    keep = False
                    break
                continue
            if f[0] == 'not_startswith':          # drop if col starts with ANY listed prefix
                cell = _ds_cell(row, f[1] - 1).upper()
                if any(cell.startswith(pfx.upper()) for pfx in f[2]):
                    keep = False
                    break
                continue
            kind, col, allowed = f
            v = _ds_cell(row, col - 1)
            if kind == 'digits':
                if ''.join(ch for ch in v if ch.isdigit()) not in allowed:
                    keep = False
                    break
            elif v.upper() not in allowed:
                keep = False
                break
        if not keep:
            continue
        strip_cols = {c - 1 for c in spec.get('strip_hash', [])}   # 1-based → 0-based
        rec = {}
        for i in range(len(row)):
            key = header[i] if i < len(header) else 'Field_{}'.format(i + 1)
            val = _ds_cell(row, i)
            if i in strip_cols:
                val = val.replace('#', '')
            rec[key] = val
        out.append(rec)
    return out, total


def _ds_handle(name, raw, delete_path, ref, processed, skipped):
    spec = _ds_match_spec(name)
    if not spec:
        skipped.append(name)
        return
    try:
        if spec.get('otm'):                            # OTM cashflows → OTM page's own logic + path
            rows = _ds_read_rows(raw)
            recs, kept, deleted, filtered = _otm_extract(rows) if len(rows) >= 2 else ([], 0, 0, 0)
            total = kept + deleted + filtered
            jp = _otm_json_path(ref)
        elif spec.get('ndfc'):                         # SETTLEMENT.xlsx → NDF Cockpit page's logic + path
            rows = _ndfc_read_rows(raw)
            recs, kept = _ndfc_extract(rows)
            total = kept
            jp = _ndfc_json_path(ref)
        elif spec.get('cog'):                          # FXO Detail → Cognos page's logic + path
            recs, kept = _cog_extract(_cog_read_rows(raw))
            total = kept
            jp = _cog_json_path(ref)
        elif spec.get('swaphyb'):                       # BANCO_UPCOMING_PAYMENTS.csv → Kapital Hybrids
            recs, total = _swaphyb_extract(raw, ref)    # filter Settlement Date = today (import date)
            jp = _ds_display_json_path(ref, _SWAPHYB_JSON)
        elif spec.get('latam'):                         # FbiRptLatamDeskPostion → Latam Desk Position
            rows, fmt = _latam_read_rows(raw)
            recs, kept, filtered, _cmap, missing = (_latam_extract(rows) if len(rows) >= 2
                                                    else ([], 0, 0, {}, []))
            total = kept + filtered
            jp = _latam_json_path(ref)
            log.info('[latam] %s lido como %s: %d linha(s), %d filtrada(s)',
                     name, fmt, kept, filtered)
            if missing:
                log.warning('[latam] colunas não encontradas no header de %s: %s',
                            name, ', '.join(missing))
        else:
            recs, total = _ds_process(raw, spec)
            jp = os.path.join(OTM_JSON_ROOT, ref.strftime('%Y'), ref.strftime('%m'), ref.strftime('%d'),
                              '{}_{}.json'.format(spec['json'], ref.strftime('%Y%m%d')))
    except Exception:
        log.warning("[ds] process failed for %s:\n%s", name, traceback.format_exc())
        skipped.append(name)
        return
    _ds_write(jp, recs, name, spec, total, processed, delete_path)
    if spec.get('latam'):                              # guarda também o arquivo de origem
        _latam_write_meta(jp, ref.strftime('%H:%M:%S'), name)
    elif spec.get('otm') or spec.get('ndfc') or spec.get('cog') or spec.get('swaphyb'):   # timestamp = import time (no in-file time)
        _ds_write_updated(jp, ref.strftime('%H:%M:%S'))
    if spec.get('opb3'):                               # operacoes file ALSO feeds the Operations B3 page
        try:                                           # use the FILTERED recs (processed rows), not the raw file;
            _opb3_side_write(recs, raw, ref, spec['key'])   # merge by source so JPM + MGT coexist
        except Exception:
            log.warning("[ds] operations-b3 side-write failed for %s:\n%s", name, traceback.format_exc())


def _ds_write(jp, recs, name, spec, total, processed, delete_path):
    os.makedirs(os.path.dirname(jp), exist_ok=True)
    # Pelo FUNIL (auditoria §335): o bump vem junto, e o espelho DuckDB fica
    # sabendo da gravação.
    _atomic_write_json(jp, recs)
    processed.append({'file': name, 'type': spec['label'], 'kept': len(recs), 'total': total})
    if delete_path:                                    # mirror the VBA Kill (folder source only)
        try:
            os.remove(delete_path)
        except OSError:
            log.warning("[ds] could not delete %s", delete_path)


# ============================================================================
#  ACCRUAL — SWAP (VCP) classification by line of business
# ----------------------------------------------------------------------------
#  Reads the uploaded "Swap-IntrumentoFincaneiro-ConsultaContratoVCPSemPU"
#  spreadsheet (headers on row 9), cleans column A (# -> ''), keeps only the two
#  house accounts in column K, then joins each contract (col A) against the
#  latest saved SWAP position JSON to read its "Código Identificador" and route
#  the row to its line-of-business table (CEM / EDG / Hybrids / Commodities).
#  Each table keeps columns A, F, K, L, N, Q, R, T.
# ============================================================================


# Fixed table columns (always shown, independent of the imported file).
# Source file column (0-based) for each fixed display column; None = placeholder
# (filled later by the grab logic / edited by hand). 0 = Código IF = col A ('#' stripped).
#   A, F, G, K, L, N, Q, R, T, —(Fator Parte), —(Fator Contra), —(Comments)


def _acc_digits(s):
    return re.sub(r'\D', '', str(s or ''))


def _accrual_lob(identifier):
    """Map a SWAP 'Código Identificador' to one of the four LOB buckets.
    Order matters: hybrid / COMM are tested before the CEM / EDG substrings.
    Accent-insensitive and tolerant of PT/EN hybrid spellings (HYB / HIB /
    HÍBRIDO), mirroring _fcst_lob."""
    s = _fcst_norm(identifier)   # lower-case + accent-stripped
    if 'hyb' in s or 'hib' in s:  return 'Hybrids'
    if 'comm' in s:               return 'Commodities'
    if 'edg' in s:                return 'EDG'
    if 'cem' in s:                return 'CEM'
    return None


def _swap_pos_latest_records(max_back=15):
    """Latest available SWAP position JSON (list-of-dicts) + its ref date 'YYYY-MM-DD'.
    Walks back from D-1 ANBIMA until the DPOSICAO-SWAP.json file exists."""
    ref = _prev_anbima_bizday(datetime.now())
    for _ in range(max_back):
        dref = ref.strftime('%y%m%d')
        path = os.path.join(B3_JSON_ROOT, 'Swap', _b3_date_subpath(dref),
                            '73760_{}_DPOSICAO-SWAP.json'.format(dref))
        if os.path.isfile(path):
            try:
                with open(path, 'r', encoding='utf-8') as fh:
                    return json.load(fh), ref.strftime('%Y-%m-%d')
            except Exception:
                log.error('[accrual] failed reading %s:\n%s', path, traceback.format_exc())
                return [], None
        ref = _prev_anbima_bizday(ref)
    return [], None


def _swap_pos_lob_map(records):
    """Build {contract -> identifier} from the SWAP position records, keyed by the
    upper-cased contract AND a digits-only fallback ('#'+digits)."""
    cmap = {}
    if not records:
        return cmap
    keys = list(records[0].keys())
    # 'Contrato' must not collide with 'Tipo de Contrato'; prefer the exact key.
    k_contract = 'Contrato' if 'Contrato' in keys else _fcst_resolve_key(
        [k for k in keys if _fcst_norm(k) != 'tipo de contrato'], ['contrato'])
    k_lob = ('Código Identificador' if 'Código Identificador' in keys
             else _fcst_resolve_key(keys, ['codigo identificador']))
    if not k_contract or not k_lob:
        return cmap
    for rec in records:
        c = str(rec.get(k_contract, '') or '').strip()
        if not c:
            continue
        ident = str(rec.get(k_lob, '') or '').strip()
        cmap.setdefault(c.upper(), ident)
        dg = _acc_digits(c)
        if dg:
            cmap.setdefault('#' + dg, ident)
    return cmap


# ── família de liquidação: movida para platform/settlement.py (§316) ─────────
# Os nomes ficam como ALIAS — features e testes seguem alcançando por
# `routes.<nome>`, e todo chamador resolve o atributo em tempo de chamada.
from apps.pages.platform import settlement as _pf_settle  # noqa: E402
_OPS_SRC_MAP = _pf_settle._OPS_SRC_MAP
_OPS_SWAP_JOIN_TOKENS = _pf_settle._OPS_SWAP_JOIN_TOKENS
_ops_src_latest_path = _pf_settle._ops_src_latest_path
_ops_settlement_counts = _pf_settle._ops_settlement_counts
_OPS_TRADE_COLS = _pf_settle._OPS_TRADE_COLS
_swadv_indexador = _pf_settle._swadv_indexador
_ops_swap_pos_terms = _pf_settle._ops_swap_pos_terms
_ops_swap_ir_rate = _pf_settle._ops_swap_ir_rate
_ops_cpty_receives = _pf_settle._ops_cpty_receives
_OPS_RECON_TOL = _pf_settle._OPS_RECON_TOL
_ops_swap_settling = _pf_settle._ops_swap_settling
_ops_swap_trade_rows = _pf_settle._ops_swap_trade_rows
_ops_recon = _pf_settle._ops_recon
_ops_ndfc_trade_rows = _pf_settle._ops_ndfc_trade_rows
_ops_opt_trade_rows = _pf_settle._ops_opt_trade_rows
_OPS_EQ_LEG_PREFIX = _pf_settle._OPS_EQ_LEG_PREFIX
_ops_eq_ref_key = _pf_settle._ops_eq_ref_key
_ops_eq_trade_key = _pf_settle._ops_eq_trade_key
_latam_equity_b3_index = _pf_settle._latam_equity_b3_index
_ops_equity_link = _pf_settle._ops_equity_link
_latam_trade_dt = _pf_settle._latam_trade_dt
_ops_le_name_keys = _pf_settle._ops_le_name_keys
_ops_is_internal_cpty = _pf_settle._ops_is_internal_cpty
_ops_trade_rows = _pf_settle._ops_trade_rows
_ops_fmt_amt = _pf_settle._ops_fmt_amt
_OPSSUM_COLS = _pf_settle._OPSSUM_COLS
_opssum_meta_path = _pf_settle._opssum_meta_path
_opssum_meta_load = _pf_settle._opssum_meta_load
_opssum_key = _pf_settle._opssum_key
_opssum_status = _pf_settle._opssum_status
_opssum_set_status = _pf_settle._opssum_set_status
_opssum_rows = _pf_settle._opssum_rows
_OPS_SOURCES = _pf_settle._OPS_SOURCES
_ops_pos_swap_found = _pf_settle._ops_pos_swap_found
_ops_batch_status = _pf_settle._ops_batch_status
_OPS_TED_LABEL = _pf_settle._OPS_TED_LABEL
_OPSADV_FAMILIES = _pf_settle._OPSADV_FAMILIES
_OPSADV_LABEL = _pf_settle._OPSADV_LABEL
_OPSADV_REQUIRED = _pf_settle._OPSADV_REQUIRED
_opsadv_block_incomplete = _pf_settle._opsadv_block_incomplete
_opsadv_blocked_header = _pf_settle._opsadv_blocked_header
_opsadv_family_drafts = _pf_settle._opsadv_family_drafts


# ── Operations B3: leitores movidos para platform/operations_b3.py (§318) ────
# Os nomes ficam como ALIAS; o _OPB3_MSG_RECIPIENTS_FILE fica AQUI (caminho
# sobre _DAILY_METRIC_DIR, superficie de patch — a platform le por routes.<nome>).
from apps.pages.platform import operations_b3 as _pf_opb3  # noqa: E402
_opb3_ev_key = _pf_opb3._opb3_ev_key
_opb3_event_rules = _pf_opb3._opb3_event_rules
_opb3_rule_hit = _pf_opb3._opb3_rule_hit
_opb3_settle_ok = _pf_opb3._opb3_settle_ok
_opb3_settle_rows = _pf_opb3._opb3_settle_rows
_OPB3_COLUMNS = _pf_opb3._OPB3_COLUMNS
_OPB3_DATE_COLS = _pf_opb3._OPB3_DATE_COLS
_OPB3_HEADER_ROW = _pf_opb3._OPB3_HEADER_ROW
_OPB3_META_KEYS = _pf_opb3._OPB3_META_KEYS
_opb3_ensure_meta = _pf_opb3._opb3_ensure_meta
_opb3_load_cached = _pf_opb3._opb3_load_cached
_opb3_load = _pf_opb3._opb3_load
_opb3_find = _pf_opb3._opb3_find
_opb3_ref_from = _pf_opb3._opb3_ref_from
_opb3_json_path = _pf_opb3._opb3_json_path
_opb3_extract = _pf_opb3._opb3_extract
_opb3_spec = _pf_opb3._opb3_spec
_opb3_map_recs = _pf_opb3._opb3_map_recs
_opb3_merge = _pf_opb3._opb3_merge
_opb3_side_write = _pf_opb3._opb3_side_write
_opb3_updated_from = _pf_opb3._opb3_updated_from
_opb3_import = _pf_opb3._opb3_import
_opb3_breakdown = _pf_opb3._opb3_breakdown
_opb3_tipo_maps = _pf_opb3._opb3_tipo_maps
_opb3_tipo_for = _pf_opb3._opb3_tipo_for
_opb3_collect = _pf_opb3._opb3_collect
_OPB3_MSG_GDT_BCC = _pf_opb3._OPB3_MSG_GDT_BCC
_OPB3_ACCT_BANCO = _pf_opb3._OPB3_ACCT_BANCO
_OPB3_ACCT_MGT = _pf_opb3._OPB3_ACCT_MGT
_OPB3_STATUS_GENERATED = _pf_opb3._OPB3_STATUS_GENERATED
_OPB3_B3_STATUS_DONE = _pf_opb3._OPB3_B3_STATUS_DONE
_opb3_msg_load_recipients = _pf_opb3._opb3_msg_load_recipients
_opb3_msg_save_recipients = _pf_opb3._opb3_msg_save_recipients
_opb3_msg_route_key = _pf_opb3._opb3_msg_route_key
_opb3_refdata_by_account = _pf_opb3._opb3_refdata_by_account
_OPB3_LEGAL_SIDES = _pf_opb3._OPB3_LEGAL_SIDES
_opb3_legal_side = _pf_opb3._opb3_legal_side
_opb3_internal_ter_map = _pf_opb3._opb3_internal_ter_map
_opb3_internal_leg = _pf_opb3._opb3_internal_leg
_opb3_internal_swapprem_map = _pf_opb3._opb3_internal_swapprem_map
_opb3_internal_trade_map = _pf_opb3._opb3_internal_trade_map
_opb3_internal_swap_map = _pf_opb3._opb3_internal_swap_map
_opb3_internal_ndfc_map = _pf_opb3._opb3_internal_ndfc_map
_opb3_events_upgrade = _pf_opb3._opb3_events_upgrade
_ops_norm_event = _pf_opb3._ops_norm_event




# ── Live Position › Swap › Characteristics ───────────────────────────────────
#  Read-only "photo" of the swap book still in custody on a reference date, from
#  the DPOSICAO-SWAP position JSON (same source as the dashboard Live Position).
#  Widgets break the count down by Tipo de Contrato, LOB, Indexador and
#  Funcionalidade; the table lists every contract with its full characteristic
#  column set. The canonical column list lives HERE (single source of truth) and
#  is shipped to the front-end so the header/data arrays can never drift apart.
#
#  Column names repeat by design (the B3 swap layout has parallel leg/index
#  blocks), so rows are emitted as POSITIONAL arrays aligned to _SWAPCHAR_LABELS
#  — a name-keyed dict would collapse the duplicates.
_SWAPCHAR_LABELS = [
    'Tipo de Contrato', 'Data', 'Contrato', 'Participante', 'CPF/CNPJ Cliente Parte',
    'Cesta Garantias Parte', 'Comissão Parte', 'Contraparte', 'CPF/CNPJ Cliente Contraparte',
    'Cesta Garantias Contraparte', 'Comissão Contraparte', 'Data início', 'Data vencimento',
    'Tipo de Adesão', 'Valor base', 'Valor Base Remanescente', 'Valor Antecipado', 'Saldo',
    'Sinal Saldo', 'Data do Saldo', 'Funcionalidade', 'Agenda de Prêmio', 'Reset', 'Observação',
    'Valor base inicial', 'Data operação termo', 'Índice Termo', 'Percentual Termo', 'PU Inicial',
    'Tipo/Classe', 'Nome Tipo/Classe', 'Denominação', 'Juros a cada', 'Expresso em',
    'Data inicio pagamento juros', 'Amortização a cada', 'Expresso em',
    'Data inicio pagamento amortização', 'Tipo de amortização', 'Percentual', 'Código índice',
    'TR Escolhida', 'Sinal Taxa', 'Taxa', 'Lim. Inferior (Floor)', 'Lim. Superior (Cap)',
    'Valor Curva Atualizado', 'Data Correção', 'Fator Original de Juros', 'Percentual',
    'Código índice', 'TR Escolhida', 'Sinal Taxa', 'Taxa', 'Lim. Inferior (Floor)',
    'Lim. Superior (Cap)', 'Valor Curva Atualizado', 'Data Correção', 'Fator Original de Juros',
    'Parte/Contraparte', 'Cupom Limpo', 'Percentual', 'Curva', 'Sinal Taxa', 'Taxa de Juros',
    'Limitador', 'Pu inicial', 'Pu atual', 'Tipo/Classe', 'Nome Tipo/Classe', 'Denominação',
    'Pu inicial', 'Pu atual', 'Tipo/Classe', 'Nome Tipo/Classe', 'Denominação', 'Cupom Limpo',
    'Data de Cotação', 'Cupom Limpo', 'Data de Cotação', 'Tipo Libor - moeda',
    'Tipo Libor - período', 'Data de Cotação', 'Variação Cambial', 'Tipo Classe',
    'Nome Tipo/Classe', 'Outros - Cotação', 'Alíquota - IR', 'Limite inferior (FLOOR) - Perc.',
    'Limite superior (CAP) - Perc.', 'Tipo Libor - moeda', 'Tipo Libor - período',
    'Data de Cotação', 'Variação Cambial', 'Tipo Classe', 'Nome Tipo/Classe', 'Outros - Cotação',
    'Alíquota - IR', 'Limite inferior (FLOOR) - Perc.', 'Limite superior (CAP) - Perc.',
    'Taxa Juros', 'Troca de Fluxo', 'Variação Cambial', 'Tipo Classe', 'Nome Tipo/Classe',
    'Outros - Cotação', 'Alíquota - IR', 'Limite inferior (FLOOR) - Perc.',
    'Limite superior (CAP) - Perc.', 'Taxa Juros', 'Troca de Fluxo', 'Variação Cambial',
    'Tipo Classe', 'Nome Tipo/Classe', 'Outros - Cotação', 'Alíquota - IR',
    'Limite inferior (FLOOR) - Perc.', 'Limite superior (CAP) - Perc.', 'Parte/Contraparte',
    'Fator/Valor/Taxa', 'Verificação', 'Data Disparo', 'Parte/Contraparte', 'Fator/Valor/Taxa',
    'Verificação', 'Data Disparo', 'Titular', 'Prêmio 1', 'Rebate', 'Liquidação do Rebate',
    'Dias Úteis após o Trigger Out', 'Prêmio 2', 'Data Exercício Prêmio 2', 'Estratégia',
    'Amortiza sem Troca de Diferencial', 'Data da Cotação - Variação Cambial',
    'Data da Cotação - Variação Cambial', 'Cotação Inicial', 'Código Commodity',
    'Media Asiática Verificação', 'Data Cotação para Ajuste', 'Cotação Inicial', 'Código Commodity',
    'Media Asiática Verificação', 'Data Cotação para Ajuste', 'Código Identificador',
]

# Funcionalidade code → clean label (no underscores/parentheses; OPCAO_ARREPEND →
# "OPCAO ARREPENDIMENTO"). Keyed by the integer code as a string ('0'..'9').
# Funcionalidade: código → texto. Cadastro `swap-funcionalidade` (/mapping).
# Header tokens (normalised) that mark a numeric/value column → format #,##0.00.
_SWAPCHAR_VALUE_TOKENS = ('valor', 'saldo', 'percentual', 'pu inicial', 'pu atual', 'taxa',
                          'lim. inferior', 'lim. superior', 'limite inferior', 'limite superior',
                          'curva atualizado', 'fator original', 'cupom limpo', 'premio', 'rebate',
                          'cotacao inicial', 'aliquota', 'fator/valor', 'outros - cotacao',
                          'variacao cambial', 'prêmio')


def _swapchar_lob(identifier):
    """Swap LOB bucket for the Characteristics widget: CEM / EDG / COMM / HYB.
    Hybrid is tested first (its id also contains 'CEM')."""
    s = _fcst_norm(identifier)
    if 'hyb' in s or 'hib' in s:
        return 'HYB'
    if 'comm' in s or 'commod' in s:
        return 'COMM'
    if 'edg' in s:
        return 'EDG'
    if 'cem' in s:
        return 'CEM'
    return 'CEM'


def _swapchar_coltype(label):
    """Formatting class for a column: date | func | amort | value | text."""
    n = _fcst_norm(label)
    if n.startswith('data'):
        return 'date'
    if n == 'funcionalidade':
        return 'func'
    if n == 'tipo de amortizacao':
        return 'amort'
    if n.startswith('sinal'):
        return 'sinal'                         # Sinal Taxa: 00 → +, 01 → -
    if n in ('codigo indice', 'codigo do indice'):
        return 'indice'                        # show Nome Curva instead of the raw code
    if n.startswith('tipo') or n.startswith('nome'):
        return 'text'
    if any(tok in n for tok in _SWAPCHAR_VALUE_TOKENS):
        return 'value'
    return 'text'


_SWAPCHAR_TYPES = [_swapchar_coltype(l) for l in _SWAPCHAR_LABELS]


def _swapchar_code_map(key, field=None):
    """{code → texto} de um cadastro de tradução (`swap-funcionalidade`,
    `swap-amortizacao`, `swap-code-labels`). O código é normalizado para o
    inteiro sem zeros à esquerda, que é como os arquivos da B3 variam ('0', '00',
    '000') — registrar as três formas seria pedir erro."""
    out = {}
    for r in _mapping_rows(key):
        if field is not None and _fcst_norm(r.get('FIELD', '')).strip() != _fcst_norm(field):
            continue
        code = str(r.get('CODE', '') or '').strip()
        digits = ''.join(ch for ch in code if ch.isdigit())
        if digits == '':
            continue
        out[str(int(digits))] = str(r.get('LABEL', '') or '')
    return out


def _swapchar_func_text(v):
    """Map a Funcionalidade cell to its clean label (strip underscores/parentheses)."""
    s = str(v or '').strip()
    if not s:
        return ''
    digits = ''.join(ch for ch in s if ch.isdigit())
    if digits:
        code = str(int(digits))
        m = _swapchar_code_map('swap-funcionalidade')
        if code in m:
            return m[code]
    t = ' '.join(s.replace('(', ' ').replace(')', ' ').replace('_', ' ').split())
    if 'ARREPEND' in t.upper():
        return 'OPCAO ARREPENDIMENTO'
    return t


# Tipo de amortização: código → texto. Cadastro `swap-amortizacao` (/mapping).
# As DUAS páginas que traduzem esse código — Characteristics e Swap Flow — leem
# daqui, para não existir a versão de uma e a versão da outra.


def _swapchar_amort_text(v):
    """Map a Tipo de amortização cell to its text label (no parentheses)."""
    s = str(v or '').strip()
    if not s:
        return ''
    m = re.search(r'\(([^)]*)\)', s)          # value already carries "NN (text)"
    if m and m.group(1).strip():
        return m.group(1).strip()
    digits = ''.join(ch for ch in s if ch.isdigit())
    if digits:
        code = str(int(digits))
        m = _swapchar_code_map('swap-amortizacao')
        if code in m:
            return m[code]
    return s


def _swapchar_value_num(v):
    """Célula numérica do arquivo de POSIÇÃO → float, ou None.

    O arquivo escreve a vírgula como separador DECIMAL ('280000000,00'), sem
    separador de milhar. É por isso que ele não pode passar pelos parsers de uso
    geral: o `_mtm_parse_num` trata a vírgula como milhar e devolve
    28.000.000.000 para esse mesmo texto — cem vezes o valor.

    Existe para que a CÉLULA e o NÚMERO CRU saiam da mesma leitura. Eram duas, e
    o Valor Base aparecia certo na tela e cem vezes maior no aviso impresso."""
    s = str(v or '').strip()
    if not s:
        return None
    try:
        return float(s.replace(' ', '').replace(',', '.'))
    except ValueError:
        return None


def _swapchar_fmt_value(v):
    """Numeric cell → #,##0.00 (1,234.56); non-numeric passes through unchanged."""
    s = str(v or '').strip()
    if not s:
        return ''
    n = _swapchar_value_num(s)
    return s if n is None else '{:,.2f}'.format(n)


def _lp_fmt_dec8(v):
    """Numeric cell → #,##0.00000000 (1,234.56789000); non-numeric passes through."""
    s = str(v or '').strip()
    if not s:
        return ''
    try:
        n = float(s.replace(' ', '').replace(',', '.'))
    except ValueError:
        return s
    return '{:,.8f}'.format(n)


def _lp_fmt_cnpj(v):
    """All-digit CPF (11) / CNPJ (14) → masked; recovers a dropped leading zero
    (12-13 digits → CNPJ, 10 → CPF). Anything else passes through unchanged."""
    s = str(v or '').strip()
    if not s or not s.isdigit():
        return '' if v is None else v
    if 12 <= len(s) <= 14:
        s = s.zfill(14)
        return '{}.{}.{}/{}-{}'.format(s[:2], s[2:5], s[5:8], s[8:12], s[12:])
    if 10 <= len(s) <= 11:
        s = s.zfill(11)
        return '{}.{}.{}-{}'.format(s[:3], s[3:6], s[6:9], s[9:])
    return v


_LP_TAXID_NAME_CACHE = {'src': None, 'map': None}


def _lp_taxid_key(v):
    """CPF/CNPJ → chave de comparação: só dígitos, com o zero à esquerda
    recuperado (12-13 → CNPJ, 10 → CPF), exatamente o que `_lp_fmt_cnpj` faz
    para exibir.

    Os DOIS lados têm de passar por aqui. O RefData guarda mascarado
    (45.985.371/0001-08) e a posição da B3 guarda só números, às vezes sem o
    zero da frente — comparar sem normalizar casa silenciosamente nada, que é a
    mesma armadilha do §197."""
    d = ''.join(ch for ch in str(v or '') if ch.isdigit())
    if not d:
        return ''
    if 12 <= len(d) <= 14:
        return d.zfill(14)
    if 10 <= len(d) <= 11:
        return d.zfill(11)
    return d


def _lp_taxid_names():
    """{chave de CPF/CNPJ → nome da contraparte}, do RefData.

    Reindexa o `_refdata_by_taxid()` — que chaveia por dígitos CRUS — pela chave
    normalizada, para o lado do cadastro sofrer o mesmo zero-fill que o lado da
    posição. É um comprehension sobre um mapa JÁ cacheado por mtime (o arquivo
    não é lido de novo), refeito só quando aquele mapa troca de objeto, que é o
    que acontece quando o RefData muda em disco."""
    base = _refdata_by_taxid()
    if _LP_TAXID_NAME_CACHE['src'] is not base:
        m = {}
        for digits, nome in base.items():
            k = _lp_taxid_key(digits)
            if k and k not in m:
                m[k] = nome
        _LP_TAXID_NAME_CACHE['src'] = base
        _LP_TAXID_NAME_CACHE['map'] = m
    return _LP_TAXID_NAME_CACHE['map']


def _lp_is_taxid(v):
    """A string é um CPF/CNPJ (cru ou mascarado) e não um nome?

    Existe porque a coluna de CPF/CNPJ da contraparte passou a carregar as DUAS
    coisas — o nome quando há cadastro, o documento quando não há — e quem lê a
    coluna programaticamente precisa saber qual das duas veio. O teste é pela
    ausência de letra: razão social com número (`3M DO BRASIL`) tem letra e
    nunca casa aqui."""
    s = str(v or '').strip()
    if not s or any(ch not in '0123456789./- ' for ch in s):
        return False
    return 10 <= len(''.join(ch for ch in s if ch.isdigit())) <= 14


def _lp_cpty_name_by_taxid(v):
    """Nome da contraparte pelo CPF/CNPJ, ou '' quando não há cadastro.

    É a resolução CRUA — sem queda para o número —, e é ela que os consumidores
    programáticos usam. O Settlement Advice de NDF Commodities precisa saber a
    diferença entre "resolveu" e "não resolveu" (§197): resolvendo, o nome é o do
    cliente por trás do omnibus; não resolvendo, o aviso cai para o nome da
    posição, como sempre fez."""
    k = _lp_taxid_key(v)
    return _lp_taxid_names().get(k, '') if k else ''


def _lp_cpty_by_taxid(v):
    """Versão de EXIBIÇÃO da coluna de CPF/CNPJ da contraparte.

    Célula vazia continua vazia. Documento SEM cadastro devolve o CPF/CNPJ
    mascarado, e não branco: o número é o único dado que a linha tem sobre a
    contraparte, e apagá-lo esconderia justamente quem falta cadastrar — a
    coluna misturada é o que denuncia a lacuna."""
    if not str(v or '').strip():
        return ''
    return _lp_cpty_name_by_taxid(v) or _lp_fmt_cnpj(v)


def _lp_bool_ptbr(raw):
    """Flag de swap → texto. Cadastro `swap-code-labels`, campo `Sim/Não`
    (00 → Sim, 01 → Não — sim, nessa ordem, é a especificação). Valor fora do
    cadastro passa direto."""
    s = str(raw or '').strip()
    if not s:
        return ''
    if s.endswith('.00'):
        s = s[:-3]
    elif s.endswith('.0'):
        s = s[:-2]
    if s.isdigit():
        s = str(int(s))
    return _swapchar_code_map('swap-code-labels', 'Sim/Não').get(s, raw)


def _lp_amort_label(raw):
    """Swap Flow "Tipo Amortização" code → label. Uses the SAME nomenclature as the
    Live Position Swap Characteristics "Tipo de amortização" column
    (cadastro `swap-amortizacao`), so both pages read identically."""
    s = str(raw or '').strip()
    if not s:
        return ''
    m = re.search(r'\(([^)]*)\)', s)          # value already carries "NN (text)"
    if m and m.group(1).strip():
        return m.group(1).strip()
    if s.endswith('.0'):                       # tolerate a "3.0" style decimal tail
        s = s[:-2]
    digits = ''.join(ch for ch in s if ch.isdigit())
    if digits:
        code = str(int(digits))
        m = _swapchar_code_map('swap-amortizacao')
        if code in m:
            return m[code]
    return raw


# Swap Index — Código de Referência Externa → Nome da Curva. Saiu do
# `SwapIndex.json` (a aba Swap Index do B3 Index Results) e virou o cadastro
# `swap-index` do /mapping — apontado para o MESMO arquivo (`file` no
# _MAPPING_DEFS), não para uma cópia. As duas telas editam o mesmo SwapIndex.json,
# então não existe a versão de uma e a versão da outra; a única diferença é que
# agora a leitura cacheia por mtime, como todos os outros cadastros.
def _swapindex_lookup():
    """{code(upper) → Nome Curva} do cadastro `swap-index`."""
    m = {}
    for rec in _mapping_rows('swap-index'):
        code = str(rec.get('Codigo Referencia Externa', '') or '').strip().upper()
        name = str(rec.get('Nome Curva', '') or '').strip()
        if code:
            m[code] = name
            m.setdefault(code.lstrip('0') or '0', name)   # tolerante a zeros à esquerda
    return m


def _swapindex_name(code):
    s = str(code or '').strip()
    if not s:
        return ''
    m = _swapindex_lookup()
    return m.get(s.upper()) or m.get(s.upper().lstrip('0') or '0') or s


def _swapchar_sinal_text(v):
    """Sinal Taxa: código → sinal. Cadastro `swap-code-labels`, campo
    `Sinal Taxa` (00 → +, 01 → -), tolerante a '0'/'1'."""
    s = str(v or '').strip()
    digits = ''.join(ch for ch in s if ch.isdigit())
    if digits != '':
        return _swapchar_code_map('swap-code-labels', 'Sinal Taxa').get(str(int(digits)), s)
    return s


def _swapchar_fmt_cell(value, ctype):
    if value in (None, ''):
        return ''
    if ctype == 'date':
        d = _fcst_parse_date(value)
        return d.strftime('%d/%m/%Y') if d else str(value)
    if ctype == 'func':
        return _swapchar_func_text(value)
    if ctype == 'amort':
        return _swapchar_amort_text(value)
    if ctype == 'sinal':
        return _swapchar_sinal_text(value)
    if ctype == 'indice':
        return _swapindex_name(value)
    if ctype == 'value':
        return _swapchar_fmt_value(value)
    return str(value)


# Columns actually shown on the page — a subset of the 146 (in file order), from
# the desk's reference layout. Values are read POSITIONALLY from the position-file
# rows (which carry all 146 fields in order), so the repeated column names still
# resolve unambiguously to the right cell.
_SWAPCHAR_DISPLAY_IDX = [
    0, 2, 3, 7, 8, 11, 12, 14, 15, 16, 17, 18, 20, 21, 22, 23,           # A,C,D,H,I,L,M,O,P,Q,R,S,U,V,W,X
    24, 25, 26, 27, 28, 31, 38, 39, 40, 42, 43,                          # Y,Z,AA,AB,AC,AF,AM,AN,AO,AQ,AR
    44, 45, 46, 48, 49, 50, 52, 53, 54, 55, 56, 58, 60, 61, 62, 63, 64,  # AS..BM (skips AV,AZ,BF,BH)
    65, 66, 67, 68, 69, 70, 71, 72, 73, 74, 75,                          # BN..BX
    76, 77, 78, 79, 126, 127, 132, 133, 134, 144, 145,                   # BY,BZ,CA,CB,DW,DX,EC,ED,EE,EO,EP
]
_SWAPCHAR_DISPLAY_LABELS = [_SWAPCHAR_LABELS[i] for i in _SWAPCHAR_DISPLAY_IDX]
# Flag columns shown as Sim/Não (01 = Não, 00 = Sim; empty stays empty).
_SWAPCHAR_BOOL_COLS = {'Reset', 'Amortiza sem Troca de Diferencial', 'Agenda de Prêmio'}
# Coluna de CPF/CNPJ que a tela mostra como NOME da contraparte (xlookup no
# RefData). É a da CONTRAPARTE — a da Parte é a nossa e continua o documento.
# Chaveada pelo RÓTULO e não pelo índice: índice errado pega a coluna vizinha
# sem erro nenhum, e aqui a vizinha é o nome simplificado.
_SWAPCHAR_CPTY_NAME_COLS = {'CPF/CNPJ Cliente Contraparte'}


def _swapchar_collect(ref):
    """Build widgets + display rows from the DPOSICAO-SWAP file for `ref` (date).
    The saved position JSON carries all 146 fields IN ORDER (headerless file parsed
    with _B3_SWAP_HEADERS), so cells are read positionally by index; only the
    _SWAPCHAR_DISPLAY_IDX subset is emitted. Missing file → empty payload (logged)."""
    widgets = {
        'total': 0,
        'tipo':  {'total': 0, 'cashflow': 0, 'bullet': 0},
        'lob':   {'total': 0, 'CEM': 0, 'EDG': 0, 'COMM': 0, 'HYB': 0},
        # VCP / Calculado breakdown — counting logic pending (user will supply).
        'index': {'total': 0, 'vcp': 0, 'calculado': 0},
        # Forward Start / Notional / Prêmio / Arrependimento / Sem Funcionalidade —
        # counting logic pending (user will supply).
        'func':  {'total': 0, 'forward_start': 0, 'notional': 0, 'premio': 0,
                  'arrependimento': 0, 'sem': 0},
    }
    dref = ref.strftime('%y%m%d')
    path = os.path.join(B3_JSON_ROOT, 'Swap', _b3_date_subpath(dref),
                        '73760_{}_DPOSICAO-SWAP.json'.format(dref))
    rows_out = []
    if not os.path.isfile(path):
        log.warning("[swapchar] no DPOSICAO-SWAP for %s; page shows 0", dref)
        return {'widgets': widgets, 'columns': _SWAPCHAR_DISPLAY_LABELS, 'rows': []}
    try:
        with open(path, encoding='utf-8') as fh:
            src = json.load(fh)
    except Exception:
        return {'widgets': widgets, 'columns': _SWAPCHAR_DISPLAY_LABELS, 'rows': []}
    if not src:
        return {'widgets': widgets, 'columns': _SWAPCHAR_DISPLAY_LABELS, 'rows': []}

    keys = list(src[0].keys())
    tipo_key = _fcst_resolve_key(keys, ['tipo de contrato', 'tipo do contrato', 'tipo contrato'])
    cpty_key = _fcst_resolve_key(keys, ['contraparte'])
    venc_key = _fcst_resolve_key(keys, ['data vencimento', 'data de vencimento'])
    id_key   = _fcst_resolve_key(keys, ['código identificador', 'codigo identificador', 'identificador'])
    func_key = _fcst_resolve_key(keys, ['funcionalidade'])
    for row in src:
        vals = list(row.values())          # all 146 fields, in file order (real file)
        full = len(vals) >= 120             # sparse mock (4 named cols) → name-resolve fallback
        tv = str(row.get(tipo_key, '') or '').strip() if tipo_key else ''
        if tv.endswith('.0'):
            tv = tv[:-2]
        if tv.isdigit():                    # '01' → '1', '02' → '2' (leading zeros)
            tv = str(int(tv))
        cid = str(row.get(id_key, '') or '') if id_key else ''
        # Sparse-mock fallback: the few present fields keyed by their 146-list index.
        sparse = {} if full else {
            0: tv,
            7: (row.get(cpty_key, '') if cpty_key else ''),
            12: (row.get(venc_key, '') if venc_key else ''),
            20: (row.get(func_key, '') if func_key else ''),
            145: cid,
        }
        disp = []
        for i in _SWAPCHAR_DISPLAY_IDX:
            raw = (vals[i] if i < len(vals) else '') if full else sparse.get(i, '')
            if i == 0:                      # Tipo de Contrato: 02 → Bullet, 01 → Cashflow
                rv = str(raw or '').strip()
                if rv.endswith('.0'):
                    rv = rv[:-2]
                if rv.isdigit():
                    rv = str(int(rv))
                disp.append('Bullet' if rv == '2' else ('Cashflow' if rv == '1'
                            else _swapchar_fmt_cell(raw, _SWAPCHAR_TYPES[i])))
            elif _SWAPCHAR_LABELS[i] in _SWAPCHAR_BOOL_COLS:
                disp.append(_lp_bool_ptbr(raw))
            elif _SWAPCHAR_LABELS[i] in _SWAPCHAR_CPTY_NAME_COLS:
                disp.append(_lp_cpty_by_taxid(raw))
            else:
                disp.append(_swapchar_fmt_cell(raw, _SWAPCHAR_TYPES[i]))
        rows_out.append(disp)
        # Widgets
        widgets['total'] += 1
        widgets['tipo']['total'] += 1
        if tv == '2':                       # 02 → Bullet, 01 → Cashflow
            widgets['tipo']['bullet'] += 1
        elif tv == '1':
            widgets['tipo']['cashflow'] += 1
        lob = _swapchar_lob(cid)
        widgets['lob']['total'] += 1
        widgets['lob'][lob] += 1
    return {'widgets': widgets, 'columns': _SWAPCHAR_DISPLAY_LABELS, 'rows': rows_out}


# ── Live Position Swap Cashflow (DFLUXO) & Premium (DAGENDAPREMIOS) ───────────
#  Same page template as Swap Characteristics, but a fixed subset of columns
#  pulled POSITIONALLY from the respective headerless position JSON (parsed with
#  the _B3_SWAP_HEADERS standard for that file). Values are formatted with the
#  same _swapchar_fmt_cell rules (dates, sign, numeric).
_SWAPFLUX_LABELS = _B3_SWAP_HEADERS['swap_fluxo']
# swap_fluxo has 26 cols; drop "Papel Parte"(4), "Papel Contraparte"(7) and the
# trailing "Data Início Composição da Taxa Contraparte"(25) → the 23 requested.
_SWAPFLUX_DISPLAY_IDX = [0, 1, 2, 3, 5, 6, 8, 9, 10, 11, 12, 13, 14, 15, 16,
                         17, 18, 19, 20, 21, 22, 23, 24]
_SWAPFLUX_TYPES = [_swapchar_coltype(l) for l in _SWAPFLUX_LABELS]

_SWAPPREM_LABELS = _B3_SWAP_HEADERS['swap_premio']
_SWAPPREM_DISPLAY_IDX = list(range(len(_SWAPPREM_LABELS)))   # all 10 columns
_SWAPPREM_TYPES = [_swapchar_coltype(l) for l in _SWAPPREM_LABELS]


def _swap_simple_collect(ref, file_tpl, labels, display_idx, types):
    """Read a headerless Swap position JSON (values in file order) and emit the
    display-column subset, formatted like Swap Characteristics. Returns
    {widgets:{total}, columns, rows}. Missing file → empty payload (logged)."""
    dref = ref.strftime('%y%m%d')
    disp_labels = [labels[i] for i in display_idx]
    empty = {'widgets': {'total': 0}, 'columns': disp_labels, 'rows': []}
    path = os.path.join(B3_JSON_ROOT, 'Swap', _b3_date_subpath(dref), file_tpl.format(dref))
    if not os.path.isfile(path):
        log.warning("[swap-simple] no %s for %s; page shows 0", file_tpl.format(dref), dref)
        return empty
    try:
        with open(path, encoding='utf-8') as fh:
            src = json.load(fh)
    except Exception:
        return empty
    if not src:
        return empty
    named = {_fcst_norm(k): k for k in src[0].keys()}   # sparse-mock name fallback
    rows_out = []
    for row in src:
        vals = list(row.values())                        # file order (dup headers → _2, _3…)
        full = len(vals) >= len(labels)
        disp = []
        for i in display_idx:
            if full:
                raw = vals[i] if i < len(vals) else ''
            else:
                key = named.get(_fcst_norm(labels[i]))
                raw = row.get(key, '') if key else ''
            if labels[i] == 'Tipo Amortização':
                disp.append(_lp_amort_label(raw))
            else:
                disp.append(_swapchar_fmt_cell(raw, types[i]))
        rows_out.append(disp)
    return {'widgets': {'total': len(rows_out)}, 'columns': disp_labels, 'rows': rows_out}


def _swap_simple_ref(request_args):
    ds = (request_args.get('date') or '').strip()
    try:
        return datetime.strptime(ds[:10], '%Y-%m-%d').date() if ds else \
            _prev_anbima_bizday(datetime.now()).date()
    except ValueError:
        return _prev_anbima_bizday(datetime.now()).date()


def _swapprem_collect(ref):
    """Swap Premium (DAGENDAPREMIOS) with an extra "Contraparte" column inserted
    right after "Nome Simplificado", joined from the DPOSICAO-SWAP position file by
    "Codigo do Contrato" (the premium file carries only the contract code)."""
    dref = ref.strftime('%y%m%d')
    labels, idx, types = _SWAPPREM_LABELS, _SWAPPREM_DISPLAY_IDX, _SWAPPREM_TYPES
    disp_labels = [labels[i] for i in idx]
    ins = disp_labels.index('Nome Simplificado') + 1
    out_labels = disp_labels[:ins] + ['Contraparte'] + disp_labels[ins:]
    empty = {'widgets': {'total': 0}, 'columns': out_labels, 'rows': []}
    path = os.path.join(B3_JSON_ROOT, 'Swap', _b3_date_subpath(dref),
                        '73760_{}_DAGENDAPREMIOS.json'.format(dref))
    if not os.path.isfile(path):
        log.warning("[swap-premium] no DAGENDAPREMIOS for %s; page shows 0", dref)
        return empty
    try:
        with open(path, encoding='utf-8') as fh:
            src = json.load(fh)
    except Exception:
        return empty
    if not src:
        return empty
    cpty_map = _swap_contract_cpty_map(dref)
    named = {_fcst_norm(k): k for k in src[0].keys()}
    rows_out = []
    for row in src:
        vals = list(row.values())
        full = len(vals) >= len(labels)
        disp = []
        for i in idx:
            if full:
                raw = vals[i] if i < len(vals) else ''
            else:
                key = named.get(_fcst_norm(labels[i]))
                raw = row.get(key, '') if key else ''
            disp.append(_swapchar_fmt_cell(raw, types[i]))
        if full:
            contrato_raw = vals[0] if vals else ''
        else:
            k0 = named.get(_fcst_norm(labels[0]))
            contrato_raw = row.get(k0, '') if k0 else ''
        disp.insert(ins, cpty_map.get(_fcst_norm_contract(contrato_raw), ''))
        rows_out.append(disp)
    return {'widgets': {'total': len(rows_out)}, 'columns': out_labels, 'rows': rows_out}


# ── Other Products › OTM Settlements ─────────────────────────────────────────
#  Replaces the legacy Excel/VBA "Settlement - OTM" import. A cashflows_*.xlsx
#  file (actually a TAB-delimited text export, opened by the VBA via OpenText
#  Tab:=True) is dropped in OTM_SOURCE_ROOT. On import we clean it exactly like
#  the macro's CleanSettlementOTM:
#    A) drop rows whose col 14 == "DELETE" (keep header),
#    B) normalise col 22 to a 4-digit text code (leading zeros),
#    C) keep only col 22 in {"0228","0123"},
#  then keep ONLY the reporting columns below and write them to
#    static/data/cache/daily settlement/YYYY/MM/DD/otm-settlement_YYYYMMDD.json
#  (today's date), deleting the consumed source file. Widgets' counting logic
#  (RATES/EQUITIES/COMMODITIES) is pending (user will supply).
OTM_SOURCE_ROOT = os.getenv('OTM_SOURCE_ROOT', os.path.join(
    Config.SHARED_DRIVE_ROOT, 'Confirmation', 'Derivativos', 'OTC Tracker', 'Settlement', 'OTM'))
OTM_JSON_ROOT = data_write('cache', 'daily settlement')
_OTM_COLUMNS = [
    'Trade Id', 'Currency', 'Amount', 'Value Date', 'Direction', 'Cpty SPN', 'Cpty Name',
    'Owner SPN', 'Trade Date', 'Asset Class', 'Owner Legal Entity', 'Owner Name',
    'Exception Type', 'Cashflow Stage', 'Trade Ref', 'Underlying', 'Product Class', 'Break Reason',
]
_OTM_DATE_COLS = {'Value Date', 'Trade Date'}
_OTM_KEEP_CODES = {'0228', '0123'}          # col 22 values to keep (CleanSettlementOTM step C)

# Maker/checker lifecycle meta stored per OTM record (keys prefixed to avoid clashing
# with the 18 data columns). Imported/added rows start 'OK'; an edit → 'Pending' (maker
# set, checker cleared); a DIFFERENT user confirms → 'OK' (checker set).
_OTM_META_KEYS = ('_ot_status', '_ot_maker', '_ot_checker', '_ot_id')


def _otm_new_id():
    return uuid.uuid4().hex[:10]


def _otm_ensure_meta(data, default_status='OK'):
    """Ensure every record has status/maker/checker/id meta. Returns True if any
    record changed (caller may persist — one-time migration for legacy JSONs)."""
    changed = False
    for rec in data:
        if not rec.get('_ot_id'):
            rec['_ot_id'] = _otm_new_id(); changed = True
        if '_ot_status' not in rec:
            rec['_ot_status'] = default_status; changed = True
        for k in ('_ot_maker', '_ot_checker'):
            if k not in rec:
                rec[k] = ''; changed = True
    return changed


@_req_cached
def _otm_load_cached(ref):
    """A leitura em si — é este resultado que o cache guarda. Ver `_otm_load`."""
    jp = _otm_json_path(ref)
    if not os.path.isfile(jp):
        return jp, None
    try:
        with open(jp, encoding='utf-8') as fh:
            data = json.load(fh) or []
    except Exception:
        return jp, None
    _otm_ensure_meta(data)
    return jp, data


def _otm_load(ref):
    """(json_path, data|None) for `ref`; ensures meta on the loaded records.

    Devolve uma CÓPIA dos registros, nunca a lista que está no cache. Os
    endpoints de add/edit/delete carregam o dia, mexem na lista e só então
    gravam (`data.remove(rec)`, `rec[c] = ...`): com o objeto do cache na mão,
    essa mutação passa a valer para todo mundo ANTES do save — e continua
    valendo quando o save FALHA. A linha some da tela de quem não pediu nada, e
    o request seguinte, que dentro do TTL recebe o mesmo objeto, grava por cima
    o estado que nunca chegou ao disco. É perda de dado sem erro nenhum.

    A cópia é rasa por registro porque toda escrita destes endpoints é escalar
    (`rec[k] = v`), e ela custa uma fração da leitura do share que o cache
    existe para poupar.
    """
    jp, data = _otm_load_cached(ref)
    return jp, (None if data is None else [dict(r) for r in data])


def _otm_save(jp, data):
    """Grava o arquivo-dia do OTM Settlements.

    O `_bump_cache_gen` está aqui porque este save NÃO passa pelo
    `_atomic_write_json` — ele escreve direto. Sem a chamada, os loaders
    decorados com `@_req_cached` que derivam deste arquivo continuariam
    servindo o resultado anterior por até `SHARED_CACHE_TTL_SECONDS`: a pessoa
    edita a linha, a tela recarrega e mostra o valor de antes, sem erro nenhum.
    """
    os.makedirs(os.path.dirname(jp), exist_ok=True)
    _atomic_write_json(jp, data)                # funil: bump + espelho (§335)


def _otm_find(data, rid):
    for rec in data:
        if str(rec.get('_ot_id', '')) == str(rid):
            return rec
    return None


def _otm_json_path(ref):
    return os.path.join(OTM_JSON_ROOT, ref.strftime('%Y'), ref.strftime('%m'), ref.strftime('%d'),
                        'otm-settlement_{}.json'.format(ref.strftime('%Y%m%d')))


# ── "Last updated" timestamp sidecar (shared by OTM + Operations B3) ──────────
# A <json>.meta.json holds {"updated": "HH:MM:SS"} — the file's own extraction time
# when the source provides one (Operations B3 = row 2 col A), else the import time.
def _ds_meta_path(json_path):
    return (json_path[:-5] if json_path.endswith('.json') else json_path) + '.meta.json'


def _ds_write_updated(json_path, hhmmss):
    try:
        _atomic_write_json(_ds_meta_path(json_path), {'updated': hhmmss or ''})
    except OSError:
        pass


def _ds_read_updated(json_path):
    mp = _ds_meta_path(json_path)
    if os.path.isfile(mp):
        try:
            with open(mp, encoding='utf-8') as fh:
                return (json.load(fh) or {}).get('updated', '')
        except Exception:
            pass
    if os.path.isfile(json_path):                    # fallback: file mtime
        return datetime.fromtimestamp(os.path.getmtime(json_path)).strftime('%H:%M:%S')
    return ''


def _otm_read_rows(path):
    """Rows (list of lists) from the cashflows file. The VBA treats it as a TAB
    text file even though it's named .xlsx; handle both a real .xlsx (zip) and a
    tab-delimited text export."""
    with open(path, 'rb') as fh:
        raw = fh.read()
    if raw[:2] == b'PK':                     # real .xlsx (zip container)
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        return [list(r) for r in ws.iter_rows(values_only=True)]
    text = raw.decode('latin-1')
    return [ln.split('\t') for ln in text.splitlines() if ln.strip()]


def _otm_extract(rows):
    """Clean + extract the OTM reporting columns from a cashflows file's rows
    (CleanSettlementOTM: drop col14==DELETE, keep col22 ∈ {0228,0123}; keep the 18
    columns by header name; Cpty Name upper-cased). Returns (records, kept,
    deleted, filtered). Shared by the OTM page and the Save Daily Settlement card."""
    header = [str(h or '').strip() for h in rows[0]]
    hnorm = [_fcst_norm(h) for h in header]

    def col_idx(name):
        n = _fcst_norm(name)
        if n in hnorm:
            return hnorm.index(n)
        for i, h in enumerate(hnorm):
            if h and (n in h or h in n):
                return i
        return None
    idx_map = {c: col_idx(c) for c in _OTM_COLUMNS}

    def cell(r, i):
        return str(r[i]).strip() if (i is not None and i < len(r) and r[i] is not None) else ''

    out, kept, deleted, filtered = [], 0, 0, 0
    for r in rows[1:]:
        if cell(r, 13).upper() == 'DELETE':          # col 14 (0-based 13)
            deleted += 1
            continue
        c22 = cell(r, 21)                            # col 22 (0-based 21)
        try:
            c22 = '{:04d}'.format(int(float(c22))) if c22 else ''
        except (ValueError, TypeError):
            pass
        if c22 not in _OTM_KEEP_CODES:
            filtered += 1
            continue
        rec = {c: cell(r, idx_map.get(c)) for c in _OTM_COLUMNS}
        rec['Cpty Name'] = rec['Cpty Name'].upper()      # store counterparty name upper-cased
        out.append(rec)
        kept += 1
    _otm_ensure_meta(out)                                # stamp status='OK' + id per imported row
    return out, kept, deleted, filtered


def _otm_import(ref=None):
    """Find cashflows_*.xlsx in OTM_SOURCE_ROOT, clean + extract the reporting
    columns, write today's JSON and delete the source. Returns a summary dict."""
    ref = ref or datetime.now()
    if not os.path.isdir(OTM_SOURCE_ROOT):
        return {'success': False, 'error': 'Source folder not found: {}'.format(OTM_SOURCE_ROOT)}
    matches = sorted(f for f in os.listdir(OTM_SOURCE_ROOT)
                     if f.lower().startswith('cashflows_') and f.lower().endswith('.xlsx'))
    if not matches:
        return {'success': False, 'error': 'No cashflows_*.xlsx found in {}'.format(OTM_SOURCE_ROOT)}
    src_path = os.path.join(OTM_SOURCE_ROOT, matches[0])
    try:
        rows = _otm_read_rows(src_path)
    except Exception:
        log.warning("[otm] read failed for %s:\n%s", src_path, traceback.format_exc())
        return {'success': False, 'error': 'Could not read {}'.format(matches[0])}
    if not rows or len(rows) < 2:
        return {'success': False, 'error': 'File {} has no data rows'.format(matches[0])}

    out, kept, deleted, filtered = _otm_extract(rows)
    jp = _otm_json_path(ref)
    os.makedirs(os.path.dirname(jp), exist_ok=True)
    _atomic_write_json(jp, out)                 # funil: bump + espelho (§335)
    _ds_write_updated(jp, ref.strftime('%H:%M:%S'))      # cashflows has no in-file time → import time
    try:
        os.remove(src_path)
    except OSError:
        log.warning("[otm] could not delete source %s", src_path)
    log.info("[otm] imported %s: kept %d (deleted %d, filtered %d) → %s",
             matches[0], kept, deleted, filtered, jp)
    return {'success': True, 'file': matches[0], 'rows': kept, 'deleted': deleted,
            'filtered': filtered, 'date': ref.strftime('%Y-%m-%d')}


def _otm_asset_bucket(asset_class):
    """Asset Class → widget bucket: COMMODITIES → commodities, EQUITIES → equities,
    INTEREST_RATE → rates (accent/case tolerant). None if unmapped."""
    ac = str(asset_class or '').upper()
    if 'COMMOD' in ac:
        return 'commodities'
    if 'EQUIT' in ac:
        return 'equities'
    if 'INTEREST' in ac or 'RATE' in ac:
        return 'rates'
    return None


def _otm_collect(ref):
    """Read the OTM JSON for `ref` (date) → display rows + widgets. Dates are
    formatted dd/mm/yyyy and Amount as #,##0.00. Widgets count DISTINCT Trade Ids
    per Asset Class (a Trade Id may span several rows)."""
    widgets = {'total': 0, 'rates': 0, 'equities': 0, 'commodities': 0}
    jp = _otm_json_path(ref)
    rows_out = []
    if os.path.isfile(jp):
        try:
            with open(jp, encoding='utf-8') as fh:
                data = json.load(fh) or []
        except Exception:
            data = []
        if _otm_ensure_meta(data) and data:              # legacy JSON w/o meta → migrate once
            try:
                _otm_save(jp, data)
            except Exception:
                pass
        buckets = {'commodities': set(), 'equities': set(), 'rates': set()}
        all_ids = set()
        for rec in data:
            row = []
            for c in _OTM_COLUMNS:
                v = rec.get(c, '')
                if c in _OTM_DATE_COLS:
                    d = _fcst_parse_date(v)
                    v = d.strftime('%d/%m/%Y') if d else (v or '')
                elif c == 'Amount':
                    v = _swapchar_fmt_value(v)
                elif c == 'Cpty Name':
                    # O nome mostrado é o do REFERENCE DATA, resolvido pelo
                    # **Cpty SPN** da própria linha (`_otm_cpty_name`: entidade
                    # nossa pelo `le-spn`, o resto pelo Reference Data, ignorando
                    # zeros à esquerda). O do arquivo é texto livre do OTM e
                    # divergia do cadastro — é o mesmo nome que o Settlement
                    # Summary usa para agrupar, então mostrar outro fazia a tela
                    # e a apuração falarem de clientes diferentes.
                    #
                    # Resolvido na LEITURA, não na importação: corrigir o
                    # Reference Data passa a valer na hora, sem reimportar o dia.
                    # Sem SPN ou sem cadastro fica o nome do arquivo, para a linha
                    # não sair anônima.
                    v = _otm_cpty_name(rec.get('Cpty SPN', '')) or v
                row.append('' if v is None else v)
            # Append maker/checker meta as the row tail: [...18 data..., status, maker, checker, id]
            row += [rec.get('_ot_status', 'OK'), rec.get('_ot_maker', ''),
                    rec.get('_ot_checker', ''), rec.get('_ot_id', '')]
            rows_out.append(row)
            tid = str(rec.get('Trade Id', '') or '').strip()
            if tid:
                all_ids.add(tid)
                bucket = _otm_asset_bucket(rec.get('Asset Class', ''))
                if bucket:
                    buckets[bucket].add(tid)
        widgets['commodities'] = len(buckets['commodities'])
        widgets['equities'] = len(buckets['equities'])
        widgets['rates'] = len(buckets['rates'])
        widgets['total'] = len(all_ids)                 # distinct Trade Ids overall
    return {'widgets': widgets, 'columns': _OTM_COLUMNS, 'rows': rows_out,
            'updated': _ds_read_updated(jp)}


def _ds_display_json_path(ref, json_key):
    """Cached daily-settlement JSON path for a non-OTM import key (br-onshore,
    eventos-swap-jpm, …), same layout _ds_handle writes to."""
    return os.path.join(OTM_JSON_ROOT, ref.strftime('%Y'), ref.strftime('%m'), ref.strftime('%d'),
                        '{}_{}.json'.format(json_key, ref.strftime('%Y%m%d')))


def _ds_value_col(name):
    """Heuristic: is this a numeric (money/rate) column that should render #,##0.00?
    IDs/accounts/dates/text are excluded so they are never thousand-separated."""
    n = _fcst_norm(name)
    if any(t in n for t in ('codigo', 'conta', 'cnpj', 'cpf', 'data', 'dt ', 'indexador',
                            'ponta', 'nome', 'tipo', 'continuidade', 'estrategia',
                            'funcionalidade', 'reset', 'agenda', 'sistema', 'direction',
                            'spn', 'cetip', 'kapital')):
        return False
    return any(t in n for t in ('valor', 'fator', 'pu ', 'percentual', 'diferenca',
                                'lim.', 'lim inf', 'lim sup', 'taxa', 'amount', 'curve'))


def _ds_display_collect(ref, json_key, columns=None, value_cols=None):
    """Generic read-only settlement view: read the cached JSON for `json_key` and
    map it to display rows. `columns` None → the file's own columns (keys); else the
    given ordered list, each resolved to a JSON key. `value_cols` None → the
    _ds_value heuristic; a set/collection → those exact names get #,##0.00."""
    jp = _ds_display_json_path(ref, json_key)
    rows_out, cols = [], list(columns) if columns else []
    if os.path.isfile(jp):
        try:
            with open(jp, encoding='utf-8') as fh:
                data = json.load(fh) or []
        except Exception:
            data = []
        if data:
            keys = list(data[0].keys())
            if not cols:
                cols = keys
            keymap = {c: (_fcst_resolve_key(keys, [c]) or c) for c in cols}
            if value_cols is None:
                is_val = _ds_value_col
            else:
                vset = {_fcst_norm(c) for c in value_cols}
                is_val = lambda c: _fcst_norm(c) in vset
            for rec in data:
                row = []
                for c in cols:
                    v = rec.get(keymap[c], '')
                    if is_val(c):
                        v = _swapchar_fmt_value(v)
                    row.append('' if v is None else v)
                rows_out.append(row)
    return {'widgets': {'total': len(rows_out)}, 'columns': cols, 'rows': rows_out,
            'updated': _ds_read_updated(jp)}


# ── Other Products › Swap › Events (Swap-InstrumentoFinanceiro-ConsultaContrato) ─
#  Read-only view of the already-processed 'eventos-swap-jpm' JSON. Only the
#  reporting columns below are shown (the file carries many more); each is
#  resolved to the JSON key by name and money/rate columns render #,##0.00 via
#  the _ds_value heuristic.
_EVENTS_COLUMNS = [
    'Código do Contrato', 'Continuidade do Contrato', 'Tipo do Contrato', 'Data de Registro',
    'Dt Última Correção', 'Dt Último Aditamento', 'Valor Base', 'Percentual Índice Termo',
    'Índice Termo', 'PU Inicial Termo', 'PU Atual Termo', 'Dt PU Atual Termo',
    'Fator de Atualização Termo', 'Dt Início', 'Dt Vencimento', 'Valor Base Atual',
    'Valor Amortizado', 'Valor Base Remanescente', 'Valor Antecipado Acumulado',
    'Código da Estratégia',
    'PARTE / Conta', 'PARTE / Nome Simplificado', 'PARTE / CPF/CNPJ', 'PARTE / Ponta',
    'PARTE / Percentual do indexador', 'PARTE / Indexador', 'PARTE / Dt Atualizacao',
    'PARTE / PU Inicial', 'PARTE / PU Atual', 'PARTE / Fator de Juros',
    'PARTE / Fator de Correção', 'PARTE / Valor Juros', 'PARTE / Diferença Juros',
    'PARTE / Valor Amortização', 'PARTE / Diferença Amortização',
    'PARTE / Valor da Curva Atualizado', 'PARTE / Diferença da Curva',
    'PARTE / Lim. Inf.', 'PARTE / Lim. Sup.',
    'CONTRAPARTE / Contraparte', 'CONTRAPARTE / Nome Simplificado', 'CONTRAPARTE / CPF/CNPJ',
    'CONTRAPARTE / Ponta', 'CONTRAPARTE / Percentual do indexador', 'CONTRAPARTE / Indexador',
    'CONTRAPARTE / Dt Atualizacao', 'CONTRAPARTE / PU Inicial', 'CONTRAPARTE / PU Atual',
    'CONTRAPARTE / Fator de Juros', 'CONTRAPARTE / Fator de Correção',
    'CONTRAPARTE / Valor Juros', 'CONTRAPARTE / Diferença Juros',
    'CONTRAPARTE / Valor Amortização', 'CONTRAPARTE / Diferença Amortização',
    'CONTRAPARTE / Valor da Curva Atualizado', 'CONTRAPARTE / Diferença da Curva',
    'CONTRAPARTE / Lim. Inf.', 'CONTRAPARTE / Lim. Sup.',
    'Agenda Prêmio', 'Reset', 'Funcionalidades', 'Parte/Contraparte(Terc. Curva)',
    'Fator/Valor/Taxa(Terc. Curva)',
]


# ── Other Products › Swap › Athena (BrazilOnshoreSettlementsWarningFile) ──────
#  Read-only view of the already-processed 'br-onshore-settlements' JSON (header
#  row 2; LAWTON MULTIMERCADO EXCLUSIVO* already excluded on import). The 9
#  requested columns; Owner curve / Counterparty curve / BRL Net Amount as #,##0.00.
_ATHENA_COLUMNS = ['CETIP ID', 'Kapital ID', 'Owner Legal Entity', 'CounterParty',
                   'SPN', 'Owner curve', 'Counterparty curve', 'BRL Net Amount', 'Direction']
_ATHENA_VALUE_COLS = {'Owner curve', 'Counterparty curve', 'BRL Net Amount'}


def _athena_settlements(ref):
    """O BrazilOnshoreSettlements do dia com o **CounterParty resolvido pelo SPN**.

    O nome que vem no arquivo é texto livre do Athena ('S T E S A L', apelido de
    mesa), e era ele que aparecia na tela e no aviso que vai ao cliente. Ao lado
    dele vem o SPN, que é identificador: `_otm_cpty_name` o resolve pelo cadastro
    `le-spn` quando é entidade nossa e pelo Reference Data quando é cliente,
    ignorando zeros à esquerda dos dois lados.

    UMA função para a página Swap Athena e para o Settlement Advice de Swap. As
    duas mostram a mesma contraparte da mesma operação, e resolver o nome em dois
    lugares é exatamente como elas passariam a discordar — que é o problema que
    esta função existe para não ter.

    SPN em branco ou sem cadastro mantém o nome do arquivo: a linha não pode sair
    anônima só porque o cadastro está incompleto.
    """
    payload = _ds_display_collect(ref, 'br-onshore-settlements',
                                  _ATHENA_COLUMNS, _ATHENA_VALUE_COLS)
    cols = payload.get('columns') or []
    if 'CounterParty' not in cols or 'SPN' not in cols:
        return payload
    ci, si = cols.index('CounterParty'), cols.index('SPN')
    for row in (payload.get('rows') or []):
        if si < len(row) and ci < len(row):
            nome = _otm_cpty_name(row[si])
            if nome:
                row[ci] = nome
    return payload


# ── Other Products › Swap › Settlement Advice ────────────────────────────────
#  A planilha de aviso de liquidação, montada na tela. Uma linha por swap que
#  liquida na data — MESMO universo do Trade Level (`_ops_swap_settling`), para
#  as duas telas nunca discordarem sobre o que liquidou.
#
#  De onde vem cada coluna, que é o que importa aqui:
#
#    Cliente ............ Athena (CounterParty); sem casamento, o Nome
#                         Simplificado do B3 — mesma regra do Trade Level, e pela
#                         mesma razão: 'INTRAGMGTFDO' é apelido de conta e não
#                         casaria com o cadastro de IR.
#    LOB ................ token do Código Identificador da posição (EDG/CEM/…)
#    Número de Contrato . Título do Operations B3
#    Data Operação ...... posição: **Data operação termo** e, só se vazia, Data
#                         início. É o que faz um forward start pagar IR pelo
#                         prazo desde o TRADE (§182).
#    Vencimento/Prazo ... posição (Data vencimento) e a diferença em dias
#    Valor Base Original  posição (Valor base)
#    Indexador Bco/Client posição: Código índice traduzido pelo cadastro
#                         `swap-index`; se der VCP, o Nome Tipo/Classe da perna
#    Curva Banco/Cliente  Athena (Owner curve · Counterparty curve)
#    Resultado Bruto .... Athena (BRL Net Amount)
#    Alíquota/Valor IR .. `_ops_swap_ir_rate` — a MESMA tabela do Trade Level
#    Valor Líquido ...... bruto menos o IR, que sempre ENCOLHE o caixa
#
#  Os três valores em dinheiro saem do MESMO registro do Athena (o arquivo de
#  aviso de liquidação), então a linha fecha: as duas curvas e o resultado são a
#  mesma fonte. Ativo e Valor Base vêm dos eventos porque o Athena não os traz.
_SWADV_COLUMNS = ['Cliente', 'LOB', 'Número de Contrato', 'Data Operação', 'Vencimento',
                  'Prazo', 'Valor Base Original', 'Indexador Banco', 'Curva Banco',
                  'Indexador Cliente', 'Curva Cliente', 'Resultado Bruto', 'Alíquota IR',
                  'Valor IR', 'Valor Líquido']
# O aviso impresso começa em "Número de Contrato": Cliente e LOB são o
# DESTINATÁRIO e o agrupamento, não conteúdo do documento que ele recebe.
_SWADV_EMAIL_FROM = 2
# Coluna que muda de nome quando o evento é prêmio (o pagamento não é vencimento
# de nada — é a parcela de prêmio do dia).
_SWADV_VENC_COL = 4
_SWADV_PREMIO_LABEL = 'Pagamento de Prêmio'


def _swadv_pct(rate):
    """Alíquota em texto ('22.50%'), ou '' quando não deu para afirmar. Vazio ≠
    0%: 0% é isenção conferida, vazio é pedido de conferência."""
    return '' if rate is None else '{:,.2f}%'.format(rate * 100.0)


def _swadv_collect(ref):
    """Linhas do Settlement Advice de SWAP para a data `ref` (datetime), com os
    números crus e os dados do destinatário ao lado das células — a tela usa as
    células, o aviso impresso usa o resto. Uma coleta só para os dois, para o que
    o cliente recebe não poder divergir do que a tela mostra."""
    _jp, opb3 = _opb3_load(ref)
    titulos, by_titulo = _ops_swap_settling(opb3)
    if not titulos:
        return []
    tipo_maps = _opb3_tipo_maps(ref)
    terms = _ops_swap_pos_terms(ref)

    # Mesma coleta da página Swap Athena, com o CounterParty já resolvido pelo
    # SPN — o nome do aviso que vai ao cliente é o do Reference Data.
    athena = _athena_settlements(ref)
    ai = {c: i for i, c in enumerate(athena.get('columns') or [])}
    by_cetip = {}
    for row in athena.get('rows') or []:
        cet = str(row[ai['CETIP ID']] if 'CETIP ID' in ai else '').strip().upper()
        if cet:
            by_cetip.setdefault(cet, row)

    # Equity: o Swap Athena é só de CEM. Sem esta rota (Operations B3 → Latam →
    # OTM) o aviso saía com o nome curto da B3 e com as três colunas de valor em
    # branco — a MESMA função que o Trade Level usa, para o documento que vai ao
    # cliente não poder discordar da tela.
    eqlink = _ops_equity_link(ref)

    # O arquivo de EVENTOS deixou de ser lido aqui: Valor Base e os indexadores
    # passaram a sair da posição, que já é lida para as datas. Uma fonte a menos
    # é um join a menos para falhar em silêncio.
    def _cell(row, idx_map, name):
        i = idx_map.get(name)
        return '' if (row is None or i is None or i >= len(row)) else str(row[i] or '').strip()

    def _dt(d):
        return d.strftime('%d/%m/%Y') if d else ''

    spn_by_name = _ndfsum_refdata_spn()
    premio = _ops_norm_event('PAGAMENTO DE PREMIO')

    out = []
    for titulo, rec in titulos:
        key = titulo.upper()
        arow = by_cetip.get(key)
        eq = eqlink.get(key) or {}
        pos = terms.get(_fcst_norm_contract(titulo).upper()) or {}
        cliente = (_cell(arow, ai, 'CounterParty')
                   or eq.get('counterparty', '')
                   or str(rec.get('Contraparte (Nome Simpl.)', '') or '').strip())
        # Perna interna não recebe aviso — o documento é endereçado ao cliente, e
        # a entidade nossa produziria um aviso para nós mesmos. Mesma regra e
        # mesma função do Trade Level; só para equity, pelo mesmo motivo de lá.
        if eq and _ops_is_internal_cpty(cliente, eq.get('spn', '')):
            continue
        op_dt = pos.get('op') or eq.get('trade_date')
        # Vencimento do aviso = a data da LIQUIDAÇÃO. É esta parcela que está
        # sendo paga hoje; o vencimento do swap só interessa quando os dois
        # coincidem. E o Prazo é a diferença entre as duas datas impressas, senão
        # o cliente confere a conta do aviso e ela não fecha.
        venc_dt = ref.date()
        prazo = (venc_dt - op_dt).days if op_dt else None

        # Em equity as três colunas de valor saem do OTM Settlements, com a regra
        # da mesa: **Curva Banco = os fluxos positivos, Curva Cliente = os
        # negativos, Resultado Bruto = a soma dos dois**. O Athena não tem essas
        # operações, e o aviso saía com as três em branco.
        curva_banco_n = (_mtm_parse_num(_cell(arow, ai, 'Owner curve')) if arow
                         else eq.get('curva_banco'))
        curva_cliente_n = (_mtm_parse_num(_cell(arow, ai, 'Counterparty curve')) if arow
                           else eq.get('curva_cliente'))
        bruto_txt = _cell(arow, ai, 'BRL Net Amount')
        bruto = _mtm_parse_num(bruto_txt) if bruto_txt else None
        if bruto is None and eq.get('settlement') is not None:
            bruto = eq['settlement']
            bruto_txt = _ops_fmt_amt(bruto)
        # A direção vem do texto do Athena (é o que a fórmula da planilha lê); o
        # sinal do Resultado Bruto só entra quando o texto falta, e assume a
        # mesma convenção do settlement — negativo é o banco pagando.
        rate = _ops_swap_ir_rate(cliente, prazo,
                                 _ops_cpty_receives(_cell(arow, ai, 'Direction'), bruto))
        ir = None if (rate is None or bruto is None) else abs(bruto) * rate
        # O IR retido sempre ENCOLHE o que se movimenta, seja qual for o sinal.
        liq = None if (bruto is None or ir is None) else (bruto - ir if bruto >= 0 else bruto + ir)

        # Prêmio quando TODOS os eventos registrados do Título são prêmio. Um
        # swap que paga prêmio e diferencial no mesmo dia é liquidação comum —
        # chamar o conjunto de "Pagamento de Prêmio" no assunto esconderia o
        # diferencial que também está na tabela.
        # `by_titulo` já vem só com os eventos registrados — não refiltrar aqui,
        # senão fica parecendo que a coleção traz eventos de fora.
        evs = {_ops_norm_event(r.get('Tipo Operação', '')) for r in by_titulo.get(key, [])}
        ref_rec = spn_by_name.get(_fcst_norm(cliente), {})
        out.append({
            'cells': [
                cliente,
                _fcst_lob(_opb3_tipo_for(rec, tipo_maps)) or '',
                titulo,
                _dt(op_dt),
                _dt(venc_dt),
                '' if prazo is None else '{:,}'.format(prazo).replace(',', '.'),
                _swapchar_fmt_value(pos.get('valor_base', '')),
                pos.get('idx_banco', ''),
                _cell(arow, ai, 'Owner curve') if arow else _ops_fmt_amt(curva_banco_n),
                pos.get('idx_cliente', ''),
                _cell(arow, ai, 'Counterparty curve') if arow else _ops_fmt_amt(curva_cliente_n),
                bruto_txt,
                _swadv_pct(rate),
                _ops_fmt_amt(ir),
                _ops_fmt_amt(liq),
            ],
            'counterparty': cliente,
            'lob': _fcst_lob(_opb3_tipo_for(rec, tipo_maps)) or '',
            'legal': _cell(arow, ai, 'Owner Legal Entity'),
            'spn': ref_rec.get('spn', '') or _cell(arow, ai, 'SPN'),
            'taxid': ref_rec.get('taxid', ''),
            'premium': bool(evs) and evs == {premio},
            'bruto': bruto, 'ir': ir, 'liquido': liq, 'rate': rate,
            # Crus para o aviso: ele imprime em BR (R$ 1.234,56), e a tela em US.
            # Reformatar o texto de uma para a outra erraria no primeiro valor
            # com separador ambíguo — do número não há como errar.
            # MESMA leitura da célula (`_swapchar_fmt_value`, logo acima): o
            # arquivo de posição usa vírgula DECIMAL, e ler esse texto com o
            # parser de uso geral multiplicava o Valor Base por cem no aviso.
            'valor_base': _swapchar_value_num(pos.get('valor_base', '')),
            'curva_banco': curva_banco_n,
            'curva_cliente': curva_cliente_n,
        })
    return out


def _swadv_email_rows(ref):
    """Linhas do aviso impresso: as mesmas do Settlement Advice, da coluna
    `Número de Contrato` em diante e com os valores em BR (`R$ #.##0,00`, o
    negativo entre parênteses). Cliente e LOB ficam de fora — são o destinatário
    e o agrupamento, não conteúdo do documento."""
    from apps.pages import otc_emails

    def money(v):
        return otc_emails._brl(v) if v is not None else ''

    out = []
    for r in _swadv_items(ref):
        c = r['cells']
        out.append(dict(r, cells=[
            c[2],                                   # Número de Contrato
            c[3],                                   # Data Operação
            c[4],                                   # Vencimento (ou Pagamento de Prêmio)
            c[5],                                   # Prazo (#.##0)
            money(r['valor_base']),
            c[7],                                   # Indexador Banco
            money(r['curva_banco']),
            c[9],                                   # Indexador Cliente
            money(r['curva_cliente']),
            money(r['bruto']),
            '' if r['rate'] is None else otc_emails._br(r['rate'] * 100.0) + '%',
            money(r['ir']),
            money(r['liquido']),
        ]))
    return out


# ── Edições manuais do aviso de Swap ─────────────────────────────────────────
#  A linha é derivada de cinco arquivos; quando um deles vem errado (ou falta), a
#  mesa precisa corrigir a célula e mandar o aviso assim mesmo. As correções
#  ficam num overlay do DIA, fora dos arquivos de origem: reimportar o batch não
#  as apaga, e nada do que veio da B3/Athena é sobrescrito.
#
#  A chave é o NÚMERO DE CONTRATO, não a posição na tela — a tabela ordena por
#  cliente e a posição muda a cada carga.
_SWADV_KEY_COL = 2
# Célula editada que também é NÚMERO: sem isto a tela mostraria o valor corrigido
# e o aviso impresso continuaria imprimindo o original, que é a divergência que
# este módulo inteiro existe para evitar.
_SWADV_NUM_FIELDS = {6: 'valor_base', 8: 'curva_banco', 10: 'curva_cliente',
                     11: 'bruto', 13: 'ir', 14: 'liquido'}


def _swadv_edits_path(ref):
    """Ao lado do overlay do Settlement Summary, na pasta do dia — mesma
    convenção do `_opssum_meta_path`."""
    return os.path.join(OTM_JSON_ROOT, ref.strftime('%Y'), ref.strftime('%m'),
                        ref.strftime('%d'),
                        'swap-settlement-advice_{}.json'.format(ref.strftime('%Y%m%d')))


def _swadv_edits_load(ref):
    fp = _swadv_edits_path(ref)
    try:
        with open(fp, encoding='utf-8') as fh:
            d = json.load(fh)
        return (fp, d) if isinstance(d, dict) else (fp, {})
    except Exception:
        return fp, {}


def _swadv_apply_edits(items, edits):
    """Aplica o overlay às linhas coletadas: some com as apagadas e sobrescreve
    as células editadas — nas células E nos números crus, que são o que o aviso
    impresso lê."""
    if not edits:
        return items
    out = []
    for r in items:
        key = str(r['cells'][_SWADV_KEY_COL] or '').strip()
        e = edits.get(key) or {}
        if e.get('deleted'):
            continue
        cells = list(r['cells'])
        rec = dict(r)
        for idx_s, val in (e.get('cells') or {}).items():
            try:
                i = int(idx_s)
            except (TypeError, ValueError):
                continue
            if not 0 <= i < len(cells):
                continue
            cells[i] = str(val)
            fld = _SWADV_NUM_FIELDS.get(i)
            # `_conf_to_float` e não `_mtm_parse_num`: a TABELA mostra em US
            # ('12,345.67') e o aviso imprime em BR ('R$ 12.345,67'), então o
            # operador digita ora um, ora outro. O parser tolerante entende os
            # dois; o outro lê '12.345,67' como 12,345 e o aviso sai com um valor
            # mil vezes menor, sem erro nenhum na tela.
            if fld:
                rec[fld] = _conf_to_float(str(val))
            elif i == 12:                      # Alíquota IR ('22.50%') → fração
                n = _conf_to_float(str(val).replace('%', ''))
                rec['rate'] = None if n is None else n / 100.0
            elif i == 0:
                rec['counterparty'] = str(val)
            elif i == 1:
                rec['lob'] = str(val)
        rec['cells'] = cells
        out.append(rec)
    return out


def _swadv_items(ref):
    """As linhas do aviso JÁ com as edições manuais do dia aplicadas. É por aqui
    que a tela e o aviso impresso passam — os dois, sempre."""
    return _swadv_apply_edits(_swadv_collect(ref), _swadv_edits_load(ref)[1])


def _swadv_rows(ref):
    """Só as células, na ordem de `_SWADV_COLUMNS` — o que a tela consome."""
    return [r['cells'] for r in _swadv_items(ref)]


def _swadv_ref_and_key(payload):
    """(ref, contrato) de um payload das ações da linha, ou (None, '')."""
    ds = str((payload or {}).get('date') or '').strip()
    try:
        ref = datetime.strptime(ds[:10], '%Y-%m-%d') if ds else datetime.now()
    except ValueError:
        ref = datetime.now()
    return ref, str((payload or {}).get('contrato') or '').strip()


def _swadv_email_headers(premium):
    """Cabeçalho da tabela do aviso: as colunas da tela a partir de `Número de
    Contrato`. No aviso de prêmio a coluna `Vencimento` vira `Pagamento de
    Prêmio` — não é vencimento de nada, é a parcela do dia."""
    cols = list(_SWADV_COLUMNS)
    if premium:
        cols[_SWADV_VENC_COL] = _SWADV_PREMIO_LABEL
    return cols[_SWADV_EMAIL_FROM:]


# ── Other Products › NDF › Settlement Advice (commodities) ───────────────────
#  A planilha de aviso do TERMO DE MERCADORIA, montada na tela. Universo:
#  Operations B3 com **Tipo Operação = RESGATE**, **Tipo Título = TER** e a coluna
#  derivada **Type = COMMODITIES** (que para TER é a Classe do Ativo Subjacente da
#  posição) — é o resgate do termo de commodity, e só ele.
#
#  De onde vem cada coluna:
#
#    Contraparte .......... Live Position NDF (Nome da Contraparte)
#    B3 ID ................ Título do Operations B3
#    Nº da Confirmação .... posição NDF: Código Identificador do contrato
#    Data de Início ....... posição NDF: Data de Emissao
#    Ativo Subjacente ..... código do subjacente → `Subjacente.json` → COMMODITY(CÓD)
#    Ptax ................. posição NDF: Data de Fixing da Moeda
#    Cotação Mercadoria ... posição NDF: Data de Fixing do Ativo Subjacente; vazia
#                           numa operação ASIÁTICA, vira "Média Fev/2027" a partir
#                           do mês/ano da 1ª data de verificação
#    Quantidade ........... posição NDF: Valor Base no registro
#    Resultado Apurado .... INTERNO: soma do OTM Settlements pelos Trade Ids cujo
#                           sufixo (depois do hífen) bate com o do Nº da Confirmação
#    IR 0,005% ............ só quando o banco PAGA (apurado < 0) e a contraparte
#                           não é LAWTON — porte da fórmula da planilha
#    Resultado Líquido .... o IR sempre ENCOLHE o caixa (mesma regra do aviso FX)
#    Settlement Net ....... net type do Reference Data
_NDFADV_COLUMNS = ['Contraparte', 'B3 ID', 'Nº da Confirmação', 'Data de Início da Operação',
                   'Ativo Subjacente', 'Ptax', 'Cotação Mercadoria', 'Quantidade da Operação',
                   'Resultado Apurado (R$)', 'IR 0,005% (R$)', 'Resultado Líquido (R$)',
                   'Settlement Net']
_NDFADV_IR_RATE = 0.00005                      # 0,005%
_NDFADV_MESES = ('Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun',
                 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez')


def _subjacente_commodity(code):
    """`Codigo do Ativo Subjacente` → 'COMMODITY(CÓDIGO)', pelo cadastro de
    Subjacente do B3 Index Results. Código sem commodity registrada volta só o
    código: melhor mostrar o que veio do arquivo do que inventar um nome."""
    code = str(code or '').strip()
    if not code:
        return ''
    name = (_subjacente_map() or {}).get(code.upper(), '')
    return '{}({})'.format(name.strip().upper(), code) if name else code


_SUBJ_CACHE = {'mtime': None, 'map': {}}


def _subjacente_map():
    """{código(upper) → Commodity} do Subjacente.json, cacheado por mtime.

    O arquivo tem ~7.800 linhas e repete o mesmo código em várias (uma por Tipo
    IF: OPC, COE, TER…). A primeira com Commodity preenchida vence — as demais
    trazem o MESMO nome, e uma linha sem Commodity não pode apagar a que tem."""
    path = os.path.join(_B3_DATA_DIR, 'Subjacente.json')
    try:
        mt = os.path.getmtime(path)
    except OSError:
        return {}
    if _SUBJ_CACHE['mtime'] != mt:
        try:
            with open(path, encoding='utf-8') as fh:
                data = json.load(fh) or []
        except Exception:
            data = []
        m = {}
        for rec in data:
            code = str(rec.get('Codigo do Ativo Subjacente', '') or '').strip().upper()
            comm = str(rec.get('Commodity', '') or '').strip()
            if code and comm and code not in m:
                m[code] = comm
        _SUBJ_CACHE['mtime'] = mt
        _SUBJ_CACHE['map'] = m
    return _SUBJ_CACHE['map']


def _ndfc_ir_exempt(client):
    """A contraparte é isenta do IR de 0,005%? Cadastro `ndfc-ir-exempt`.

    A MESMA lista serve o Settlement Advice e o Trade Level: são o mesmo imposto
    sobre a mesma operação, e duas listas divergiriam sem erro nenhum — uma tela
    reteria e a outra não."""
    cn = _fcst_norm(client).strip()
    if not cn:
        return False
    for row in _mapping_rows('ndfc-ir-exempt'):
        pat = _fcst_norm(row.get('CLIENT', '')).strip()
        if not pat:
            continue
        if cn.startswith(pat) if 'starts' in _fcst_norm(row.get('MATCH', '')) else cn == pat:
            return True
    return False


_REFDATA_TAXID_CACHE = {'mtime': None, 'map': {}}
_REFDATA_SPN_CACHE = {'mtime': None, 'map': {}}


def _spn_key(v):
    """SPN comparável. O OTM grava '1234567', o Reference Data às vezes
    '1234567.0' (veio de planilha) e às vezes com zero à esquerda. Só os dígitos,
    sem os zeros da frente — comparar as strings deixava metade sem casar.

    O rabo `.0` sai ANTES de tirar a pontuação: depois, aquele zero viraria mais
    um dígito no fim e '1234567.0' deixaria de casar com '1234567'."""
    s = str(v or '').strip()
    if s.endswith('.0'):
        s = s[:-2]
    return re.sub(r'\D', '', s).lstrip('0')


_REFDATA_TRIPLE_CACHE = {'mtime': None, 'rows': []}


def _refdata_records():
    """As linhas CRUAS do Reference Data — DB-first (fase 3, HANDOFF §328).

    O `reference_data.db` responde quando o manifest prova que ele reflete o
    `RefData.json` atual; senão vale o JSON de sempre (e o espelho é avisado
    para se curar). Os três índices derivados abaixo continuam cacheados pelo
    MTIME DO JSON — que é exatamente a chave do contrato de frescor, então as
    duas fontes respondem a mesma pergunta."""
    try:
        from apps.pages import duck_read
        rows = duck_read.refdata_rows()
    except Exception:                                       # noqa: BLE001
        rows = None
    if rows is not None:
        return rows
    path = os.path.join(_B3_DATA_DIR, 'RefData.json')
    try:
        with open(path, encoding='utf-8') as fh:
            data = json.load(fh) or []
    except Exception:                                       # noqa: BLE001
        data = []
    return data if isinstance(data, list) else []


def _refdata_triples():
    """[{name, spn, taxid}] do RefData.json, cacheado por mtime.

    É o que alimenta o autocompletar do cadastro MT300: escolhido QUALQUER um dos
    três, os outros dois se preenchem. Sem isso a mesa copiava SPN e CNPJ à mão
    de outra tela — e um dígito errado ali não dá erro, só faz a operação sumir
    do e-mail (o casamento é por esses campos).

    Uma linha só com o nome não serve para completar nada, então entra apenas
    quem tem nome E pelo menos um dos dois identificadores."""
    path = os.path.join(_B3_DATA_DIR, 'RefData.json')
    try:
        mt = os.path.getmtime(path)
    except OSError:
        return []
    if _REFDATA_TRIPLE_CACHE['mtime'] != mt:
        data = _refdata_records()
        vistos, out = set(), []
        for rec in data:
            nome = str(rec.get('COUNTERPARTY', '') or '').strip()
            spn = str(rec.get('SPN', '') or '').strip()
            taxid = str(rec.get('TAX ID', '') or '').strip()
            if not nome or not (spn or taxid):
                continue
            chave = (nome.upper(), spn, re.sub(r'\D', '', taxid))
            if chave in vistos:
                continue
            vistos.add(chave)
            out.append({'name': nome, 'spn': spn, 'taxid': taxid})
        out.sort(key=lambda r: r['name'].upper())
        _REFDATA_TRIPLE_CACHE['mtime'] = mt
        _REFDATA_TRIPLE_CACHE['rows'] = out
    return _REFDATA_TRIPLE_CACHE['rows']


def _refdata_by_spn():
    """{SPN → COUNTERPARTY} do RefData.json, cacheado por mtime."""
    path = os.path.join(_B3_DATA_DIR, 'RefData.json')
    try:
        mt = os.path.getmtime(path)
    except OSError:
        return {}
    if _REFDATA_SPN_CACHE['mtime'] != mt:
        data = _refdata_records()
        m = {}
        for rec in data:
            spn = _spn_key(rec.get('SPN', ''))
            name = str(rec.get('COUNTERPARTY', '') or '').strip()
            if spn and name and spn not in m:
                m[spn] = name
        _REFDATA_SPN_CACHE['mtime'] = mt
        _REFDATA_SPN_CACHE['map'] = m
    return _REFDATA_SPN_CACHE['map']


def _otm_cpty_name(spn):
    """Razão social da contraparte a partir do **Cpty SPN** do OTM Settlements.

    O SPN é o identificador que o OTM carrega na própria linha do fluxo, e é o
    único que existe igual dos dois lados. Antes o nome do swap saía do
    `CounterParty` do Athena e o de commodities do `Nome da Contraparte` da
    posição — dois textos livres, escritos por sistemas diferentes, que
    divergiam em pontuação e sufixo e faziam o MESMO cliente virar duas linhas
    no Settlement Summary.

    Duas fontes, nesta ordem:
      1. cadastro **`le-spn`** — se o SPN é de uma entidade nossa, o nome é o
         `Reference Data Name` dela. Entidade própria não está no Reference Data
         como contraparte, então procurá-la lá devolveria vazio;
      2. **Reference Data** por SPN.

    Vazio quando não acha: quem chama mantém o nome que já tinha, para a linha
    não sair anônima."""
    k = _spn_key(spn)
    if not k:
        return ''
    for r in _mapping_rows('le-spn'):
        if _spn_key(r.get('SPN', '')) == k:
            return (str(r.get('NAME', '') or '').strip()
                    or str(r.get('LE', '') or '').strip())
    return _refdata_by_spn().get(k, '')


@_once_per_request
def _refdata_by_taxid():
    """{CNPJ só dígitos → COUNTERPARTY} do RefData.json, cacheado por mtime.

    `@_once_per_request` porque as CINCO telas de Live Position resolvem o nome
    da contraparte por aqui uma vez POR LINHA (`_lp_cpty_by_taxid`): o cache por
    mtime já evitava reler o arquivo, mas não o `stat` que decide se ele mudou —
    e esse stat ficava dentro do laço. Medido: 1,00 stat por linha. No share
    isso é uma ida à rede por linha, e a tela levava minutos sem erro nenhum.

    Comparação por DÍGITOS: o RefData guarda mascarado (45.985.371/0001-08) e a
    posição da B3 guarda só números. Comparar as strings não casaria nada."""
    path = os.path.join(_B3_DATA_DIR, 'RefData.json')
    try:
        mt = os.path.getmtime(path)
    except OSError:
        return {}
    if _REFDATA_TAXID_CACHE['mtime'] != mt:
        data = _refdata_records()
        m = {}
        for rec in data:
            digits = ''.join(ch for ch in str(rec.get('TAX ID', '') or '') if ch.isdigit())
            name = str(rec.get('COUNTERPARTY', '') or '').strip()
            if digits and name and digits not in m:
                m[digits] = name
        _REFDATA_TAXID_CACHE['mtime'] = mt
        _REFDATA_TAXID_CACHE['map'] = m
    return _REFDATA_TAXID_CACHE['map']


def _b3_is_omnibus(account):
    """A Conta Contraparte é uma conta guarda-chuva? Cadastro `b3-accounts`.

    Quem responde é o TIPO DA CONTA, não estar na tabela: o cadastro passou a
    listar TODAS as contas B3 das nossas entidades, e a PRÓPRIA identifica quem
    é — é a nossa. Só CLIENT 1 e CLIENT 2 são guarda-chuva, e nelas o nome que
    vem da B3 é o do titular; o cliente sai do CNPJ.

    Comparação por dígitos: a conta aparece ora `73760.10-2`, ora `7376010 2`."""
    d = ''.join(ch for ch in str(account or '') if ch.isdigit())
    if not d:
        return False
    for row in _mapping_rows('b3-accounts'):
        rd = ''.join(ch for ch in str(row.get('ACCOUNT', '') or '') if ch.isdigit())
        if rd and rd == d:
            return _b3_account_type(row.get('ACCOUNT TYPE', '')) in _B3_CLIENT_ACCOUNT_TYPES
    return False


def _b3_account_row(account):
    """A linha do cadastro `b3-accounts` daquela conta, ou None.

    É a resposta para "esta conta é de uma entidade NOSSA?" — estar no cadastro
    é ser interna, porque a tabela lista as contas B3 das nossas entidades
    (Banco, MGT, Lawton, Atacama…) e nada mais. A comparação é por DÍGITOS: a
    mesma conta aparece `73760.10-2` num arquivo e `7376010 2` no outro, e
    comparar string casa silenciosamente nada."""
    d = _acc_digits(account)
    if not d:
        return None
    for row in _mapping_rows('b3-accounts'):
        if _acc_digits(row.get('ACCOUNT', '')) == d:
            return row
    return None


def _b3_msg_view_use(account):
    """A mensageria do Operations B3 sai na visão desta conta?

    Devolve `True` (gera), `False` (não gera) — e `True` também para a conta que
    NÃO está no cadastro, que é a conta de terceiro: a mensagem que a mesa manda
    hoje sai da visão do Banco contra o cliente, e travá-la por falta de linha
    calaria a rotina inteira em qualquer instância que não tivesse aberto o
    /mapping.

    O que a coluna decide é a PONTA da liquidação intragrupo. O mesmo pagamento
    chega pelos dois arquivos, espelhado (a visão do Banco e a da outra
    entidade); as duas virando e-mail, o time cobraria duas vezes o que é uma
    coisa só. Era uma regra escrita no código — casa == MGT e contraparte ==
    Banco —, e ela só conhecia esse par: a visão de Lawton e a de Atacama
    passavam direto."""
    row = _b3_account_row(account)
    if row is None:
        return True
    return not _fcst_norm(row.get('MESSAGING', '')).strip().startswith('disreg')


def _b3_account_le(account):
    """Legal Entity dona da conta no cadastro `b3-accounts` ('' se não é nossa).
    É por aqui que a mensageria sabe que a contraparte é o Lawton ou a Atacama
    sem depender da grafia do Nome Simplificado no arquivo da B3."""
    row = _b3_account_row(account)
    return str((row or {}).get('LE', '') or '').strip().upper()


def _b3_account_refdata_name(account):
    """Nome da entidade dona da conta como ele está no Reference Data ('' se não
    cadastrado). O Nome Simplificado ao lado é o apelido de 20 caracteres da B3
    e não serve para endereçar documento nenhum."""
    row = _b3_account_row(account)
    return str((row or {}).get('REFDATA NAME', '') or '').strip()


def _b3_participant_name(le):
    """Nome Simplificado da entidade `le` no cadastro `b3-accounts` — o que vai
    no campo Participante do header dos arquivos TER.

    Era um dicionário fixo no código, e a mesma resposta estava escrita em dois
    lugares (aqui e no `source_note` do File Interpreter). A conta PRÓPRIA vem
    primeiro por ser a da entidade; qualquer linha da LE serve, já que o nome é
    o da entidade e não o da conta. Sem cadastro devolve '' — quem chama é que
    decide o que fazer, e o arquivo para a B3 não sai com o campo em branco."""
    alvo = _fcst_norm(le).strip()
    if not alvo:
        return ''
    fallback = ''
    for row in _mapping_rows('b3-accounts'):
        if _fcst_norm(row.get('LE', '')).strip() != alvo:
            continue
        nome = str(row.get('SIMPLIFIED NAME', '') or '').strip()
        if not nome:
            continue
        if _b3_account_type(row.get('ACCOUNT TYPE', '')) == 'OWN':
            return nome
        fallback = fallback or nome
    return fallback


def _ndfc_split_by_commodity(client):
    """A contraparte recebe um aviso POR COMMODITY? Cadastro `ndfc-advice-split`.

    Fora da lista, um aviso pode trazer alumínio e café na mesma tabela."""
    cn = _fcst_norm(client).strip()
    if not cn:
        return False
    for row in _mapping_rows('ndfc-advice-split'):
        pat = _fcst_norm(row.get('CLIENT', '')).strip()
        if not pat:
            continue
        if cn.startswith(pat) if 'starts' in _fcst_norm(row.get('MATCH', '')) else cn == pat:
            return True
    return False


def _ndfc_ir(apurado, client):
    """IR de 0,005% do Termo de Mercadoria: só quando o BANCO paga (apurado < 0)
    e a contraparte não é isenta. Arredondado a 2 casas, como a planilha."""
    if apurado is None or apurado >= 0 or _ndfc_ir_exempt(client):
        return 0.0
    return round(abs(apurado) * _NDFADV_IR_RATE, 2)


def _ndfadv_media_label(date_str):
    """'Média Fev/2027' a partir de uma data de verificação. Só o mês e o ano: a
    cotação da asiática é a média do período, não a de um dia."""
    d = _fcst_parse_date(date_str)
    if not d:
        return ''
    return 'Média {}/{}'.format(_NDFADV_MESES[d.month - 1], d.year)


@_req_cached
def _ndfadv_otm_by_suffix(ref):
    """({sufixo do Trade Id → Σ Amount}, {sufixo → Cpty SPN}) do OTM do dia.

    O Trade Id do OTM e o Nº da Confirmação carregam o MESMO identificador depois
    do hífen; o prefixo difere entre os sistemas. Por isso o join é pelo sufixo —
    e a soma é necessária porque um trade aparece em várias linhas de fluxo.

    O SPN sai junto porque é a mesma linha: é dele que sai o nome da contraparte
    do aviso (`_otm_cpty_name`), e uma segunda leitura do arquivo só para buscá-lo
    abriria espaço para as duas discordarem."""
    _jp, otm = _otm_load(ref)
    out, spn = {}, {}
    for rec in (otm or []):
        tid = str(rec.get('Trade Id', '') or '').strip().upper()
        if not tid:
            continue
        suf = tid.rsplit('-', 1)[-1]
        if not suf:
            continue
        amt = _conf_to_float(rec.get('Amount'))
        if amt is not None:
            out[suf] = out.get(suf, 0.0) + amt
        if _spn_key(rec.get('Cpty SPN', '')) and suf not in spn:
            spn[suf] = str(rec.get('Cpty SPN', '') or '').strip()
    return out, spn


def _ndfadv_collect(ref):
    """Linhas do NDF Settlement Advice (commodities) para a data `ref`."""
    # Peneirado pelo `opb3-events` UMA vez, no topo: a lista de Títulos e a soma
    # do Settlement B3 lá embaixo têm de enxergar as mesmas linhas. Filtrar só a
    # seleção deixava a operação cancelada fora da tabela e dentro do total.
    opb3 = _opb3_settle_rows(ref)
    if not opb3:
        return []
    tipo_maps = _opb3_tipo_maps(ref)

    titulos, seen = [], set()
    for rec in opb3:
        titulo = str(rec.get('Título', '') or '').strip()
        if not titulo or titulo.upper() in seen:
            continue
        # O EVENTO que liquida sai do cadastro `opb3-events` (o filtro já rodou
        # em `_opb3_settle_rows`), não de um teste aqui. O 'resgate' estava fixo
        # neste ponto e era a segunda resposta para a pergunta que o cadastro
        # existe para responder — a mesa registrava um evento novo lá e o aviso
        # continuava cego a ele, sem erro nenhum. O que fica é o recorte de
        # PRODUTO, que é o assunto desta tela.
        if 'ter' not in _fcst_norm(rec.get('Tipo Título', '')):
            continue
        if 'commodit' not in _fcst_norm(_opb3_tipo_for(rec, tipo_maps)):
            continue
        seen.add(titulo.upper())
        titulos.append((titulo, rec))
    if not titulos:
        return []

    # Posição NDF pela TELA do Live Position: as datas já vêm dd/mm/yyyy e as
    # colunas de média asiática já vêm resolvidas do bloco posicional.
    lp = _lpndf_collect(ref)
    li = {c: i for i, c in enumerate(lp.get('columns') or [])}
    by_contract = {}
    for row in lp.get('rows') or []:
        k = str(row[li['Contrato']] if 'Contrato' in li else '').strip().upper()
        if k:
            by_contract.setdefault(k, row)

    otm_by_suffix, otm_spn = _ndfadv_otm_by_suffix(ref)
    spn_by_name = _ndfsum_refdata_spn()
    cpd = _cpd_load()

    def _lcell(row, name):
        i = li.get(name)
        return '' if (row is None or i is None or i >= len(row)) else str(row[i] or '').strip()

    out = []
    for titulo, rec in titulos:
        lrow = by_contract.get(titulo.upper())
        # Contraparte: numa conta OMNIBUS o nome que vem da B3 é o do titular do
        # guarda-chuva, não o do cliente — quem é o cliente sai do CNPJ da
        # posição, procurado no RefData. Fora do omnibus, o nome da posição vale.
        # Errar isso manda o aviso de liquidação para o cliente errado.
        conf = _lcell(lrow, 'Codigo Identificador')
        suf = conf.rsplit('-', 1)[-1].upper() if conf else ''
        # 1ª fonte: o **Cpty SPN** do OTM. É um identificador na própria linha do
        # fluxo, e resolve o omnibus de graça — a linha do OTM é do trade, não da
        # conta guarda-chuva.
        cliente = _otm_cpty_name(otm_spn.get(suf, '')) if suf else ''
        # A coluna 'CPF/CNPJ da Contraparte' do Live Position **já resolve** o
        # nome no RefData, e é a mesma pergunta que se faz aqui — deixá-la
        # respondida em dois lugares é deixá-los discordarem. Ela só volta como
        # DOCUMENTO quando não há cadastro, e nesse caso o omnibus não resolve,
        # exatamente como antes (o `.get(cnpj, '')` devolvia '' e caía para o
        # nome da posição).
        doc = _lcell(lrow, 'CPF/CNPJ da Contraparte')
        if not cliente and _b3_is_omnibus(rec.get('Conta Contraparte', '')) and doc:
            cliente = '' if _lp_is_taxid(doc) else doc
        cliente = (cliente
                   or _lcell(lrow, 'Nome da Contraparte')
                   or str(rec.get('Contraparte (Nome Simpl.)', '') or '').strip())

        # Cotação Mercadoria: a data única do fixing do subjacente; vazia (o caso
        # da asiática), o mês/ano da 1ª data de verificação.
        cot = _lcell(lrow, 'Data de Fixing do Ativo Subjacente')
        if not cot:
            cot = _ndfadv_media_label(_lcell(lrow, 'Média Asiática (data) 1'))

        apurado = otm_by_suffix.get(suf) if suf else None
        ir = _ndfc_ir(apurado, cliente)
        # O IR retido sempre ENCOLHE o que se movimenta (regra do aviso de FX).
        liq = None if apurado is None else (apurado - ir if apurado >= 0 else apurado + ir)

        # SPN: o do OTM quando existe (é o mesmo que deu o nome); senão, o de
        # sempre, pelo nome no Reference Data. Guardar o do OTM evita o caminho
        # de volta nome → SPN, que erra em toda diferença de pontuação.
        ref_rec = spn_by_name.get(_fcst_norm(cliente), {})
        spn = str(otm_spn.get(suf, '') or '').strip() or ref_rec.get('spn', '')
        net_type = _ndfsum_net_type(_cpd_find(cpd, spn) if spn else None)

        # Settlement B3: soma de TODAS as linhas daquele Título no Operations B3
        # (o caixa do dia), mesma regra do Trade Level de swap.
        b3_vals = [_conf_to_float(r.get('Valor')) for r in opb3
                   if str(r.get('Título', '') or '').strip().upper() == titulo.upper()]
        b3_vals = [v for v in b3_vals if v is not None]
        out.append({
            'cells': [
                cliente,
                titulo,
                conf,
                _lcell(lrow, 'Data de Emissao'),
                _subjacente_commodity(_lcell(lrow, 'Codigo do Ativo Subjacente')),
                _lcell(lrow, 'Data de Fixing da Moeda'),
                cot,
                _lcell(lrow, 'Valor Base no registro'),
                _ops_fmt_amt(apurado),
                _ops_fmt_amt(ir),
                _ops_fmt_amt(liq),
                net_type,
            ],
            'counterparty': cliente,
            'legal': _lcell(lrow, 'Nome da Parte'),
            'spn': spn,
            'taxid': ref_rec.get('taxid', ''),
            'net_type': net_type,
            'commodity': _subjacente_commodity(_lcell(lrow, 'Codigo do Ativo Subjacente')),
            'b3_id': titulo, 'internal_id': conf,
            'apurado': apurado, 'ir': ir, 'liquido': liq,
            'b3': sum(b3_vals) if b3_vals else None,
        })
    return out


# O aviso impresso começa em "B3 ID": Contraparte é o destinatário e
# Settlement Net é o critério de quebra — nenhum dos dois é conteúdo do
# documento que o cliente recebe.
_NDFADV_EMAIL_FROM = 1
_NDFADV_EMAIL_DROP = ('Settlement Net',)


def _ndfadv_email_headers():
    return [c for c in _NDFADV_COLUMNS[_NDFADV_EMAIL_FROM:] if c not in _NDFADV_EMAIL_DROP]


def _ndfadv_email_rows(ref):
    """Linhas do aviso impresso: as mesmas da tela, sem Contraparte e sem
    Settlement Net, e com os valores em BR (`R$ #.##0,00`, negativo entre
    parênteses). Formatadas a partir do NÚMERO, não do texto da tela — o
    caminho inverso erraria no primeiro valor com separador ambíguo."""
    from apps.pages import otc_emails
    keep = [i for i, c in enumerate(_NDFADV_COLUMNS)
            if i >= _NDFADV_EMAIL_FROM and c not in _NDFADV_EMAIL_DROP]
    money_at = {_NDFADV_COLUMNS.index(c) for c in
                ('Resultado Apurado (R$)', 'IR 0,005% (R$)', 'Resultado Líquido (R$)')}
    num_by_col = {_NDFADV_COLUMNS.index('Resultado Apurado (R$)'): 'apurado',
                  _NDFADV_COLUMNS.index('IR 0,005% (R$)'): 'ir',
                  _NDFADV_COLUMNS.index('Resultado Líquido (R$)'): 'liquido'}
    out = []
    for r in _ndfadv_collect(ref):
        cells = []
        for i in keep:
            if i in money_at:
                v = r.get(num_by_col[i])
                cells.append(otc_emails._brl(v) if v is not None else '')
            else:
                cells.append(r['cells'][i])
        out.append(dict(r, cells=cells))
    return out


# ── Other Products › Option › Settlement Advice ──────────────────────────────
#  Irmão do aviso de Termo de Mercadoria, e de propósito: o caminho até o valor é
#  o MESMO (Operations B3 → posição → OTM Settlements pelo sufixo do Trade Id),
#  e por isso ele reusa `_ndfadv_otm_by_suffix`, `_ndfc_ir` e `_ndfadv_media_label`
#  em vez de recopiar a conta — as duas telas têm de imprimir o mesmo imposto
#  sobre o mesmo caixa.
#
#  O que muda é ONDE a posição guarda os dois identificadores: a posição de
#  opções não tem `Contrato` nem `Codigo Identificador`.
#
#      NDF (DPOSICAO-TER)        Opção (DPOSICAO)
#      ───────────────────────   ─────────────────────────────
#      Contrato ............. →  Código IF                (B3 ID)
#      Codigo Identificador . →  Combinação de operações  (Nº da Confirmação)
#
#  Quais linhas do Operations B3 entram é do cadastro `opb3-events` (Tipo Título
#  = OPC), como nas outras duas telas — aqui não há teste de evento no código.
_OPTADV_COLUMNS = ['Contraparte', 'B3 ID', 'Nº da Confirmação', 'Data de Início da Operação',
                   'Ativo Subjacente', 'Ptax', 'Cotação Ativo Subjacente',
                   'Quantidade da Operação', 'Resultado Apurado (R$)', 'IR 0,005% (R$)',
                   'Resultado Líquido (R$)']
# Opção de MOEDA não tem cotação de subjacente a imprimir: o subjacente É a taxa
# de câmbio, que já sai na coluna Ptax ao lado. A célula vai 'N/A' e não vazia —
# vazia se lê como "faltou o dado" e manda alguém procurar o que não existe.
_OPTADV_FX_NA = 'N/A'


def _optadv_product_label(classe):
    """O que vai entre parênteses no ASSUNTO do aviso de opção.

    Commodities e taxa de câmbio têm nome próprio no documento ('Opção de
    Commodities', 'Opção de Taxas de Câmbio'); qualquer outra classe entra com o
    texto que a posição de opções escreve na coluna `Classe do ativo subjacente`
    (AÇÕES, ÍNDICE…). É de propósito que não há de-para de classes aqui: a B3
    acrescenta classe sem avisar, e um mapa fechado mandaria a classe nova para um
    rótulo genérico — dizer o nome que veio no arquivo é sempre verdade."""
    if _optadv_is_commodity(classe):
        return 'Opção de Commodities'
    if _optadv_is_fx(classe):
        return 'Opção de Taxas de Câmbio'
    c = str(classe or '').strip()
    return 'Opção de {}'.format(c) if c else 'Opção'


def _optadv_is_commodity(classe):
    """A classe do ativo subjacente da opção é COMMODITIES?

    Teste de TOKEN, no espírito do `_fcst_lob`: é a classe — não a contraparte,
    não o book — que decide se o Ativo Subjacente sai como
    `MERCADORIA(SUBJACENTE)` ou só como o subjacente."""
    return 'commodit' in _fcst_norm(classe)


def _optadv_is_fx(classe):
    """A classe do ativo subjacente da opção é TAXA DE CÂMBIO?

    Mesma família de teste do `_optadv_is_commodity`, e por token: a B3 escreve
    essa classe ora como MOEDA, ora como TAXA DE CÂMBIO, dependendo do arquivo —
    comparar o texto inteiro simplesmente não casaria, sem erro nenhum."""
    s = _fcst_norm(classe)
    return any(tok in s for tok in ('cambio', 'moeda', 'currency'))


def _optadv_subjacente(classe, mercadoria, subjacente):
    """Coluna `Ativo Subjacente` do aviso de opção.

    Em COMMODITIES a posição já diz a mercadoria por extenso (PETROLEO, MILHO,
    SOJA) na coluna `Tipo de Mercadoria`, então a célula sai no mesmo formato do
    termo — `MERCADORIA(SUBJACENTE)` —, mas sem passar pelo `Subjacente.json`:
    aqui a resposta está na própria linha.

    Fora de commodities (câmbio, equity) NÃO há mercadoria nenhuma, e a célula é
    só o `Ativo subjacente / Moeda base`. Repetir a classe entre parênteses
    diria duas vezes a mesma coisa."""
    subj = str(subjacente or '').strip()
    merc = str(mercadoria or '').strip()
    if not _optadv_is_commodity(classe) or not merc:
        return subj
    return '{}({})'.format(merc.upper(), subj) if subj else merc.upper()


_OPTADV_PRM_AMBIGUO = object()          # sentinela: dois trades na mesma chave


def _optadv_cog_key(v):
    """Chave de casamento entre o `Athena ID` do Cognos e a `Combinação de
    operações` da posição de opções.

    O MESMO identificador, com um separador diferente em cada sistema: o Cognos
    escreve com HÍFEN e a B3 com UNDERLINE. Comparar o texto cru não casava nada —
    e um filtro que não casa nada não dá erro, só deixa a coluna de valor vazia
    para toda opção de câmbio.

    Os DOIS lados passam por aqui, e não só o do Cognos: assim não importa qual
    deles trouxe o hífen no dia em que um dos dois arquivos mudar de convenção."""
    return str(v or '').strip().upper().replace('-', '_')


def _optadv_cognos_prm(ref):
    """({identificador → PRM Amount ASSINADO}, {identificador → Counterparty SPN})
    do Cognos de `ref`, para a opção de TAXA DE CÂMBIO.

    O valor da opção de câmbio NÃO está no OTM Settlements: ele é o prêmio, e
    quem o tem é o FXO Detail — o arquivo da página Cognos. O SINAL sai da coluna
    `Direction`: RECEIVE é o banco recebendo (+), PAY é o banco pagando (−). O
    relatório traz o módulo, e sem o sinal metade dos avisos sairia com a direção
    invertida — o cliente lendo "vamos debitar" onde o banco é quem paga.

    O **SPN sai junto** porque é a mesma linha, e é ele que dá o NOME da
    contraparte: em opção de câmbio não existe linha de OTM, então o nome caía no
    apelido de conta da B3 ('INTRAGLAWTONFDO') em vez do cadastro. É a mesma razão
    de o `_ndfadv_otm_by_suffix` devolver o Cpty SPN ao lado do valor — uma segunda
    leitura do arquivo só para buscá-lo abriria espaço para as duas discordarem.

    A chave é o **`Athena ID` INTEIRO** (só o hífen trocado por underline, ver
    `_optadv_cog_key`), casado com a `Combinação de operações` da posição de
    opções — e é por essa linha da posição que se chega ao B3 ID (`Código IF`).
    Este join é mais simples que o do OTM Settlements de propósito: lá os dois
    sistemas escrevem o mesmo número com prefixos diferentes e o casamento é pelo
    SUFIXO; aqui as duas colunas carregam o mesmo identificador, então comparar o
    valor inteiro é o certo — e um sufixo casaria trades diferentes que por acaso
    terminam igual.

    Chave repetida com valores DIFERENTES é descartada em vez de resolvida por
    desempate: um valor plausível no aviso do cliente errado é pior do que uma
    célula vazia pedindo conferência."""
    _jp, data = _cog_load(ref)
    out, spn = {}, {}
    for rec in (data or []):
        k = _optadv_cog_key(rec.get('Athena ID', ''))
        if not k:
            continue
        # SPN: primeiro não vazio vence. As várias linhas de um trade são do mesmo
        # cliente, e uma delas vir sem SPN não pode apagar o nome — e ele é
        # colhido ANTES do teste de valor, porque a linha sem PRM Amount ainda
        # sabe de quem é o trade.
        if _spn_key(rec.get('Counterparty SPN', '')) and k not in spn:
            spn[k] = str(rec.get('Counterparty SPN', '') or '').strip()
        v = _conf_to_float(rec.get('PRM Amount'))
        if v is None:
            continue
        direction = _fcst_norm(rec.get('Direction', '')).strip()
        if direction.startswith('pay'):
            v = -abs(v)
        elif direction.startswith('receive'):
            v = abs(v)
        # Direction em branco ou desconhecida: o valor entra como veio. Zerar a
        # linha esconderia o prêmio; inventar um sinal inventaria a direção.
        if k in out and out[k] != v:
            out[k] = _OPTADV_PRM_AMBIGUO
        elif k not in out:
            out[k] = v
    return out, spn


def _optadv_prm_for(prm, conf):
    """PRM Amount da opção de câmbio pela `Combinação de operações` da posição.

    Nada casando (ou chave repetida com valores diferentes) devolve None — a
    célula fica vazia e pede conferência, que é o desfecho desejado."""
    k = _optadv_cog_key(conf)
    if not k:
        return None
    v = prm.get(k)
    if v is _OPTADV_PRM_AMBIGUO:
        log.warning('[opt-advice] PRM Amount repetido com valores diferentes para o '
                    'Athena ID %s — a célula fica vazia em vez de escolher um dos dois', k)
        return None
    return v


_OPTADV_IR_COL = 9                             # 'IR 0,005% (R$)'
_OPTADV_LIQ_COL = 10                           # 'Resultado Líquido (R$)'


def _optadv_apply_ir(items):
    """Aplica o IR de 0,005% às linhas do aviso de opção, IN PLACE.

    O imposto da opção NÃO é da linha, e é aqui que ele difere do termo de
    mercadoria (onde cada contrato paga o seu):

    1. **Só o PAGAMENTO DE PRÊMIO paga.** Recompra e exercício não têm IR — é o
       que a flag `premium` da linha carrega (e ela exige que TODOS os eventos
       aprovados do Título sejam prêmio, ver `_optadv_collect`).
    2. **A base é o NET por contraparte × data de liquidação**, não a operação.
       Todo o `_optadv_collect` é de UMA data, então o net é por contraparte
       dentro dele. Cobrar por linha faria dois prêmios que se anulam pagarem
       imposto sobre um caixa que não existe.
    3. **Só quando o net é positivo PARA A CONTRAPARTE**, isto é, quando o banco
       paga (`net < 0`, que é a ótica de todos os valores desta tela). Net a favor
       do banco não retém nada.

    O net entra só com as linhas de prêmio: incluir o exercício aqui faria uma
    operação sem IR mexer no IR de outra, o que contradiz a regra 1.

    A isenção continua saindo do MESMO cadastro `ndfc-ir-exempt` do termo — é o
    mesmo imposto sobre o mesmo cliente, e uma segunda lista divergiria com uma
    tela retendo e a outra não.

    O imposto é calculado sobre o net e depois RATEADO **só pelas linhas que
    PAGAM**, na proporção do módulo de cada uma e com a sobra de arredondamento na
    última. Nada é decidido por linha — quem decide se há imposto, e quanto, é o
    net; a linha só carrega a parte da retenção que sai com ELA. Espalhar o
    rateio também pelas linhas que recebem fazia a retenção andar para os DOIS
    lados: o líquido de um recebimento encolhia por um imposto que não era dele,
    as duas metades quase se anulavam, e o rodapé do aviso — que soma a coluna —
    imprimia `(28.884,13)` onde o net com imposto é `(28.882,73)`. De um lado só,
    a coluna fecha com o rodapé e o rodapé fecha com o net, que é a mesma conta
    do Pay/Rec (HANDOFF §205: net Pay −219.047,36 → −219.036,41)."""
    grupos = {}
    for r in items:
        if not r.get('premium') or r.get('apurado') is None:
            continue
        grupos.setdefault(_fcst_norm(r.get('counterparty', '')).strip(), []).append(r)

    for cliente, linhas in grupos.items():
        net = sum(r['apurado'] for r in linhas)
        if net >= 0 or _ndfc_ir_exempt(linhas[0].get('counterparty', '')):
            continue
        total = round(abs(net) * _NDFADV_IR_RATE, 2)
        if not total:
            continue
        # Só o lado que paga. `net < 0` garante que ele existe.
        alvo = [r for r in linhas if r['apurado'] < 0]
        base = sum(abs(r['apurado']) for r in alvo)
        acc = 0.0
        for i, r in enumerate(alvo):
            if i == len(alvo) - 1:
                r['ir'] = round(total - acc, 2)          # a sobra fecha o total
            else:
                r['ir'] = round(total * abs(r['apurado']) / base, 2) if base else 0.0
                acc += r['ir']

    # Resultado Líquido e as duas células — o IR retido sempre ENCOLHE o que se
    # movimenta, qualquer que seja o sinal (mesma regra do aviso de FX). Como o
    # rateio agora só toca o lado que paga, a linha que recebe sai com `0,00` de
    # imposto e o líquido igual ao apurado: não houve retenção sobre ela.
    for r in items:
        ap, ir = r.get('apurado'), r.get('ir') or 0.0
        r['ir'] = ir
        r['liquido'] = None if ap is None else (ap - ir if ap >= 0 else ap + ir)
        r['cells'][_OPTADV_IR_COL] = _ops_fmt_amt(ir)
        r['cells'][_OPTADV_LIQ_COL] = _ops_fmt_amt(r['liquido'])
    return items


def _optadv_collect(ref):
    """Linhas do Option Settlement Advice (commodities, equities e câmbio) da
    data `ref`."""
    # Peneirado pelo `opb3-events` UMA vez, no topo — mesma razão do aviso de
    # termo: a lista de Títulos e a soma do Settlement B3 lá embaixo têm de
    # enxergar as mesmas linhas.
    opb3 = _opb3_settle_rows(ref)
    if not opb3:
        return []

    titulos, seen = [], set()
    for rec in opb3:
        titulo = str(rec.get('Título', '') or '').strip()
        if not titulo or titulo.upper() in seen:
            continue
        if 'opc' not in _fcst_norm(rec.get('Tipo Título', '')):
            continue
        seen.add(titulo.upper())
        titulos.append((titulo, rec))
    if not titulos:
        return []

    # Posição de opções pela TELA do Live Position: as datas já vêm dd/mm/yyyy e
    # o bloco posicional das médias asiáticas já vem resolvido em colunas.
    lp = _lpopt_collect(ref)
    li = {c: i for i, c in enumerate(lp.get('columns') or [])}
    by_codigo_if = {}
    for row in lp.get('rows') or []:
        k = str(row[li['Código IF']] if 'Código IF' in li else '').strip().upper()
        if k:
            by_codigo_if.setdefault(k, row)

    otm_by_suffix, otm_spn = _ndfadv_otm_by_suffix(ref)
    # O MESMO elo que o SWAP de equity usa (`_ops_equity_link`): Operations B3
    # (Título) → Latam Desk Position → OTM Settlements. Ele é o plano B do valor
    # e do SPN quando o caminho normal — o sufixo da Combinação de operações —
    # não resolve. Ver o comentário no cálculo do `apurado`.
    eqlink = _ops_equity_link(ref)
    # Prêmio e SPN do FXO Detail — só a opção de TAXA DE CÂMBIO os consulta, mas o
    # arquivo é lido UMA vez, aqui: dentro do laço ele seria reaberto por linha.
    cog_prm, cog_spn = _optadv_cognos_prm(ref)
    spn_by_name = _ndfsum_refdata_spn()
    cpd = _cpd_load()

    def _lcell(row, name):
        i = li.get(name)
        return '' if (row is None or i is None or i >= len(row)) else str(row[i] or '').strip()

    out = []
    for titulo, rec in titulos:
        lrow = by_codigo_if.get(titulo.upper())
        conf = _lcell(lrow, 'Combinação de operações')
        suf = conf.rsplit('-', 1)[-1].upper() if conf else ''
        # A classe vem PRIMEIRO: ela decide de onde saem o valor, o SPN e as duas
        # colunas de data. Resolvê-la no meio do laço já fez a linha de câmbio
        # buscar o valor numa fonte e o nome noutra.
        classe = _lcell(lrow, 'Classe do ativo subjacente')
        e_fx = _optadv_is_fx(classe)

        # Contraparte pelo **SPN**, sempre que houver um: é um identificador na
        # própria linha do fluxo, resolve o omnibus de graça (numa conta
        # guarda-chuva o nome que vem da B3 é o do titular, não o do cliente) e
        # devolve a razão social do cadastro em vez do apelido de conta.
        #
        # A fonte do SPN depende do produto, e é a MESMA que deu o valor: o OTM em
        # commodities e equity, o **Cognos** em taxa de câmbio. Sem o lado do
        # Cognos a opção de câmbio não tinha SPN nenhum e caía no Nome Simplificado
        # da B3 — a Lawton aparecia como 'INTRAGLAWTONFDO' no Settlement Summary,
        # com o SPN dela já mapeado no `le-spn`.
        # O elo de equity, resolvido UMA vez por linha: ele responde pelo valor
        # e pelo SPN, e procurá-lo duas vezes abriria espaço para os dois virem
        # de trades diferentes.
        # A chave é o Título em MAIÚSCULA, a mesma forma que o swap usa
        # (`key = titulo.upper()`): o elo é indexado assim, e consultá-lo com
        # outra grafia não casaria nada, em silêncio.
        eq = None if e_fx else (eqlink.get(str(titulo or '').strip().upper()) or {})
        cliente_spn = (cog_spn.get(_optadv_cog_key(conf), '') if e_fx
                       else (otm_spn.get(suf, '') if suf else ''))
        if not cliente_spn and eq:
            cliente_spn = str(eq.get('spn', '') or '')
        cliente = _otm_cpty_name(cliente_spn) if cliente_spn else ''
        # Mesma leitura do aviso de termo: a coluna 'CPF/CNPJ Cliente
        # Contraparte' do Live Position Option JÁ resolve o nome no RefData, e
        # só volta como DOCUMENTO quando não há cadastro — caso em que o omnibus
        # não resolve e a linha cai para o nome da posição, como sempre.
        doc = _lcell(lrow, 'CPF/CNPJ Cliente Contraparte')
        if not cliente and _b3_is_omnibus(rec.get('Conta Contraparte', '')) and doc:
            cliente = '' if _lp_is_taxid(doc) else doc
        cliente = (cliente
                   or _lcell(lrow, 'Contraparte (Nome simplificado)')
                   or str(rec.get('Contraparte (Nome Simpl.)', '') or '').strip())

        # As duas colunas de data trocam de papel na opção de CÂMBIO, e é a mesma
        # razão nas duas: lá o subjacente É a taxa de câmbio.
        #
        #                    Ptax                              Cotação Ativo Subj.
        #   câmbio    fixing do ATIVO SUBJACENTE               N/A
        #   as demais fixing da MOEDA do ativo subjacente      fixing do subjacente
        #
        # Em câmbio, a PTAX que o aviso imprime é justamente o fixing do
        # subjacente; a coluna de fixing da moeda é de conversão e não diz nada ali.
        # E não há cotação de subjacente a imprimir — ela vai 'N/A', que é
        # diferente de vazia ("faltou o dado").
        #
        # O teste é "é câmbio?" e não "é commodity ou equity?": as três classes são
        # o universo, e perguntar pela exceção deixa uma classe nova cair no lado
        # que TEM cotação — o outro jeito a mandaria para 'N/A' calada.
        fix_subj = _lcell(lrow, 'Data de fixing do ativo subjacente')
        if e_fx:
            ptax, cot = fix_subj, _OPTADV_FX_NA
        else:
            ptax = _lcell(lrow, 'Data de fixing da moeda do ativo subjacente')
            # Vazio é o caso da ASIÁTICA: não há data única, e o que vale é o
            # mês/ano da 1ª data de verificação (a cotação é a média do período).
            cot = fix_subj or _ndfadv_media_label(_lcell(lrow, 'Média Asiática (data) 1'))

        # Resultado Apurado: em commodities e equity é a soma do OTM Settlements
        # pelo SUFIXO do identificador (a regra do aviso de termo); em TAXA DE
        # CÂMBIO é o PRM Amount do Cognos, casado pelo `Athena ID` INTEIRO contra
        # a Combinação de operações. Duas fontes porque são dois caixas diferentes
        # — a opção de câmbio liquida o PRÊMIO, e ele não passa pelo OTM.
        if e_fx:
            apurado = _optadv_prm_for(cog_prm, conf)
        else:
            apurado = otm_by_suffix.get(suf) if suf else None
            # A OPÇÃO DE EQUITY costuma cair aqui. O caminho normal casa o sufixo
            # da `Combinação de operações` com o do Trade Id do OTM, e depende de
            # a Live Position de Opção ter esse campo preenchido — o que não
            # acontece na opção de ação. Sem valor, a linha aparecia no Trade
            # Level com a célula vazia e SUMIA do Settlement Summary, que
            # descarta quem não tem o que liquidar (`_opssum_rows`).
            #
            # O plano B é o elo do SWAP de equity, pelo Título da B3. Ele vem
            # DEPOIS e não antes porque o sufixo é um join direto: quando existe,
            # é o mais confiável dos dois.
            if apurado is None and eq:
                apurado = eq.get('settlement')
        # IR fica para DEPOIS do laço: na opção ele não é da linha, é do NET por
        # contraparte (ver `_optadv_apply_ir`). Aqui a linha nasce sem imposto.
        ir, liq = 0.0, apurado

        ref_rec = spn_by_name.get(_fcst_norm(cliente), {})
        # O SPN da fonte (o MESMO que deu o nome) vence o caminho de volta
        # nome → SPN pelo Reference Data, que erra em toda diferença de pontuação.
        spn = str(cliente_spn or '').strip() or ref_rec.get('spn', '')
        net_type = _ndfsum_net_type(_cpd_find(cpd, spn) if spn else None)
        subj = _optadv_subjacente(classe, _lcell(lrow, 'Tipo de Mercadoria'),
                                  _lcell(lrow, 'Ativo subjacente / Moeda base'))

        # Settlement B3: soma de TODAS as linhas daquele Título no Operations B3
        # (o caixa do dia) — mesma regra do termo e do Trade Level de swap.
        mine = [r for r in opb3
                if str(r.get('Título', '') or '').strip().upper() == titulo.upper()]
        b3_vals = [v for v in (_conf_to_float(r.get('Valor')) for r in mine) if v is not None]
        # Prêmio: TODOS os eventos aprovados daquele Título são pagamento de
        # prêmio. `evs == {premio}` e não "algum é prêmio", que é a mesma regra do
        # aviso de Swap — um Título que traz prêmio E liquidação no mesmo dia não
        # é um aviso de prêmio, e rotulá-lo assim esconderia a liquidação que está
        # na mesma tabela.
        evs = {_ops_norm_event(r.get('Tipo Operação', '')) for r in mine}
        out.append({
            'cells': [
                cliente,
                titulo,
                conf,
                _lcell(lrow, 'Data Início'),
                subj,
                ptax,
                cot,
                _lcell(lrow, 'Quantidade') or _lcell(lrow, 'Valor base'),
                _ops_fmt_amt(apurado),
                _ops_fmt_amt(ir),
                _ops_fmt_amt(liq),
            ],
            'counterparty': cliente,
            'legal': _lcell(lrow, 'Parte (Nome simplificado)'),
            'spn': spn,
            'taxid': ref_rec.get('taxid', ''),
            'net_type': net_type,
            # LOB = a CLASSE do subjacente como a B3 a escreve (COMMODITIES,
            # EQUITIES, MOEDA). É o mesmo texto que a coluna Type do Operations
            # B3 mostra para OPC, e não um de-para novo: traduzi-lo aqui criaria
            # um segundo vocabulário para a mesma coluna.
            'lob': classe.strip().upper(),
            'underlying': subj,
            # Rótulo do produto no assunto do aviso, e `premium` — os dois são
            # lidos pelo gerador de e-mail e entram na CHAVE do agrupamento, então
            # um documento nunca mistura classe nem junta prêmio com liquidação.
            'product_label': _optadv_product_label(classe),
            'premium': bool(evs) and evs == {_ops_norm_event('PAGAMENTO DE PREMIO')},
            'b3_id': titulo, 'internal_id': conf,
            'apurado': apurado, 'ir': ir, 'liquido': liq,
            'b3': sum(b3_vals) if b3_vals else None,
        })
    # O IR é do NET por contraparte, então só dá para calculá-lo com as linhas do
    # dia todas na mão — por isso ele é um passe depois do laço, e não parte dele.
    return _optadv_apply_ir(out)


# ── Edições manuais do aviso de Opção ────────────────────────────────────────
#  Mesmo desenho do aviso de Swap (`_swadv_edits_*`): a linha é derivada de três
#  arquivos e, quando um deles vem errado, a mesa corrige a célula e manda o
#  aviso assim mesmo. As correções ficam num overlay do DIA, fora dos arquivos de
#  origem — reimportar o batch não as apaga.
#
#  A chave é o **B3 ID**, não a posição na tela: a tabela ordena por cliente e a
#  posição muda a cada carga.
_OPTADV_KEY_COL = 1
# Célula editada que também é NÚMERO: sem isto a tela mostraria o valor corrigido
# e o Settlement Summary continuaria somando o original.
_OPTADV_NUM_FIELDS = {8: 'apurado', 9: 'ir', 10: 'liquido'}


def _optadv_edits_path(ref):
    return os.path.join(OTM_JSON_ROOT, ref.strftime('%Y'), ref.strftime('%m'),
                        ref.strftime('%d'),
                        'option-settlement-advice_{}.json'.format(ref.strftime('%Y%m%d')))


def _optadv_edits_load(ref):
    fp = _optadv_edits_path(ref)
    try:
        with open(fp, encoding='utf-8') as fh:
            d = json.load(fh)
        return (fp, d) if isinstance(d, dict) else (fp, {})
    except Exception:
        return fp, {}


def _optadv_items(ref):
    """As linhas do aviso JÁ com as edições manuais do dia aplicadas. É por aqui
    que a tela e o Trade Level passam — os dois, sempre."""
    edits = _optadv_edits_load(ref)[1]
    items = _optadv_collect(ref)
    if not edits:
        return items
    out = []
    for r in items:
        key = str(r['cells'][_OPTADV_KEY_COL] or '').strip()
        e = edits.get(key) or {}
        if e.get('deleted'):
            continue
        cells, rec = list(r['cells']), dict(r)
        for idx_s, val in (e.get('cells') or {}).items():
            try:
                i = int(idx_s)
            except (TypeError, ValueError):
                continue
            if not 0 <= i < len(cells):
                continue
            cells[i] = str(val)
            fld = _OPTADV_NUM_FIELDS.get(i)
            # `_conf_to_float` porque o operador digita ora US ('12,345.67'),
            # ora BR ('12.345,67') — o parser tolerante entende os dois.
            if fld:
                rec[fld] = _conf_to_float(str(val))
            elif i == 0:
                rec['counterparty'] = str(val)
        rec['cells'] = cells
        out.append(rec)
    return out


def _optadv_ref_and_key(payload):
    ds = str((payload or {}).get('date') or '').strip()
    try:
        ref = datetime.strptime(ds[:10], '%Y-%m-%d') if ds else datetime.now()
    except ValueError:
        ref = datetime.now()
    return ref, str((payload or {}).get('contrato') or '').strip()


# O aviso impresso começa em "B3 ID": Contraparte é o DESTINATÁRIO, não conteúdo
# do documento que ele recebe — mesma regra do aviso de termo.
_OPTADV_EMAIL_FROM = 1


def _optadv_email_headers():
    return list(_OPTADV_COLUMNS[_OPTADV_EMAIL_FROM:])


def _optadv_email_rows(ref):
    """Linhas do aviso impresso de opção: as mesmas da tela, sem Contraparte, e
    com os valores em BR (`R$ #.##0,00`, negativo entre parênteses). Formatadas a
    partir do NÚMERO e não do texto da tela — o caminho inverso erraria no
    primeiro valor com separador ambíguo."""
    from apps.pages import otc_emails
    money_at = {_OPTADV_COLUMNS.index(c): fld for c, fld in
                (('Resultado Apurado (R$)', 'apurado'), ('IR 0,005% (R$)', 'ir'),
                 ('Resultado Líquido (R$)', 'liquido'))}
    out = []
    for r in _optadv_items(ref):
        cells = []
        for i in range(_OPTADV_EMAIL_FROM, len(_OPTADV_COLUMNS)):
            if i in money_at:
                v = r.get(money_at[i])
                cells.append(otc_emails._brl(v) if v is not None else '')
            else:
                cells.append(r['cells'][i])
        # `commodity` é o que o gerador lê para quebrar o aviso por mercadoria e
        # para desambiguar o assunto. Na opção o equivalente é o ATIVO
        # SUBJACENTE, que é o mesmo texto que a coluna mostra.
        #
        # `family` diz ao gerador de e-mail QUE aviso é este. Ele serve o termo e
        # a opção com a mesma função, e o aviso de PRÊMIO de opção é o único que
        # imprime a tabela sem as colunas de IR e Resultado Líquido — deduzir
        # isso do rótulo do produto ('Opção de …') amarraria a regra a um texto
        # que existe para ser lido por gente.
        out.append(dict(r, cells=cells, commodity=r.get('underlying', ''),
                        family='option'))
    return out


# ── Other Products › Swap › Kapital Hybrids (BANCO_UPCOMING_PAYMENTS.csv) ─────
#  Comma-delimited upcoming-payments file. On import (_swaphyb_extract) only rows
#  whose Settlement Date = today are kept (the file has two Settlement Date
#  columns, dd/mmm/yyyy and mm/dd/yyyy — same date, either matches). The page
#  collapses the file's per-leg rows into one row per trade, keyed by Kapital ID
#  (Trade Confirmation ID): Owner curve = Σ positive Amounts, Counterparty curve =
#  Σ negative Amounts (kept negative), BRL Net Amount = Owner + Counterparty — the
#  net cashflow (all #,##0.00). The Cetip ID is pulled from mapping_swap-hyb.json
#  (Kapital ID → b3_id).
_SWAPHYB_COLUMNS = ['Kapital ID', 'Cetip ID', 'Trade Date', 'Settlement Date',
                    'Stream Notional', 'Stream Notional Currency', 'Coupon Rate', 'Currency',
                    'DCF', 'Counterparty SPN', 'Counterparty Name',
                    'Owner curve', 'Counterparty curve', 'BRL Net Amount']


def _swaphyb_read_rows(raw):
    """BANCO_UPCOMING_PAYMENTS.csv → list of rows (comma-delimited, quoted fields
    honoured so amounts like "1,234.56" survive)."""
    import csv as _csv
    text = raw.decode('utf-8-sig', errors='replace') if isinstance(raw, (bytes, bytearray)) else str(raw)
    return list(_csv.reader(io.StringIO(text)))


def _swaphyb_num(v):
    """US-formatted amount ('-7,012,145.46', '(123.45)') → float, or None."""
    s = str(v or '').strip()
    if not s:
        return None
    neg = s.startswith('(') and s.endswith(')')
    if neg:
        s = s[1:-1]
    s = s.replace(',', '').replace(' ', '')     # drop thousands separators / spaces
    try:
        f = float(s)
    except ValueError:
        return None
    return -f if neg else f


def _swaphyb_parse_date(v):
    """Settlement/Trade date cell (dd/mmm/yyyy or mm/dd/yyyy, plus a few variants)
    → date, or None."""
    s = str(v or '').strip()
    if not s:
        return None
    for fmt in ('%d/%b/%Y', '%d-%b-%Y', '%d/%B/%Y', '%m/%d/%Y', '%Y-%m-%d', '%d/%m/%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _swaphyb_extract(raw, ref):
    """Parse the CSV, keep only rows whose Settlement Date = the import date, and
    return per-leg dicts (page-oriented keys). Aggregation happens at display time.
    Returns (records, total_data_rows)."""
    rows = _swaphyb_read_rows(raw)
    if len(rows) < 2:
        return [], 0
    # Locate the header row (bank files sometimes carry a title/blank line first).
    hidx = 0
    for i, r in enumerate(rows[:15]):
        low = [_fcst_norm(str(c)) for c in r]
        if any('confirmation' in c for c in low) or \
           (any('settlement date' in c for c in low) and any(c == 'amount' for c in low)):
            hidx = i
            break
    norm = [_fcst_norm(str(h)) for h in rows[hidx]]

    def first(pred):
        for i, n in enumerate(norm):
            if pred(n):
                return i
        return None

    def all_idx(pred):
        return [i for i, n in enumerate(norm) if pred(n)]

    i_trade  = first(lambda n: n == 'trade')
    i_kap    = first(lambda n: 'confirmation' in n)              # Trade Confirmation ID = Kapital ID
    i_tdate  = first(lambda n: n == 'trade date')
    i_settle = all_idx(lambda n: n == 'settlement date' or n.startswith('settlement date'))
    i_notl   = first(lambda n: n == 'stream notional')
    i_notlc  = first(lambda n: n == 'stream notional currency')
    i_coupon = first(lambda n: n == 'coupon rate' or 'coupon' in n)
    i_ccy    = first(lambda n: n == 'currency')
    i_dcf    = first(lambda n: n == 'dcf')
    i_spn    = first(lambda n: 'spn' in n)
    i_cpty   = first(lambda n: 'counterparty' in n and 'name' in n)
    i_amt    = first(lambda n: n == 'amount')

    today = ref.date() if hasattr(ref, 'date') else ref

    def cell(row, i):
        return str(row[i]).strip() if (i is not None and i < len(row) and row[i] is not None) else ''

    out, total = [], 0
    for row in rows[hidx + 1:]:
        if not any(str(c).strip() for c in row):
            continue
        total += 1
        sdate = None
        for si in i_settle:                       # either Settlement Date column may match
            d = _swaphyb_parse_date(cell(row, si))
            if d:
                sdate = d
                break
        if sdate != today:                        # keep ONLY today's settlements
            continue
        tdate = _swaphyb_parse_date(cell(row, i_tdate))
        out.append({
            'Trade': cell(row, i_trade),
            'Kapital ID': cell(row, i_kap),
            'Trade Date': tdate.strftime('%d/%m/%Y') if tdate else cell(row, i_tdate),
            'Settlement Date': sdate.strftime('%d/%m/%Y'),
            'Stream Notional': cell(row, i_notl),
            'Stream Notional Currency': cell(row, i_notlc),
            'Coupon Rate': cell(row, i_coupon),
            'Currency': cell(row, i_ccy),
            'DCF': cell(row, i_dcf),
            'Counterparty SPN': cell(row, i_spn),
            'Counterparty Name': cell(row, i_cpty),
            'Amount': cell(row, i_amt),
        })
    return out, total


def _swaphyb_kap_to_cetip():
    """{Kapital ID (hybrids_id) → Cetip ID (b3_id)} from mapping_swap-hyb.json."""
    path = data_path('mapping_swap-hyb.json')
    out = {}
    try:
        with open(path, encoding='utf-8') as fh:
            for rec in json.load(fh) or []:
                k = str(rec.get('hybrids_id', '') or '').strip()
                if k:
                    out[k] = str(rec.get('b3_id', '') or '').strip()
    except Exception:
        pass
    return out


def _swaphyb_collect(ref):
    """Read the cached JSON and collapse per-leg rows into one row per trade
    (Kapital ID). Owner curve = Σ positive Amounts, Counterparty curve = Σ negative
    Amounts (kept negative), BRL Net Amount = Owner + Counterparty (net cashflow).
    Cetip ID from the mapping."""
    jp = _ds_display_json_path(ref, _SWAPHYB_JSON)
    rows_out = []
    if os.path.isfile(jp):
        try:
            with open(jp, encoding='utf-8') as fh:
                data = json.load(fh) or []
        except Exception:
            data = []
        cet = _swaphyb_kap_to_cetip()
        groups, order = {}, []
        for rec in data:
            kap = str(rec.get('Kapital ID', '') or '').strip()
            if kap not in groups:
                groups[kap] = {'first': rec, 'owner': 0.0, 'cpty': 0.0}
                order.append(kap)
            amt = _swaphyb_num(rec.get('Amount', ''))
            if amt is not None:
                if amt > 0:
                    groups[kap]['owner'] += amt
                elif amt < 0:
                    groups[kap]['cpty'] += amt
        for kap in order:
            g = groups[kap]
            r = g['first']
            owner, cpty = g['owner'], g['cpty']
            net = owner + cpty                     # net cashflow (cpty is the negative sum)
            rows_out.append([
                kap, cet.get(kap, ''),
                r.get('Trade Date', ''), r.get('Settlement Date', ''),
                r.get('Stream Notional', ''), r.get('Stream Notional Currency', ''),
                r.get('Coupon Rate', ''), r.get('Currency', ''), r.get('DCF', ''),
                r.get('Counterparty SPN', ''), r.get('Counterparty Name', ''),
                '{:,.2f}'.format(owner), '{:,.2f}'.format(cpty), '{:,.2f}'.format(net),
            ])
    # Sort by Kapital ID A→Z (accent-insensitive); blank goes last.
    rows_out.sort(key=lambda x: (str(x[0]).strip() == '', _fcst_norm(str(x[0]))))
    return {'widgets': {'total': len(rows_out)}, 'columns': list(_SWAPHYB_COLUMNS),
            'rows': rows_out, 'updated': _ds_read_updated(jp)}


# ── Other Products › Swap › VCP (AVISO DE INEXISTENCIA DE PU) ─────────────────
#  Cross-join. Base rows come from the Operations B3 JSON, keeping only
#  Tipo Operação = "AVISO DE INEXISTENCIA DE PU" (Título → Código do Contrato,
#  Conta → PARTE / Conta). Each contract is joined to the Events file
#  (eventos-swap-jpm) to pull PARTE / Indexador and the counterparty account /
#  CNPJ / indexer. The counterparty NAME is then resolved from RefData.json by
#  B3 account — or by Tax ID when the account is the shared 73760.10-2 omnibus
#  (many clients share it). PARTE / Fator and CONTRAPARTE/ Fator are left blank
#  (formula pending). Fator columns render #,##0.00 once populated.
_VCP_COLUMNS = ['Contraparte', 'Código do Contrato', 'PARTE / Conta',
                'PARTE / Indexador', 'PARTE / Fator', 'CONTRAPARTE / Conta',
                'CONTRAPARTE / CPF/CNPJ', 'CONTRAPARTE / Indexador', 'CONTRAPARTE/ Fator']
_VCP_VALUE_COLS = {'PARTE / Fator', 'CONTRAPARTE/ Fator'}
_VCP_AVISO_TYPE = 'AVISO DE INEXISTENCIA DE PU'
_VCP_SHARED_ACCT = '73760102'                          # 73760.10-2 omnibus (digits only)


def _vcp_refdata_maps():
    """(by_account, by_taxid) → COUNTERPARTY name, keyed on digits-only B3 ACCOUNT
    and TAX ID. by_account skips the shared 73760.10-2 omnibus (ambiguous — many
    clients share it; those resolve by Tax ID instead)."""
    by_acct, by_taxid = {}, {}
    data = _refdata_records()                     # DB-first (fase 3)
    for rec in data:
        name = str(rec.get('COUNTERPARTY', '') or '').strip()
        if not name:
            continue
        acct = _acc_digits(rec.get('B3 ACCOUNT', ''))
        if acct and acct != _VCP_SHARED_ACCT:
            by_acct.setdefault(acct, name)
        tax = _acc_digits(rec.get('TAX ID', ''))
        if tax:
            by_taxid.setdefault(tax, name)
    return by_acct, by_taxid


def _vcp_events_map(ref):
    """{contract-digits → leg dict} from the Events file JSON (eventos-swap-jpm):
    PARTE / Indexador and the CONTRAPARTE account / CPF-CNPJ / Indexador, resolved
    by header name so a small layout drift still matches."""
    jp = _ds_display_json_path(ref, 'eventos-swap-jpm')
    out = {}
    if not os.path.isfile(jp):
        return out
    try:
        with open(jp, encoding='utf-8') as fh:
            data = json.load(fh) or []
    except Exception:
        return out
    if not data:
        return out
    keys = list(data[0].keys())
    k_contract = _fcst_resolve_key(keys, ['codigo do contrato', 'código do contrato',
                                          'codigo if', 'código if', 'contrato',
                                          'instrumento financeiro'])
    k_parte_ix = _fcst_resolve_key(keys, ['parte / indexador', 'indexador parte', 'parte indexador'])
    # The counterparty account lives in the "CONTRAPARTE / Contraparte" column
    # (parallel to the PARTE side's "PARTE / Conta"), not "CONTRAPARTE / Conta".
    k_cpty_conta = _fcst_resolve_key(keys, ['contraparte / contraparte', 'contraparte / conta',
                                            'conta contraparte', 'contraparte conta'])
    k_cpty_cnpj = _fcst_resolve_key(keys, ['contraparte / cpf/cnpj', 'contraparte / cnpj',
                                           'cpf/cnpj contraparte', 'cnpj contraparte',
                                           'contraparte cnpj', 'contraparte / cpf'])
    k_cpty_ix = _fcst_resolve_key(keys, ['contraparte / indexador', 'indexador contraparte',
                                         'contraparte indexador'])
    for rec in data:
        cc = _acc_digits(rec.get(k_contract, '')) if k_contract else ''
        if not cc:
            continue
        out.setdefault(cc, {
            'parte_ix': str(rec.get(k_parte_ix, '') or '').strip() if k_parte_ix else '',
            'cpty_conta': str(rec.get(k_cpty_conta, '') or '').strip() if k_cpty_conta else '',
            'cpty_cnpj': str(rec.get(k_cpty_cnpj, '') or '').strip() if k_cpty_cnpj else '',
            'cpty_ix': str(rec.get(k_cpty_ix, '') or '').strip() if k_cpty_ix else '',
        })
    return out


def _vcp_collect(ref):
    """VCP display rows: Operations B3 'AVISO DE INEXISTENCIA DE PU' entries joined
    to the Events file legs and the RefData counterparty name."""
    _, ops = _opb3_load(ref)
    rows_out = []
    if ops:
        by_acct, by_taxid = _vcp_refdata_maps()
        events = _vcp_events_map(ref)
        aviso = _fcst_norm(_VCP_AVISO_TYPE)
        for rec in ops:
            if _fcst_norm(rec.get('Tipo Operação', '')) != aviso:
                continue
            contrato = str(rec.get('Título', '') or '').strip()
            parte_conta = str(rec.get('Conta', '') or '').strip()
            ev = events.get(_acc_digits(contrato), {})
            cpty_conta = ev.get('cpty_conta', '')
            cpty_cnpj = ev.get('cpty_cnpj', '')
            acct_dig = _acc_digits(cpty_conta)
            if acct_dig and acct_dig != _VCP_SHARED_ACCT:      # dedicated account → by account
                name = by_acct.get(acct_dig, '')
            else:                                              # shared 73760.10-2 omnibus → by Tax ID
                name = by_taxid.get(_acc_digits(cpty_cnpj), '')
            rows_out.append([name, contrato, parte_conta, ev.get('parte_ix', ''), '',
                             cpty_conta, cpty_cnpj, ev.get('cpty_ix', ''), ''])
    # Sort by Contraparte A→Z (accent-insensitive); rows with no name go last.
    rows_out.sort(key=lambda r: (r[0] == '', _fcst_norm(r[0])))
    return {'widgets': {'total': len(rows_out)}, 'columns': list(_VCP_COLUMNS),
            'rows': rows_out, 'updated': _ds_read_updated(_opb3_json_path(ref))}


# ── OTM maker/checker CRUD — every change/insert is persisted to the day's JSON ──
def _otm_ref_from(payload):
    ds = str((payload or {}).get('date', '') or '').strip()
    try:
        return datetime.strptime(ds[:10], '%Y-%m-%d') if ds else datetime.now()
    except ValueError:
        return datetime.now()


# ══════════════════════════════════════════════════════════════════════════════
#  Other Products › Latam Desk Position
#  Source: FbiRptLatamDeskPostion-NY-* dropped in the Settlements folder (the file
#  name really is "Postion" — a typo in the report; the PAGE is "Position").
#  Substitui o import VBA legado: a macro aplica AutoFilter <> "" no campo 62,
#  copia as linhas visíveis, limpa, aplica <> "" no campo 63 e ANEXA — ou seja,
#  mantém a linha quando a coluna 62 (BJ, CLEARING_TRD_ID_IN…) OU a 63 (BK,
#  CLEARING_TRD_ID_CLN…) está preenchida. Aqui cada linha entra UMA vez (a macro
#  duplicaria a linha que tem as duas colunas preenchidas).
#  O relatório NÃO é diário: a página abre no ÚLTIMO JSON disponível
#  (_latam_latest_ref), não no de hoje.
# ══════════════════════════════════════════════════════════════════════════════
LATAM_SOURCE_ROOT = os.getenv('LATAM_SOURCE_ROOT', SETTLEMENTS_ROOT)
LATAM_JSON_ROOT = OTM_JSON_ROOT
_LATAM_JSON_BASE = 'latam-desk-position'
_LATAM_FILE_PREFIX = 'fbirptlatamdeskpo'        # cobre "Postion" e um futuro "Position"

# (label exibido na página, candidatos de header no arquivo).
# O primeiro candidato é o nome inteiro; o segundo é o texto como aparece
# truncado na planilha e serve de PREFIXO — várias colunas do relatório só foram
# vistas cortadas, então 'CLEARING_TRD_ID_IN' precisa achar
# 'CLEARING_TRD_ID_INTERNAL' sem que ninguém tenha de digitar o nome exato.
_LATAM_COLUMNS = [
    ('Instrument_ID',        ('Instrument_ID', 'Instrument_')),
    ('Instrument_Name',      ('Instrument_Name',)),
    ('RIC',                  ('RIC',)),
    ('Instrument_Currency',  ('Instrument_Currency', 'Instrument_Curren')),
    ('Imnt_Ccy_Issuer',      ('Imnt_Ccy_Issuer', 'Imnt_Ccy_Iss')),
    ('FX Rate',              ('FX Rate', 'FX_Rate')),
    ('Maturity_Date',        ('Maturity_Date',)),
    ('Type',                 ('Type', 'Ty')),
    ('Subtype',              ('Subtype', 'Subty')),
    ('Underlying_Name',      ('Underlying_Name',)),
    ('Underlying Currency',  ('Underlying Currency', 'Underlying_Currency', 'Underlying Curr')),
    ('Country_Name',         ('Country_Name',)),
    ('Deal_Ref',             ('Deal_Ref', 'Deal_R')),
    ('Deal_ID',              ('Deal_ID',)),
    ('Counterparty',         ('Counterparty',)),
    ('Counterparty_SPN',     ('Counterparty_SPN',)),
    ('Counterparty_Type',    ('Counterparty_Type',)),
    ('Trade_Date',           ('Trade_Date',)),
    ('CALLPUT',              ('CALLPUT',)),
    ('Strike',               ('Strike',)),
    ('Legal_Entity',         ('Legal_Entity',)),
    ('id',                   ('id',)),
    ('CLEARING_TRADE_ID',    ('CLEARING_TRADE_ID', 'CLEARING_TRADE_')),
    ('CLEARING_TRD_ID_INT',  ('CLEARING_TRD_ID_INT', 'CLEARING_TRD_ID_IN')),
    ('CLEARING_TRD_ID_CLNT', ('CLEARING_TRD_ID_CLNT', 'CLEARING_TRD_ID_CLN')),
    ('TOTAL_PREMIUM',        ('TOTAL_PREMIUM',)),
    ('PREMIUM_SETT',         ('PREMIUM_SETT',)),
    ('LAST_PMT_DATE',        ('LAST_PMT_DATE',)),
    ('REBATE',               ('REBATE',)),
    ('REBATE_SCHEDULE',      ('REBATE_SCHEDULE',)),
    ('REBATE_PNL',           ('REBATE_PNL', 'REBATE_PN')),
    ('BARRIER_TYPE',         ('BARRIER_TYPE', 'BARRIER_TYP')),
    ('BARRIER_SCHEDULE',     ('BARRIER_SCHEDULE',)),
    ('CONTRACT_LEVEL',       ('CONTRACT_LEVEL',)),
    ('INITIAL_PRICE',        ('INITIAL_PRICE',)),
    ('START_SPOT',           ('START_SPOT',)),
    ('OPTION_EXERCISE_TYPE', ('OPTION_EXERCISE_TYPE', 'OPTION_EXERCISE_TYP')),
    ('UNDERLYING_RIC',       ('UNDERLYING_RIC', 'UNDERLYING_R')),
]
_LATAM_LABELS = [c[0] for c in _LATAM_COLUMNS]
# Toda coluna de data do relatório vira dd/mm/yyyy (o arquivo mistura
# '2030-01-16 00:00:00.0' com '20260108').
_LATAM_DATE_COLS = {'Maturity_Date', 'Trade_Date', 'PREMIUM_SETT', 'LAST_PMT_DATE',
                    'REBATE_SCHEDULE', 'BARRIER_SCHEDULE'}
# Sentinela de "sem data": o relatório grava o epoch (1969-12-31 19:00 em EST /
# 1970-01-01 em UTC) no lugar de vazio. Essas datas NÃO entram — a célula fica ''.
_LATAM_EPOCH = {(1969, 12, 31), (1970, 1, 1)}
# Colunas 62/63 da macro (BJ/BK, 1-based) — fallback posicional quando o header
# dos dois CLEARING_TRD_ID_* não é reconhecido.
_LATAM_FILTER_COLS = (62, 63)

# Meta de maker/checker por registro (prefixo '_lt_' para não colidir com as
# colunas do relatório). Importado/adicionado nasce 'OK'; editar → 'Pending'
# (maker gravado, checker limpo); OUTRO usuário confirma → 'OK'.
_LATAM_META_KEYS = ('_lt_status', '_lt_maker', '_lt_checker', '_lt_id')


def _latam_norm(s):
    """Header comparável: minúsculo, só letras e dígitos (assim 'FX Rate',
    'FX_Rate' e 'fx rate' são o mesmo header)."""
    s = unicodedata.normalize('NFKD', str(s or '')).encode('ascii', 'ignore').decode('ascii')
    return re.sub(r'[^a-z0-9]', '', s.lower())


def _latam_col_map(header):
    """{label → índice da coluna} para o header de um arquivo. Duas passadas:
    (1) header IGUAL a um dos candidatos; (2) header que COMEÇA com o candidato,
    do candidato mais longo para o mais curto — assim 'Counterparty_Type' fica com
    a sua coluna antes de 'Counterparty' tentar o prefixo. Cada coluna do arquivo
    é usada por um label só."""
    hn = [_latam_norm(h) for h in header]
    out, used = {}, set()

    def claim(label, pred):
        for i, h in enumerate(hn):
            if i in used or not h or not pred(h):
                continue
            out[label] = i
            used.add(i)
            return True
        return False

    for label, cands in _LATAM_COLUMNS:                      # (1) exato
        for cand in cands:
            n = _latam_norm(cand)
            if n and claim(label, lambda h, n=n: h == n):
                break
    pending = [(cand, label) for label, cands in _LATAM_COLUMNS
               if label not in out for cand in cands]
    pending.sort(key=lambda t: -len(_latam_norm(t[0])))
    for cand, label in pending:                              # (2) prefixo
        n = _latam_norm(cand)
        if label in out or not n:
            continue
        claim(label, lambda h, n=n: h.startswith(n))
    return out


def _latam_date(v):
    """Data do relatório → 'dd/mm/yyyy'. Aceita '2030-01-16 00:00:00.0',
    '20260108', dd/mm/yyyy e datetime; devolve '' para o sentinela epoch
    (1969-12-31 / 1970-01-01) e o texto original quando não é data nenhuma."""
    if v is None:
        return ''
    if hasattr(v, 'year') and hasattr(v, 'month') and hasattr(v, 'day'):
        d = v
    else:
        s = str(v).strip()
        if not s:
            return ''
        d = _fcst_parse_date(s)
        if d is None:
            return s
    if (d.year, d.month, d.day) in _LATAM_EPOCH:
        return ''
    return '{:02d}/{:02d}/{:04d}'.format(d.day, d.month, d.year)


def _latam_new_id():
    return uuid.uuid4().hex[:10]


def _latam_ensure_meta(data, default_status='OK'):
    """Garante status/maker/checker/id em todo registro. True se algo mudou (o
    caller pode persistir — migração de JSON antigo, sem meta)."""
    changed = False
    for rec in data:
        if not rec.get('_lt_id'):
            rec['_lt_id'] = _latam_new_id(); changed = True
        if '_lt_status' not in rec:
            rec['_lt_status'] = default_status; changed = True
        for k in ('_lt_maker', '_lt_checker'):
            if k not in rec:
                rec[k] = ''; changed = True
    return changed


def _latam_pick_source(names, root):
    """Dos candidatos `FbiRptLatamDeskPostion-NY-*`, o do relatório MAIS RECENTE —
    mtime primeiro, nome decrescente no empate.

    O relatório é reemitido no mesmo dia, e quando ele é, a pasta passa a ter DOIS
    arquivos: o consumido de manhã só é apagado quando alguma linha entrou, e o novo
    chega ao lado. Escolher `sorted(...)[0]` (o primeiro em ordem alfabética) lia o
    ANTIGO e regravava o JSON do dia com a posição da manhã — import "com sucesso",
    N linhas, e a tela sem a atualização. A ordem do `os.listdir` no caminho do Save
    Daily Settlement era pior ainda: sem ordem nenhuma, o vencedor dependia do
    sistema de arquivos, e os dois caminhos podiam discordar sobre qual é o
    relatório do dia.

    Devolve (escolhido, preteridos) — os preteridos ficam em disco de propósito
    (apagar um arquivo que não foi lido destrói a única cópia) e são registrados no
    log, porque pasta com dois relatórios é o estado que produziu o bug."""
    cands = [n for n in (names or []) if n.lower().startswith(_LATAM_FILE_PREFIX)]
    if not cands:
        return None, []
    def _mtime(n):
        try:
            return os.path.getmtime(os.path.join(root, n))
        except OSError:
            return 0.0
    cands.sort(key=lambda n: (_mtime(n), n), reverse=True)
    if len(cands) > 1:
        log.warning('[latam] %d relatórios em %s — lendo o mais recente (%s); ignorados: %s',
                    len(cands), root, cands[0], ', '.join(cands[1:]))
    return cands[0], cands[1:]


def _latam_json_path(ref):
    return os.path.join(LATAM_JSON_ROOT, ref.strftime('%Y'), ref.strftime('%m'), ref.strftime('%d'),
                        '{}_{}.json'.format(_LATAM_JSON_BASE, ref.strftime('%Y%m%d')))


@_req_cached
def _latam_load_cached(ref):
    """A leitura em si — é este resultado que o cache guarda. Ver `_latam_load`."""
    jp = _latam_json_path(ref)
    if not os.path.isfile(jp):
        return jp, None
    try:
        with open(jp, encoding='utf-8') as fh:
            data = json.load(fh) or []
    except Exception:
        return jp, None
    _latam_ensure_meta(data)
    return jp, data


def _latam_load(ref):
    """(json_path, data|None) para `ref`, com a meta garantida nos registros.

    Devolve uma CÓPIA dos registros, nunca a lista que está no cache. Os
    endpoints de add/edit/delete carregam o dia, mexem na lista e só então
    gravam (`data.remove(rec)`, `rec[c] = ...`): com o objeto do cache na mão,
    essa mutação passa a valer para todo mundo ANTES do save — e continua
    valendo quando o save FALHA. A linha some da tela de quem não pediu nada, e
    o request seguinte, que dentro do TTL recebe o mesmo objeto, grava por cima
    o estado que nunca chegou ao disco. É perda de dado sem erro nenhum.

    A cópia é rasa por registro porque toda escrita destes endpoints é escalar
    (`rec[k] = v`), e ela custa uma fração da leitura do share que o cache
    existe para poupar.
    """
    jp, data = _latam_load_cached(ref)
    return jp, (None if data is None else [dict(r) for r in data])


def _latam_save(jp, data):
    """Grava o arquivo-dia do Latam Desk Position.

    O `_bump_cache_gen` está aqui porque este save NÃO passa pelo
    `_atomic_write_json` — ele escreve direto. Sem a chamada, os loaders
    decorados com `@_req_cached` que derivam deste arquivo continuariam
    servindo o resultado anterior por até `SHARED_CACHE_TTL_SECONDS`: a pessoa
    edita a linha, a tela recarrega e mostra o valor de antes, sem erro nenhum.
    """
    os.makedirs(os.path.dirname(jp), exist_ok=True)
    _atomic_write_json(jp, data)                # funil: bump + espelho (§335)


def _latam_find(data, rid):
    for rec in data:
        if str(rec.get('_lt_id', '')) == str(rid):
            return rec
    return None


def _latam_all_dates():
    """Datas (datetime) que TÊM JSON de Latam Desk Position, da mais nova para a
    mais antiga. O relatório não é diário, então é isso que a página usa para
    saber o que existe."""
    root = os.path.normpath(LATAM_JSON_ROOT)
    if not os.path.isdir(root):
        return []
    pref = _LATAM_JSON_BASE + '_'
    out = []
    for _dirpath, _dirs, files in os.walk(root):
        for f in files:
            if not (f.startswith(pref) and f.endswith('.json')) or f.endswith('.meta.json'):
                continue
            try:
                out.append(datetime.strptime(f[len(pref):-5], '%Y%m%d'))
            except ValueError:
                continue
    return sorted(set(out), reverse=True)


def _latam_latest_ref():
    """Data do último JSON disponível (None se nunca foi importado)."""
    dates = _latam_all_dates()
    return dates[0] if dates else None


def _latam_write_meta(jp, hhmmss, fname=''):
    """Sidecar <json>.meta.json com a hora do import e o arquivo de origem — a
    página mostra os dois porque o relatório pode ser de qualquer dia."""
    try:
        _atomic_write_json(_ds_meta_path(jp),
                           {'updated': hhmmss or '', 'file': fname or ''})
    except OSError:
        pass


def _latam_read_meta(jp):
    mp = _ds_meta_path(jp)
    if os.path.isfile(mp):
        try:
            with open(mp, encoding='utf-8') as fh:
                d = json.load(fh) or {}
            return str(d.get('updated', '') or ''), str(d.get('file', '') or '')
        except Exception:
            pass
    return _ds_read_updated(jp), ''


def _latam_sniff_format(raw):
    """Formato REAL do arquivo, pelo conteúdo e não pela extensão. O relatório
    chega como `FbiRptLatamDeskPostion-NY-....xls`, mas '.xls' aí é só o nome:
    pode ser um .xls binário de verdade (OLE2/BIFF), um xlsx (zip), uma TABELA
    HTML ou um texto delimitado — geradores desse tipo de relatório usam os
    quatro (o cashflows do OTM, por exemplo, é texto com nome .xlsx). A macro
    nunca precisou distinguir porque o `Workbooks.Open` do Excel fareja sozinho."""
    if raw[:2] == b'PK':
        return 'xlsx'
    if raw[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1':      # OLE2 compound file
        return 'xls'
    head = raw[:2048].lstrip().lower()
    if head[:1] == b'<' and (b'<table' in raw[:65536].lower() or b'<html' in head):
        return 'html'
    return 'text'


def _latam_read_xls(raw):
    """.xls binário (BIFF) via xlrd — import tardio, no mesmo espírito do reportlab:
    sem a lib o erro diz o que instalar, em vez de ler bytes binários como texto e
    devolver zero linha silenciosamente."""
    try:
        import xlrd
    except ImportError:
        raise RuntimeError('Arquivo .xls binário (Excel 97-2003): a lib xlrd não está '
                           'instalada. Rode "pip install xlrd" na instância ou salve o '
                           'relatório como .xlsx / texto.')
    book = xlrd.open_workbook(file_contents=raw)
    sh = book.sheet_by_index(0)
    out = []
    for r in range(sh.nrows):
        row = []
        for c in range(sh.ncols):
            cell = sh.cell(r, c)
            v = cell.value
            if cell.ctype == xlrd.XL_CELL_DATE:              # serial → ISO, que _latam_date entende
                y, mo, d, hh, mi, ss = xlrd.xldate_as_tuple(v, book.datemode)
                v = '{:04d}-{:02d}-{:02d}'.format(y, mo, d) if y else ''
            elif cell.ctype == xlrd.XL_CELL_NUMBER:
                v = str(int(v)) if float(v).is_integer() else repr(v)   # SPN 281808.0 → '281808'
            elif cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                v = ''
            row.append(v)
        out.append(row)
    return out


def _latam_read_html(raw):
    """Tabela HTML salva com nome de planilha (formato comum nesses relatórios):
    pega a PRIMEIRA <table> e devolve <tr>/<td> como linhas e colunas."""
    from html.parser import HTMLParser

    class _Table(HTMLParser):
        def __init__(self):
            HTMLParser.__init__(self, convert_charrefs=True)
            self.rows, self._row, self._cell, self._done = [], None, None, False

        def handle_starttag(self, tag, attrs):
            if self._done:
                return
            if tag == 'tr':
                self._row = []
            elif tag in ('td', 'th') and self._row is not None:
                self._cell = []
            elif tag == 'br' and self._cell is not None:
                self._cell.append(' ')

        def handle_endtag(self, tag):
            if self._done:
                return
            if tag in ('td', 'th') and self._cell is not None:
                self._row.append(''.join(self._cell).strip())
                self._cell = None
            elif tag == 'tr' and self._row is not None:
                self.rows.append(self._row)
                self._row = None
            elif tag == 'table' and self.rows:
                self._done = True            # só a primeira tabela

        def handle_data(self, data):
            if self._cell is not None:
                self._cell.append(data)

    try:
        text = raw.decode('utf-8')
    except UnicodeDecodeError:
        text = raw.decode('latin-1')
    p = _Table()
    p.feed(text)
    return p.rows


def _latam_read_text(raw):
    """Texto delimitado. `_ds_read_rows` assume TAB (como o OpenText da macro), mas
    este relatório vem de um extrator externo: se o header partido por TAB der uma
    coluna só, o separador é outro — testa ; , e | antes de desistir. Sem isso o
    arquivo "carrega" como uma coluna gigante, as colunas 62/63 ficam vazias e o
    filtro descarta TODAS as linhas, deixando a página vazia sem erro nenhum."""
    try:
        text = raw.decode('utf-8-sig')
    except UnicodeDecodeError:
        text = raw.decode('latin-1')
    lines = [ln for ln in text.splitlines() if ln.strip()]
    if not lines:
        return []
    best, best_n = '\t', len(lines[0].split('\t'))
    if best_n < 2:
        for sep in (';', ',', '|'):
            n = len(lines[0].split(sep))
            if n > best_n:
                best, best_n = sep, n
    return [ln.split(best) for ln in lines]


def _latam_read_rows(raw):
    """(linhas, formato detectado). O formato volta para a resposta do import: com
    ele, "0 linhas" deixa de ser um mistério — diz se o arquivo foi lido como
    planilha, HTML ou texto."""
    fmt = _latam_sniff_format(raw)
    if fmt == 'xlsx':
        return _ds_read_rows(raw), fmt
    if fmt == 'xls':
        return _latam_read_xls(raw), fmt
    if fmt == 'html':
        return _latam_read_html(raw), fmt
    return _latam_read_text(raw), fmt


def _latam_extract(rows):
    """Limpa + extrai as colunas do relatório Latam Desk Position (filtro da macro
    nas colunas 62/63; datas em dd/mm/yyyy com o epoch virando ''). Devolve
    (registros, kept, filtered, col_map, labels_não_encontrados). Compartilhado
    pela página e pelo card Save Daily Settlement."""
    header = [str(h or '').strip() for h in (rows[0] if rows else [])]
    cmap = _latam_col_map(header)
    # A macro filtra por POSIÇÃO (campos 62 e 63). Preferimos o header resolvido —
    # sobrevive a uma coluna inserida no relatório — e caímos na posição da macro
    # quando o header dos dois CLEARING_TRD_ID_* não foi reconhecido.
    i_int = cmap.get('CLEARING_TRD_ID_INT', _LATAM_FILTER_COLS[0] - 1)
    i_cln = cmap.get('CLEARING_TRD_ID_CLNT', _LATAM_FILTER_COLS[1] - 1)
    out, kept, filtered = [], 0, 0
    for r in rows[1:]:
        if not any(_ds_cell(r, i) for i in range(len(r))):
            continue                                   # linha totalmente vazia
        if not (_ds_cell(r, i_int) or _ds_cell(r, i_cln)):
            filtered += 1
            continue
        rec = {}
        for label in _LATAM_LABELS:
            i = cmap.get(label)
            v = _ds_cell(r, i) if i is not None else ''
            rec[label] = _latam_date(v) if label in _LATAM_DATE_COLS else v
        out.append(rec)
        kept += 1
    _latam_ensure_meta(out)                            # status='OK' + id por linha importada
    missing = [lb for lb in _LATAM_LABELS if lb not in cmap]
    return out, kept, filtered, cmap, missing


def _latam_import(ref=None):
    """Acha o relatório MAIS RECENTE (`_latam_pick_source`) em LATAM_SOURCE_ROOT,
    extrai e grava o JSON da data de referência. O arquivo lido é apagado depois —
    como a macro faz —, mas SÓ quando alguma linha entrou: apagar um arquivo que
    não foi lido (formato inesperado, header diferente) destruiria a única cópia
    antes de dar para investigar. Os demais candidatos ficam em disco intactos e
    voltam na resposta em `ignored`."""
    ref = ref or datetime.now()
    if not os.path.isdir(LATAM_SOURCE_ROOT):
        return {'success': False, 'error': 'Source folder not found: {}'.format(LATAM_SOURCE_ROOT)}
    chosen, ignored = _latam_pick_source(os.listdir(LATAM_SOURCE_ROOT), LATAM_SOURCE_ROOT)
    if not chosen:
        return {'success': False,
                'error': 'No FbiRptLatamDeskPostion-NY-* found in {}'.format(LATAM_SOURCE_ROOT)}
    matches = [chosen]
    src = os.path.join(LATAM_SOURCE_ROOT, chosen)
    try:
        with open(src, 'rb') as fh:
            rows, fmt = _latam_read_rows(fh.read())
    except Exception as exc:
        log.warning('[latam] read failed for %s:\n%s', src, traceback.format_exc())
        return {'success': False, 'error': '{}: {}'.format(matches[0], exc)}
    if not rows or len(rows) < 2:
        return {'success': False,
                'error': 'File {} has no data rows (lido como {})'.format(matches[0], fmt)}

    recs, kept, filtered, cmap, missing = _latam_extract(rows)
    jp = _latam_json_path(ref)
    _latam_save(jp, recs)
    _latam_write_meta(jp, ref.strftime('%H:%M:%S'), matches[0])
    log.info('[latam] imported %s (%s): kept %d (filtered %d) → %s',
             matches[0], fmt, kept, filtered, jp)
    if missing:
        log.warning('[latam] colunas não encontradas no header: %s', ', '.join(missing))
    # Consome o arquivo, como a macra fazia — mas SÓ quando alguma linha entrou:
    # apagar um arquivo que não foi lido (formato inesperado, header diferente)
    # destruiria a única cópia antes de dar para investigar.
    deleted = False
    if kept:
        try:
            os.remove(src)
            deleted = True
        except OSError:
            log.warning('[latam] could not delete source %s', src)
    return {'success': True, 'file': matches[0], 'rows': kept, 'filtered': filtered,
            'missing': missing, 'header_cols': len(rows[0]), 'read': len(rows) - 1,
            'format': fmt, 'deleted': deleted, 'ignored': ignored,
            'date': ref.strftime('%Y-%m-%d'), 'date_fmt': ref.strftime('%d/%m/%Y')}


def _latam_collect(ref):
    """JSON de `ref` → linhas de exibição + widgets. As datas são reformatadas
    aqui também (defensivo: JSON gravado antes desta regra, ou linha inserida à
    mão, continua saindo dd/mm/yyyy e sem o epoch)."""
    widgets = {'calls': 0, 'puts': 0, 'counterparties': 0, 'total': 0}
    jp = _latam_json_path(ref)
    rows_out = []
    if os.path.isfile(jp):
        try:
            with open(jp, encoding='utf-8') as fh:
                data = json.load(fh) or []
        except Exception:
            data = []
        if _latam_ensure_meta(data) and data:              # JSON legado sem meta → migra uma vez
            try:
                _latam_save(jp, data)
            except Exception:
                pass
        cptys = set()
        for rec in data:
            row = []
            for c in _LATAM_LABELS:
                v = rec.get(c, '')
                if c in _LATAM_DATE_COLS:
                    v = _latam_date(v)
                row.append('' if v is None else v)
            row += [rec.get('_lt_status', 'OK'), rec.get('_lt_maker', ''),
                    rec.get('_lt_checker', ''), rec.get('_lt_id', '')]
            rows_out.append(row)
            cp = str(rec.get('CALLPUT', '') or '').strip().upper()[:1]
            if cp == 'C':
                widgets['calls'] += 1
            elif cp == 'P':
                widgets['puts'] += 1
            name = str(rec.get('Counterparty', '') or '').strip().upper()
            if name:
                cptys.add(name)
        widgets['counterparties'] = len(cptys)
        widgets['total'] = len(data)
    updated, fname = _latam_read_meta(jp)
    return {'widgets': widgets, 'columns': _LATAM_LABELS, 'rows': rows_out,
            'updated': updated, 'file': fname}


# ── Latam Desk maker/checker CRUD — toda alteração é persistida no JSON do dia ──
def _latam_ref_from(payload):
    ds = str((payload or {}).get('date', '') or '').strip()
    try:
        return datetime.strptime(ds[:10], '%Y-%m-%d') if ds else (_latam_latest_ref() or datetime.now())
    except ValueError:
        return _latam_latest_ref() or datetime.now()


# ══════════════════════════════════════════════════════════════════════════════
#  NDF Cockpit — Daily Settlement › NDF › Cockpit
#  Source: SETTLEMENT.xlsx (header on ROW 4 starting at column B, data from ROW 5).
#  Modelled on OTM Settlements: per-day JSON, maker/checker CRUD, glass Add/Edit.
#  Columns kept (the file has more; only these): the 16 below (by header name).
#  Formatting: DT_* dates m/d/yyyy → dd/mm/yyyy; notional/tax/settlement → #,##0.00;
#  VL_FORWARD_RATE → 0.000000; LEGAL & NM_COUNTERPARTY → UPPER (stored upper-cased).
# ══════════════════════════════════════════════════════════════════════════════
NDFC_SOURCE_ROOT = os.getenv('NDFC_SOURCE_ROOT', SETTLEMENTS_ROOT)
NDFC_JSON_ROOT = OTM_JSON_ROOT
_NDFC_COLUMNS = [
    'LEGAL', 'NM_COUNTERPARTY', 'ID_SOURCE_DEAL', 'DT_DEAL', 'CD_CETIP_RETURN', 'DT_SETTLEMENT',
    'CCY_NOTIONAL_LC', 'VL_NOTIONAL_LC', 'CCY_NOTIONAL_FC', 'VL_NOTIONAL_FC',
    'VL_STRIKE_PRICE', 'VL_FORWARD_RATE', 'PUBLISHER',
    'VL_TAX_INCOME',
    'ID_DEAL', '[PROD] Cockpit.SETTLEMENT', 'NB_BANK', 'CD_BRANCH', 'CD_BANK_ACCOUNT',
]
_NDFC_HEADER_ROW = 4                                  # 1-based (data from row 5)
_NDFC_DATE_COLS  = {'DT_DEAL', 'DT_SETTLEMENT'}
_NDFC_VALUE_COLS = {'VL_NOTIONAL_LC', 'VL_NOTIONAL_FC', 'VL_TAX_INCOME', '[PROD] Cockpit.SETTLEMENT'}
_NDFC_FWD_COLS   = {'VL_FORWARD_RATE'}
_NDFC_UPPER_COLS = {'LEGAL', 'NM_COUNTERPARTY'}
_NDFC_META_KEYS  = ('_nc_status', '_nc_maker', '_nc_checker', '_nc_id')


def _ndfc_json_path(ref):
    return os.path.join(NDFC_JSON_ROOT, ref.strftime('%Y'), ref.strftime('%m'), ref.strftime('%d'),
                        'ndf-cockpit_{}.json'.format(ref.strftime('%Y%m%d')))


def _ndfc_fmt_date(v):
    """SETTLEMENT dates are US m/d/yyyy (e.g. 7/6/2026 = 6 Jul) → dd/mm/yyyy."""
    s = str(v or '').strip()
    if not s:
        return ''
    for fmt in ('%m/%d/%Y', '%m/%d/%y'):
        try:
            return datetime.strptime(s.split(' ')[0], fmt).strftime('%d/%m/%Y')
        except ValueError:
            continue
    d = _fcst_parse_date(s)                            # fallback (yyyymmdd / dd/mm/yyyy / …)
    return d.strftime('%d/%m/%Y') if d else s


# Piso de casas do forward rate. Antes era também o TETO ('{:.6f}'), e o arquivo
# traz mais: a taxa aparecia arredondada na tela enquanto o Fixing Rate era
# calculado com a precisão inteira — o valor mostrado não explicava o resultado.
# Piso e não valor fixo: acima dele valem as casas que o arquivo tiver, e abaixo
# a taxa curta não encolhe ('5.4' → '5.40000000'), senão a coluna deixa de ler
# como taxa. Oito é o mínimo pedido pela mesa para conferir o fixing — quando o
# arquivo traz só seis, as duas últimas saem zero, e esse zero é informação: diz
# que a precisão que falta está na ORIGEM, não na tela.
_NDFC_FWD_MIN_DEC = 8


def _ndfc_text_decimals(v):
    """Casas decimais que o TEXTO traz. É o que o arquivo escreveu — não uma
    escolha nossa —, e a vírgula decimal conta igual ao ponto."""
    s = str(v or '').strip().replace(' ', '').replace(',', '.')
    return len(s.split('.', 1)[1]) if '.' in s else 0


def _ndfc_fmt_fwd(v):
    """Forward rate com TODAS as casas do arquivo (mínimo 6); vírgula decimal
    tolerada; texto não numérico passa inteiro."""
    s = str(v or '').strip()
    if not s:
        return ''
    try:
        n = float(s.replace(' ', '').replace(',', '.'))
    except ValueError:
        return s
    return '{:.{}f}'.format(n, max(_NDFC_FWD_MIN_DEC, _ndfc_text_decimals(s)))


def _ndfc_ensure_meta(data, default_status='OK'):
    changed = False
    for rec in data:
        if not rec.get('_nc_id'):
            rec['_nc_id'] = _otm_new_id(); changed = True
        if '_nc_status' not in rec:
            rec['_nc_status'] = default_status; changed = True
        for k in ('_nc_maker', '_nc_checker'):
            if k not in rec:
                rec[k] = ''; changed = True
    return changed


def _ndfc_load(ref):
    jp = _ndfc_json_path(ref)
    if not os.path.isfile(jp):
        return jp, None
    try:
        with open(jp, encoding='utf-8') as fh:
            data = json.load(fh) or []
    except Exception:
        return jp, None
    _ndfc_ensure_meta(data)
    return jp, data


def _ndfc_save(jp, data):
    os.makedirs(os.path.dirname(jp), exist_ok=True)
    _atomic_write_json(jp, data)                # funil: bump + espelho (§335)


def _ndfc_find(data, rid):
    for rec in data:
        if str(rec.get('_nc_id', '')) == str(rid):
            return rec
    return None


def _ndfc_read_rows(raw):
    """Rows from SETTLEMENT.xlsx — real .xlsx (zip) or tab-delimited text fallback."""
    if raw[:2] == b'PK':
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        return [list(r) for r in ws.iter_rows(values_only=True)]
    return [ln.split('\t') for ln in raw.decode('latin-1').splitlines()]


def _ndfc_extract(rows):
    """Header on row 4, data from row 5. Keep the 16 columns by name; LEGAL &
    NM_COUNTERPARTY stored UPPER; meta stamped 'OK'. Returns (records, kept)."""
    hidx = _NDFC_HEADER_ROW - 1
    if len(rows) <= hidx:
        return [], 0
    header = [str(h or '').strip() for h in rows[hidx]]
    hnorm = [_fcst_norm(h) for h in header]

    def col_idx(name):
        n = _fcst_norm(name)
        if n in hnorm:
            return hnorm.index(n)
        for i, h in enumerate(hnorm):
            if h and (n in h or h in n):
                return i
        return None
    idx_map = {c: col_idx(c) for c in _NDFC_COLUMNS}

    def cell(r, i):
        return str(r[i]).strip() if (i is not None and i < len(r) and r[i] is not None) else ''

    out = []
    for r in rows[hidx + 1:]:
        if not any(cell(r, i) for i in range(len(r))):
            continue                                  # skip fully-blank lines
        rec = {c: cell(r, idx_map.get(c)) for c in _NDFC_COLUMNS}
        for c in _NDFC_UPPER_COLS:
            rec[c] = rec.get(c, '').upper()
        out.append(rec)
    _ndfc_ensure_meta(out)
    return out, len(out)


def _ndfc_import(ref=None):
    """Find SETTLEMENT*.xlsx in NDFC_SOURCE_ROOT, extract, write today's JSON, delete source."""
    ref = ref or datetime.now()
    if not os.path.isdir(NDFC_SOURCE_ROOT):
        return {'success': False, 'error': 'Source folder not found: {}'.format(NDFC_SOURCE_ROOT)}
    matches = sorted(f for f in os.listdir(NDFC_SOURCE_ROOT)
                     if f.lower().startswith('settlement') and f.lower().endswith('.xlsx'))
    if not matches:
        return {'success': False, 'error': 'No SETTLEMENT*.xlsx found in {}'.format(NDFC_SOURCE_ROOT)}
    src_path = os.path.join(NDFC_SOURCE_ROOT, matches[0])
    try:
        with open(src_path, 'rb') as fh:
            rows = _ndfc_read_rows(fh.read())
    except Exception:
        log.warning("[ndfc] read failed for %s:\n%s", src_path, traceback.format_exc())
        return {'success': False, 'error': 'Could not read {}'.format(matches[0])}
    out, kept = _ndfc_extract(rows)
    jp = _ndfc_json_path(ref)
    _ndfc_save(jp, out)
    _ds_write_updated(jp, ref.strftime('%H:%M:%S'))
    try:
        os.remove(src_path)
    except OSError:
        log.warning("[ndfc] could not delete source %s", src_path)
    return {'success': True, 'file': matches[0], 'rows': kept, 'date': ref.strftime('%Y-%m-%d')}


def _ndfc_num(v):
    try:
        return float(str(v or '').replace(' ', '').replace(',', '.'))
    except ValueError:
        return 0.0


# Sentinel emitted when a blank CD_CETIP_RETURN has no match in the Live Position
# NDF file — the front-end renders it as a warning badge ("Missing B3 ID").
_NDFC_MISSING_B3 = '__MISSING_B3_ID__'
# Second-chance rescue via Operations B3 (Resgate/TER): BRL tolerance between the
# B3 "Valor" and the cockpit's [PROD] Cockpit.SETTLEMENT.
_NDFC_OPB3_VAL_TOL = 5.0

# Sisbacen numeric currency code → ISO 3-letter (keys with leading zeros
# stripped). Codes not mapped here fall through as the raw number so ops can
# spot the gap and extend the table.
_NDFC_SISBACEN_CCY = {
    '55': 'DKK', '65': 'NOK', '70': 'SEK', '150': 'AUD', '165': 'CAD',
    '220': 'USD', '245': 'NZD', '425': 'CHF', '470': 'JPY', '540': 'GBP',
    '706': 'ARS', '715': 'CLP', '720': 'COP', '741': 'MXN', '745': 'UYU',
    '785': 'ZAR', '790': 'BRL', '795': 'CNY', '978': 'EUR',
}


def _ndfc_ccy_from_sisbacen(code):
    """'00220' / '220' → 'USD'; unmapped codes return the raw (stripped) code."""
    digits = _acc_digits(code)
    if not digits:
        return ''
    key = digits.lstrip('0') or '0'
    return _NDFC_SISBACEN_CCY.get(key, key)


def _ndfc_valnum(v):
    """Value string → float or None. Accepts BR ('1.234.567,89' / '1234,89') and
    US ('1234567.89') formats — with a comma present, dots are thousands."""
    s = str(v or '').strip()
    if not s:
        return None
    try:
        if ',' in s:
            return float(s.replace('.', '').replace(',', '.'))
        return float(s)
    except ValueError:
        return None


def _ndfc_b3_maps(ref):
    """Lookup maps from the newest Live Position NDF file (DPOSICAO-TER, walking
    back from D-1 ANBIMA of `ref`):
      ident_map: right-14 of "Codigo Identificador" (athena id registered at B3)
                 → "Contrato" (CETIP contract)
      pub_map:   "Contrato" → publisher: "BACEN" when "Nome do Feeder" is BACEN,
                 otherwise the "Tela funcao Consulta" value.
      contr_map: "Contrato" → {'fwd', 'base', 'cnpj', 'ccy_fc', 'ccy_lc', 'pos',
                 'conta_parte', 'conta_cparte'}:
                 Taxa Forward, Valor Base no registro and digits-only CPF/CNPJ da
                 Contraparte (used by the Operations B3 rescue to double-check a
                 candidate contract), plus Simbolo da Moeda (→ CCY_NOTIONAL_FC)
                 and Codigo Sisbacen da Moeda Cotada mapped to ISO-3
                 (→ CCY_NOTIONAL_LC).
    """
    ident_map, pub_map, contr_map = {}, {}, {}
    try:
        path, _ = _ndf_ter_path(_prev_anbima_bizday(ref))
        if not path:
            return ident_map, pub_map, contr_map
        with open(path, encoding='utf-8') as fh:
            data = json.load(fh) or []
        if not data:
            return ident_map, pub_map, contr_map
        keys, _seen = [], set()
        for rec in data:
            for k in rec.keys():
                if k not in _seen:
                    _seen.add(k); keys.append(k)
        k_ident = _fcst_resolve_key(keys, ('Codigo Identificador',))
        k_contr = _fcst_resolve_key(keys, ('Contrato',))
        k_feed  = _fcst_resolve_key(keys, ('Nome do Feeder',))
        k_tela  = _fcst_resolve_key(keys, ('Tela funcao Consulta',))
        k_fwd   = _fcst_resolve_key(keys, ('Taxa Forward',))
        k_base  = _fcst_resolve_key(keys, ('Valor Base no registro',))
        k_cnpj  = _fcst_resolve_key(keys, ('CPF/CNPJ da Contraparte',))
        k_sym   = _fcst_resolve_key(keys, ('Simbolo da Moeda',))
        k_cot   = _fcst_resolve_key(keys, ('Codigo Sisbacen da Moeda Cotada',))
        k_pos   = _fcst_resolve_key(keys, ('Descricao da posicao do Participante',))
        k_cparte = _fcst_resolve_key(keys, ('Codigo da Parte',))
        k_ccpty  = _fcst_resolve_key(keys, ('Codigo da Contraparte',))
        k_emis  = _fcst_resolve_key(keys, ('Data de Emissao',))
        k_venc  = _fcst_resolve_key(keys, ('Data de Vencimento',))
        for rec in data:
            contrato = str(rec.get(k_contr, '') or '').strip() if k_contr else ''
            if not contrato:
                continue
            ident = str(rec.get(k_ident, '') or '').strip() if k_ident else ''
            if ident:
                ident_map[ident[-14:]] = contrato
            feeder = str(rec.get(k_feed, '') or '').strip() if k_feed else ''
            tela = str(rec.get(k_tela, '') or '').strip() if k_tela else ''
            pub_map[contrato.upper()] = 'BACEN' if feeder.upper() == 'BACEN' else tela
            contr_map[contrato.upper()] = {
                'fwd':  _ndfc_valnum(rec.get(k_fwd, '')) if k_fwd else None,
                'base': _ndfc_valnum(rec.get(k_base, '')) if k_base else None,
                'cnpj': _acc_digits(rec.get(k_cnpj, '')) if k_cnpj else '',
                'ccy_fc': str(rec.get(k_sym, '') or '').strip() if k_sym else '',
                'ccy_lc': _ndfc_ccy_from_sisbacen(rec.get(k_cot, '')) if k_cot else '',
                'pos': str(rec.get(k_pos, '') or '').strip() if k_pos else '',
                'conta_parte': str(rec.get(k_cparte, '') or '').strip() if k_cparte else '',
                'conta_cparte': str(rec.get(k_ccpty, '') or '').strip() if k_ccpty else '',
                'emissao': str(rec.get(k_emis, '') or '').strip() if k_emis else '',
                'venc': str(rec.get(k_venc, '') or '').strip() if k_venc else '',
            }
    except Exception:
        log.warning('[ndfc] live-position lookup failed:\n%s', traceback.format_exc())
    return ident_map, pub_map, contr_map


def _ndfc_opb3_resgates(ref):
    """(valor, Título) pairs from the day's Operations B3 rows whose Tipo
    Operação = Resgate and Tipo Título = TER — candidate CETIP contracts for
    cockpit rows whose CD_CETIP_RETURN is still unresolved.

    Peneirado pelo `opb3-events`: uma operação cancelada não é candidata a
    contrato de nada — casá-la aqui amarraria a linha do Cockpit a um Título que
    a B3 já desfez."""
    out = []
    try:
        ops = _opb3_settle_rows(ref)
        for r in (ops or []):
            if _fcst_norm(str(r.get('Tipo Operação', ''))) != 'resgate':
                continue
            if _fcst_norm(str(r.get('Tipo Título', ''))) != 'ter':
                continue
            valor = _ndfc_valnum(r.get('Valor'))
            titulo = str(r.get('Título', '') or '').strip()
            if valor is not None and titulo:
                out.append((valor, titulo))
    except Exception:
        log.warning('[ndfc] opb3 resgates load failed:\n%s', traceback.format_exc())
    return out


def _ndfc_strike_calc(rec, lp, is_cross):
    """VL_STRIKE_PRICE display-time calc (port of the ops Excel formula):
    strike = VL_FORWARD_RATE ± |SETTLEMENT| / VL_NOTIONAL_FC, where the sign is
    + when the settlement direction agrees with the participant's TER position
    (COMPRADOR with positive settlement, or VENDEDOR with negative), − otherwise.
    Only valid for CCY×BRL pairs — a cross-currency NDF settles in BRL but quotes
    the forward in the cross, so the division is in the wrong unit; those show
    '-' until the BRL→quote-ccy conversion is defined."""
    if is_cross:
        return '-'
    pos = _fcst_norm(lp.get('pos', ''))
    if pos not in ('comprador', 'vendedor'):
        return '-'
    settle = _ndfc_valnum(rec.get('[PROD] Cockpit.SETTLEMENT', ''))
    fwd = _ndfc_valnum(rec.get('VL_FORWARD_RATE', ''))
    notional = _ndfc_valnum(rec.get('VL_NOTIONAL_FC', ''))
    if settle is None or fwd is None or not notional:
        return '-'
    ratio = settle / notional
    add = (ratio > 0 and pos == 'comprador') or (ratio < 0 and pos == 'vendedor')
    delta = abs(settle) / notional
    # Mesma precisão do forward que o gerou: o fixing é `forward ± delta`, então
    # mostrar menos casas que o forward esconderia justamente a diferença que a
    # conta produziu, e mostrar mais inventaria dígitos que nenhuma das entradas
    # tem. O cálculo em si sempre foi feito com o valor cheio.
    dec = max(_NDFC_FWD_MIN_DEC, _ndfc_text_decimals(rec.get('VL_FORWARD_RATE', '')))
    return '{:.{}f}'.format(fwd + delta if add else fwd - delta, dec)


def _ndfc_opb3_rescue(rec, resgates, contr_map, by_taxid):
    """Second-chance CD_CETIP_RETURN: find an Operations B3 Resgate/TER whose
    Valor ≈ [PROD] Cockpit.SETTLEMENT (±_NDFC_OPB3_VAL_TOL BRL), then confirm the
    candidate Título against Live Position NDF (forward rate and Valor Base no
    registro must match the cockpit row) and against RefData (the contraparte's
    CNPJ must resolve to the cockpit's NM_COUNTERPARTY). Returns the confirmed
    contract or ''."""
    settle = _ndfc_valnum(rec.get('[PROD] Cockpit.SETTLEMENT', ''))
    fwd = _ndfc_valnum(rec.get('VL_FORWARD_RATE', ''))
    base = _ndfc_valnum(rec.get('VL_NOTIONAL_FC', ''))
    name = _fcst_norm(str(rec.get('NM_COUNTERPARTY', '') or '').strip())
    if settle is None or fwd is None or base is None or not name:
        return ''
    for valor, titulo in resgates:
        if abs(valor - settle) > _NDFC_OPB3_VAL_TOL:
            continue
        lp = contr_map.get(titulo.upper())
        if not lp:
            continue
        if lp['fwd'] is None or round(lp['fwd'], 6) != round(fwd, 6):
            continue
        if lp['base'] is None or abs(lp['base'] - base) > 0.01:
            continue
        ref_name = by_taxid.get(lp['cnpj'], '') if lp['cnpj'] else ''
        if not ref_name or _fcst_norm(ref_name) != name:
            continue
        return titulo
    return ''


def _ndfc_collect(ref):
    """Read the NDF Cockpit JSON for `ref` → display rows (formatted) + widgets."""
    widgets = {'total': 0, 'counterparties': 0, 'notional': '0.00', 'settlement': '0.00'}
    jp = _ndfc_json_path(ref)
    rows_out = []
    if os.path.isfile(jp):
        try:
            with open(jp, encoding='utf-8') as fh:
                data = json.load(fh) or []
        except Exception:
            data = []
        if _ndfc_ensure_meta(data) and data:          # legacy JSON w/o meta → migrate once
            try:
                _ndfc_save(jp, data)
            except Exception:
                pass
        ident_map, pub_map, contr_map = _ndfc_b3_maps(ref)
        resgates = _ndfc_opb3_resgates(ref)
        _, by_taxid = _vcp_refdata_maps()
        cpties, sum_notional, sum_settle = set(), 0.0, 0.0
        for rec in data:
            # Only JPM legal entities are shown (BANCO J.P. MORGAN…, JPMORGAN
            # CHASE…): rows under other legals (e.g. LAWTON…) don't belong to
            # this desk's settlement. Punctuation/spacing-insensitive ("J.P." ≡
            # "JP"). Blank LEGAL stays visible so a hand-added row never
            # vanishes silently. Display-time only — the JSON keeps every row.
            legal = re.sub(r'[^A-Z0-9]', '', str(rec.get('LEGAL', '') or '').upper())
            if legal and not (legal.startswith('BANCOJP') or legal.startswith('JPMORGANCHASE')):
                continue
            # Display-time lookups vs Live Position NDF: fill a blank
            # CD_CETIP_RETURN from ID_SOURCE_DEAL (right-14 → Contrato), and
            # derive PUBLISHER from the matched contract's feeder. When the id
            # lookup misses, try the Operations B3 Resgate/TER rescue before
            # flagging the row as Missing B3 ID.
            cd = str(rec.get('CD_CETIP_RETURN', '') or '').strip()
            if not cd:
                src = str(rec.get('ID_SOURCE_DEAL', '') or '').strip()
                cd = ident_map.get(src[-14:], '') if src else ''
                if not cd:
                    cd = _ndfc_opb3_rescue(rec, resgates, contr_map, by_taxid)
                cd_display = cd or _NDFC_MISSING_B3
            else:
                cd_display = cd
            publisher = str(rec.get('PUBLISHER', '') or '').strip() or pub_map.get(cd.upper(), '')
            lp = contr_map.get(cd.upper(), {}) if cd else {}
            lk_fc, lk_lc = lp.get('ccy_fc', ''), lp.get('ccy_lc', '')
            # Cross-currency NDF (neither leg is BRL): the TER file carries the
            # legs swapped relative to the cockpit, so invert the looked-up pair
            # (e.g. USD/EUR → EUR/USD). Manually typed values are left alone.
            if lk_fc and lk_lc and lk_fc.upper() != 'BRL' and lk_lc.upper() != 'BRL':
                lk_fc, lk_lc = lk_lc, lk_fc
            ccy_fc = str(rec.get('CCY_NOTIONAL_FC', '') or '').strip() or lk_fc
            ccy_lc = str(rec.get('CCY_NOTIONAL_LC', '') or '').strip() or lk_lc
            strike = str(rec.get('VL_STRIKE_PRICE', '') or '').strip()
            if not strike:
                is_cross = bool(ccy_fc and ccy_lc
                                and ccy_fc.upper() != 'BRL' and ccy_lc.upper() != 'BRL')
                strike = _ndfc_strike_calc(rec, lp, is_cross)
            row = []
            for c in _NDFC_COLUMNS:
                v = rec.get(c, '')
                if c == 'CD_CETIP_RETURN':
                    v = cd_display
                elif c == 'PUBLISHER':
                    v = publisher
                elif c == 'CCY_NOTIONAL_FC':
                    v = ccy_fc
                elif c == 'CCY_NOTIONAL_LC':
                    v = ccy_lc
                elif c == 'VL_STRIKE_PRICE':
                    v = strike
                elif c in _NDFC_DATE_COLS:
                    v = _ndfc_fmt_date(v)
                elif c in _NDFC_FWD_COLS:
                    v = _ndfc_fmt_fwd(v)
                elif c in _NDFC_VALUE_COLS:
                    v = _swapchar_fmt_value(v)
                row.append('' if v is None else v)
            row += [rec.get('_nc_status', 'OK'), rec.get('_nc_maker', ''),
                    rec.get('_nc_checker', ''), rec.get('_nc_id', '')]
            rows_out.append(row)
            nm = str(rec.get('NM_COUNTERPARTY', '') or '').strip()
            if nm:
                cpties.add(nm)
            sum_notional += _ndfc_num(rec.get('VL_NOTIONAL_LC', ''))
            sum_settle += _ndfc_num(rec.get('[PROD] Cockpit.SETTLEMENT', ''))
        widgets['total'] = len(rows_out)
        widgets['counterparties'] = len(cpties)
        widgets['notional'] = '{:,.2f}'.format(sum_notional)
        widgets['settlement'] = '{:,.2f}'.format(sum_settle)
    # Headers are displayed UPPER; positional order matches _NDFC_COLUMNS, which
    # the row add/edit endpoints rely on when mapping cells back by index.
    return {'widgets': widgets, 'columns': [c.upper() for c in _NDFC_COLUMNS], 'rows': rows_out,
            'updated': _ds_read_updated(jp)}


def _ndfc_ref_from(payload):
    ds = str((payload or {}).get('date', '') or '').strip()
    try:
        return datetime.strptime(ds[:10], '%Y-%m-%d') if ds else datetime.now()
    except ValueError:
        return datetime.now()


# ══════════════════════════════════════════════════════════════════════════════
#  NDF › Other Publisher — Daily Settlement › NDF › Other Publisher
#  Derived (read-only) view, VCP-style: the base rows are the day's Operations B3
#  entries with Tipo Operação = PENDENTE_CAMBIO, each joined to the other NDF
#  pages by the CETIP contract (the B3 "Título"):
#    CLIENT           ← that Cockpit row's NM_COUNTERPARTY
#    B3 ID            ← Operations B3 "Título"
#    ATHENA ID        ← NDF Cockpit ID_SOURCE_DEAL of the row whose CD_CETIP_RETURN
#                       is this contract (Cockpit resolution reused as-is, so a
#                       contract recovered by athena id or by the Operations B3
#                       rescue counts here too)
#    CCY FC           ← that Cockpit row's CCY_NOTIONAL_FC
#    TX PARIDADE      ← that Cockpit row's VL_STRIKE_PRICE, shown 0.00000000
#    CCY LC           ← always BRL (the quoted leg of a PENDENTE_CAMBIO is local)
#    TX COTADA        ← always 1.00000000
#    CONTA PARTE      ← Live Position NDF "Codigo da Parte" for the contract
#    CONTA CONTRAPARTE← Live Position NDF "Codigo da Contraparte"
#  Values are recomputed on every load; only the maker/checker meta is persisted
#  (per day, keyed by B3 ID), so a confirmation survives a reload.
# ══════════════════════════════════════════════════════════════════════════════
_NDFOP_COLUMNS = ['CLIENT', 'B3 ID', 'ATHENA ID', 'CCY FC', 'TX PARIDADE',
                  'CCY LC', 'TX COTADA', 'CONTA PARTE', 'CONTA CONTRAPARTE']
_NDFOP_OP_TYPE = 'PENDENTE_CAMBIO'
_NDFOP_TX_COTADA = '1.00000000'
_NDFOP_CCY_LC = 'BRL'
# B3 ID is the row identity (and the meta key), so it is never overridable by an edit.
_NDFOP_KEY_COL = _NDFOP_COLUMNS.index('B3 ID')


def _ndfop_key(v):
    """Comparison key tolerant to accent/case/separator ('PENDENTE_CAMBIO',
    'Pendente Câmbio' and 'pendente-cambio' all collapse to the same token)."""
    return re.sub(r'[^a-z0-9]', '', _fcst_norm(str(v or '')))


def _ndfop_fmt8(v):
    """Rate → 0.00000000 (8 decimals). Non-numeric (e.g. the Cockpit's '-' when
    the strike can't be computed) passes through untouched."""
    n = _ndfc_valnum(v)
    return '{:.8f}'.format(n) if n is not None else str(v or '').strip()


def _ndfop_meta_path(ref):
    return os.path.join(NDFC_JSON_ROOT, ref.strftime('%Y'), ref.strftime('%m'), ref.strftime('%d'),
                        'ndf-other-publisher_{}.json'.format(ref.strftime('%Y%m%d')))


def _ndfop_meta_load(ref):
    """{B3 ID → {status, maker, checker, cells, deleted}} — the persisted overlay.
    `cells` holds manual edits (index-aligned to _NDFOP_COLUMNS, only the entries
    that differ from the derived value) and `deleted` hides the row."""
    p = _ndfop_meta_path(ref)
    try:
        with open(p, encoding='utf-8') as fh:
            data = json.load(fh) or {}
        return p, (data if isinstance(data, dict) else {})
    except Exception:
        return p, {}


def _ndfop_meta_save(path, meta):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    _atomic_write_json(path, meta)              # funil: bump + espelho (§335)


def _ndfop_cockpit_index(ref):
    """{CD_CETIP_RETURN (upper) → {'athena', 'strike', 'client', 'ccy_fc'}} built from
    the NDF Cockpit DISPLAY rows, so the contract shown here is exactly the one the
    Cockpit shows (including the ones it recovered via athena id or the Operations B3
    rescue)."""
    out = {}
    try:
        payload = _ndfc_collect(ref)
        cols = payload.get('columns') or []
        i_cd = cols.index('CD_CETIP_RETURN')
        i_ath = cols.index('ID_SOURCE_DEAL')
        i_stk = cols.index('VL_STRIKE_PRICE')
        i_cli = cols.index('NM_COUNTERPARTY')
        i_fc = cols.index('CCY_NOTIONAL_FC')
        for row in payload.get('rows') or []:
            cd = str(row[i_cd] or '').strip()
            if not cd or cd == _NDFC_MISSING_B3:
                continue
            out.setdefault(cd.upper(), {'athena': str(row[i_ath] or '').strip(),
                                        'strike': row[i_stk],
                                        'client': str(row[i_cli] or '').strip(),
                                        'ccy_fc': str(row[i_fc] or '').strip()})
    except Exception:
        log.warning('[ndf-other-publisher] cockpit index failed:\n%s', traceback.format_exc())
    return out


def _ndfop_collect(ref):
    _, ops = _opb3_load(ref)
    cockpit = _ndfop_cockpit_index(ref)
    _, _, contr = _ndfc_b3_maps(ref)
    _, meta = _ndfop_meta_load(ref)
    want = _ndfop_key(_NDFOP_OP_TYPE)
    rows_out = []
    for rec in (ops or []):
        if _ndfop_key(rec.get('Tipo Operação', '')) != want:
            continue
        b3 = str(rec.get('Título', '') or '').strip()
        if not b3:
            continue
        m = meta.get(b3, {})
        if m.get('deleted'):
            continue
        ck = cockpit.get(b3.upper(), {})
        lp = contr.get(b3.upper(), {})
        cells = [
            ck.get('client', ''),
            b3,
            ck.get('athena', ''),
            ck.get('ccy_fc', ''),
            _ndfop_fmt8(ck.get('strike', '')),
            _NDFOP_CCY_LC,
            _NDFOP_TX_COTADA,
            lp.get('conta_parte', ''),
            lp.get('conta_cparte', ''),
        ]
        # A manual edit wins over the derived value, so a fix made here survives
        # the next load even if the upstream file still disagrees. B3 ID is the row
        # identity and therefore not overridable.
        for i, v in enumerate(m.get('cells') or []):
            if i != _NDFOP_KEY_COL and i < len(cells) and str(v or '').strip():
                cells[i] = str(v).strip()
        rows_out.append(cells + [m.get('status', 'New'), m.get('maker', ''),
                                 m.get('checker', ''), b3])
    return {'widgets': {'total': len(rows_out)}, 'columns': list(_NDFOP_COLUMNS),
            'rows': rows_out, 'updated': _ds_read_updated(_opb3_json_path(ref))}


# ── Send to Conecta (Batch Conecta TAXA files) ────────────────────────────────
#  One positional line per row. The record layout (field order, labels, widths
#  and the fixed literals TER/1/0015/000) lives in the File Interface template
#  `taxacambioter` — editing it on /file-interpreter changes the file and the
#  double-click preview without touching code. This code only computes the
#  per-row values, ported from the Excel column formulas: random 10-digit
#  internal number (MID(RAND();3;10)), bank participant 73760009, counterparty
#  account normalised to 8 digits (41007 → 00041007), rates as 4-digit integer
#  + 8-digit decimals. Rows against Lawton (C.Parte 00041007) are mirrored into
#  a second file with Participante ↔ C.Parte swapped — the Lawton-side view of
#  the same trade:
#    TAXA_BANCO.txt  — header TER  00015JPMORGANBM··········yyyymmdd00001 + lines
#    TAXA_LAWTON.txt — header TER  00015INTRAGLAWTONFDO·····yyyymmdd00001 + the
#                      swapped lines (TCO_LAWTON-style participant header)
#  Both written to CONECTA_NEW_PATH like the New Deals send-conecta files.
_NDFOP_FI_KEY = 'taxacambioter'   # File Interface template key
_NDFOP_FI_PAGE = '/ndf-other-publisher'
_NDFOP_PARTICIPANT = '73760009'   # bank participant account
_NDFOP_LAWTON = '00041007'        # Lawton account → triggers the mirrored file
_NDFOP_FI_ERROR = 'File Interpreter template missing/invalid — check /file-interpreter'


def _ndfop_acct8(v):
    """Account → 8-digit code: digits only, left-padded with zeros
    (41007 → 00041007, 73760.10-2 → 73760102)."""
    d = re.sub(r'\D', '', str(v or ''))
    return d.zfill(8) if d else ''


def _ndfop_rate12(v):
    """Rate → positional 12 chars: 4-digit integer part + 8-digit decimals, no
    separator (5.55 → 000555000000, 1 → 000100000000). '' when not usable."""
    n = _ndfc_valnum(v)
    if n is None or n < 0:
        return ''
    ip, dec = '{:.8f}'.format(n).split('.')
    return (ip.zfill(4) + dec) if len(ip) <= 4 else ''


def _ndfop_fi_block(block_id):
    """One block of the `taxacambioter` File Interface template. Missing
    template/block raises ValueError — a B3 file must never fall back to a
    hardcoded layout in silence."""
    tpl = _fi_tpl_cached(_NDFOP_FI_KEY)
    for b in (tpl or {}).get('blocks', []):
        if b.get('id') == block_id:
            return b
    raise ValueError('file-interpreter template missing: {}/{}'.format(_NDFOP_FI_KEY, block_id))


def _ndfop_conecta_fields(cells, swap=False):
    """[(label, value)] of one Conecta line for a display row (indexed by
    _NDFOP_COLUMNS). swap=True builds the Lawton-side view: Participante and
    C.Parte trade places, everything else identical.

    The line itself comes from `_fi_build_line` over the `registro` block —
    labels, order, widths and fixed literals are the template's. The pairs are
    sliced BACK from that line (field width each; a generator value longer
    than its format keeps its own length — the engine never truncates), so the
    preview shows exactly the bytes the file gets."""
    idx = {c: i for i, c in enumerate(_NDFOP_COLUMNS)}
    part, cparte = _NDFOP_PARTICIPANT, _ndfop_acct8(cells[idx['CONTA CONTRAPARTE']])
    if swap:
        part, cparte = cparte, part
    values = {'4': ''.join(random.choices('0123456789', k=10)),
              '5': part,
              '7': cparte,
              '8': str(cells[idx['B3 ID']] or '').strip(),
              '10': _ndfop_rate12(cells[idx['TX PARIDADE']]),
              '11': _ndfop_rate12(cells[idx['TX COTADA']])}
    line = _fi_build_line(_NDFOP_FI_KEY, 'registro', values, page_url=_NDFOP_FI_PAGE)
    fields, pos = [], 0
    for f in _ndfop_fi_block('registro').get('fields', []):
        w = _fi_width(f.get('format')) or 0
        if str(_fi_field_src(f, _NDFOP_FI_PAGE).get('source', '')) != 'Fixed':
            w = max(w, len(values.get(_fi_seq_key(f.get('seq')), '')))
        fields.append((f.get('field', ''), line[pos:pos + w]))
        pos += w
    return fields


def _ndfop_conecta_header(participant):
    """File header (line type 0) via the template's `header` block —
    participant name per view (JPMORGANBM / INTRAGLAWTONFDO), system date."""
    return _fi_build_line(_NDFOP_FI_KEY, 'header',
                          {'4': participant,
                           '5': datetime.today().strftime('%Y%m%d')},
                          page_url=_NDFOP_FI_PAGE)


def _ndfop_rows_by_id(ref):
    """{row id → data cells} of the day's display rows (overrides applied)."""
    n = len(_NDFOP_COLUMNS)
    return {str(r[n + 3]): r[:n] for r in _ndfop_collect(ref)['rows']}


# ══════════════════════════════════════════════════════════════════════════════
#  Cognos — Daily Settlement › Other Products › Option › Cognos
#  Source: "FXO Detail - Beta.xlsx" (header on ROW 1). Modelled on OTM Settlements:
#  per-day JSON, maker/checker CRUD, glass Add/Edit. Keeps the reporting columns
#  below (by header name). Date columns are yyyy-mm-dd in the file → dd/mm/yyyy
#  on the front end. Also fed by the Save Daily Settlement Files card (cog flag).
# ══════════════════════════════════════════════════════════════════════════════
COG_SOURCE_ROOT = os.getenv('COG_SOURCE_ROOT', SETTLEMENTS_ROOT)
COG_JSON_ROOT = OTM_JSON_ROOT
_COG_COLUMNS = [
    'Athena ID', 'Branch', 'MAP STAT', 'TSS Contract NO', 'CNPTY', 'Counterparty SPN',
    'Counterparty Name', 'Buy Sell', 'Call Put Indicator', 'Call Currency', 'Call Amount',
    'Put Currency', 'Put Amount', 'Strike Rate', 'Client Type', 'USD Amount', 'PRM DUE DT',
    'PRM Currency', 'PRM Amount', 'Expiry Date From', 'Expiry Date To', 'TRN DT', 'Trade Date',
    'Event Trade Date', 'OPT STRT DT', 'OPT END DT', 'OPT SET DT', 'FX RATE IND', 'ATH SET CUR',
    'ATH SET AMT', 'Delivery Style', 'CPTY caid', 'Client Type 2', 'Source System',
    'SWIFT BIC / Address', 'Athena Instr Name', 'Direction',
]
# The FXO file has TWO "Client Type" columns — the 2nd is resolved positionally and
# stored under 'Client Type 2' (its display header is normalised back to 'Client Type').
_COG_DUP_HEADER = {'Client Type 2': 'Client Type'}
_COG_DATE_COLS = {'PRM DUE DT', 'Expiry Date From', 'Expiry Date To', 'TRN DT', 'Trade Date',
                  'Event Trade Date', 'OPT STRT DT', 'OPT END DT', 'OPT SET DT'}
# Colunas de VALOR: a tela as imprime como #,##0.00. A lista vai no payload para
# a página não manter uma segunda cópia — coluna acrescentada aqui e esquecida
# lá apareceria como número cru, sem erro nenhum.
#
# `Strike Rate` fica de fora de propósito: é TAXA, não valor. Duas casas fariam
# dois strikes diferentes aparecerem iguais na tela, que é o mesmo cuidado que a
# recon de FXO já toma com esse campo.
_COG_VALUE_COLS = {'Call Amount', 'Put Amount', 'USD Amount', 'PRM Amount', 'ATH SET AMT'}
_COG_META_KEYS = ('_cg_status', '_cg_maker', '_cg_checker', '_cg_id')


def _cog_new_id():
    return uuid.uuid4().hex[:10]


def _cog_ensure_meta(data, default_status='OK'):
    changed = False
    for rec in data:
        if not rec.get('_cg_id'):
            rec['_cg_id'] = _cog_new_id(); changed = True
        if '_cg_status' not in rec:
            rec['_cg_status'] = default_status; changed = True
        for k in ('_cg_maker', '_cg_checker'):
            if k not in rec:
                rec[k] = ''; changed = True
    return changed


def _cog_json_path(ref):
    return os.path.join(COG_JSON_ROOT, ref.strftime('%Y'), ref.strftime('%m'), ref.strftime('%d'),
                        'cognos_{}.json'.format(ref.strftime('%Y%m%d')))


def _cog_load(ref):
    jp = _cog_json_path(ref)
    if not os.path.isfile(jp):
        return jp, None
    try:
        with open(jp, encoding='utf-8') as fh:
            data = json.load(fh) or []
    except Exception:
        return jp, None
    _cog_ensure_meta(data)
    return jp, data


def _cog_save(jp, data):
    os.makedirs(os.path.dirname(jp), exist_ok=True)
    _atomic_write_json(jp, data)                # funil: bump + espelho (§335)


def _cog_find(data, rid):
    for rec in data:
        if str(rec.get('_cg_id', '')) == str(rid):
            return rec
    return None


def _cog_read_rows(path_or_raw):
    """Rows from the FXO Detail file — real .xlsx (zip) or tab-delimited text."""
    raw = path_or_raw if isinstance(path_or_raw, (bytes, bytearray)) else open(path_or_raw, 'rb').read()
    if raw[:2] == b'PK':
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        return [list(r) for r in ws.iter_rows(values_only=True)]
    return [ln.split('\t') for ln in raw.decode('latin-1').splitlines() if ln.strip()]


def _cog_extract(rows):
    """Keep the Cognos reporting columns from the FXO Detail rows (header row 1),
    resolving duplicate 'Client Type' positionally. Returns (records, kept)."""
    if not rows or len(rows) < 2:
        return [], 0
    if str((rows[0][0] if rows[0] else '') or '').strip().lower().startswith('no data available'):
        return [], 0
    header = [str(h or '').strip() for h in rows[0]]
    hnorm = [_fcst_norm(h) for h in header]

    used = set()

    def col_idx(name):
        target = _COG_DUP_HEADER.get(name, name)
        n = _fcst_norm(target)
        for i, h in enumerate(hnorm):        # exact first, honouring already-used indices
            if h == n and i not in used:
                used.add(i)
                return i
        for i, h in enumerate(hnorm):        # substring fallback
            if h and (n in h or h in n) and i not in used:
                used.add(i)
                return i
        return None
    idx_map = {c: col_idx(c) for c in _COG_COLUMNS}

    def cell(r, i):
        return str(r[i]).strip() if (i is not None and i < len(r) and r[i] is not None) else ''

    out = []
    for r in rows[1:]:
        if not any(cell(r, i) for i in range(len(r))):
            continue
        out.append({c: cell(r, idx_map.get(c)) for c in _COG_COLUMNS})
    _cog_ensure_meta(out)
    return out, len(out)


# ── Operations B3 ────────────────────────────────────────────────────────────
#  Curated view of the B3 "Operações" file (tab-delimited). Header on ROW 5, data
#  from row 6 — the JSON keeps only the reporting columns below (by header name).
#  The extraction time is in ROW 2, COLUMN A (HH:MM:SS) → shown as the page's
#  "last updated" timestamp. Import mirrors the Save Daily Settlement Files read
#  logic. JSON: static/data/cache/daily settlement/YYYY/MM/DD/operations-b3_YYYYMMDD.json
OPB3_SOURCE_ROOT = os.getenv('OPB3_SOURCE_ROOT', SETTLEMENTS_ROOT)


# ── Operations B3 maker/checker CRUD — every change/insert persisted to the JSON ──


# ── Operations B3 › Mensageria (time do piloto) ──────────────────────────────
#  Um e-mail por quebra Tipo × Contraparte × Tipo Operação das linhas com
#  Modalidade de Liquidação Bilateral*/Bruta*. Drafts .eml (X-Unsent) com o
#  shell padrão dos avisos; destinatários vêm dos cards CEM / Equities.
_OPB3_MSG_RECIPIENTS_FILE = os.path.join(_DAILY_METRIC_DIR, 'operations_b3_mensageria_recipients.json')


# ── Live Position › NDF ──────────────────────────────────────────────────────
#  Read-only view of the NDF book from the DPOSICAO-TER position JSON (the TER
#  file ships with its own header, so columns resolve by NAME). Besides the
#  reporting columns below, rows whose "Tipo Media Asiática" == ARITMETICA carry
#  a block of per-date columns right after it (yyyymmdd) — those are appended as
#  dynamic columns (only yyyymmdd values are kept, shown dd/mm/yyyy).
_LPNDF_COLUMNS = [
    'Codigo da Parte', 'Nome da Parte', 'CPF/CNPJ do Participante', 'Codigo da Contraparte',
    'Nome da Contraparte', 'CPF/CNPJ da Contraparte', 'Contrato', 'Classe do Ativo Subjacente',
    'Data de Emissao', 'Data de Inicio de Vigencia', 'Data de Vencimento', 'Valor Base no registro',
    'Valor Base atual', 'Taxa Forward', 'Taxa de Cambio', 'Descricao da posicao do Participante',
    'Tipo Indicador Financeiro', 'Tipo do Contrato', 'Situacao do Contrato', 'Codigo Sisbacen da Moeda',
    'Simbolo da Moeda', 'Codigo Sisbacen da Moeda Cotada', 'Codigo do Ativo Subjacente',
    'Data de Fixing do Ativo Subjacente', 'Data de Fixing da Moeda', 'Codigo da Cotacao',
    'Cotacao de Moeda para o Vencimento', 'Tipo de Cotacao', 'Cross-Rate', 'Codigo da Paridade Cross',
    'Data de fixing da Paridade Cross', 'Codigo do Boletim', 'Horario do Boletim', 'Taxa de Paridade',
    'Nome do Feeder', 'Tela funcao Consulta', 'Tipo de Cotacao Moeda', 'Tipo de Paridade',
    'Data de Avaliacao', 'Indicador de Termo a Termo', 'Data de Fixacao',
    'Forma de Atualizacao da taxa a termo', 'Valor ou Percentual Negociado', 'Ajustar Taxa',
    'Responsavel pelo Ajuste de Taxa', 'Data inicial para o ajuste da taxa',
    'Data final para o ajuste da taxa', 'Data do Ajuste da Taxa', 'Cotacao para Fixing',
    'Atualizar Valor Base?', 'Cotacao Inicial', 'Valor Base Original', 'Premio a ser pago pelo',
    'Valor do Premio', 'Modalidade de Liquidacao', 'Data de Liquidacao do Premio',
    'Premio em Moeda Estrangeira', 'Data de Fixing da Moeda do Premio', 'Codigo Identificador',
    'Data de Alteracao', 'Valor Antecipado', 'Criterio de Apuracao', 'Taxa a Termo em Reais',
    'Data da Cotacao Inicial', 'Tipo Media Asiatico',
]
# Date columns → dd/mm/yyyy; value columns → #,##0.00 (derived from the names above;
# "Horario do Boletim" is a time, not a date, so it stays raw).
_LPNDF_DATE_COLS = {c for c in _LPNDF_COLUMNS if c.startswith('Data ')}
_LPNDF_VALUE_COLS = {c for c in _LPNDF_COLUMNS if c.startswith('Valor ')}
# Rate columns → US format (dot decimal), no unnecessary leading zeros, rounded to 8 dp.
_LPNDF_RATE_COLS = {'Taxa Forward', 'Taxa de Cambio'}
# Columns shown as #,##0.00000000 (thousands + fixed 8 decimals).
_LPNDF_DEC8_COLS = {'Taxa de Paridade', 'Cotacao Inicial'}
# Coluna de CPF/CNPJ mostrada como NOME da contraparte (xlookup no RefData).
# Só a da CONTRAPARTE: o 'CPF/CNPJ do Participante' é a NOSSA perna, e trocá-lo
# pelo nome faria a coluna repetir o 'Nome da Parte' que já está ao lado.
_LPNDF_CPTY_NAME_COLS = {'CPF/CNPJ da Contraparte'}
# Max number of Asian-average fixing-date columns to surface (a contract can have
# 100+ fixings; cap to keep the table sane).
_LPNDF_MAX_ASIAN = 60
# Asian-average fixing dates are a FIXED positional block in dposicao.ter: the
# first date is source column CW (0-based index 100) and each subsequent date is
# +3 columns (CZ, DC, …) — every date is followed by a blank + a "0" companion.
_LPNDF_ASIAN_START = 100   # Excel column CW (0-based)
_LPNDF_ASIAN_STEP = 3      # CW → CZ → DC → … (skip the blank + "0" companions)


def _lpndf_fmt_rate(v):
    """BR rate string like '000000000006,02650000' → US '6.0265' (dot decimal, no
    leading zeros, rounded to 8 decimals, trailing zeros trimmed). Non-numeric →
    returned untouched."""
    s = str(v or '').strip()
    if not s:
        return ''
    try:
        f = float(s.replace('.', '').replace(',', '.'))   # drop dot thousands, comma → decimal
    except ValueError:
        return s
    out = ('%.8f' % round(f, 8)).rstrip('0').rstrip('.')
    return out if out else '0'


# ── New Deals: motor movido para platform/new_deals.py (§319) ────────────────
# Os nomes ficam como ALIAS; os caminhos de cache, o _fxo_refdata_by_spn e o
# _generic_nd_cfg ficam AQUI (superficie de patch e chamada interna — §316).
from apps.pages.platform import new_deals as _pf_nd  # noqa: E402


def _mdea_record_rebooks(rebooks, now):
    """Gancho platform → vertical: o pull do NDF (platform/new_deals) grava os
    pares vanilla ↔ FWD Start na vertical mdea — a "única entrada de fora" do
    §10. A platform não importa feature (fronteira da seção 10 do
    check_soc_layers); quem conhece as verticais é a casca, então a travessia
    mora aqui, com o import atrasado porque os entrypoints só são importados no
    fim deste arquivo."""
    from apps.pages.features.mdea import entrypoint as _mdea
    _mdea.record_rebooks(rebooks, now)
_ndf_ter_path = _pf_nd._ndf_ter_path
_find_deal_in_cache = _pf_nd._find_deal_in_cache
_nd_fix_underlying_marker = _pf_nd._nd_fix_underlying_marker
_deal_matches = _pf_nd._deal_matches
_find_fxo = _pf_nd._find_fxo
_FXO_MONTHS_EN = _pf_nd._FXO_MONTHS_EN
_fxo_ccy = _pf_nd._fxo_ccy
_fxo_num = _pf_nd._fxo_num
_fxo_date_dmy = _pf_nd._fxo_date_dmy
_fxo_refdata_by_accronym = _pf_nd._fxo_refdata_by_accronym
_FXO_FIELD_ORDER = _pf_nd._FXO_FIELD_ORDER
_fxo_order_deal = _pf_nd._fxo_order_deal
_fxo_deal_from_row = _pf_nd._fxo_deal_from_row
_fxo_persist_deals = _pf_nd._fxo_persist_deals
_ND_AMEND_SKIP = _pf_nd._ND_AMEND_SKIP
_ND_AMEND_COSMETIC = _pf_nd._ND_AMEND_COSMETIC
_ND_AMEND_COSMETIC_BY_PRODUCT = _pf_nd._ND_AMEND_COSMETIC_BY_PRODUCT
_ND_AMEND_KEEP_STATUS = _pf_nd._ND_AMEND_KEEP_STATUS
_nd_amend_entity = _pf_nd._nd_amend_entity
_nd_amend_same_entity = _pf_nd._nd_amend_same_entity
_nd_amend_flat = _pf_nd._nd_amend_flat
_nd_amend_is_economic = _pf_nd._nd_amend_is_economic
_nd_api_amend = _pf_nd._nd_api_amend
_nd_amend_index = _pf_nd._nd_amend_index
_nd_amend_find = _pf_nd._nd_amend_find
_nd_amend_register = _pf_nd._nd_amend_register
_nd_cancel_in_file = _pf_nd._nd_cancel_in_file
_fxo_persist_new_deals = _pf_nd._fxo_persist_new_deals
_fxo_deals_from_api_records = _pf_nd._fxo_deals_from_api_records
_fxo_api_pull = _pf_nd._fxo_api_pull
_FXO_API_POLL_MIN = _pf_nd._FXO_API_POLL_MIN
_fxo_api_scheduler_started = _pf_nd._fxo_api_scheduler_started
_fxo_api_scheduler_lock = _pf_nd._fxo_api_scheduler_lock
_fxo_api_scheduler_loop = _pf_nd._fxo_api_scheduler_loop
_fxo_api_start_scheduler = _pf_nd._fxo_api_start_scheduler
_ndf_api_norm = _pf_nd._ndf_api_norm
_ndf_api_get = _pf_nd._ndf_api_get
_api_rec_is_cancelled = _pf_nd._api_rec_is_cancelled
_ndf_flat = _pf_nd._ndf_flat
_fxo_internal_cpty_upgrade = _pf_nd._fxo_internal_cpty_upgrade
_fxo_book_disregard_upgrade = _pf_nd._fxo_book_disregard_upgrade
_ndf_weak_leg = _pf_nd._ndf_weak_leg
_ndf_le_from_location = _pf_nd._ndf_le_from_location
_ndf_le_from_accronym = _pf_nd._ndf_le_from_accronym
_ndf_le_accronyms = _pf_nd._ndf_le_accronyms
_ndf_accronym_variants = _pf_nd._ndf_accronym_variants
_ndf_le_row = _pf_nd._ndf_le_row
_ndf_le_refdata = _pf_nd._ndf_le_refdata
_ndf_ref_by_accronym = _pf_nd._ndf_ref_by_accronym
_ndf_api_key = _pf_nd._ndf_api_key
_ndf_interbook_rules = _pf_nd._ndf_interbook_rules
_ndf_is_interbook = _pf_nd._ndf_is_interbook
_ndf_deal_from_api = _pf_nd._ndf_deal_from_api
_generic_nd_persist_new_deals = _pf_nd._generic_nd_persist_new_deals
_NDF_REBOOK_LOOKBACK_MONTHS = _pf_nd._NDF_REBOOK_LOOKBACK_MONTHS
_ndf_rebook_key = _pf_nd._ndf_rebook_key
_ndf_fwdstart_cached_keys = _pf_nd._ndf_fwdstart_cached_keys
_ndf_drop_fwdstart_rebooks = _pf_nd._ndf_drop_fwdstart_rebooks
_ndf_api_pull = _pf_nd._ndf_api_pull
_NDF_API_POLL_MIN = _pf_nd._NDF_API_POLL_MIN
_ndf_api_scheduler_started = _pf_nd._ndf_api_scheduler_started
_ndf_api_scheduler_lock = _pf_nd._ndf_api_scheduler_lock
_ndf_api_scheduler_loop = _pf_nd._ndf_api_scheduler_loop
_ndf_api_start_scheduler = _pf_nd._ndf_api_start_scheduler
_find_ndf_deal_in_cache = _pf_nd._find_ndf_deal_in_cache
_ndf_comm_ter_lines = _pf_nd._ndf_comm_ter_lines
_GENERIC_ND_PC_TYPE = _pf_nd._GENERIC_ND_PC_TYPE
_GENERIC_ND_MC_SOURCE = _pf_nd._GENERIC_ND_MC_SOURCE
_generic_nd_pending_status = _pf_nd._generic_nd_pending_status
_generic_nd_pc_trigger = _pf_nd._generic_nd_pc_trigger
_find_generic_nd_deal = _pf_nd._find_generic_nd_deal
_generic_nd_reenrich = _pf_nd._generic_nd_reenrich
_ndf_publisher_row = _pf_nd._ndf_publisher_row
_NDF_NOTES_BACEN = _pf_nd._NDF_NOTES_BACEN
_ndf_publisher_is_bacen = _pf_nd._ndf_publisher_is_bacen
_ndf_publisher_codes = _pf_nd._ndf_publisher_codes
_ndf_publisher_fonte_info = _pf_nd._ndf_publisher_fonte_info
_generic_ndf_ter_line = _pf_nd._generic_ndf_ter_line
_nd_lawton_mirror = _pf_nd._nd_lawton_mirror
_nd_lawton_sig = _pf_nd._nd_lawton_sig
_ND_MAPPING_ERRORABLE = _pf_nd._ND_MAPPING_ERRORABLE
_generic_nd_mapping_candidates = _pf_nd._generic_nd_mapping_candidates


def _lpndf_collect(ref, exact=False):
    widgets = {'total': 0, 'vanilla': 0, 'other_publisher': 0, 't0': 0, 'commodities': 0}
    path, _src_dref = _ndf_ter_path(ref, exact=exact)
    columns = list(_LPNDF_COLUMNS)
    rows_out = []
    if path:
        try:
            with open(path, encoding='utf-8') as fh:
                data = json.load(fh) or []
        except Exception:
            data = []
        if data:
            # Ordered UNION of keys across ALL records — an Asian (ARITMETICA) row
            # carries a big block of per-date columns that a vanilla row lacks, so
            # keying off data[0] alone would drop them. First-seen order is kept.
            keys, _seen = [], set()
            for rec in data:
                for k in rec.keys():
                    if k not in _seen:
                        _seen.add(k); keys.append(k)

            def resolve(name):
                n = _fcst_norm(name)
                low = [(k, _fcst_norm(k)) for k in keys]
                for k, kn in low:
                    if kn == n:
                        return k                       # exact (accent/case-insensitive) — preferred
                # Fallback: only when the file header is a SUPERSET of the target
                # (n ⊆ kn). The reverse (a short generic key like "Contrato" inside
                # "Tipo do Contrato") would mis-map, so it is intentionally excluded;
                # pick the closest (shortest) candidate.
                cands = [(k, kn) for k, kn in low if kn and n in kn]
                if cands:
                    return min(cands, key=lambda t: len(t[1]))[0]
                return None
            idx = {c: resolve(c) for c in _LPNDF_COLUMNS}

            def is_yyyymmdd(s):
                s = str(s or '').strip()
                return len(s) == 8 and s.isdigit()

            # Asian-average fixing dates: fixed positional block starting at source
            # column CW (index 100), each date +3 columns (CW, CZ, DC, …). Date N ←
            # keys[100 + 3*(N-1)]; stop at the first grid slot no record has a date in
            # (end of the fixing block). CW is always "Média Asiática (data) 1". Same
            # label text as the Live Position Option page.
            date_keys = []
            p = _LPNDF_ASIAN_START
            while p < len(keys) and len(date_keys) < _LPNDF_MAX_ASIAN:
                k = keys[p]
                if not any(is_yyyymmdd(rec.get(k, '')) for rec in data):
                    break
                date_keys.append(k)
                p += _LPNDF_ASIAN_STEP
            asian_labels = ['Média Asiática (data) {}'.format(i + 1) for i in range(len(date_keys))]
            columns = list(_LPNDF_COLUMNS) + asian_labels

            # Widgets: mesma classificação dos cards do NDF Summary
            # (FX + Tipo do Contrato + Código da Cotação), porém SEM o filtro de
            # Data de Vencimento — conta sobre toda a posição viva (Live Position).
            # Resolve tipo/cotação aceitando as duas grafias (do/de), igual ao
            # NDF Summary, para não zerar os buckets se o header vier com "de".
            fx_norm = _fcst_norm('TAXAS DE CAMBIO')
            comm_norm = _fcst_norm('COMMODITIES')
            classe_key = idx['Classe do Ativo Subjacente']
            tipo_key = resolve('Tipo do Contrato') or resolve('Tipo de Contrato')
            cot_key = resolve('Codigo da Cotacao') or resolve('Codigo de Cotacao')

            for rec in data:
                row = []
                for c in _LPNDF_COLUMNS:
                    v = rec.get(idx[c], '') if idx[c] else ''
                    if c in _LPNDF_DATE_COLS:
                        d = _fcst_parse_date(v)
                        v = d.strftime('%d/%m/%Y') if d else (v or '')
                    elif c in _LPNDF_VALUE_COLS:
                        v = _swapchar_fmt_value(v)
                    elif c in _LPNDF_RATE_COLS:
                        v = _lpndf_fmt_rate(v)
                    elif c in _LPNDF_DEC8_COLS:
                        v = _lp_fmt_dec8(v)
                    elif c in _LPNDF_CPTY_NAME_COLS:
                        v = _lp_cpty_by_taxid(v)
                    row.append('' if v is None else v)
                for k in date_keys:                    # one dd/mm/yyyy column per fixing date
                    raw = str(rec.get(k, '') or '').strip()
                    if is_yyyymmdd(raw):
                        d = _fcst_parse_date(raw)
                        row.append(d.strftime('%d/%m/%Y') if d else '')
                    else:
                        row.append('')
                rows_out.append(row)

                # NDFs de câmbio (TAXAS DE CAMBIO) entram na classificação; o
                # Tipo do Contrato + Código da Cotação decidem o bucket:
                #   SISBACEN + Cotação = 0   → T+0
                #   SISBACEN + Cotação <> 0  → Vanilla
                #   FEEDER                   → Other Publisher
                # Commodities → contagem simples por Classe do Ativo Subjacente.
                classe_val = _fcst_norm(str(rec.get(classe_key, ''))) if classe_key else ''
                if classe_val == fx_norm:
                    tp = _fcst_norm(str(rec.get(tipo_key, ''))) if tipo_key else ''
                    if tp == 'sisbacen':
                        cot = str(rec.get(cot_key, '') if cot_key else '').strip().replace(',', '.')
                        try:
                            is_zero = float(cot) == 0
                        except ValueError:
                            is_zero = True            # vazio / não-numérico → tratado como 0
                        if is_zero:
                            widgets['t0'] += 1
                        else:
                            widgets['vanilla'] += 1
                    elif tp == 'feeder':
                        widgets['other_publisher'] += 1
                elif classe_val == comm_norm:
                    widgets['commodities'] += 1
            widgets['total'] = len(data)      # Total = contagem da posição viva (todas as linhas)
    # `source_date` é a data do ARQUIVO que foi lido, que nem sempre é a data
    # pedida: sem `exact`, o resolvedor anda para trás até achar posição. Quem
    # monta uma série precisa saber disso — senão o dia sem arquivo entra na
    # planilha com o conteúdo do dia anterior e a data de hoje.
    return {'widgets': widgets, 'columns': columns, 'rows': rows_out,
            'source_date': _b3_dref_to_iso(_src_dref)}


# ── Live Position › Option ───────────────────────────────────────────────────
#  Read-only view of the Option book from the DPOSICAO position JSON (the file
#  ships with its own header, so columns resolve by NAME — same engine as NDF).
#  If the file carries a trailing block of per-date columns (yyyymmdd) not mapped
#  to any reporting column below, they are appended as dynamic "Média Asiática"
#  columns (shown dd/mm/yyyy), mirroring the NDF page.
_LPOPT_COLUMNS = [
    'Código IF', 'Tipo de Opção', 'Combinação de operações', 'Posição da Parte', 'Parte (Conta)',
    'Parte (Nome simplificado)', 'CPF/CNPJ Cliente Parte', 'Contraparte (Conta)',
    'Contraparte (Nome simplificado)', 'CPF/CNPJ Cliente Contraparte', 'Data Registro', 'Data Início',
    'Data Vencimento', 'Classe do ativo subjacente', 'Classe do ativo VCP', 'Tipo de Mercadoria',
    'Opção quanto', 'Cotação para opção quanto', 'Ativo subjacente / Moeda base',
    'Moeda do ativo / Moeda cotada', 'Quantidade', 'Quantidade Antecipada',
    'Strike/Limitador/Barreiras em Valor ou Percentual?', 'Strike (percentual)', 'Strike (valor)',
    'Valor base', 'Tipo de Exercício', 'Situação do contrato', 'Data Referência para valores em %',
    'Data de fixing do ativo subjacente', 'Tipo de cotação', 'Data de fixing da moeda do ativo subjacente',
    'Strike/Limitador/Barreiras em Reais', 'Limitador', 'Limitador (percentual)',
    'Periodicidade de Verificação das Barreiras', 'Rebate Unitário', 'Barreira de KI',
    'Direção Barreira de KI', 'Barreira KI (percentual)', 'Tipo de Cotação p/ Verificação de Barreira de KI',
    'Barreira de KO', 'Direção Barreira de KO', 'Barreira KO (percentual)',
    'Tipo de Cotação p/ Verificação de Barreira de KO', 'Liquidação do Rebate de KO',
    'Situação Barreira KI', 'Situação Barreira KO', 'Média Asiática', 'Quantidade de datas de verificação',
    'Modalidade de liquidação do prêmio', 'Prêmio Unitário', 'Valor financeiro total do prêmio',
    'Valor financeiro Total do Rebate', 'Data de Liquidação do Prêmio', 'Data da Última Antecipação',
    'Proteção contra Proventos', 'Informações complementares', 'Cotação Inicial do Ativo (em reais)',
    'Trigger Proporção',
]
# Date columns → dd/mm/yyyy; value columns → #,##0.00 (derived from the names above).
_LPOPT_DATE_COLS = {c for c in _LPOPT_COLUMNS if c.startswith('Data ')}
_LPOPT_VALUE_COLS = {c for c in _LPOPT_COLUMNS if c.startswith('Valor ')}
# 8-decimal (#,##0.00000000), 2-decimal (#,##0.00) and CPF/CNPJ-masked columns.
_LPOPT_DEC8_COLS = {'Strike (valor)', 'Prêmio Unitário', 'Cotação Inicial do Ativo (em reais)'}
_LPOPT_DEC2_COLS = {'Quantidade', 'Quantidade Antecipada'}
# CPF/CNPJ mascarado: só a coluna da PARTE, que é a nossa perna. A da
# CONTRAPARTE virou o NOME dela (xlookup no RefData) — ver `_LPOPT_CPTY_NAME_COLS`.
_LPOPT_CNPJ_COLS = {'CPF/CNPJ Cliente Parte'}
_LPOPT_CPTY_NAME_COLS = {'CPF/CNPJ Cliente Contraparte'}
# Max number of dynamic Asian-average fixing-date columns to surface (cap).
_LPOPT_MAX_ASIAN = 60
# Asian-average fixing dates are a FIXED positional block in dposicao.opc: the
# first date is source column CC (0-based index 80) and each subsequent date is
# +3 columns (CF, CI, …) — every date is followed by a blank + a "0" companion.
_LPOPT_ASIAN_START = 80    # Excel column CC (0-based)
_LPOPT_ASIAN_STEP = 3      # CC → CF → CI → … (skip the blank + "0" companions)


def _opt_dposicao_path(ref, max_back=10, exact=False):
    """Newest existing Option DPOSICAO (path, dref) walking back from `ref`.

    `exact=True` não anda para trás — ver `_ndf_ter_path`."""
    cur = ref
    for _ in range(1 if exact else max_back):
        dref = cur.strftime('%y%m%d')
        p = os.path.join(B3_JSON_ROOT, 'Option', _b3_date_subpath(dref),
                         '73760_{}_DPOSICAO.json'.format(dref))
        if os.path.isfile(p):
            return p, dref
        cur = _prev_anbima_bizday(cur)
    return None, None


def _lpopt_collect(ref, exact=False):
    widgets = {'total': 0, 'a': 0, 'b': 0, 'c': 0}
    path, _src_dref = _opt_dposicao_path(ref, exact=exact)
    columns = list(_LPOPT_COLUMNS)
    rows_out = []
    if path:
        try:
            with open(path, encoding='utf-8') as fh:
                data = json.load(fh) or []
        except Exception:
            data = []
        if data:
            # Ordered UNION of keys across ALL records (a row may carry extra
            # per-date columns another row lacks). First-seen order is kept.
            keys, _seen = [], set()
            for rec in data:
                for k in rec.keys():
                    if k not in _seen:
                        _seen.add(k); keys.append(k)

            def resolve(name):
                n = _fcst_norm(name)
                low = [(k, _fcst_norm(k)) for k in keys]
                for k, kn in low:
                    if kn == n:
                        return k                       # exact (accent/case-insensitive) — preferred
                cands = [(k, kn) for k, kn in low if kn and n in kn]
                if cands:
                    return min(cands, key=lambda t: len(t[1]))[0]
                return None
            idx = {c: resolve(c) for c in _LPOPT_COLUMNS}

            def is_yyyymmdd(s):
                s = str(s or '').strip()
                return len(s) == 8 and s.isdigit()

            # Asian-average fixing dates: fixed positional block starting at source
            # column CC (index 80), each date +3 columns (CC, CF, CI, …). Date N ←
            # keys[80 + 3*(N-1)]; stop at the first grid slot no record has a date in
            # (end of the fixing block). CC is always "Média Asiática (data) 1".
            date_keys = []
            p = _LPOPT_ASIAN_START
            while p < len(keys) and len(date_keys) < _LPOPT_MAX_ASIAN:
                k = keys[p]
                if not any(is_yyyymmdd(rec.get(k, '')) for rec in data):
                    break
                date_keys.append(k)
                p += _LPOPT_ASIAN_STEP
            asian_labels = ['Média Asiática (data) {}'.format(i + 1) for i in range(len(date_keys))]
            columns = list(_LPOPT_COLUMNS) + asian_labels

            for rec in data:
                row = []
                for c in _LPOPT_COLUMNS:
                    v = rec.get(idx[c], '') if idx[c] else ''
                    if c in _LPOPT_DATE_COLS:
                        d = _fcst_parse_date(v)
                        v = d.strftime('%d/%m/%Y') if d else (v or '')
                    elif c in _LPOPT_VALUE_COLS:
                        v = _swapchar_fmt_value(v)
                    elif c in _LPOPT_DEC8_COLS:
                        v = _lp_fmt_dec8(v)
                    elif c in _LPOPT_DEC2_COLS:
                        v = _swapchar_fmt_value(v)
                    elif c in _LPOPT_CNPJ_COLS:
                        v = _lp_fmt_cnpj(v)
                    elif c in _LPOPT_CPTY_NAME_COLS:
                        v = _lp_cpty_by_taxid(v)
                    row.append('' if v is None else v)
                for k in date_keys:
                    raw = str(rec.get(k, '') or '').strip()
                    if is_yyyymmdd(raw):
                        d = _fcst_parse_date(raw)
                        row.append(d.strftime('%d/%m/%Y') if d else '')
                    else:
                        row.append('')
                rows_out.append(row)
            widgets['total'] = len(data)
    # `source_date` é a data do ARQUIVO que foi lido, que nem sempre é a data
    # pedida: sem `exact`, o resolvedor anda para trás até achar posição. Quem
    # monta uma série precisa saber disso — senão o dia sem arquivo entra na
    # planilha com o conteúdo do dia anterior e a data de hoje.
    return {'widgets': widgets, 'columns': columns, 'rows': rows_out,
            'source_date': _b3_dref_to_iso(_src_dref)}


# ── Daily Settlement › NDF › Summary ─────────────────────────────────────────
#  Page modelled on Other Products Summary: header cards from the latest
#  DPOSICAO-TER position JSON + two tables now auto-populated per reference date
#  (see _ndfsum_collect): Trade Level mirrors the NDF Cockpit display rows and
#  Settlement Summary nets them per counterparty. Manual Add-row remains for
#  ad-hoc lines; nothing on this page is persisted.


# ── NDF Summary data: Trade Level from the Cockpit + Settlement Summary ──────
#  Trade Level mirrors the NDF Cockpit DISPLAY rows (so the athena-id and
#  Operations B3 contract rescues apply here for free), joined with the day's
#  Operations B3 by Título for the B3 settlement leg. Settlement Summary nets
#  the trades per counterparty according to the CounterpartyDetails net type
#  and fills the default PAY/RECEIVE bank account. Everything is derived on
#  each load — this endpoint persists nothing.
_NDFSUM_TOL = 5.0                    # |SETTLEMENT − SETTLEMENT B3| tolerance (BRL)


def _ndfsum_refdata_spn():
    """{normalized counterparty name → {'spn', 'taxid'}} from RefData.json
    (recon join rule: first record per name wins)."""
    out = {}
    try:
        with open(os.path.join(_B3_DATA_DIR, 'RefData.json'), encoding='utf-8') as fh:
            data = json.load(fh) or []
    except (IOError, json.JSONDecodeError):
        data = []
    for rec in (data if isinstance(data, list) else []):
        nm = _fcst_norm(str(rec.get('COUNTERPARTY', '') or ''))
        spn = str(rec.get('SPN', '') or '').strip()
        if nm and spn and nm not in out:
            out[nm] = {'spn': spn, 'taxid': str(rec.get('TAX ID', '') or '').strip()}
    return out


def _ndfsum_net_type(rec_cpd):
    """Approved NET.value for a CounterpartyDetails record; anything unconfigured
    or not yet checked falls back to Total Net (same safe rule as the recon)."""
    net = (rec_cpd or {}).get('NET') or {}
    val = str(net.get('value', '') or '').strip()
    status = str(net.get('status', '') or '').strip() or 'Active'
    return val if (val in ('Total Net', 'Pay/Rec', 'No Net') and status == 'Active') else 'Total Net'


# Per-day overlay (only thing this page persists): {counterparty name →
# {'status': 'Generated', 'maker', 'at'}} written when the settlement notices
# are generated, so the Generated pill survives reloads. Lives next to the
# cockpit day JSON.
def _ndfsum_meta_path(ref):
    return os.path.join(NDFC_JSON_ROOT, ref.strftime('%Y'), ref.strftime('%m'), ref.strftime('%d'),
                        'ndf-summary_{}.json'.format(ref.strftime('%Y%m%d')))


def _ndfsum_meta_load(ref):
    path = _ndfsum_meta_path(ref)
    try:
        with open(path, encoding='utf-8') as fh:
            data = json.load(fh)
        return path, (data if isinstance(data, dict) else {})
    except Exception:
        return path, {}


def _ndfsum_meta_save(path, meta):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    _atomic_write_json(path, meta)              # funil: bump + espelho (§335)


def _ndfsum_default_account(banking, slot_key):
    """Conta default aprovada (slot.current) de um slot DEFAULT_PAY/RECEIVE."""
    slot = (banking or {}).get(slot_key) or {}
    return next((a for a in (banking or {}).get('ACCOUNTS', [])
                 if a.get('id') and a.get('id') == slot.get('current')), None)


def _ndfsum_account_fmt(banking, direction):
    """Conta default do CLIENTE para a linha → 'BCO: 341 | AG: 0910 | CC: 967'.

    `direction` é a visão do BANCO (coluna Direction do Settlement Summary);
    os defaults do Reference Data são visão da CONTRAPARTE — por isso o
    cruzamento: banco PAY = cliente recebe → DEFAULT_RECEIVE; banco RECEIVE =
    cliente paga → DEFAULT_PAY. Só default aprovado (slot.current); sem
    default → ''."""
    slot_key = 'DEFAULT_PAY' if direction == 'RECEIVE' else 'DEFAULT_RECEIVE'
    acc = _ndfsum_default_account(banking, slot_key)
    if not acc:
        return ''
    bank = re.sub(r'\D', '', str(acc.get('bank', '') or ''))
    return 'BCO: {} | AG: {} | CC: {}'.format(bank.zfill(3) if bank else '',
                                              str(acc.get('agency', '') or '').strip(),
                                              str(acc.get('account', '') or '').strip())


def _ndfsum_obs_auto(banking):
    """Observação automática do Settlement Summary: classifica as contas
    default do cliente como Internal (banco 376 / JPMorgan) ou External.
    Ambas internas → 'Pay and Receive Internal'; nenhuma → 'Pay and Receive
    External'; mista → 'Pay Internal | Receive External' (ou o inverso).
    Slot sem default aprovado fica fora do rótulo; sem nenhum default → ''."""
    def _internal(acc):
        bank = str(acc.get('bank', '') or '').upper()
        digits = re.sub(r'\D', '', bank)
        return digits.lstrip('0') == '376' or 'JPMORGAN' in re.sub(r'[^A-Z]', '', bank)

    parts = {}
    for label, slot_key in (('Pay', 'DEFAULT_PAY'), ('Receive', 'DEFAULT_RECEIVE')):
        acc = _ndfsum_default_account(banking, slot_key)
        if acc:
            parts[label] = _internal(acc)
    if not parts:
        return ''
    if len(parts) == 2 and parts['Pay'] == parts['Receive']:
        return 'Pay and Receive ' + ('Internal' if parts['Pay'] else 'External')
    return ' | '.join('{} {}'.format(label, 'Internal' if internal else 'External')
                      for label, internal in parts.items())


def _ndfsum_fx_map(ref):
    """{CETIP contract (upper) → 'vanilla' | 't0' | 'other'} for the FX NDFs
    (Classe do Ativo Subjacente = TAXAS DE CAMBIO) maturing on `ref`, read from
    the latest Live Position NDF (DPOSICAO-TER). Vanilla = SISBACEN with a
    non-zero Código da Cotação, T+0 = SISBACEN with Cotação 0 (same-day fixing),
    Other Publisher = FEEDER.

    This is the whitelist that keeps the B3 reconciliation honest: Operations B3
    mixes FX NDFs, commodity NDFs and others, so only contracts in this map count
    on the B3 side — otherwise the totals could never tie out."""
    out = {}
    path, _ = _ndf_ter_path(_prev_anbima_bizday(datetime.now()))
    if not path:
        return out
    try:
        with open(path, encoding='utf-8') as fh:
            data = json.load(fh) or []
    except Exception:
        return out
    if not data:
        return out
    fx = _fcst_norm('TAXAS DE CAMBIO')
    want_mat = ref.strftime('%Y%m%d')
    tipo_key = contr_key = venc_key = classe_key = cot_key = None
    for k in data[0].keys():
        kn = _fcst_norm(k)
        if kn in ('tipo do contrato', 'tipo de contrato'):
            tipo_key = k
        elif kn == 'contrato':
            contr_key = k
        elif kn == 'data de vencimento':
            venc_key = k
        elif kn == 'classe do ativo subjacente':
            classe_key = k
        elif kn in ('codigo da cotacao', 'codigo de cotacao'):
            cot_key = k

    def _is_zero_cot(rec):
        v = str(rec.get(cot_key, '') if cot_key else '').strip().replace(',', '.')
        try:
            return float(v) == 0
        except ValueError:
            return True                          # empty / non-numeric → treated as 0
    for rec in data:
        if classe_key and _fcst_norm(str(rec.get(classe_key, ''))) != fx:
            continue
        d = _fcst_parse_date(rec.get(venc_key, '')) if venc_key else None
        if not (d and d.strftime('%Y%m%d') == want_mat):
            continue
        contrato = str(rec.get(contr_key, '') or '').strip().upper() if contr_key else ''
        if not contrato:
            continue
        tipo = _fcst_norm(str(rec.get(tipo_key, ''))) if tipo_key else ''
        if tipo == 'feeder':
            out[contrato] = 'other'
        else:
            out[contrato] = 't0' if _is_zero_cot(rec) else 'vanilla'
    return out


def _ndfsum_money(n):
    """US thousands (#,##0.00) with parentheses for negatives — the settlement
    convention the reconciliation cards / the reference spreadsheet use."""
    s = '{:,.2f}'.format(abs(n))
    return '({})'.format(s) if n < -0.005 else s


def _ndfsum_b3_legs(ops):
    """Lado B3 do Trade Level, indexado por TRÊS chaves do mais específico ao
    mais frouxo: (Título, Conta, Conta Contraparte), (Título, Conta) e Título.

    As duas primeiras existem por causa do INTRAGRUPO. O mesmo negócio JPM × MGT
    chega pelos DOIS arquivos de casa que alimentam o Operations B3 (o do Banco e
    o da MGT), espelhado: uma linha com Conta 73760.00-9 / Conta Contraparte
    04880.00-6 e o valor de uma ponta, outra com as contas invertidas e o sinal
    trocado. Procurando só pelo Título, quem decidia o sinal era a ordem de
    chegada no arquivo — metade dos intragrupo saía com o Settlement B3 invertido
    contra a coluna SETTLEMENT e a diferença dava o dobro do valor.

    Devolve (resgates, resto): o Resgate é a liquidação propriamente dita e vence
    qualquer outro Tipo Operação do mesmo Título."""
    strong, weak = {}, {}
    for r in (ops or []):
        titulo = str(r.get('Título', '') or '').strip().upper()
        if not titulo:
            continue
        conta = _acc_digits(r.get('Conta', ''))
        ccp = _acc_digits(r.get('Conta Contraparte', ''))
        dst = strong if _fcst_norm(str(r.get('Tipo Operação', ''))) == 'resgate' else weak
        for k in ((titulo, conta, ccp), (titulo, conta), (titulo,)):
            dst.setdefault(k, r.get('Valor', ''))
    return strong, weak


def _ndfsum_b3_val(legs, titulo, casa, cpty_acc):
    """Valor B3 do contrato pela ótica de `casa` (a conta da nossa entidade).

    Vai da chave mais específica para a mais frouxa e, dentro de cada uma, o
    Resgate vence. Chave com pedaço vazio é pulada: `casa` só é conhecida quando
    o LEGAL classifica, e `cpty_acc` só quando a contraparte é a OUTRA entidade
    JPM — que é exatamente o caso em que as duas visões existem."""
    strong, weak = legs
    for k in ((titulo, casa, cpty_acc), (titulo, casa), (titulo,)):
        if not all(k):
            continue
        for d in (strong, weak):
            if k in d:
                return d[k]
    return ''


def _ndfsum_collect(ref):
    ci = {c: i for i, c in enumerate(_NDFC_COLUMNS)}
    # Operations B3 settlement leg, já peneirado pelo cadastro `opb3-events` —
    # uma operação CANCELADA: COMANDADA continua no arquivo com o valor cheio e
    # entrava tanto no Settlement B3 da linha quanto nos cards de conciliação.
    ops = _opb3_settle_rows(ref)
    legs = _ndfsum_b3_legs(ops)

    def _ter_date(v):
        d = _fcst_parse_date(v)
        return d.strftime('%d/%m/%Y') if d else ''

    # B3 × internal reconciliation for the header cards. The FX whitelist (from
    # the Live Position NDF) is the universe on BOTH sides; per category:
    #   internal (JP) = NDF Cockpit rows whose contract is in it → count + Σ SETTLEMENT
    #   B3 (CETIP)    = Operations B3 Resgate rows for those contracts → count + Σ Valor
    fx_map = _ndfsum_fx_map(ref)
    recon_acc = {k: {'b3_count': 0, 'b3_value': 0.0, 'int_count': 0, 'int_value': 0.0}
                 for k in ('vanilla', 't0', 'other', 'total')}

    _, _, contr_map = _ndfc_b3_maps(ref)
    trade, raws = [], []
    for row in _ndfc_collect(ref)['rows']:
        b3 = str(row[ci['CD_CETIP_RETURN']] or '').strip()
        if b3 == _NDFC_MISSING_B3:
            b3 = ''
        lp = contr_map.get(b3.upper(), {}) if b3 else {}
        trade_date = _ter_date(lp.get('emissao', ''))
        settle_date = _ter_date(lp.get('venc', ''))
        # Qual das duas visões da B3 é a NOSSA: a conta de casa sai do LEGAL da
        # linha e, no intragrupo, a conta da contraparte sai do nome dela — que é
        # a outra entidade JPM. Sem isso o sinal vinha da ordem do arquivo.
        casa = _opb3_legal_side(row[ci['LEGAL']])
        cpty_acc = _opb3_legal_side(row[ci['NM_COUNTERPARTY']])
        raw_b3 = _ndfsum_b3_val(legs, b3.upper(), casa, cpty_acc) if b3 else ''
        # Cockpit DISPLAY cells are US-formatted (#,##0.00 via _swapchar_fmt_value)
        # → parse with the US parser; the raw Operations B3 Valor keeps the same
        # BR/US-tolerant parse the resgates reader uses.
        settle_n = _mtm_parse_num(row[ci['[PROD] Cockpit.SETTLEMENT']])
        b3_n = _ndfc_valnum(raw_b3)
        diff = settle_n - b3_n if (settle_n is not None and b3_n is not None) else None
        ok = diff is not None and -_NDFSUM_TOL < diff < _NDFSUM_TOL
        trade.append({
            'cells': [row[ci['LEGAL']], row[ci['NM_COUNTERPARTY']],
                      row[ci['ID_SOURCE_DEAL']], b3, trade_date, settle_date,
                      row[ci['VL_NOTIONAL_FC']], row[ci['CCY_NOTIONAL_FC']],
                      row[ci['VL_FORWARD_RATE']], row[ci['[PROD] Cockpit.SETTLEMENT']],
                      _swapchar_fmt_value(raw_b3) if str(raw_b3 or '').strip() else '',
                      row[ci['VL_STRIKE_PRICE']], row[ci['VL_TAX_INCOME']]],
            'diff': '{:,.2f}'.format(diff) if diff is not None else '',
            'ok': ok,
        })
        # Internal (Cockpit) side of the reconciliation: only FX NDFs settling on
        # ref (contract in the whitelist), bucketed by its category.
        cat = fx_map.get(b3.upper()) if b3 else None
        if cat:
            sv = settle_n or 0.0
            recon_acc[cat]['int_count'] += 1
            recon_acc[cat]['int_value'] += sv
            recon_acc['total']['int_count'] += 1
            recon_acc['total']['int_value'] += sv
        cpty = str(row[ci['NM_COUNTERPARTY']] or '').strip()
        if cpty and settle_n is not None:
            raws.append({
                'counterparty': cpty,
                'legal': str(row[ci['LEGAL']] or '').strip(),
                'athena': str(row[ci['ID_SOURCE_DEAL']] or '').strip(),
                'trade_date': trade_date,
                'notional_fc': abs(_mtm_parse_num(row[ci['VL_NOTIONAL_FC']]) or 0.0),
                'ccy': str(row[ci['CCY_NOTIONAL_FC']] or '').strip(),
                'settlement': settle_n,
                'tax': abs(_mtm_parse_num(row[ci['VL_TAX_INCOME']]) or 0.0),
            })

    spn_by_name = _ndfsum_refdata_spn()
    cpd = _cpd_load()
    _, sum_meta = _ndfsum_meta_load(ref)
    groups = {}
    for r in raws:
        groups.setdefault(r['counterparty'], []).append(r)
    summary = []
    for cpty in sorted(groups, key=_fcst_norm):
        items = groups[cpty]
        ref_rec = spn_by_name.get(_fcst_norm(cpty), {})
        spn = ref_rec.get('spn', '')
        rec_cpd = _cpd_find(cpd, spn) if spn else None
        net_type = _ndfsum_net_type(rec_cpd)
        for r in items:                     # net type + IDs flow to the notices
            r['net_type'] = net_type
            r['spn'] = spn
            r['taxid'] = ref_rec.get('taxid', '')
        # Per-trade cash considering tax: the IR withheld always SHRINKS the cash
        # actually moving — sign-independent, so it is right regardless of whether
        # the cockpit signs the settlement from the bank's or the client's point
        # of view (tax is only present on client gains).
        vals = [(r['settlement'] - r['tax'] if r['settlement'] >= 0
                 else r['settlement'] + r['tax']) for r in items]
        recv = sum(v for v in vals if v > 0)
        pay = sum(v for v in vals if v < 0)
        total = recv + pay
        if net_type == 'Total Net':
            # One netted figure on the side of the final result only.
            recv, pay = (total, 0.0) if total >= 0 else (0.0, total)
        direction = 'RECEIVE' if total >= 0 else 'PAY'
        banking = _bank_norm((rec_cpd or {}).get('BANKING'))
        summary.append({
            'counterparty': cpty,
            'status': (sum_meta.get(cpty) or {}).get('status') or 'New',
            'receive': '{:,.2f}'.format(recv) if recv else '',
            'pay': '{:,.2f}'.format(pay) if pay else '',
            'net_type': net_type,
            'direction': direction,
            'account': _ndfsum_account_fmt(banking, direction),
            # Observação manual (overlay do dia) prevalece; sem ela entra a
            # classificação automática Internal/External das contas default.
            'obs': (sum_meta.get(cpty) or {}).get('obs') or _ndfsum_obs_auto(banking),
        })

    # B3 side: Operations B3 settlement (Resgate) rows for FX-whitelisted
    # contracts only — commodity/other NDFs that also live in Operations B3 are
    # excluded because their Título is not in fx_map.
    for r in (ops or []):
        if _fcst_norm(str(r.get('Tipo Operação', ''))) != 'resgate':
            continue
        cat = fx_map.get(str(r.get('Título', '') or '').strip().upper())
        if not cat:
            continue
        val = _ndfc_valnum(r.get('Valor')) or 0.0
        recon_acc[cat]['b3_count'] += 1
        recon_acc[cat]['b3_value'] += val
        recon_acc['total']['b3_count'] += 1
        recon_acc['total']['b3_value'] += val

    recon = {}
    for k, a in recon_acc.items():
        recon[k] = {
            'b3_count': a['b3_count'], 'b3_value': _ndfsum_money(a['b3_value']),
            'int_count': a['int_count'], 'int_value': _ndfsum_money(a['int_value']),
            'diff_value': _ndfsum_money(a['int_value'] - a['b3_value']),
            # Matched only when both the operation count AND the value totals agree
            # (value within the same ±tolerance the Trade Level uses per trade).
            'matched': (a['b3_count'] == a['int_count']
                        and abs(a['b3_value'] - a['int_value']) <= _NDFSUM_TOL),
        }

    return {'trade': trade, 'summary': summary, 'email_trades': raws, 'recon': recon}


# TED release request (TEDs button on the Settlement Summary) — fixed recipients.
_TED_EMAIL_TO = ['brazil.otc.ops@jpmorgan.com', 'brazil.otc.settlements@jpmorgan.com']


def _ted_ssi_attachment(cpty):
    """Newest file inside the counterparty's Electronic Inventory SSI folder
    (ELECTRONIC_INVENTORY_ROOT/<cpty>/SSI), or None when the folder is missing
    or empty."""
    try:
        folder = _ei_actual_dir_name(_ei_sanitize(cpty))
        ssi_dir = os.path.join(ELECTRONIC_INVENTORY_ROOT, folder, 'SSI')
        if not os.path.isdir(ssi_dir):
            return None
        files = [os.path.join(ssi_dir, f) for f in os.listdir(ssi_dir)
                 if os.path.isfile(os.path.join(ssi_dir, f))]
        return max(files, key=os.path.getmtime) if files else None
    except Exception:
        return None


# ============================================================================
#  MtM — Swap Mark-to-Market by line of business (+ COE)
#  Swap file  "…ConsultaInfoDerivativosSemAtualMID" → CEM / EDG / Hybrids /
#             Commodities. Cols A,C,D,E,F,H,K; house account 73760.00-9 in col D;
#             classified via the latest SWAP position (same join as Accrual).
#             The file lists contracts PENDING MtM update, so 'Valor MTM' has no
#             source column and starts blank (K → 'Data Vencimento').
#  COE  file  "Swap-COE-ConsultaMTMCOE" → COE table. Cols A,B,C,D; col G reference
#             date must equal the last ANBIMA business day of the PENULTIMATE month.
#  Disk : MTM_JSON_ROOT/YYYY/MM/DD/mtm_swap_YYYYMMDD.json
#  Source folder: MTM_SOURCE_ROOT\YYYY\mm. Month\DD
# ============================================================================

# Source column (0-based) per fixed header: A=0, C=2, D=3, E=4, F=5, G=6, K=10.
# 'Nome Simplificado Contraparte' comes from col G (6). 'Valor MTM' (pending →
# blank) and 'Comments' (manual) have no source.


# Position of 'Valor MTM' / 'Comments' within a swap-book data row / a COE row.

# CEM MtM values file "VCP_CETIP_MTM": A=Trade Name, B=Counterparty Name,
# C=CETIP ID, D=MTM in BRL. Keep rows where B <> our own GEM-Rates side, join
# C (CETIP ID) to the CEM book's Código IF, D → rounded 2dp (signed).


def _mtm_norm_party(s):
    """Normalize a counterparty label for a lenient match: strip quotes + accents,
    drop ALL whitespace, lowercase. Robust to leading/trailing/inner spacing
    variations and accents (e.g. 'Bco J.P. Morgan … RATES ' with a trailing space)."""
    s = str(s or '').strip().strip("'").strip('"')
    s = ''.join(ch for ch in unicodedata.normalize('NFKD', s) if not unicodedata.combining(ch))
    return re.sub(r'\s+', '', s).lower()


# Our own GEM-Rates side (normalized) — these rows are excluded from the CEM join.


def _mtm_parse_num(s):
    """Parse a US/en amount like "-1,802,855.646864" (comma thousands, dot decimal,
    optional surrounding quotes) → float, or None. This is the format the page stores
    (Valor MTM = '{:,.2f}')."""
    s = str(s or '').strip().strip("'").strip('"').replace(',', '').strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


# Hybrids MtM values file "Stream_level_MTM": col A = Trade Name, col E (idx 4) =
# 'MTM in scaling currency'. SUMIF col E grouped by Trade Name, resolve the
# mapping_swap-hyb.json B3 ID and set the Hybrids row (Código IF = B3 ID).


# ---------------------------------------------------------------------------
#  MtM — fixed-width Conecta file generation (Send batch / Validation)
#  Header (control) line: tipo-linha '0'; data lines: tipo-linha '1'.
#  Rows use the datepicker date; headers use TODAY (system date).
#  Files saved to CONECTA_NEW_PATH and the day's MTM source folder.
# ---------------------------------------------------------------------------
# Intragroup fund accounts that register a mirror line (opposite sign) against the
# Banco book. Atacama (85398.00-5) and Lawton (00041.00-7) trade ONLY vs Banco.
# MGT (04880.00-6) faces only Banco + external clients — never Lawton/Atacama — so
# it is NOT a mirror counterparty here.
# Fixed counterparty file per book: EDG→Atacama, CEM/Hybrids→Lawton
# (MtM_ATACAMA-EDG, MtM_LAWTON-CEM, MtM_LAWTON-HYB).
# A ESTRUTURA da linha MID (ordem, larguras e os literais Fixed — MID, tipos de
# linha, 0848 e os Notionals em branco) sai do cadastro do File Interface via
# _fi_build_line; aqui ficam só os valores calculados, e os rótulos do preview
# vêm dos `field` do template. Template quebrado → ValueError (nada de arquivo
# meio montado). O COE (0475) ainda não tem cadastro e segue montado à mão.


# ── MtM Validation / End Process (EOM) — e-mail to Brazil OTC Ops ──────────────
#  From otc.tracker@jpmorgan.com → brazil.otc.ops@jpmorgan.com.
#  Validation: generate all book files (CEM/EDG/Hybrids swap + COE), attach the
#  Lawton/Atacama view files. End Process: summary of 'Check' rows (recon) with
#  their comments, or a 'no divergence' notice when there are none.


# ── MtM Recon: match the B3 ConsultaInformacoesAtualizMID file against the page ──
#  File name (source folder): Swap-MID-ConsultaInformacoesAtualizMID.<ext>.
#  Header row = 8, data from row 9. Col A = contract ID ('#' removed so it matches
#  the page's Código IF). Col D = house account (73760.00-9)
#  filter. Col G = registered MtM (signed). A page row whose Valor MTM equals the
#  file value → Success (green pill); a divergence → Check (red pill, tooltip = file
#  value) and its Comments field is unlocked.


# ── Accrual JSON helpers (load/save a specific day's file) ───────────────────
# Row layout: [ ...data cells..., status, maker, checker, id ]  (id is last)


def _accrual_parse_date(s):
    try:
        return datetime.strptime(str(s)[:10], '%Y-%m-%d').strftime('%Y%m%d')
    except Exception:
        return None


# ── CEM / EDG factor enrichment (translated & hardened from the Alteryx VBA) ──
#  The bank's view is LE 228. For every CETIP ID (= the accrual "Código IF") we
#  decide which factor feeds the PARTE and the CONTRAPARTE side, then fill the two
#  factor columns ONLY when that side's indexer is VCP (otherwise '-'). A contract
#  that sits in the table but has no factor for a VCP side is flagged 'Missing
#  Accrual'. CEM derives the LE from the workbook's 'Kapital CETIP' sheet
#  (Kapital → LE); EDG already ships A=CETIP / B=Fator Parte / C=Fator Contraparte.


# Factor file kind → (LOB table, parser). HYB ships its factors in cols L/M.


# ── CETIP SWAP "Atualização de PU/Fator" — batch file generation ─────────────
#  One accrual row → 1/2/4 fixed-width records (per updater × VCP leg). Larger
#  account = role/curve "01", smaller = "00" — divergência do manual mantida de
#  propósito (produção fica como está; documentada no notes do cadastro).
#  Updaters = our own group participants (prefixes below); an external bank
#  counterparty is not an updater (bank view only). Records are split by VIEW
#  into ACCRUAL_<VIEW>-<LOB>.txt in the Batch Conecta folder.
#  A ESTRUTURA da linha (ordem, larguras e os literais Fixed — SWAP, tipos de
#  linha, 0015, Tipo de Atualização 00 e o PU em branco de 22 espaços) sai do
#  cadastro do File Interface via _fi_build_line; aqui ficam só os valores
#  calculados. Template quebrado → ValueError (nada de arquivo meio montado).


def _send_accrual_validation_email(subject, html, logo_path, attach_paths):
    """SMTP-only e-mail of the EOM accrual validation to OTC Ops, attaching the
    Lawton/Atacama files. The HTML and logo path are resolved by the caller (so this
    can run in a background thread without a Flask app context). Best-effort."""
    from email.mime.image import MIMEImage
    from email.mime.base import MIMEBase
    from email import encoders
    try:
        msg = MIMEMultipart('mixed')
        msg['Subject'] = subject
        msg['From'] = SHARED_MAILBOX
        msg['To'] = CETIP_OTC_OPS_EMAIL

        related = MIMEMultipart('related')
        alt = MIMEMultipart('alternative')
        alt.attach(MIMEText('Accrual EOM validation files attached.', 'plain', 'utf-8'))
        alt.attach(MIMEText(html, 'html', 'utf-8'))
        related.attach(alt)
        if logo_path:
            with open(logo_path, 'rb') as f:
                img = MIMEImage(f.read())
            img.add_header('Content-ID', '<otc_logo>')
            img.add_header('Content-Disposition', 'inline', filename='logo.png')
            related.attach(img)
        _attach_email_gradient(related)
        msg.attach(related)

        for path in attach_paths:
            try:
                with open(path, 'rb') as f:
                    part = MIMEBase('application', 'octet-stream')
                    part.set_payload(f.read())
                encoders.encode_base64(part)
                part.add_header('Content-Disposition', 'attachment', filename=os.path.basename(path))
                msg.attach(part)
            except Exception:
                log.warning('[accrual] could not attach %s:\n%s', path, traceback.format_exc())

        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=20) as server:
            server.sendmail(SHARED_MAILBOX, [CETIP_OTC_OPS_EMAIL], msg.as_string())
        log.info('[accrual] validation e-mail sent to %s', CETIP_OTC_OPS_EMAIL)
        return True
    except Exception as e:
        log.error('[accrual] validation e-mail FAILED:\n%s', traceback.format_exc())
        return '{}: {}'.format(type(e).__name__, e)


# ── Recon: match the 'operacoes' return file against the saved accrual factors ──
#  operacoes layout: headers on row 5, data from row 6. Filter col B (index 1) to the
#  two house accounts AND col E == 'REGISTRO DE PU/FATOR'. col H (index 7) = título
#  (= Código IF), col P (index 15) = registered factor (BR decimal comma).
#  Simple match: gather all registered factors per Código IF; a VCP leg's factor
#  (Fator Parte / Fator Contraparte) is OK when it appears among them — else Check.


# ── End Process: final EOM status e-mail to OTC Ops (cc Middle Office) ────────
_ACC_ENDPROC_CC = ['renato.montoza@jpmorgan.com', 'danilo.camposfonseca@jpmchase.com']


def _send_accrual_endprocess_email(subject, html, logo_path):
    """SMTP-only final-status e-mail to OTC Ops, cc the Middle Office. Best-effort."""
    from email.mime.image import MIMEImage
    try:
        msg = MIMEMultipart('mixed')
        msg['Subject'] = subject
        msg['From'] = SHARED_MAILBOX
        msg['To'] = CETIP_OTC_OPS_EMAIL
        msg['Cc'] = ', '.join(_ACC_ENDPROC_CC)
        related = MIMEMultipart('related')
        alt = MIMEMultipart('alternative')
        alt.attach(MIMEText('Accrual Swap EOM final status.', 'plain', 'utf-8'))
        alt.attach(MIMEText(html, 'html', 'utf-8'))
        related.attach(alt)
        if logo_path:
            with open(logo_path, 'rb') as f:
                img = MIMEImage(f.read())
            img.add_header('Content-ID', '<otc_logo>')
            img.add_header('Content-Disposition', 'inline', filename='logo.png')
            related.attach(img)
        _attach_email_gradient(related)
        msg.attach(related)
        recipients = [CETIP_OTC_OPS_EMAIL] + _ACC_ENDPROC_CC
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=20) as server:
            server.sendmail(SHARED_MAILBOX, recipients, msg.as_string())
        log.info('[accrual] end-process e-mail sent to %s (cc %s)', CETIP_OTC_OPS_EMAIL, _ACC_ENDPROC_CC)
        return True
    except Exception:
        log.error('[accrual] end-process e-mail FAILED:\n%s', traceback.format_exc())
        return False


@blueprint.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('pages_blueprint.sign_in_page'))


@blueprint.route('/users-roles')
def users_roles():
    if not session.get('authenticated'):
        return redirect(url_for('pages_blueprint.sign_in_page'))
    # User management is admin-only; hide it (404) from everyone else so a
    # non-admin can't enumerate every SID/role/status/email.
    if not _session_is_admin():
        return render_template('pages/error-404.html'), 404
    users = get_all_users()
    role_groups = get_role_groups()
    role_display = {k: v['display'] for k, v in ROLE_META.items()}
    return render_template('pages/users-roles.html', segment='users-roles',
                           users=users, role_groups=role_groups, role_display=role_display)


# ── Page Access (admin-only) ──────────────────────────────────────────────────
@blueprint.route('/page-access')
def page_access():
    if not session.get('authenticated'):
        return redirect(url_for('pages_blueprint.sign_in_page'))
    if not _session_is_admin():
        return redirect('/dashboard')
    return render_template('pages/page-access.html', segment='page-access',
                           users=get_all_users(), is_master=_session_is_master(),
                           cp_cards=_CONTROL_PANEL_CARDS)


@blueprint.route('/api/me/access', methods=['GET'])
def api_my_access():
    """Current user's access, for the sidebar + admin-link visibility. Master →
    everything; an unconfigured admin → everything; a configured user (admin or
    not) → their allowlist; an unconfigured user → full access."""
    if not session.get('authenticated'):
        return jsonify({'success': False}), 401
    if _session_is_master():
        return jsonify({'success': True, 'is_admin': True, 'configured': False, 'pages': []})
    is_admin_role = (session.get('user_role') or '').upper() == 'ADMIN'
    configured, allowed = _get_page_access(session.get('user_sid', ''))
    # is_admin here only tells the sidebar JS to skip hiding — true for master or an
    # unrestricted admin; a configured (restricted) user gets their pages hidden.
    return jsonify({'success': True, 'is_admin': is_admin_role and not configured,
                    'configured': configured, 'pages': sorted(allowed)})


def _target_needs_master(sid):
    """A target whose access only the master may change: admins and other masters."""
    if (sid or '').strip().upper() in _MASTER_SIDS:
        return True
    target = get_user_by_sid((sid or '').strip().upper())
    return bool(target and (target.get('Role') or '').upper() == 'ADMIN')


@blueprint.route('/api/page-access/<sid>', methods=['GET', 'POST'])
def api_page_access(sid):
    if not session.get('authenticated'):
        return jsonify({'success': False}), 401
    if not _session_is_admin():
        return jsonify({'success': False, 'message': 'Admins only.'}), 403
    sid = (sid or '').strip()
    if not sid:
        return jsonify({'success': False, 'message': 'Missing SID'}), 400
    if request.method == 'GET':
        configured, allowed = _get_page_access(sid)
        master_target = sid.upper() in _MASTER_SIDS
        return jsonify({'success': True, 'configured': configured, 'pages': sorted(allowed),
                        'master_target': master_target,
                        'admin_target': _target_needs_master(sid) and not master_target,
                        'locked': _target_needs_master(sid) and not _session_is_master()})
    # Only the master can change the access of an admin or another master.
    if _target_needs_master(sid) and not _session_is_master():
        return jsonify({'success': False,
                        'message': "Only the master can change an admin's access."}), 403
    data = request.get_json(silent=True) or {}
    pages = data.get('pages')
    if not isinstance(pages, list):
        return jsonify({'success': False, 'message': 'pages must be a list'}), 400
    # Only persist real controllable items — nav pages or Control Panel card tokens.
    clean = [u for u in (str(p) for p in pages) if u in _NAV_URLS or u in _CP_CARD_TOKENS]
    _set_page_access(sid, clean)
    _create_notification(session.get('user_sid', ''), session.get('user_name', ''),
                         'Page Access Updated', 'Users',
                         sid + ' — ' + str(len(clean)) + ' page' + ('' if len(clean) == 1 else 's'),
                         target_role='ADMIN')
    return jsonify({'success': True, 'pages': sorted(set(clean))})


@blueprint.route('/api/users/update', methods=['POST'])
def api_update_user():
    if not session.get('authenticated'):
        return jsonify({"success": False, "message": "Not authenticated"}), 401

    # Changing a user's role/status is an admin-only action. Without this check
    # any authenticated user could POST their own SID with role=ADMIN and
    # escalate to full admin on next login.
    if not _session_is_admin():
        return jsonify({"success": False, "message": "Only Admin users can modify accounts."}), 403

    data = request.get_json()
    sid = data.get('sid', '').strip().upper()
    new_role = data.get('role', '').strip()
    new_status = data.get('status', '').strip()
    new_position = data.get('position', '').strip()

    valid_roles = {'', 'ADMIN', 'BO', 'FO', 'MO', 'INSTITUTIONAL', 'HUB'}
    valid_statuses = {'Active', 'Inactive', 'Pending'}
    valid_positions = {'', 'Consultant', 'Intern', 'Analyst', 'Associate', 'Senior Associate', 'VP', 'ED', 'MD'}

    if not sid:
        return jsonify({"success": False, "message": "SID is required."}), 400
    # The master account can only be modified by the master.
    if sid in _MASTER_SIDS and not _session_is_master():
        return jsonify({"success": False, "message": "Only the master can modify the master account."}), 403
    # Granting ADMIN, or modifying an existing admin/master account, requires the
    # master (mirrors the page-access trust model in _target_needs_master).
    if (new_role == 'ADMIN' or _target_needs_master(sid)) and not _session_is_master():
        return jsonify({"success": False, "message": "Only the master can grant or modify Admin accounts."}), 403
    if new_role not in valid_roles:
        return jsonify({"success": False, "message": "Invalid role."}), 400
    if new_status not in valid_statuses:
        return jsonify({"success": False, "message": "Invalid status."}), 400
    if new_position not in valid_positions:
        return jsonify({"success": False, "message": "Invalid position."}), 400

    user = get_user_by_sid(sid)
    if not user:
        return jsonify({"success": False, "message": "User not found."}), 404

    prev_status = user.get("Status", "Pending")
    update_user_role_status(sid, new_role, new_status, new_position)

    # Send activation email when Pending -> Active
    if prev_status == 'Pending' and new_status == 'Active':
        first_name = user["Name"].split()[0] if user["Name"] else sid
        send_account_activated_email(user["Email"], first_name)

    _create_notification(
        session.get('user_sid', ''), session.get('user_name', ''),
        'User Updated', 'Users',
        sid + ' — Role: ' + new_role + ' | Status: ' + new_status
    )
    return jsonify({"success": True, "message": "User updated successfully."})


@blueprint.route('/api/users/delete', methods=['POST'])
def api_delete_user():
    if not session.get('authenticated'):
        return jsonify({"success": False, "message": "Not authenticated"}), 401

    if not _session_is_admin():
        return jsonify({"success": False, "message": "Only Admin users can delete accounts."}), 403

    data = request.get_json()
    sid = data.get('sid', '').strip().upper()

    if not sid:
        return jsonify({"success": False, "message": "SID is required."}), 400

    if sid == session.get('user_sid', '').upper():
        return jsonify({"success": False, "message": "You cannot delete your own account."}), 400

    # The master account can only be removed by the master.
    if sid in _MASTER_SIDS and not _session_is_master():
        return jsonify({"success": False, "message": "Only the master can delete the master account."}), 403

    if not get_user_by_sid(sid):
        return jsonify({"success": False, "message": "User not found."}), 404

    deleted_name = get_user_by_sid(sid).get('Name', sid)
    conn = get_db_connection()
    try:
        conn.execute("DELETE FROM verification_codes WHERE SID = ?", [sid])
        conn.execute("DELETE FROM users WHERE SID = ?", [sid])
        conn.commit()
    finally:
        conn.close()
    # O usuário sumiu do banco; a allowlist dele não pode ficar em memória para
    # o caso de o mesmo SID ser recadastrado dentro do TTL.
    _page_access_forget(sid)

    _create_notification(
        session.get('user_sid', ''), session.get('user_name', ''),
        'User Deleted', 'Users', deleted_name + ' (' + sid + ')'
    )
    return jsonify({"success": True, "message": "User deleted successfully."})


@blueprint.route('/user-info')
def user_info():
    if not session.get('authenticated'):
        return jsonify({"error": "Not authenticated"}), 401
    return jsonify({
        "sid": session.get('user_sid'),
        "name": session.get('user_name'),
        "email": session.get('user_email'),
        "role": session.get('user_role'),
        "client_ip": get_client_ip()
    })


# ==============================================================================
# ROTAS — CACHE DE DEALS (New Deals › Options › Commodities)
# ==============================================================================



# O parse de datas mora em `platform/dates.py` — aliases.
from apps.pages.platform import dates as _pf_dates  # noqa: E402

_parse_date_any = _pf_dates._parse_date_any
_parse_deal_date = _pf_dates._parse_deal_date




# ── Pending Confirmation (DuckDB-backed) ──────────────────────────────────────
# Data lives in three standalone DuckDB files (built by
# scripts/import_pending_confirmation.py). The smart filter's Status chip picks
# which one: Ok → ok, Backlog → backlog, anything else (Pending / non-Ok) →
# pending. All other chips filter the loaded rows via _deal_matches.
# A pasta sai do Config (`DATABASE_DIR`), nunca do diretório do pacote: é ela
# que muda quando os bancos vão para o share, e montar o caminho aqui deixaria
# esta tela lendo o banco local enquanto o resto do app lê o do share.
_PC_DB_DIR = Config.DATABASE_DIR


# Erros de fórmula do Excel, como o openpyxl os entrega (texto) quando lê o valor
# em cache de uma célula com fórmula quebrada.
_XL_ERROR_TEXT = {'#NULL!', '#N/A', '#REF!', '#VALUE!', '#DIV/0!', '#NAME?', '#NUM!',
                  '#SPILL!', '#CALC!', '#GETTING_DATA'}




# ── Pending Confirmation — populate on New Deals Success+mapped ────────────────
# When a New Deals NDF Comm / Opt Comm / FXO deal becomes Status 'Success' and is
# mapped, it is a fresh outstanding confirmation → insert it into the pending DB.
# (Intragroup deals go to Intrag instead, so they are skipped here — this trigger
# complements the existing Intrag ones.)

def _intrag_engine():
    """Gancho para os gravadores da vertical Intrag (features/intrag): os saves
    do New Deals espelham a operação intragrupo para os day-files dela. Import
    atrasado — os entrypoints só são importados no fim deste arquivo.

    Aponta para o `commands` desde a separação interna (§321): os quatro nomes
    que o New Deals chama (`_save_intrag_ndf_entry`,
    `_save_intrag_ndf_moeda_entry`, `_maybe_save_intrag_opt` e
    `_maybe_save_intrag_fxo`) são todos de ESCRITA. O nome da função fica como
    está: ele é a fronteira que o New Deals conhece."""
    from apps.pages.features.intrag import commands
    return commands


# ──────────────────────────────────────────────────────────────────────────
# Horário de Brasília — mora em `platform/anbima.py` (fatia platform/); alias.
# ──────────────────────────────────────────────────────────────────────────
from apps.pages.platform import anbima as _pf_anbima  # noqa: E402

_BR_TZ = _pf_anbima._BR_TZ
_br_now = _pf_anbima._br_now


# ──────────────────────────────────────────────────────────────────────────
# Janela dos schedulers de importação
# ──────────────────────────────────────────────────────────────────────────
# As três rotinas que trazem operação de fora — a API de NDF, a de FXO e a
# varredura do box de commodities — só valem enquanto a mesa opera. O INTERVALO
# de cada uma continua sendo o dela (20 min, 60 min, 30 min); o que a janela
# decide é se aquele tique faz alguma coisa. Fora dela, cada poll era uma ida à
# Athena — ou uma abertura do Outlook — para importar zero operação.
#
# A janela é em horário de BRASÍLIA (`_br_now`), como todo agendamento do app:
# a instância do time não roda necessariamente em BRT, e uma janela medida no
# relógio do servidor abriria e fecharia na hora errada, em silêncio.
_IMPORT_POLL_WINDOW = os.getenv('IMPORT_POLL_WINDOW', '08:00-20:00')


def _parse_hhmm_window(txt):
    """'08:00-20:00' → (480, 1200), em minutos desde a meia-noite.

    Devolve None quando não dá para entender o valor — e aí a janela fica
    SEMPRE ABERTA, que é o comportamento anterior: um `.env` malformado não
    pode desligar a importação do dia sem ninguém pedir."""
    def _min(parte):
        h, _, mi = str(parte).strip().partition(':')
        h, mi = int(h), int(mi or 0)
        if not (0 <= h <= 23 and 0 <= mi <= 59):
            raise ValueError(parte)
        return h * 60 + mi

    try:
        ini, _, fim = str(txt or '').strip().partition('-')
        return _min(ini), _min(fim)
    except (ValueError, TypeError):
        log.warning('[scheduler] IMPORT_POLL_WINDOW inválida (%r) — os '
                    'schedulers de importação rodam o dia inteiro', txt)
        return None


_IMPORT_WINDOW = _parse_hhmm_window(_IMPORT_POLL_WINDOW)


def _import_window_open(now=None):
    """Estamos dentro da janela de importação?

    As duas pontas são INCLUSIVAS — "das 8h às 20h" tem de deixar passar o tique
    das 20h em ponto. Janela com o fim antes do começo (`20:00-08:00`) atravessa
    a meia-noite, em vez de nunca abrir."""
    if not _IMPORT_WINDOW:
        return True
    ini, fim = _IMPORT_WINDOW
    agora = now or _br_now()
    cur = agora.hour * 60 + agora.minute
    return (ini <= cur <= fim) if ini <= fim else (cur >= ini or cur <= fim)


def _import_window_label():
    """Como a janela aparece no log de subida de cada scheduler — é ali que se
    descobre por que o poll das 6h da manhã não importou nada."""
    if not _IMPORT_WINDOW:
        return '24h'
    return '{:02d}:{:02d}-{:02d}:{:02d}'.format(
        _IMPORT_WINDOW[0] // 60, _IMPORT_WINDOW[0] % 60,
        _IMPORT_WINDOW[1] // 60, _IMPORT_WINDOW[1] % 60)




# Sobe com o app (ver `_schedule_on_start`). Com o reloader do Werkzeug a
# fábrica roda no supervisor e no filho, mas a partida é idempotente e a thread
# extra do supervisor fica dormindo — não vale gatear por WERKZEUG_RUN_MAIN.
_schedule_on_start('pending-confirmation', _pc_start_scheduler)




# ── cola da esteira (_mc_*): movida para platform/manual_confirmation.py (§317)
# Os nomes ficam como ALIAS — o _pc_save_from_deal, as features e os testes
# seguem alcancando por `routes.<nome>`, resolvido em tempo de chamada.
from apps.pages.platform import manual_confirmation as _pf_mc  # noqa: E402
_MC_CONFIRMATION_SOURCES = _pf_mc._MC_CONFIRMATION_SOURCES
_COMMODITY_SOURCES = _pf_mc._COMMODITY_SOURCES
_lob_for_source = _pf_mc._lob_for_source
_MC_JPM_SOURCES = _pf_mc._MC_JPM_SOURCES
_MC_NOTIONAL_CCY_FIELD = _pf_mc._MC_NOTIONAL_CCY_FIELD
_mc_notional_ccy = _pf_mc._mc_notional_ccy
_mc_legal_entity = _pf_mc._mc_legal_entity
_mc_save_from_deal = _pf_mc._mc_save_from_deal
_mc_conf_trade_keys = _pf_mc._mc_conf_trade_keys
_mc_ei_link = _pf_mc._mc_ei_link
_mc_stamp_generated = _pf_mc._mc_stamp_generated
_mc_counts = _pf_mc._mc_counts
_MC_DOCS_TTL = _pf_mc._MC_DOCS_TTL
_MC_DOCS_CACHE = _pf_mc._MC_DOCS_CACHE
_MC_DOCS_LOCK = _pf_mc._MC_DOCS_LOCK
_mc_folder_files = _pf_mc._mc_folder_files
_mc_folder_pdfs = _pf_mc._mc_folder_pdfs
_MC_MAIL_TOKENS = _pf_mc._MC_MAIL_TOKENS
_mc_folder_emails = _pf_mc._mc_folder_emails
_MC_SUBJECT_CACHE = _pf_mc._MC_SUBJECT_CACHE
_MC_SUBJECT_LOCK = _pf_mc._MC_SUBJECT_LOCK
_mc_email_subject = _pf_mc._mc_email_subject
_mc_confirmation_docs = _pf_mc._mc_confirmation_docs
_mc_sync_email_subjects = _pf_mc._mc_sync_email_subjects
_mc_flush_email_subjects = _pf_mc._mc_flush_email_subjects
_MC_STAGE_ROLE = _pf_mc._MC_STAGE_ROLE
_MC_STAGE_NOTIFY_ROLES = _pf_mc._MC_STAGE_NOTIFY_ROLES
_mc_notify_roles = _pf_mc._mc_notify_roles
_mc_can_validate = _pf_mc._mc_can_validate
_mc_stage_denied = _pf_mc._mc_stage_denied
_MC_GENERATE_PRODUCTS = _pf_mc._MC_GENERATE_PRODUCTS
_mc_generate_url = _pf_mc._mc_generate_url
_mc_pc_sync = _pf_mc._mc_pc_sync






def _fxo_refdata_by_spn():
    """SPN (leading-zeros stripped) → RefData record, for client/taxid/acronym lookup."""
    out = {}
    try:
        with open(os.path.join(_B3_DATA_DIR, 'RefData.json'), encoding='utf-8') as fh:
            data = json.load(fh)
        for rec in (data if isinstance(data, list) else []):
            key = _norm_spn(rec.get('SPN', ''))
            if key:
                out[key] = rec
    except (IOError, json.JSONDecodeError):
        pass
    return out




def _api_ref_date(value=None):
    """Data de referência do pull da API → datetime. Aceita o formato do input
    date do navegador ('YYYY-MM-DD'), o da própria API ('YYYYMMDD') e o de tela
    ('DD/MM/YYYY'). Vazio ou inválido cai em hoje, que é o default do campo."""
    s = str(value or '').strip()
    if s:
        for fmt in ('%Y-%m-%d', '%Y%m%d', '%d/%m/%Y'):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue
        log.warning('[api-pull] data de referência inválida (%r) — usando hoje', s)
    return datetime.now()


def _api_ref_suffix(ref_dt):
    """' (ref DD/MM/YYYY)' quando o pull não é do dia — entra na notificação para
    ninguém confundir import retroativo com importação do dia."""
    if ref_dt.date() == datetime.now().date():
        return ''
    return ' (ref {})'.format(ref_dt.strftime('%d/%m/%Y'))




_schedule_on_start('opt-fxo', _fxo_api_start_scheduler)


# ──────────────────────────────────────────────────────────────────────────
# NDF — Athena getTrades API import (FWD Start / Other Publisher / Vanilla)
# ──────────────────────────────────────────────────────────────────────────
# One pull, three destinations. Routing per record:
#   1. Instrument Type = FXForwardStartNDF → FWD Start — EXCEPT when the
#      Strike Set Date is today (reference date): the trade will be cancelled
#      and re-booked as vanilla later, so it must not be imported at all.
#   2. Otherwise, o cadastro decide: a linha do publisher em /mapping ›
#      Publisher × B3 (NDF) com NOTES = BACEN → Vanilla; qualquer outro NOTES
#      (e publisher sem linha) → Other Publisher.
#   3. Everything else → Vanilla.
# Same new-only persistence rule as the FXO API pull: a poll or re-import
# never overwrites a deal already in the day file (key Deal+Client).



# ── Mappings (página Mapping, seção Data Base) ────────────────────────────────
#  De-paras que viviam hardcoded no código viram arquivos JSON editáveis pela
#  tela /mapping. Cada tipo tem colunas fixas e um SEED com exatamente os
#  valores que estavam no código: na primeira leitura o arquivo nasce com eles,
#  então o comportamento não muda até alguém editar na tela. O loader cacheia
#  por mtime — edição vale na requisição seguinte, sem restart do servidor.
_MAPPINGS_DIR = data_write('mappings')

# Legal Entities do fluxo de NDF/FXO. Uma lista só, usada pelos dois mappings que
# têm a LE como chave (le-accronym e le-spn), para não divergirem.
_MAP_LE_OPTIONS = ['', 'JPM', 'MGT', 'LAWTON']

# A ATACAMA entra só no le-spn: o le-accronym continua com as três entidades que
# já tinham cadastro de accronym/settlement location. Se ela também passar a ter
# accronym, o caminho é mover 'ATACAMA' para _MAP_LE_OPTIONS e apagar esta lista.
_MAP_LE_SPN_OPTIONS = _MAP_LE_OPTIONS + ['ATACAMA']

# Campos da API getTrades que podem formar o par do filtro interbook. Os nomes
# são os da API já normalizados (espaço/underscore → espaço, maiúsculas), do
# jeito que _ndf_api_norm devolve.
_MAP_INTERBOOK_FIELDS = [
    'OTHER BOOK', 'TRADING BOOK', 'SETTLEMENT LOCATION', 'END COUNTERPARTY',
    'COUNTERPARTY', 'PUBLISHER', 'INSTRUMENT TYPE', 'INSTRUMENT',
    'QUANTITY CURRENCY', 'OTHER QUANTITY UNITS', 'TYPE', 'DEAL NAME',
]


def _interbook_upgrade(rows):
    """Converte linhas do formato antigo (coluna RULE com o par fixo, FIELD A/B
    guardando só os valores) para o par genérico campo+valor. Vale para arquivos
    de instâncias que ainda não pegaram o novo formato."""
    out = []
    for r in rows:
        if 'RULE' not in r:
            out.append(r)
            continue
        cpty = str(r.get('RULE') or '').strip().upper().startswith(('END', 'CPTY'))
        out.append({
            'FIELD A': 'END COUNTERPARTY' if cpty else 'OTHER BOOK',
            'VALUE A': str(r.get('FIELD A') or ''),
            'FIELD B': 'TRADING BOOK' if cpty else 'SETTLEMENT LOCATION',
            'VALUE B': str(r.get('FIELD B') or ''),
            'BOTH WAYS': 'YES' if cpty else '',
        })
    return out


# Produtos que a API Athena atende — são os valores do parâmetro `product` do
# getTrades, e não as páginas. NDF é UM produto que alimenta TRÊS páginas
# (Vanilla, Other Publisher e FWD Start): quem separa as três é o roteamento pelo
# publisher e pelo Instrument Type (§166), não o endereço. Linha com PRODUCT em
# branco vale para qualquer produto daquele uso.
_MAP_API_PRODUCTS = ['', 'NDF', 'FXO', 'Commodities', 'Swaps']

# Domínios das três colunas do Operations B3 que decidem se a linha entra numa
# apuração de liquidação. Tipo Título é fechado (são os quatro tipos de título
# que a B3 registra); Tipo Operação e Status B3 são ABERTOS — as listas abaixo
# são sugestões no campo, não uma trava, porque a B3 acrescenta valores sem
# avisar e um `select` fecharia a porta para o cadastro do valor novo.
_MAP_OPB3_TITULOS = ['', 'TER', 'OPC', 'SWAP', 'COE']
_MAP_OPB3_OPERACOES = [
    'PAGAMENTO DE DIF. AMORTIZACAO', 'PAGAMENTO DE DIF. DE JUROS',
    'PAGAMENTO DE PREMIO', 'RESGATE', 'RESGATE ANTECIPADO',
    'REGISTRO', 'PAGAMENTO DE EVENTO', 'VENCIMENTO ANTECIPADO',
]
_MAP_OPB3_STATUSES = [
    'CANCELADA: COMANDADA', 'FINALIZADA', 'PENDENTE DE LIQUIDACAO FINANCEIRA',
    'PENDENTE DE CONFIRMACAO', 'PENDENTE DE AUTORIZACAO', 'REGISTRADA',
]
_MAP_OPB3_USE = ['Consider', 'Disregard']

# As linhas com que o cadastro nasce. Todo Settlement Advice de Other Products
# (Swap, NDF Commodities e Opção) pergunta a ELE quais eventos entram no aviso —
# e, por consequência, no Settlement Summary e nos cards de reconciliação. A
# seleção do termo estava FIXA no `_ndfadv_collect` (só RESGATE), o que era a
# segunda resposta para a mesma pergunta.
_MAP_OPB3_SEED = (
    # As três do swap, como estavam em `swap-b3-events`. RESGATE e RESGATE
    # ANTECIPADO ficam de fora de propósito: são vencimento/antecipação,
    # não pagamento de diferencial.
    {'TIPO TITULO': 'SWAP', 'TIPO OPERACAO': 'PAGAMENTO DE DIF. AMORTIZACAO',
     'STATUS B3': '', 'USE': 'Consider', 'NOTES': ''},
    {'TIPO TITULO': 'SWAP', 'TIPO OPERACAO': 'PAGAMENTO DE DIF. DE JUROS',
     'STATUS B3': '', 'USE': 'Consider', 'NOTES': ''},
    {'TIPO TITULO': 'SWAP', 'TIPO OPERACAO': 'PAGAMENTO DE PREMIO',
     'STATUS B3': '', 'USE': 'Consider', 'NOTES': ''},
    # Termo de mercadoria e opção: o RESGATE é a liquidação — no termo, o
    # vencimento; na opção, o exercício. É o que o código fixava para o TER e o
    # que passa a valer para o OPC, agora cadastrável pelos dois.
    {'TIPO TITULO': 'TER', 'TIPO OPERACAO': 'RESGATE', 'STATUS B3': '',
     'USE': 'Consider', 'NOTES': 'Liquidação do termo de mercadoria'},
    {'TIPO TITULO': 'OPC', 'TIPO OPERACAO': 'RESGATE', 'STATUS B3': '',
     'USE': 'Consider', 'NOTES': 'Liquidação/exercício da opção'},
    # O prêmio da opção é caixa do dia como qualquer outro, e o aviso o distingue
    # sozinho (assunto prefixado com "(Pagamento de Prêmio)"). No SWAP ele é um
    # Consider próprio pela mesma razão.
    {'TIPO TITULO': 'OPC', 'TIPO OPERACAO': 'PAGAMENTO DE PREMIO', 'STATUS B3': '',
     'USE': 'Consider', 'NOTES': 'Pagamento de prêmio da opção'},
    # A operação cancelada continua no arquivo da B3 com o valor cheio.
    # Somá-la é contar um caixa que não vai acontecer.
    {'TIPO TITULO': '', 'TIPO OPERACAO': '', 'STATUS B3': 'CANCELADA: COMANDADA',
     'USE': 'Disregard', 'NOTES': 'Cancelada na B3 — fora de toda liquidação'},
)




# Quem valida a confirmação: REQUESTED = precisa passar por essa mesa; EXEMPT =
# não precisa. Fechado de propósito — um terceiro valor cairia no ramo do EXEMPT
# e tiraria a mesa da esteira sem ninguém perceber.
_MAP_MC_VALIDATION = ['REQUESTED', 'EXEMPT']

_ATHENA_GETTRADES = ('https://athena-app.jpmchase.net/FXCASH/brazil-trade-data-api'
                     '/api/v1/getTrades?product={}&date=YYYYMMDD')

_API_LINKS_SEED = (
    {'USE': 'New Deals', 'PRODUCT': 'NDF', 'URL': _ATHENA_GETTRADES.format('NDF'),
     'NOTES': 'Alimenta Vanilla, Other Publisher e FWD Start — a página sai do roteamento'},
    {'USE': 'New Deals', 'PRODUCT': 'FXO', 'URL': _ATHENA_GETTRADES.format('FXO'),
     'NOTES': ''},
    {'USE': 'New Deals', 'PRODUCT': 'Commodities', 'URL': _ATHENA_GETTRADES.format('Commodities'),
     'NOTES': 'Sem consumidor: o pull de commodities ainda vem do box'},
    {'USE': 'New Deals', 'PRODUCT': 'Swaps', 'URL': _ATHENA_GETTRADES.format('Swaps'),
     'NOTES': 'Sem consumidor ainda'},
    {'USE': 'Unwinds', 'PRODUCT': '', 'URL': '',
     'NOTES': 'Preencher com a URL de unwinds quando ela existir'},
    # Recon FXO: outro Athena — o relatório EOD do bob-reports, não o getTrades.
    # A data fica no CAMINHO (AAAA-MM-DD), que é justamente o caso para o qual o
    # placeholder existe: nenhum parâmetro de query alcança ali.
    {'USE': 'Recon FXO', 'PRODUCT': 'FXO',
     'URL': ('http://athena-reports.jpmchase.net:8080/bob-reports/'
             'YYYY-MM-DD/EOD/GEM_OFFICIAL_TRD/FXOEODReport/brazil_fxo_trades.csv'),
     'NOTES': 'Relatório EOD que a reconciliação de FXO compara com a posição B3'},
)




# O cabeçalho do relatório EOD de FXO da Athena, na ordem em que ele vem. Serve
# SÓ para o dropdown do cadastro `fxo-book-disregard`: o motor aceita o nome que
# estiver gravado e o procura no arquivo, então esta lista é conveniência de
# tela, não regra. É por isso que ela pode envelhecer sem quebrar nada — uma
# coluna nova no relatório só precisa entrar aqui para ficar escolhível.
_ATHENA_FXO_COLUMNS = [
    # Vazia na frente, como o `_MAP_OPB3_TITULOS`: é ela que deixa o critério 2 e
    # o 3 realmente EM BRANCO. Sem a opção vazia o dropdown nasce na primeira
    # coluna da lista, e a linha ficaria carregando um `ReportDate` que ninguém
    # escolheu — inofensivo hoje (o par só conta com valor preenchido), mas é
    # cadastro dizendo o que não foi dito.
    '',
    'ReportDate', 'ProductID', 'LegalEntity', 'ClientID', 'TradeDate', 'StartDate',
    'MaturityDate', 'IndexCode', 'BasketIndicator', 'OptionType', 'OptionIndicator',
    'DealID', 'OriginalPrincipalAmount', 'NotionalAmount', 'Quantity',
    'GuaranteeIndicator', 'MTMAmountBRL', 'MTMAmountOriginalCCY', 'Premium',
    'PremiumCurrency', 'TransactionType', 'OptionStyle', 'Strike', 'MatchingDealID',
    'CounterpartyName', 'MatchingCounterpartySPN', 'MatchingCounterpartyName',
    'Portfolio', 'SettlementLocation', 'FixingDate', 'SettlementDate', 'Days2Fixing',
    'QuantityCurrency', 'OtherQuantityUnit', 'OtherQuantityCurrency', 'ExpirationCut',
    'Barrier', 'BarrierDirection', 'BarrierInOut', 'RebateValue', 'PayTime',
    'AverageType', 'FirstFixingDate', 'NumberofFixing', 'INT_EXT',
]


# Os tipos de conta da B3, na grafia em que são gravados. É código, não rótulo:
# a comparação é feita sobre ele, e por isso o `select` do cadastro tem esta
# lista e nada mais. CLIENT 1 e CLIENT 2 são as contas GUARDA-CHUVA — a que
# aparece na linha do Operations B3 não diz quem é o cliente.
_B3_ACCOUNT_TYPES = ('OWN', 'CLIENT 1', 'CLIENT 2')
_B3_CLIENT_ACCOUNT_TYPES = ('CLIENT 1', 'CLIENT 2')

# Grafias que valem pelo mesmo tipo. A tabela nasceu em português (PRÓPRIA /
# CLIENTE 1 / CLIENTE 2) e é assim que a mesa a lê no documento da B3; quem
# digitar dessa forma na tela tem de ser entendido, senão a linha deixa de valer
# sem erro nenhum — uma conta de cliente que não é reconhecida como guarda-chuva
# passa a identificar o cliente pelo nome do titular do omnibus.
_B3_ACCOUNT_TYPE_ALIASES = {
    'PROPRIA': 'OWN', 'CONTA PROPRIA': 'OWN', 'OWN ACCOUNT': 'OWN', 'HOUSE': 'OWN',
    'CLIENTE 1': 'CLIENT 1', 'CLIENTE1': 'CLIENT 1', 'CLIENT1': 'CLIENT 1',
    'CLIENTE 2': 'CLIENT 2', 'CLIENTE2': 'CLIENT 2', 'CLIENT2': 'CLIENT 2',
}

# A mensageria do Operations B3 sai na visão de UMA conta nossa (a coluna Conta
# da linha), e a mesma liquidação chega pelas duas pontas quando o negócio é
# intragrupo. Esta coluna é quem decide de que visão a mensagem sai.
_B3_MSG_USES = ('Consider', 'Disregard')


def _b3_account_type(value):
    """Tipo de conta na grafia canônica (`_B3_ACCOUNT_TYPES`) ou ''.

    Cego a caixa, acento e espaço repetido: `Própria`, `PROPRIA` e `própria `
    são o mesmo tipo."""
    txt = unicodedata.normalize('NFKD', str(value or '').strip().upper())
    txt = ''.join(c for c in txt if not unicodedata.combining(c))
    txt = ' '.join(txt.split())
    if txt in _B3_ACCOUNT_TYPES:
        return txt
    return _B3_ACCOUNT_TYPE_ALIASES.get(txt, '')


def _b3_accounts_upgrade(rows):
    """Formato antigo do `b3-omnibus-account` (só ACCOUNT + NOTES) → as contas
    B3 completas, e a grafia do tipo normalizada.

    A tabela antiga listava SÓ as contas guarda-chuva — estar nela era, por si
    só, a resposta. Aqui o que responde é a coluna TIPO, então uma linha sem
    tipo tem de virar CLIENT 1: lida como PRÓPRIA, ela deixaria de mandar o app
    procurar o cliente pelo CNPJ e o aviso de liquidação sairia endereçado ao
    titular do omnibus.

    A coluna MESSAGING nasce depois das outras, então o arquivo que já está em
    disco vem sem ela. Preencher com o SEED (pelas duas chaves, conta e LE) é o
    que mantém a regra que a mesa pediu — Banco assina, MGT/Lawton/Atacama não —
    na instância que já abriu a tela uma vez. Um default cego 'Consider' faria a
    mensagem intragrupo sair pelas duas pontas; um 'Disregard' cego a faria não
    sair de nenhuma, e a segunda é pior: some sem erro nenhum."""
    seed_by_acct, seed_by_le = {}, {}
    for s in _MAPPING_DEFS['b3-accounts']['seed']:
        d = _acc_digits(s['ACCOUNT'])
        if d:
            seed_by_acct[d] = s['MESSAGING']
        seed_by_le.setdefault(_fcst_norm(s['LE']).strip(), s['MESSAGING'])

    out = []
    for r in rows:
        if not isinstance(r, dict):
            continue
        r = dict(r)
        tem_colunas = any(k in r for k in ('LE', 'SIMPLIFIED NAME', 'ACCOUNT TYPE'))
        tipo = _b3_account_type(r.get('ACCOUNT TYPE', ''))
        if not tipo and not tem_colunas:
            tipo = 'CLIENT 1'          # linha do cadastro antigo: era omnibus
        r['LE'] = str(r.get('LE', '') or '').strip()
        r['SIMPLIFIED NAME'] = str(r.get('SIMPLIFIED NAME', '') or '').strip()
        r['ACCOUNT'] = str(r.get('ACCOUNT', '') or '').strip()
        r['ACCOUNT TYPE'] = tipo
        r['REFDATA NAME'] = str(r.get('REFDATA NAME', '') or '').strip()
        if 'MESSAGING' not in r or not str(r.get('MESSAGING') or '').strip():
            r['MESSAGING'] = (seed_by_acct.get(_acc_digits(r['ACCOUNT']))
                              or seed_by_le.get(_fcst_norm(r['LE']).strip())
                              # Conta nossa que o seed não conhece: a visão dela
                              # gera, que é o que acontecia antes de a coluna
                              # existir (só a visão MGT era descartada).
                              or 'Consider')
        r['NOTES'] = str(r.get('NOTES', '') or '').strip()
        out.append(r)
    return out


def _api_links_upgrade(rows):
    """Traz para o formato com PRODUCT os arquivos gravados antes da coluna.

    A linha antiga de New Deals era uma só, com `product=NDF` no endereço e o
    produto sendo reescrito pela rotina. Ela vira a linha do NDF, e as dos outros
    produtos entram do seed — senão a instância que já abriu a tela ficaria com o
    FXO sem endereço cadastrado, que foi exatamente o que a mesa notou faltando.
    """
    for r in rows:
        if isinstance(r, dict) and 'PRODUCT' not in r:
            r['PRODUCT'] = 'NDF' if _use_key_py(r.get('USE')) == 'newdeals' else ''
    have = {(_use_key_py(r.get('USE')), str(r.get('PRODUCT') or '').strip().upper())
            for r in rows if isinstance(r, dict)}
    for s in _API_LINKS_SEED:
        if (_use_key_py(s['USE']), s['PRODUCT'].upper()) not in have:
            rows.append(dict(s))
    return rows


def _use_key_py(value):
    """Mesma normalização de `athena_api._use_key` ('New Deals' ≡ 'new_deals')."""
    return re.sub(r'[^a-z]', '', str(value or '').lower())


# Razão social de cada Legal Entity no Reference Data (ditada pela mesa). A busca
# é NORMALIZADA (`_pc_norm`: sem acento, sem pontuação, minúsculas), então
# 'BANCO J.P MORGAN S/A' e 'BANCO J.P MORGAN S.A' são o mesmo nome — o que está
# aqui é a grafia do Reference Data. MGT e ATACAMA ainda não têm linha no
# Reference Data: enquanto não tiverem, a perna interna delas cai no SPN
# cadastrado e, sem ele, em Missing Counterparty — que é o pedido certo, "vá
# cadastrar", em vez de uma contraparte inventada.
_LE_SPN_SEED = (
    {'LE': 'JPM',     'NAME': 'BANCO J.P MORGAN S.A',                        'SPN': '', 'NOTES': ''},
    {'LE': 'MGT',     'NAME': 'JPMORGAN CHASE BANK, N.A. - SAO PAULO BRANCH', 'SPN': '', 'NOTES': ''},
    {'LE': 'LAWTON',  'NAME': 'LAWTON MULTIMERCADO EXCLUSIVO',                'SPN': '', 'NOTES': ''},
    {'LE': 'ATACAMA', 'NAME': '',                                             'SPN': '', 'NOTES': ''},
)


def _le_spn_upgrade(rows):
    """Garante uma linha por Legal Entity e traz a razão social para os arquivos
    gravados antes da coluna NAME existir.

    Roda na LEITURA: a instância que já abriu a tela de mapping tem o arquivo em
    disco e nunca mais receberia o seed. Só preenche o que **não existe** — linha
    ausente ou coluna ausente. Quem apagou o nome pela tela fica com o nome
    apagado (a tela grava a chave vazia), senão o cadastro brigaria com o
    usuário a cada leitura.
    """
    seen = {}
    for r in rows:
        if not isinstance(r, dict):
            continue
        seen.setdefault(str(r.get('LE') or '').strip().upper(), r)
    for s in _LE_SPN_SEED:
        row = seen.get(s['LE'])
        if row is None:
            rows.append(dict(s))
        elif 'NAME' not in row and s['NAME']:
            row['NAME'] = s['NAME']
    return rows


# O cadastro de quem valida cada tipo mora no `manual_conf` — é regra da esteira,
# e é ELE quem a lê a cada linha do Monitor. Enquanto o seed e o `upgrade` viviam
# aqui, o `manual_conf._mapping_rows` lia o arquivo CRU: quem nunca abriu a tela
# de /mapping rodava com os nomes velhos, e o SWAP CORPORATE caía no DEFAULT_RULE
# (OTC + MO) — a regra errada, porque nele o FO também valida.
#
# Os apelidos ficam porque é por eles que o `_MAPPING_DEFS` abaixo se registra.
_MC_VALIDATION_SEED = _mc_mod.VALIDATION_SEED
_mc_validation_upgrade = _mc_mod.validation_upgrade

# O prazo de cada mesa mora no mesmo lugar, e pela mesma razão: quem o lê a cada
# linha do Monitor é o `manual_conf`, não esta tela.
_MC_SLA_SEED = _mc_mod.SLA_SEED
_mc_sla_upgrade = _mc_mod.sla_upgrade
_MC_STAGES = [_mc_mod.STAGE_OTC, _mc_mod.STAGE_MO, _mc_mod.STAGE_FO]


# Valor do campo "Tipo de Cotação" no arquivo de Opção. Era o literal '5' nos
# dois builders (servidor e navegador); vira o default da coluna QUOTE TYPE OPT.
_B3_QUOTE_OPT_DEFAULT = '5'


def _commodities_b3_quote_defaults(rows):
    """Preenche as três colunas de cotação com EXATAMENTE o que o código fazia.

    Chamada no seed e no upgrade. Usa `setdefault`: só escreve onde a chave nem
    existe — linha antiga, gravada antes das colunas. Coluna que a tela gravou
    vazia continua vazia (e o consumidor cai no mesmo default), senão o cadastro
    brigaria com quem apagou o valor de propósito.

    A coluna FIXED QUOTE foi APOSENTADA (§252): ela só escolhia entre F/340 e
    A/358 quando as colunas de cotação estavam vazias, e esses valores agora
    vivem nas próprias colunas. Num arquivo antigo o flag ainda é lido AQUI —
    para materializar o F/340 das linhas YES antes de o flag sumir — e então
    removido da linha.
    """
    for r in rows:
        if not isinstance(r, dict):
            continue
        fixed = str(r.pop('FIXED QUOTE', '') or '').strip().upper() == 'YES'
        r.setdefault('QUOTE TYPE NDF', 'F' if fixed else 'A')
        r.setdefault('QUOTE TYPE OPT', _B3_QUOTE_OPT_DEFAULT)
        r.setdefault('INFO SOURCE', '340' if fixed else '358')
    return rows


def _commodities_b3_upgrade(rows):
    """Migra as linhas gravadas antes da notação de padrão (§164).

    Antes, a coluna B3 CODE de uma linha PREFIX guardava só o prefixo ('XB',
    'C ') e o mês/ano era concatenado no código-fonte; o FCPO era SPECIAL com o
    código inteiro embutido em três lugares diferentes. Agora o padrão inteiro
    mora na coluna.

    Roda na LEITURA, então uma instância que já tem o arquivo em disco passa a
    ver o formato novo sem script de migração. Idempotente: uma linha que já
    tenha aspas não é tocada.
    """
    for r in rows:
        if not isinstance(r, dict):
            continue
        typ = str(r.get('TYPE') or '').strip().upper()
        mkt = str(r.get('MARKET') or '').strip().upper()
        code = str(r.get('B3 CODE') or '')

        # FCPO deixou de ser SPECIAL: o padrão expressa o código inteiro.
        if mkt == 'FCPO_BURSA_MYR' and typ == 'SPECIAL':
            r['TYPE'] = 'PREFIX'
            r['B3 CODE'] = 'KO"MY"BNMK'
            continue

        # BRT_IPE: os dois códigos eram literais no código-fonte e a linha do
        # cadastro vinha VAZIA. Preencher aqui — e não só no seed — é o que faz a
        # regra valer na instância que já tem o arquivo em disco; sem isto a
        # asiática ficaria sem código nenhum, que é pior do que o CO1-2 fixo que
        # havia antes. Só preenche o que está vazio: quem editar pela tela manda.
        if mkt == 'BRT_IPE' and typ == 'SPECIAL':
            if not code.strip():
                r['B3 CODE'] = 'CO"MY"'
            if not str(r.get('B3 CODE FAR') or '').strip():
                r['B3 CODE FAR'] = 'CO1-2'
            continue

        if 'PREFIX' not in typ or '"' in code or not code:
            continue
        # Formato antigo: o mês/ano vinha sempre no fim, e o espaço era literal.
        r['B3 CODE'] = code.replace(' ', '_') + '"MY"'

    # TRADE TYPE (§251): linha sem a coluna vale para os dois tipos (BOTH) —
    # exceto as que sempre foram regra de UM tipo só: a SPECIAL do BRT_IPE é da
    # ASIÁTICA (a vanilla morava no ramo de código e ganha a própria linha
    # PREFIX abaixo) e o PREFIX do WTI vira da VANILLA (a asiática ganha a
    # linha FIXED CL1 — §252). Preencher só o que está vazio: quem editou pela
    # tela manda.
    arquivo_antigo = any(isinstance(r, dict) and not str(r.get('TRADE TYPE') or '').strip()
                         for r in rows)
    brt_special_migrada = wti_prefix_migrado = False
    tem_brt_vanilla = tem_wti_asian = False
    for r in rows:
        if not isinstance(r, dict):
            continue
        mkt = str(r.get('MARKET') or '').strip().upper()
        typ = str(r.get('TYPE') or '').strip().upper()
        tt = str(r.get('TRADE TYPE') or '').strip().upper()
        if not tt:
            if mkt == 'BRT_IPE' and typ == 'SPECIAL':
                r['TRADE TYPE'] = 'ASIAN'
                brt_special_migrada = True
            elif mkt == 'WTI_NYMEX' and 'PREFIX' in typ:
                r['TRADE TYPE'] = 'VANILLA'
                wti_prefix_migrado = True
            else:
                r['TRADE TYPE'] = 'BOTH'
        tt_final = str(r.get('TRADE TYPE') or '').strip().upper()
        if mkt == 'BRT_IPE' and tt_final in ('VANILLA', 'BOTH'):
            tem_brt_vanilla = True
        if mkt == 'WTI_NYMEX' and tt_final in ('ASIAN', 'BOTH'):
            tem_wti_asian = True
    # As linhas do tipo que faltou só entram quando a migração acabou de
    # acontecer (arquivo anterior à coluna) e nada cobre aquele tipo — num
    # arquivo já migrado, a ausência é decisão de quem editou, não formato
    # antigo, e não volta.
    if brt_special_migrada and not tem_brt_vanilla:
        rows.append({'TYPE': 'PREFIX', 'MARKET': 'BRT_IPE', 'TRADE TYPE': 'VANILLA',
                     'B3 CODE': 'CO"MY"', 'B3 CODE FAR': '',
                     'HOLIDAY CALENDAR': 'IPE'})
    if wti_prefix_migrado and not tem_wti_asian:
        rows.append({'TYPE': 'FIXED', 'MARKET': 'WTI_NYMEX', 'TRADE TYPE': 'ASIAN',
                     'B3 CODE': 'CL1', 'B3 CODE FAR': '',
                     'HOLIDAY CALENDAR': 'NYMEX'})
    # Os PTS* saem junto com a migração (§252): eram linhas sem MARKET que só
    # existiam para carregar o flag FIXED QUOTE, aposentado. Só no arquivo
    # ANTIGO — quem recadastrar um PTS depois disso está mandando.
    if arquivo_antigo:
        _pts = ('PTS005', 'PTS002', 'PTS006', 'PTS003')
        rows[:] = [r for r in rows
                   if not (isinstance(r, dict)
                           and not str(r.get('MARKET') or '').strip()
                           and str(r.get('B3 CODE') or '').strip().upper() in _pts)]
    return _commodities_b3_quote_defaults(rows)


_MAPPING_DEFS = {
    # Aba movida do Index B3 — edita o MESMO BaseMoeda.json que os previews de
    # NDF já leem para o código de moeda (CODIGO DE CADASTRO). Absorveu o antigo
    # mapping currency-codes: ATHENA CODE (código da moeda no getTrades → o
    # SIMBOLO é o ISO), WEAK (cotação invertida vs BRL) e INV DECIMALS (casas do
    # 1/rate no arquivo Conecta) viraram colunas daqui.
    'currency-base': {
        'label': 'Currency Base',
        'file': os.path.join(_MAPPINGS_DIR, 'BaseMoeda.json'),
        'columns': [
            {'key': 'TIPO', 'label': 'Type'},
            {'key': 'TIPO DE INDICADOR', 'label': 'Indicator Type'},
            {'key': 'CAMPO', 'label': 'Field'},
            {'key': 'DESCRICAO DO CAMPO', 'label': 'Field Description'},
            {'key': 'CODIGO DE CADASTRO', 'label': 'Registration Code'},
            {'key': 'SIMBOLO', 'label': 'Currency Code'},
            {'key': 'TIPO COTACAO', 'label': 'Quotation Type'},
            {'key': 'CASAS DECIMAIS', 'label': 'Decimal Places'},
            {'key': 'ATHENA CODE', 'label': 'Athena Code'},
            {'key': 'WEAK', 'label': 'Weak Ccy', 'type': 'select', 'options': ['', 'YES']},
            {'key': 'INV DECIMALS', 'label': 'Inverse Decimals'},
        ],
        'seed': [],
    },
    # Filtro interbook do import da API NDF. Cada linha é um par livre: escolha
    # QUAIS campos da API formam a dupla (Field A/Field B) e os valores que
    # caracterizam a perna interna. BOTH WAYS = YES também casa com os valores
    # trocados entre os dois campos — é o caso de End Counterparty × Trading
    # Book, em que a mesma operação chega uma vez por ponta.
    'interbook-ndf': {
        'label': 'Interbook API (NDF)',
        'columns': [
            {'key': 'FIELD A', 'label': 'Field A', 'type': 'select', 'options': _MAP_INTERBOOK_FIELDS},
            {'key': 'VALUE A', 'label': 'Value A'},
            {'key': 'FIELD B', 'label': 'Field B', 'type': 'select', 'options': _MAP_INTERBOOK_FIELDS},
            {'key': 'VALUE B', 'label': 'Value B'},
            {'key': 'BOTH WAYS', 'label': 'Both Ways', 'type': 'select', 'options': ['', 'YES']},
        ],
        'upgrade': _interbook_upgrade,
        'seed': (
            [{'FIELD A': 'OTHER BOOK', 'VALUE A': b, 'FIELD B': 'SETTLEMENT LOCATION',
              'VALUE B': l, 'BOTH WAYS': ''} for b, l in (
                ('GN ON BRL',             'BRAZIL'),
                ('JB ON BRL',             'BRAZIL'),
                ('JB LAWTON BRL',         'LAWTON'),
                ('LM-FWDECOMBRR FXC',     'BRAZIL'),
                ('BR ON - LN LAWTON NDF', 'LAWTON'),
                ('CLIENT FX NDF LAWTON',  'LAWTON'),
                ('DERIV NDF BJPM FXC',    'BRAZIL'),
                ('GN NDF BJPM',           'BRAZIL'),
                ('LM-FXECOMBRR FXC',      'BRAZIL'),
                ('JB NDF BJPM',           'BRAZIL'),
            )] +
            [{'FIELD A': 'END COUNTERPARTY', 'VALUE A': 'DERIV NDF BJPM FXC',
              'FIELD B': 'TRADING BOOK', 'VALUE B': 'GN NDF BJPM', 'BOTH WAYS': 'YES'}]
        ),
    },
    # Market da Athena → Código do Ativo Subjacente B3 (commodities).
    #
    #   FIXED   = código fechado, sai como está.
    #   PREFIX  = PADRÃO: texto fixo + mês/ano do contrato + texto fixo.
    #   SPECIAL = o código é calculado no código-fonte porque depende de algo que
    #             não é de-para. Sobrou só o BRT_IPE (vanilla × asian).
    #
    # Notação da coluna B3 CODE quando o tipo é PREFIX (ver `split_b3_pattern` em
    # otc_boxparse.py, espelhada nos dois JS):
    #
    #   "MY"  → onde entram a letra do mês e o último dígito do ano. Vai entre
    #           ASPAS porque um código pode ter M e Y como texto fixo — sem a
    #           marca não daria para saber qual é qual.
    #   _     → um ESPAÇO no código emitido. O milho na B3 é 'C ' COM o espaço, e
    #           espaço no fim de um campo é invisível na tela e some num trim
    #           distraído; o sublinhado torna-o visível.
    #
    #   XB"MY"     → XBZ7        C_"MY" → 'C Z7'        KO"MY"BNMK → KOZ7BNMK
    #
    # HOLIDAY CALENDAR = calendário de feriados do market.
    #
    # TIPO DE COTAÇÃO (§177): eram literais no código e agora são cadastro —
    # a coluna guarda o **código do layout**, o que vai no arquivo. São duas
    # colunas porque os DOIS LAYOUTS TÊM DOMÍNIOS DIFERENTES: o de Termo usa
    # letra ('F'/'A'), o de Opção usa número ('5'), e a mesma commodity é
    # negociada nos dois. INFO SOURCE segue a mesma ideia (340/358) e existe só
    # no arquivo de Termo. Coluna vazia = o default histórico (A / 5 / 358).
    #
    # O flag FIXED QUOTE foi APOSENTADO (§252): ele só escolhia F/340 vs A/358
    # quando as colunas de cotação estavam vazias, e esses valores estão
    # materializados nas colunas desde o §177. O `_commodities_b3_quote_defaults`
    # ainda o lê de arquivo antigo (para materializar antes de removê-lo da
    # linha). Os PTS* saíram junto — linhas sem MARKET que só carregavam o flag.
    'commodities-b3': {
        'label': 'Commodities × B3 Code',
        'columns': [
            {'key': 'TYPE', 'label': 'Type', 'type': 'select', 'options': ['FIXED', 'PREFIX', 'SPECIAL']},
            {'key': 'MARKET', 'label': 'Market (Athena)'},
            # A linha vale para qual tipo de trade: BOTH (ou em branco) para os
            # dois, VANILLA/ASIAN restringe. Permite um market ter códigos
            # diferentes por tipo — o BRT_IPE é o caso: a linha SPECIAL
            # (near/far) é só da asiática; a vanilla tem linha PREFIX CO"MY".
            {'key': 'TRADE TYPE', 'label': 'Trade Type (blank = BOTH)', 'type': 'select',
             'options': ['BOTH', 'VANILLA', 'ASIAN']},
            {'key': 'B3 CODE', 'label': 'B3 Code / Prefix'},
            {'key': 'B3 CODE FAR', 'label': 'B3 Code — far contract (SPECIAL only)'},
            {'key': 'HOLIDAY CALENDAR', 'label': 'Holiday Calendar'},
            {'key': 'QUOTE TYPE NDF', 'label': 'Tipo de Cotação — NDF (blank = A)'},
            {'key': 'QUOTE TYPE OPT', 'label': 'Tipo de Cotação — Opção (blank = 5)'},
            {'key': 'INFO SOURCE', 'label': 'Fonte de Informação — NDF (blank = 358)'},
        ],
        'upgrade': _commodities_b3_upgrade,
        'seed': _commodities_b3_quote_defaults(
            [{'TYPE': 'FIXED', 'MARKET': m, 'TRADE TYPE': 'BOTH', 'B3 CODE': c, 'HOLIDAY CALENDAR': h, 'FIXED QUOTE': q} for m, c, h, q in (
                ('MPB_LME', 'LOPBDY', 'LME', ''), ('MCU_LME', 'LOCADY', 'LME', ''),
                ('MAL_LME', 'LOAHDY', 'LME', ''), ('MZN_LME', 'LOZSDY', 'LME', ''),
                ('MSN_LME', 'LOSNDY', 'LME', ''), ('MNI_LME', 'LONIDY', 'LME', ''),
                ('FO_0.5%_ROT_BRG_FOB', 'NAEB0011', 'PLATTS-EUROPE', 'YES'),
                ('FO_0.5%_SING_FOB', 'NACX0005', 'PLATTS-ASIA', 'YES'),
                ('MAL_MW_PREMIUM', 'PMMUAKE0', 'LME', ''),
                ('BRT_DTD', 'PCRUDTB1', 'PLATTS-EUROPE', ''),
                ('NG_NYMEX', 'NG1', 'NYMEX', ''), ('MFE_TSI', 'PFATIOCH', 'PLATTS-ASIA', ''),
                ('COAL_HCC_FOB_AUS_TSI', 'PMTCLAUS', 'PLATTS-ASIA', 'YES'),
            )] +
            [{'TYPE': 'PREFIX', 'MARKET': m, 'TRADE TYPE': 'BOTH', 'B3 CODE': c, 'HOLIDAY CALENDAR': h, 'FIXED QUOTE': ''} for m, c, h in (
                ('HU_RBOB_NYMEX', 'XB"MY"', 'NYMEX'), ('HO_NYMEX', 'HO"MY"', 'NYMEX'),
                ('SB_ICE', 'SB"MY"', 'ICEAGS'), ('C_CBOT', 'C_"MY"', 'CBY_AGS'),
                ('S_CBOT', 'S_"MY"', 'CBY_AGS'), ('BO_CBOT', 'BO"MY"', 'CBY_AGS'),
                ('CC_ICE', 'CC"MY"', 'ICEAGS'), ('W_CBOT', 'W_"MY"', 'CBY_AGS'),
                ('SM_CBOT', 'SM"MY"', 'CBY_AGS'), ('CT_ICE', 'CT"MY"', 'ICEAGS'),
                ('KC_ICE', 'KC"MY"', 'ICEAGS'),
                # Era SPECIAL com o código no código-fonte — e as duas cópias JS
                # discordavam ('KOZ7BNMK' no otc-fileupload, '.KOZ7BNMK F' no
                # deals-processing-table). Agora é um padrão só, aqui. §164
                ('FCPO_BURSA_MYR', 'KO"MY"BNMK', 'BURSA'),
            )] +
            # WTI também é um código por tipo (§252): a vanilla segue o padrão
            # de contrato (WTI"MY") e a asiática usa o contínuo CL1, literal.
            [{'TYPE': 'PREFIX', 'MARKET': 'WTI_NYMEX', 'TRADE TYPE': 'VANILLA',
              'B3 CODE': 'WTI"MY"', 'HOLIDAY CALENDAR': 'NYMEX', 'FIXED QUOTE': ''},
             {'TYPE': 'FIXED', 'MARKET': 'WTI_NYMEX', 'TRADE TYPE': 'ASIAN',
              'B3 CODE': 'CL1', 'HOLIDAY CALENDAR': 'NYMEX', 'FIXED QUOTE': ''}] +
            # BRT_IPE tem DUAS linhas, separadas pelo Trade Type (§251):
            #   · SPECIAL, só ASIAN — near (CO"MY") quando o contrato é o mês
            #     seguinte à liquidação, far (CO1-2) a partir de dois meses (§212);
            #   · PREFIX, só VANILLA — CO"MY" padrão, como qualquer prefixo.
            # O resultado é o mesmo de quando a lógica vanilla morava no ramo
            # SPECIAL, mas agora cada tipo tem a SUA linha cadastrável.
            [{'TYPE': 'SPECIAL', 'MARKET': 'BRT_IPE', 'TRADE TYPE': 'ASIAN', 'B3 CODE': 'CO"MY"',
              'B3 CODE FAR': 'CO1-2', 'HOLIDAY CALENDAR': 'IPE', 'FIXED QUOTE': ''},
             {'TYPE': 'PREFIX', 'MARKET': 'BRT_IPE', 'TRADE TYPE': 'VANILLA', 'B3 CODE': 'CO"MY"',
              'HOLIDAY CALENDAR': 'IPE', 'FIXED QUOTE': ''}]
            # Os PTS* saíram (§252): eram linhas sem MARKET que só existiam para
            # carregar o flag FIXED QUOTE, aposentado junto.
        ),
    },
    # Publisher (feeder) da Athena → códigos B3 do TER das páginas Other
    # Publisher e FWD Start.
    #
    # COMO A LINHA É ESCOLHIDA (_ndf_publisher_row):
    #   1. nome exato — PUBLISHER igual ao publisher inteiro;
    #   2. token — só para linhas COM Match Tokens, porque a Athena manda o
    #      publisher composto ('PTAX|USB|WMR|4' → REUTERS - WMR).
    # Deixar Match Tokens EM BRANCO faz a linha casar SÓ pelo texto completo,
    # sem variação — é o que permite 'PTAX' e 'PTAX|BRR|PTAX' coexistirem como
    # cadastros independentes.
    #
    # FONTE INFO é o campo "Fonte de Informação" do arquivo (PTAX puro = 0,
    # demais feeders = 1) e, no FWD Start, define o Boletim (0 → 3, senão 1).
    #
    # NOTES não é só documentação: **NOTES = BACEN manda a operação para a
    # página Vanilla**; qualquer outro valor (e publisher sem linha) vai para
    # Other Publisher. Ver `_ndf_publisher_is_bacen` (§166).
    #
    # Publisher sem linha: Fonte de Informação 1, os dois códigos de consulta em
    # branco e roteamento para Other Publisher.
    'publisher-ndf': {
        'label': 'Publisher × B3 (NDF)',
        'columns': [
            {'key': 'PUBLISHER', 'label': 'Publisher (Athena)'},
            {'key': 'TOKENS', 'label': 'Match Tokens (blank = exact match only)'},
            {'key': 'FONTE INFO', 'label': 'Fonte de Informação', 'type': 'select', 'options': ['1', '0']},
            {'key': 'FONTE CONSULTA', 'label': 'Fonte de Consulta'},
            {'key': 'TELA CONSULTA', 'label': 'Tela ou Função de Consulta'},
            {'key': 'NOTES', 'label': 'Notes (BACEN → Vanilla)'},
        ],
        'seed': [
            {'PUBLISHER': 'PTAX', 'TOKENS': '', 'FONTE INFO': '0',
             'FONTE CONSULTA': '', 'TELA CONSULTA': '', 'NOTES': 'BACEN'},
            {'PUBLISHER': 'BFIX 4PM LONDON', 'TOKENS': 'BFIX', 'FONTE INFO': '1',
             'FONTE CONSULTA': '2', 'TELA CONSULTA': '14399', 'NOTES': 'BLOOMBERG'},
            {'PUBLISHER': 'OBSERVADO/BCENTRAL CL', 'TOKENS': 'BCENTRAL, OBSERVADO',
             'FONTE INFO': '1', 'FONTE CONSULTA': '5', 'TELA CONSULTA': '11703', 'NOTES': 'OUTROS'},
            {'PUBLISHER': 'PEN SBSP/BCRP', 'TOKENS': 'SBSP, BCRP', 'FONTE INFO': '1',
             'FONTE CONSULTA': '5', 'TELA CONSULTA': '11683', 'NOTES': 'OUTROS'},
            {'PUBLISHER': 'REUTERS - WMR', 'TOKENS': 'WMR', 'FONTE INFO': '1',
             'FONTE CONSULTA': '0', 'TELA CONSULTA': '247', 'NOTES': 'REUTERS'},
            {'PUBLISHER': 'TRM COP', 'TOKENS': 'TRM', 'FONTE INFO': '1',
             'FONTE CONSULTA': '5', 'TELA CONSULTA': '11682', 'NOTES': 'OUTROS'},
        ],
    },
    # Legal Entity das páginas de NDF × o que a API manda. Duas formas de casar,
    # nesta ordem: ACCRONYM (o End Counterparty exato, para o código que não
    # segue padrão nenhum) e SETTLEMENT LOCATION (era o de-para fixo
    # BRAZIL→JPM / JPMCBB→MGT no código). Linha só com LE + ACCRONYM é o uso
    # normal: 'autofill' faz a tela copiar a Settlement Location da linha que já
    # existe para aquela LE quando você escolhe a LE no dropdown.
    # O accronym sufixado por entidade ('CMBB-LAW') NÃO precisa de linha aqui nem
    # no Reference Data: o lookup de contraparte tenta o código exato e depois o
    # accronym sem o último trecho depois do hífen, então uma linha de cadastro
    # atende Banco, Lawton e MGT.
    'le-accronym': {
        'label': 'Legal Entity × Accronym',
        'columns': [
            {'key': 'LE', 'label': 'Legal Entity', 'type': 'select',
             'options': _MAP_LE_OPTIONS, 'autofill': 'SETTLEMENT LOCATION'},
            {'key': 'ACCRONYM', 'label': 'Accronym'},
            {'key': 'SETTLEMENT LOCATION', 'label': 'Settlement Location'},
        ],
        'seed': [
            {'LE': 'JPM', 'ACCRONYM': '', 'SETTLEMENT LOCATION': 'BRAZIL'},
            {'LE': 'MGT', 'ACCRONYM': '', 'SETTLEMENT LOCATION': 'JPMCBB'},
            {'LE': 'LAWTON', 'ACCRONYM': '', 'SETTLEMENT LOCATION': 'LAWTON'},
        ],
    },
    # Identidade de cada Legal Entity: a RAZÃO SOCIAL como está no Reference Data
    # e o SPN da NOSSA ponta — coisa diferente do SPN da contraparte, que vem do
    # Reference Data pelo accronym (§147/§148).
    #
    # É por este cadastro que uma **perna interna** ganha SPN, Client e Tax ID:
    # quando o End Counterparty da API é nome de book (ex. 'LM-FWDECOMBRR FXC'),
    # o mapping Legal Entity × Accronym diz de qual entidade ele é, e a RAZÃO
    # SOCIAL daqui acha a linha da entidade no Reference Data. Antes só se tentava
    # o accronym, e como book não está no Reference Data a linha ficava com os
    # três campos vazios (§174).
    #
    # O SPN continua servindo de última tentativa: sem razão social e sem
    # accronym, ele preenche ao menos a coluna SPN. As linhas nascem com a razão
    # social ditada pela mesa e SEM SPN — SPN inventado sairia num arquivo para a
    # B3 como se fosse cadastro.
    'le-spn': {
        'label': 'Legal Entity × SPN',
        'columns': [
            {'key': 'LE', 'label': 'Legal Entity', 'type': 'select', 'options': _MAP_LE_SPN_OPTIONS},
            {'key': 'NAME', 'label': 'Reference Data Name'},
            {'key': 'SPN', 'label': 'SPN'},
            {'key': 'NOTES', 'label': 'Notes'},
        ],
        'upgrade': _le_spn_upgrade,
        'seed': list(_LE_SPN_SEED),
    },
    # Bancos oferecidos no editor de contraparte (Reference Data, duplo clique →
    # contas BANKING/PAY-REC). ID = código COMPE de 3 dígitos; ISPB e TAX ID
    # (CNPJ mascarado) vêm do cadastro público do Bacen. O front monta o rótulo
    # 'NAME - ID' que a página já usa.
    'bank-name': {
        'label': 'Bank Name',
        'columns': [
            {'key': 'ID', 'label': 'ID (3-digit code)'},
            {'key': 'NAME', 'label': 'Bank Name'},
            {'key': 'ISPB', 'label': 'ISPB'},
            {'key': 'TAX ID', 'label': 'Tax ID'},
        ],
        'seed': [
            {'ID': '001', 'NAME': 'BANCO DO BRASIL S/A',        'ISPB': '00000000', 'TAX ID': '00.000.000/0001-91'},
            {'ID': '033', 'NAME': 'BANCO SANTANDER S/A',        'ISPB': '90400888', 'TAX ID': '90.400.888/0001-42'},
            {'ID': '217', 'NAME': 'BANCO JOHN DEERE S/A',       'ISPB': '91884981', 'TAX ID': '91.884.981/0001-32'},
            {'ID': '237', 'NAME': 'BANCO BRADESCO S/A',         'ISPB': '60746948', 'TAX ID': '60.746.948/0001-12'},
            {'ID': '341', 'NAME': 'BANCO ITAU S/A',             'ISPB': '60701190', 'TAX ID': '60.701.190/0001-04'},
            {'ID': '376', 'NAME': 'BANCO JP MORGAN S/A',        'ISPB': '33172537', 'TAX ID': '33.172.537/0001-98'},
            {'ID': '745', 'NAME': 'BANCO CITIBANK S/A',         'ISPB': '33479023', 'TAX ID': '33.479.023/0001-80'},
            {'ID': '755', 'NAME': 'BOFA MERRILL LYNCH BM S/A',  'ISPB': '62073200', 'TAX ID': '62.073.200/0001-21'},
        ],
    },
    # Moeda Base × Taxa de Conversão das confirmações de Opção de Câmbio
    # (Anexo I do documento Asian, colunas "Taxa de Conversão" e "Tipo de Taxa
    # de Conversão"). O Anexo II define uma taxa por moeda — "USD PTAX" é a de
    # venda do dólar pelo Bacen, "ARS MAE" a média do mercado eletrônico
    # argentino, e há moeda com mais de uma taxa possível (ARS MAE × ARS WMCO),
    # o que faz disto cadastro e não constante. Nasce só com o USD, que é a
    # linha do documento-modelo; moeda sem linha cai como aviso no painel, em
    # vez de sair em branco na confirmação sem ninguém ver.
    'fxo-conv-rate': {
        'label': 'FXO Conversion Rate',
        'columns': [
            {'key': 'MOEDA BASE', 'label': 'Base Currency'},
            {'key': 'TAXA DE CONVERSAO', 'label': 'Conversion Rate'},
            {'key': 'TIPO', 'label': 'Rate Type', 'type': 'select',
             'options': ['', 'Venda', 'Compra']},
        ],
        'seed': [
            {'MOEDA BASE': 'USD', 'TAXA DE CONVERSAO': 'USD PTAX', 'TIPO': 'Venda'},
        ],
    },
    # Contrapartes cujo aviso de liquidação de NDF de Moeda leva a Ficha de
    # Liquidação TAMBÉM em PDF anexo (mesmo conteúdo do cartão branco do e-mail).
    # Era a tupla _NDF_PDF_COUNTERPARTIES em otc_emails.py, herdada da macro
    # legada (CommodiXchange): cada cliente novo exigia mexer no código.
    #
    # O nome casa pelo NORMALIZADO (sem acento, caixa alta, espaços colapsados,
    # travessão vira hífen — `_ndf_pdf_norm`), então diferença de grafia entre o
    # que se digita aqui e o Reference Data não quebra o match. O que precisa
    # bater é a razão social, não o accronym.
    #
    # Cadastro VAZIO significa "ninguém leva PDF" e é respeitado. O consumidor só
    # volta para a lista histórica quando o arquivo não existe (instância que
    # ainda não abriu a tela) ou está ilegível — ver `_ndf_pdf_set`.
    'ndf-pdf-cpty': {
        'label': 'Settlement PDF (NDF Advice)',
        'columns': [
            {'key': 'COUNTERPARTY', 'label': 'Counterparty (RefData name)'},
            {'key': 'NOTES', 'label': 'Notes'},
        ],
        'seed': [
            {'COUNTERPARTY': 'ABB AUTOMACAO LTDA', 'NOTES': ''},
            {'COUNTERPARTY': 'ABB ELETRIFICACAO LTDA - FILIAL 0003', 'NOTES': ''},
            {'COUNTERPARTY': 'ABB ELETRIFICACAO LTDA', 'NOTES': ''},
            {'COUNTERPARTY': 'HITACHI ENERGY BRASIL LTDA', 'NOTES': ''},
            {'COUNTERPARTY': 'PHINIA DO BRASIL PRODUTOS AUTOMOTIVOS LTDA', 'NOTES': ''},
            {'COUNTERPARTY': 'VEOLIA WATER TECHNOLOGIES AND SOLUTIONS BRASIL '
                             'TRATAMENTO DE AGUAS LTDA', 'NOTES': ''},
        ],
    },
    # Curvas de swap Athena × B3 — cadastro pronto para os fluxos de swap; nasce
    # vazio porque não havia de-para hardcoded no código.
    'swap-curves': {
        'label': 'Swap Curves (Athena × B3)',
        'columns': [
            {'key': 'ATHENA CURVE', 'label': 'Athena Curve'},
            {'key': 'B3 CURVE', 'label': 'B3 Curve / Code'},
        ],
        'seed': [],
    },
    # Lista de arquivos que a rotina Save CETIP Files (Control Panel) considera.
    # TYPE é a CHAVE que liga a linha ao comportamento no código (_CETIP_BEHAVIOUR:
    # exportação do JSON, atualização do VCP, anexos de e-mail) — renomear o TYPE
    # desliga esse comportamento e o arquivo passa a ser só copiado/renomeado.
    # SOURCE e DEST usam YYMMDD para marcar onde a data fica; a data que entra no
    # nome salvo vem do nome do arquivo de origem, e a data do card escolhe a
    # pasta do dia. EXTRA DEST = cópia extra numa pasta de rede (o Alteryx tinha
    # uma segunda saída plana para .OPC e .TER).
    'cetip-files': {
        'label': 'CETIP Files (Save Routine)',
        'columns': [
            {'key': 'TYPE', 'label': 'File Type'},
            {'key': 'SOURCE', 'label': 'Source Name (YYMMDD = date)'},
            {'key': 'DEST', 'label': 'Saved As (YYMMDD = date)'},
            {'key': 'EXTRA DEST', 'label': 'Extra Copy Folder'},
        ],
        'seed': _CETIP_FILES_SEED,
    },
    # Endereços da API Athena, um por USO. Era a constante BASE_URL +
    # TRADES_ENDPOINT em athena_api.py: trocar o endpoint (versão nova, migração
    # de host, apontar para UAT) exigia mexer no código e reiniciar o servidor.
    #
    # YYYYMMDD marca onde entra a Data de Referência. Ele NÃO é o único caminho:
    # `product` e `date` do query string são sempre reescritos com o que a rotina
    # pediu (`build_url`), porque quem sabe qual produto está sendo puxado é o
    # código — um `product=NDF` esquecido na linha traria NDF para a página de
    # FXO sem ninguém ver. O placeholder serve para a data que fica no CAMINHO.
    #
    # A linha de Unwinds nasce SEM URL de propósito: não existe rotina de unwind
    # ainda, e semear um endereço não conferido faria a primeira rotina a nascer
    # chamar um endpoint inventado. Sem URL, o consumidor falha dizendo que falta
    # cadastro; o New Deals, esse sim, cai no endereço histórico.
    'api-links': {
        'label': 'API Links',
        'columns': [
            {'key': 'USE', 'label': 'Usage', 'type': 'select',
             'options': ['New Deals', 'Unwinds', 'Recon FXO']},
            {'key': 'PRODUCT', 'label': 'Product (blank = any)', 'type': 'select',
             'options': _MAP_API_PRODUCTS},
            {'key': 'URL', 'label': 'URL (YYYYMMDD = reference date)'},
            {'key': 'NOTES', 'label': 'Notes'},
        ],
        'upgrade': _api_links_upgrade,
        'seed': list(_API_LINKS_SEED),
    },
    # Quem valida a confirmação de cada produto, por Produto × LOB. Era regra de
    # boca: termo e opção de commodities, FXO e NDF FWD Start passam por OTC e
    # MO; swap e opção de EDG passam por OTC, MO e FO.
    #
    # LOB em branco é CORINGA do produto — a maioria valida igual em toda LOB, e
    # exigir uma linha por LOB faria a tela pedir cadastro a cada LOB nova.
    #
    # MO e FO correm em PARALELO, não em fila: as duas validam depois do OTC, e a
    # confirmação só fecha quando as duas pedidas responderem.
    'manual-conf-validation': {
        'label': 'Manual Confirmations — Validation Trail',
        'upgrade': _mc_validation_upgrade,
        'columns': [
            {'key': 'PRODUCT', 'label': 'Produto', 'type': 'select',
             'options': list(_CONFIRMATION_TYPES)},
            {'key': 'LOB', 'label': 'LOB (blank = any)'},
            {'key': 'OTC', 'label': 'OTC', 'type': 'select', 'options': _MAP_MC_VALIDATION},
            {'key': 'MO', 'label': 'MO', 'type': 'select', 'options': _MAP_MC_VALIDATION},
            {'key': 'FO', 'label': 'FO', 'type': 'select', 'options': _MAP_MC_VALIDATION},
            {'key': 'NOTES', 'label': 'Notes'},
        ],
        # Os produtos são os TIPOS DE CONFIRMAÇÃO (`_CONFIRMATION_TYPES`), os
        # mesmos do Confirmation Type do Electronic Inventory — uma linha por
        # tipo, na ordem da lista. As três páginas de NDF têm cada uma a sua:
        # elas geram documentos diferentes, e um 'NDF' genérico obrigava a
        # adivinhar qual delas produziu a confirmação.
        #
        # A opção de EDG é a linha FXO × LOB EDG — o que era 'OPTION EDG', um
        # produto que nenhuma linha do banco jamais teve.
        'seed': [dict(s) for s in _MC_VALIDATION_SEED],
    },
    # O PRAZO de cada mesa, em dias ÚTEIS contados da DATA DA OPERAÇÃO — é ele
    # que acende o verde/amarelo/vermelho do Confirmations Monitor e que torna a
    # justificativa obrigatória quando a validação sai fora do prazo.
    #
    # Duas coisas que a tabela não diz sozinha: o prazo corre do TRADE, não da
    # data em que a confirmação foi gerada (gerar o documento com atraso não
    # compra tempo novo), e os prazos NÃO se somam — MO e FO correm em paralelo
    # depois do OTC, os dois contados do mesmo trade date. Os dias são úteis pelo
    # calendário ANBIMA, senão a confirmação de sexta-feira nasceria atrasada na
    # segunda sem ninguém ter deixado de trabalhar.
    'manual-conf-sla': {
        'label': 'Manual Confirmations — SLA',
        'upgrade': _mc_sla_upgrade,
        'columns': [
            {'key': 'STAGE', 'label': 'Mesa', 'type': 'select', 'options': _MC_STAGES},
            {'key': 'BIZDAYS', 'label': 'Prazo (dias úteis do trade date)'},
            {'key': 'NOTES', 'label': 'Notes'},
        ],
        'seed': [dict(s) for s in _MC_SLA_SEED],
    },
    # Reconciliação de FXO: a perna INTERNA. Ela chega à Athena com o
    # nome da mesa (o book), enquanto a CETIP registra o código do fundo — sem
    # tradução, toda operação intragrupo sai NOK.
    #
    # INVERT DIRECTION separa os dois casos, e a diferença importa:
    #   'No'  → só o nome muda, e a troca vale SEMPRE;
    #   'Yes' → a perna vem espelhada (o Buy/Sell também está invertido), e a
    #           troca vale só quando Ctpty e JPM Dir estão os DOIS NOK. Aplicar
    #           essa sempre inverteria a direção de operações que estavam certas.
    'fxo-internal-cpty': {
        'label': 'FXO Recon — Internal Counterparty',
        'upgrade': _fxo_internal_cpty_upgrade,
        'columns': [
            {'key': 'ATHENA NAME', 'label': 'Nome na Athena'},
            {'key': 'CETIP CODE', 'label': 'Código na CETIP'},
            {'key': 'INVERT DIRECTION', 'label': 'Perna espelhada (inverte Buy/Sell)',
             'type': 'select', 'options': ['No', 'Yes']},
            # `Disregard` tira do batimento as linhas da Athena com aquele
            # CounterpartyName, ANTES do match. É para a perna interna que não
            # tem par na CETIP: mantida, ela vira `Unmatched Athena` todo dia.
            {'key': 'USE', 'label': 'No batimento', 'type': 'select',
             'options': ['Consider', 'Disregard']},
            {'key': 'NOTES', 'label': 'Notes'},
        ],
        'seed': [
            {'ATHENA NAME': 'LAWTON MULTIMERCADO EXCLUSIVO FUNDO DE INVESTIMENTO - LABAY LAWTON',
             'CETIP CODE': 'INTRAGLAWTONFDO', 'INVERT DIRECTION': 'No',
             'USE': 'Consider', 'NOTES': 'Nome do fundo por extenso na Athena'},
            {'ATHENA NAME': 'BCO J.P. MORGAN S.A. 2768 - GEM BR - EXPENSES & CASH MGMT',
             'CETIP CODE': 'INTRAGLAWTONFDO', 'INVERT DIRECTION': 'Yes',
             'USE': 'Disregard',
             'NOTES': 'Conta interna GEM — sai do batimento, não tem par na CETIP'},
        ],
    },
    # A MESMA exclusão do cadastro acima, por OUTRO identificador que não o nome
    # da contraparte: a perna interbook é a mesa contra a mesa, não tem registro
    # na CETIP e viraria `Unmatched Athena` todo dia.
    #
    # A linha é uma CONJUNÇÃO de até três critérios `coluna = valor`, com a
    # coluna escolhida num dropdown do cabeçalho real do relatório: um critério
    # preenchido tira tudo que tem aquele valor naquela coluna, dois tiram só o
    # que tem os dois, três só o que tem os três. Par com coluna ou valor em
    # branco não conta, e a linha SEM critério nenhum é ignorada — sem nada a
    # exigir ela apagaria o lado da Athena da recon.
    #
    # A coluna é escolhida, e não fixada no código, porque a primeira versão
    # deste cadastro supôs `TRADING BOOK` / `OTHER BOOK` — nomes que o relatório
    # não tem. Quem sabe em que coluna mora cada valor é quem opera.
    'fxo-book-disregard': {
        'label': 'FXO Recon — Athena Rows to Disregard',
        'upgrade': _fxo_book_disregard_upgrade,
        'columns': [
            {'key': 'COLUMN 1', 'label': 'Column 1', 'type': 'select',
             'options': _ATHENA_FXO_COLUMNS},
            {'key': 'VALUE 1', 'label': 'Value 1'},
            {'key': 'COLUMN 2', 'label': 'Column 2 (optional)', 'type': 'select',
             'options': _ATHENA_FXO_COLUMNS},
            {'key': 'VALUE 2', 'label': 'Value 2'},
            {'key': 'COLUMN 3', 'label': 'Column 3 (optional)', 'type': 'select',
             'options': _ATHENA_FXO_COLUMNS},
            {'key': 'VALUE 3', 'label': 'Value 3'},
            {'key': 'NOTES', 'label': 'Notes'},
        ],
        'seed': [
            {'COLUMN 1': 'Portfolio', 'VALUE 1': 'FSLTVNCT VNLA CETIP LAWTON',
             'COLUMN 2': 'CounterpartyName', 'VALUE 2': 'BRL_FXO LAWTON',
             'COLUMN 3': 'INT_EXT', 'VALUE 3': 'INTERBOOK',
             'NOTES': 'Perna interbook da Lawton — não tem par na CETIP'},
        ],
    },
    # Quais linhas do Operations B3 entram numa apuração de liquidação — de
    # qualquer produto. Era `swap-b3-events`, só com o Tipo Operação do swap;
    # hoje a linha é uma REGRA sobre as três colunas que decidem isso (Tipo
    # Título, Tipo Operação e Status B3) mais o que fazer com ela.
    #
    # Campo em branco é CORINGA ('vale para qualquer'), então a regra do
    # cancelamento se escreve numa linha só e vale para TER, OPC, SWAP e COE de
    # uma vez.
    #
    # Precedência (ver `_opb3_settle_ok`): Disregard vence sempre; depois, um
    # Tipo Título que tenha ao menos um Consider vira LISTA BRANCA — só o que
    # está registrado entra. Tipo Título sem nenhum Consider não é filtrado.
    #
    # O casamento ignora caixa, acento e pontuação: 'CANCELADA: COMANDADA',
    # 'CANCELADA:COMANDADA' e 'Cancelada Comandada' são a mesma coisa, e
    # 'AMORTIZAÇÃO' com cedilha casa com a linha registrada sem ele.
    'opb3-events': {
        'label': 'Operations B3 Events',
        'columns': [
            {'key': 'TIPO TITULO', 'label': 'Tipo Título (blank = any)', 'type': 'select',
             'options': _MAP_OPB3_TITULOS},
            {'key': 'TIPO OPERACAO', 'label': 'Tipo Operação (blank = any)',
             'type': 'datalist', 'options': _MAP_OPB3_OPERACOES},
            {'key': 'STATUS B3', 'label': 'Status B3 (blank = any)',
             'type': 'datalist', 'options': _MAP_OPB3_STATUSES},
            {'key': 'USE', 'label': 'Consider / Disregard', 'type': 'select',
             'options': _MAP_OPB3_USE},
            {'key': 'NOTES', 'label': 'Notes'},
        ],
        'seed': list(_MAP_OPB3_SEED),
        'upgrade': _opb3_events_upgrade,
    },
    # IR do swap, parte 1: as EXCEÇÕES por cliente, testadas antes de tudo. Vêm
    # do IF encadeado da planilha de avisos — bancos e as duas entidades JPM.
    # 'Starts with' existe por causa do LEFT(A12;5)="BANCO", que isenta qualquer
    # razão social começada em BANCO sem listá-las uma a uma.
    'swap-ir-client': {
        'label': 'Swap IR — Client Exceptions',
        'columns': [
            {'key': 'CLIENT', 'label': 'Client (Counterparty)'},
            {'key': 'MATCH', 'label': 'Match', 'type': 'select',
             'options': ['Exact', 'Starts with']},
            {'key': 'RATE', 'label': 'Rate %'},
        ],
        'seed': [
            {'CLIENT': 'JPMORGAN CHASE BANK, N.A', 'MATCH': 'Exact', 'RATE': '0'},
            {'CLIENT': 'LAWTON MULTIMERCADO EXCLUSIVO FUNDO DE INVESTIMENTO '
                       'CREDITO PRIVADO', 'MATCH': 'Exact', 'RATE': '0'},
            {'CLIENT': 'BANCO', 'MATCH': 'Starts with', 'RATE': '0'},
            {'CLIENT': 'J.P. MORGAN OVERSEAS CAPITAL LLC', 'MATCH': 'Exact', 'RATE': '10'},
        ],
    },
    # IR do swap, parte 2: a tabela regressiva por PRAZO, aplicada só quando quem
    # recebe é a contraparte. A linha com UP TO DAYS vazio é o "acima de tudo" —
    # e é ela que fecha o vão da planilha, onde `IF(E12>721;15%)` deixava o prazo
    # 721 exato sem resposta (a fórmula devolvia FALSE).
    'swap-ir-term': {
        'label': 'Swap IR — Term Brackets',
        'columns': [
            {'key': 'UP TO DAYS', 'label': 'Up to (days · blank = above all)'},
            {'key': 'RATE', 'label': 'Rate %'},
        ],
        'seed': [
            {'UP TO DAYS': '180', 'RATE': '22.5'},
            {'UP TO DAYS': '360', 'RATE': '20'},
            {'UP TO DAYS': '720', 'RATE': '17.5'},
            {'UP TO DAYS': '',    'RATE': '15'},
        ],
    },
    # Nome do banker → e-mail. É o Cc do e-mail de coleta de assinatura: o
    # `BANKER` do Reference Data traz o GRUPO por extenso ('Fulano e Sicrano'), e
    # é esta lista que resolve cada nome num endereço.
    #
    # A lista era o `signature_collection_bankers.json`, mantida à mão com 58
    # nomes: banker novo só entrava por commit, e enquanto isso o e-mail saía sem
    # ele no Cc. Virou cadastro e o arquivo foi para `mappings/bankers-email.json`
    # — o caminho padrão da chave, junto dos outros 26 —, então não há `file`
    # aqui. O `seed` existe para a instância que perder o arquivo.
    'bankers-email': {
        'label': 'Bankers E-mails',
        'columns': [
            {'key': 'BANKER', 'label': 'Banker'},
            {'key': 'EMAIL', 'label': 'E-mail'},
        ],
        'seed': [
            {'BANKER': 'Ana Katayama', 'EMAIL': 'ana.katayama@jpmchase.com'},
            {'BANKER': 'Andre Regula', 'EMAIL': 'andre.regula@jpmorgan.com'},
            {'BANKER': 'Andre Schwartzman', 'EMAIL': 'andre.schwartzman@jpmorgan.com'},
            {'BANKER': 'Andreia Choi', 'EMAIL': 'andreia.choi@jpmorgan.com'},
            {'BANKER': 'Antonio Mariz', 'EMAIL': 'antonio.v.mariz@jpmchase.com'},
            {'BANKER': 'Arthur Pego', 'EMAIL': 'arthur.pego@jpmorgan.com'},
            {'BANKER': 'Bruno Mello', 'EMAIL': 'bruno.mello@jpmorgan.com'},
            {'BANKER': 'Camila Bludeni', 'EMAIL': 'camila.bludeni@jpmorgan.com'},
            {'BANKER': 'Carolina Pires', 'EMAIL': 'carolinasena.o.pires@jpmorgan.com'},
            {'BANKER': 'Claudio Junior', 'EMAIL': 'claudio.m.junior@jpmorgan.com'},
            {'BANKER': 'Daniel Pineschi', 'EMAIL': 'daniel.pineschi@jpmchase.com'},
            {'BANKER': 'Debora Darin', 'EMAIL': 'debora.darin@jpmorgan.com'},
            {'BANKER': 'Diogo Yoshinaga', 'EMAIL': 'diogo.yoshinaga@jpmorgan.com'},
            {'BANKER': 'Edoardo Freschet', 'EMAIL': 'edoardo.freschet@jpmorgan.com'},
            {'BANKER': 'Erik Pontes', 'EMAIL': 'erik.p.pontes@jpmorgan.com'},
            {'BANKER': 'Fabiano Fernandes', 'EMAIL': 'fabiano.fernandes@jpmchase.com'},
            {'BANKER': 'Felipe Esper', 'EMAIL': 'felipe.esper@jpmorgan.com'},
            {'BANKER': 'Felipe Ferraz', 'EMAIL': 'felipe.ferraz@jpmorgan.com'},
            {'BANKER': 'Fernando Moreira', 'EMAIL': 'fernando.a.moreira@jpmorgan.com'},
            {'BANKER': 'Gabriel Mendes', 'EMAIL': 'gabriel.mendes@jpmorgan.com'},
            {'BANKER': 'Gabriel Sousa', 'EMAIL': 'gabriel.j.sousa@jpmorgan.com'},
            {'BANKER': 'Giovana Alves', 'EMAIL': 'giovana.ds.alves@jpmchase.com'},
            {'BANKER': 'Giulia Menegon', 'EMAIL': 'giulia.menegon@jpmorgan.com'},
            {'BANKER': 'Guilherme Rissi', 'EMAIL': 'guilherme.rissi@jpmorgan.com'},
            {'BANKER': 'Isabela Bacchi', 'EMAIL': 'isabela.p.bacchi@jpmorgan.com'},
            {'BANKER': 'Isabela Ramos', 'EMAIL': 'isabela.ramos@jpmchase.com'},
            {'BANKER': 'Isabella Brunele', 'EMAIL': 'isabella.brunele@jpmchase.com'},
            {'BANKER': 'Isabella Giovanelli', 'EMAIL': 'isabella.giovanelli@jpmorgan.com'},
            {'BANKER': 'Joao Camargo', 'EMAIL': 'joao.camargo@jpmorgan.com'},
            {'BANKER': 'Joao Sousa', 'EMAIL': 'joao.sousa@jpmorgan.com'},
            {'BANKER': 'Julia Chohfi', 'EMAIL': 'julia.chohfi@jpmorgan.com'},
            {'BANKER': 'Lais Zacarias', 'EMAIL': 'lais.zacarias@jpmorgan.com'},
            {'BANKER': 'Liana Pollastri', 'EMAIL': 'liana.pollastri@jpmorgan.com'},
            {'BANKER': 'Lucca Maciel', 'EMAIL': 'lucca.maciel@jpmorgan.com'},
            {'BANKER': 'Luciana Filoni', 'EMAIL': 'luciana.filoni@jpmorgan.com'},
            {'BANKER': 'Luciana Furtado', 'EMAIL': 'luciana.furtado@jpmorgan.com'},
            {'BANKER': 'Lucianna Lorenzo', 'EMAIL': 'lucianna.lorenzo@jpmorgan.com'},
            {'BANKER': 'Marcelo Afonseca', 'EMAIL': 'marcelo.afonseca@jpmorgan.com'},
            {'BANKER': 'Marcilio Zanoni', 'EMAIL': 'marcilio.zanonijunior@jpmorgan.com'},
            {'BANKER': 'Michel Maluf', 'EMAIL': 'michel.berbari@jpmorgan.com'},
            {'BANKER': 'Nathalia Ferreira', 'EMAIL': 'nathalia.ferreira@jpmorgan.com'},
            {'BANKER': 'Nathalia Ramos', 'EMAIL': 'nathalia.ramos@jpmorgan.com'},
            {'BANKER': 'Nicolas Belmonte', 'EMAIL': 'nicolas.belmonte@jpmorgan.com'},
            {'BANKER': 'Nicolas Guevara', 'EMAIL': 'nicolas.a.guevara@jpmorgan.com'},
            {'BANKER': 'Nikolas Dorto', 'EMAIL': 'nikolas.dorto@jpmorgan.com'},
            {'BANKER': 'Paulo Samelo', 'EMAIL': 'paulo.samelo@jpmchase.com'},
            {'BANKER': 'Rafael Matos', 'EMAIL': 'rafael.matos@jpmchase.com'},
            {'BANKER': 'Rafaela Negrão', 'EMAIL': 'rafaela.negrao@jpmorgan.com'},
            {'BANKER': 'Raone Turco', 'EMAIL': 'raone.turco@jpmorgan.com'},
            {'BANKER': 'Rhadur Domingos', 'EMAIL': 'rhadur.domingos@jpmorgan.com'},
            {'BANKER': 'Roberto Michels', 'EMAIL': 'roberto.michels@jpmorgan.com'},
            {'BANKER': 'Rodrigo Carmo', 'EMAIL': 'rodrigo.carmo@jpmorgan.com'},
            {'BANKER': 'Rodrigo Choi', 'EMAIL': 'rodrigo.h.choi@jpmorgan.com'},
            {'BANKER': 'Thaina Picado', 'EMAIL': 'thaina.picado@jpmchase.com'},
            {'BANKER': 'Thiago Eloy', 'EMAIL': 'thiago.eloy@jpmorgan.com'},
            {'BANKER': 'Thiago Martinez', 'EMAIL': 'thiago.martinez@jpmorgan.com'},
            {'BANKER': 'Vinicius Almeida', 'EMAIL': 'vinicius.almeida@jpmorgan.com'},
            {'BANKER': 'Willian Melo', 'EMAIL': 'willian.hara@jpmorgan.com'},
        ],
    },
    'swap-index': {
        'label': 'Swap Index — B3 Code',
        # MESMO arquivo da aba Swap Index do B3 Index Results — não uma cópia.
        # `file` aponta o cadastro para lá, então as duas telas editam o mesmo
        # SwapIndex.json e não há como divergirem. As colunas são as chaves do
        # próprio arquivo (inclusive STATUS/MAKER/CHECKER, declaradas para que um
        # POST do /mapping, que reescreve o arquivo inteiro, não as apague).
        'file': data_write('SwapIndex.json'),
        'columns': [
            {'key': 'Codigo Referencia Externa', 'label': 'B3 Code'},
            {'key': 'Nome Curva', 'label': 'Curve Name'},
            {'key': 'Nome Categoria', 'label': 'Category'},
            {'key': 'STATUS', 'label': 'Status'},
            {'key': 'MAKER', 'label': 'Maker'},
            {'key': 'CHECKER', 'label': 'Checker'},
        ],
        'seed': [
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': 'ACO', 'Nome Curva': 'ACOES', 'Nome Categoria': 'ACOES'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': 'COM', 'Nome Curva': 'COMMODITIES', 'Nome Categoria': 'COMMODITIES'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': 'C04', 'Nome Curva': 'OURO Ajuste', 'Nome Categoria': 'COMMODITIES'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': 'IBL', 'Nome Curva': 'IBOVESPA LIQUIDACAO', 'Nome Categoria': 'INDICES'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': 'IBC', 'Nome Curva': 'IBOVESPA LIQUIDACAO CONTINUO', 'Nome Categoria': 'INDICES'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': 'IBO', 'Nome Curva': 'IBOVESPA FECHAMENTO', 'Nome Categoria': 'INDICES'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': 'C19', 'Nome Curva': 'IPCA', 'Nome Categoria': 'INDICES DE PRECOS'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': 'C09', 'Nome Curva': 'IGP-M', 'Nome Categoria': 'INDICES DE PRECOS'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': 'C10', 'Nome Curva': 'IGP-DI', 'Nome Categoria': 'INDICES DE PRECOS'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': 'C16', 'Nome Curva': 'INPC', 'Nome Categoria': 'INDICES DE PRECOS'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': 'C21', 'Nome Curva': 'IPCA-NC', 'Nome Categoria': 'INDICES DE PRECOS'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': 'C23', 'Nome Curva': 'TJLP', 'Nome Categoria': 'JUROS'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': 'C52', 'Nome Curva': 'DI 360D', 'Nome Categoria': 'JUROS'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': 'C20', 'Nome Curva': 'TR', 'Nome Categoria': 'JUROS'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': 'C60', 'Nome Curva': 'PREFIXADO 360D', 'Nome Categoria': 'JUROS'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': 'C66', 'Nome Curva': 'PRE LINEAR 360D', 'Nome Categoria': 'JUROS'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': 'C99', 'Nome Curva': 'PREFIXADO 252D', 'Nome Categoria': 'JUROS'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': 'C03', 'Nome Curva': 'DI', 'Nome Categoria': 'JUROS'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': 'C01', 'Nome Curva': 'SELIC', 'Nome Categoria': 'JUROS'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': 'C65', 'Nome Curva': 'PREFIXADO 365D', 'Nome Categoria': 'JUROS'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': 'T12', 'Nome Curva': 'TSFR12M', 'Nome Categoria': 'JUROS INTERNACIONAIS'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': 'T06', 'Nome Curva': 'TSFR6M', 'Nome Categoria': 'JUROS INTERNACIONAIS'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': 'SFR', 'Nome Curva': 'SOFR', 'Nome Categoria': 'JUROS INTERNACIONAIS'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': 'T03', 'Nome Curva': 'TSFR3M', 'Nome Categoria': 'JUROS INTERNACIONAIS'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': 'T01', 'Nome Curva': 'TSFR1M', 'Nome Categoria': 'JUROS INTERNACIONAIS'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': 'C29', 'Nome Curva': 'TJMI', 'Nome Categoria': 'JUROS INTERNACIONAIS'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': '425', 'Nome Curva': 'FRANCO SUICO', 'Nome Categoria': 'TAXAS DE CAMBIO'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': '796', 'Nome Curva': 'RENMINBI HONG KON', 'Nome Categoria': 'TAXAS DE CAMBIO'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': '930', 'Nome Curva': 'WON/COREIA SUL', 'Nome Categoria': 'TAXAS DE CAMBIO'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': '642', 'Nome Curva': 'NOVA LIRA/TURQUIA', 'Nome Categoria': 'TAXAS DE CAMBIO'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': '450', 'Nome Curva': 'GUARANI/PARAGUAI', 'Nome Categoria': 'TAXAS DE CAMBIO'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': '745', 'Nome Curva': 'PESO/URUGUAIO', 'Nome Categoria': 'TAXAS DE CAMBIO'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': '820', 'Nome Curva': 'RIAL/ARAB SAUDITA', 'Nome Categoria': 'TAXAS DE CAMBIO'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': '165', 'Nome Curva': 'DOLAR CANADENSE', 'Nome Categoria': 'TAXAS DE CAMBIO'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': 'USD', 'Nome Curva': 'DOLAR COMERCIAL EXPONENCIAL', 'Nome Categoria': 'TAXAS DE CAMBIO'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': '100', 'Nome Curva': 'DINAR/KWAIT', 'Nome Categoria': 'TAXAS DE CAMBIO'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': '055', 'Nome Curva': 'COROA DINAM/DINAM', 'Nome Categoria': 'TAXAS DE CAMBIO'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': '660', 'Nome Curva': 'NOVO-SOL', 'Nome Categoria': 'TAXAS DE CAMBIO'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': '795', 'Nome Curva': 'IUAN RENMIMBI/CHI', 'Nome Categoria': 'TAXAS DE CAMBIO'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': '785', 'Nome Curva': 'RANDE/AFRICA SUL', 'Nome Categoria': 'TAXAS DE CAMBIO'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': '860', 'Nome Curva': 'RUPIA/INDIA', 'Nome Categoria': 'TAXAS DE CAMBIO'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': 'CHF', 'Nome Curva': 'FRANCO SUICO BCE', 'Nome Categoria': 'TAXAS DE CAMBIO'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': '129', 'Nome Curva': 'EURO WMR', 'Nome Categoria': 'TAXAS DE CAMBIO'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': '706', 'Nome Curva': 'PESO ARGENTINO', 'Nome Categoria': 'TAXAS DE CAMBIO'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': '470', 'Nome Curva': 'IENE', 'Nome Categoria': 'TAXAS DE CAMBIO'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': '540', 'Nome Curva': 'LIBRA ESTERLINA', 'Nome Categoria': 'TAXAS DE CAMBIO'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': '245', 'Nome Curva': 'DOLAR/NOVA ZELAND', 'Nome Categoria': 'TAXAS DE CAMBIO'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': '145', 'Nome Curva': 'DIRHAM/EMIR.ARABE', 'Nome Categoria': 'TAXAS DE CAMBIO'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': '137', 'Nome Curva': 'FRANCO SUICO WMR', 'Nome Categoria': 'TAXAS DE CAMBIO'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': '150', 'Nome Curva': 'DOLAR AUSTRALIANO', 'Nome Categoria': 'TAXAS DE CAMBIO'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': '070', 'Nome Curva': 'COROA SUECA/SUECI', 'Nome Categoria': 'TAXAS DE CAMBIO'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': '978', 'Nome Curva': 'EURO/COM.EUROPEIA', 'Nome Categoria': 'TAXAS DE CAMBIO'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': '640', 'Nome Curva': 'NOVO DOLAR/TAIWAN', 'Nome Categoria': 'TAXAS DE CAMBIO'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': '828', 'Nome Curva': 'RINGGIT/MALASIA', 'Nome Categoria': 'TAXAS DE CAMBIO'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': '800', 'Nome Curva': 'RIAL/CATAR', 'Nome Categoria': 'TAXAS DE CAMBIO'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': '026', 'Nome Curva': 'BOLIVAR FORTE/VEN', 'Nome Categoria': 'TAXAS DE CAMBIO'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': '345', 'Nome Curva': 'FORINT/HUNGRIA', 'Nome Categoria': 'TAXAS DE CAMBIO'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': '471', 'Nome Curva': 'IENE WMR REAIS', 'Nome Categoria': 'TAXAS DE CAMBIO'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': '220', 'Nome Curva': 'DOLAR DOS EUA', 'Nome Categoria': 'TAXAS DE CAMBIO'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': '741', 'Nome Curva': 'PESO MEXICANO', 'Nome Categoria': 'TAXAS DE CAMBIO'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': 'EUR', 'Nome Curva': 'EURO BCE', 'Nome Categoria': 'TAXAS DE CAMBIO'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': '205', 'Nome Curva': 'DOLAR/HONG-KONG', 'Nome Categoria': 'TAXAS DE CAMBIO'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': '015', 'Nome Curva': 'BATH/TAILANDIA', 'Nome Categoria': 'TAXAS DE CAMBIO'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': '075', 'Nome Curva': 'COROA TCHECA', 'Nome Categoria': 'TAXAS DE CAMBIO'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': '735', 'Nome Curva': 'PESO/FILIPINAS', 'Nome Categoria': 'TAXAS DE CAMBIO'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': '506', 'Nome Curva': 'NOVO LEU/ROMENIA', 'Nome Categoria': 'TAXAS DE CAMBIO'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': 'U30', 'Nome Curva': 'DOLAR DOS EUA 30/360', 'Nome Categoria': 'TAXAS DE CAMBIO'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': '715', 'Nome Curva': 'PESO CHILENO', 'Nome Categoria': 'TAXAS DE CAMBIO'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': '195', 'Nome Curva': 'DOLAR/CINGAPURA', 'Nome Categoria': 'TAXAS DE CAMBIO'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': '830', 'Nome Curva': 'RUBLO/RUSSIA', 'Nome Categoria': 'TAXAS DE CAMBIO'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': '260', 'Nome Curva': 'DONGUE/VIETNAN', 'Nome Categoria': 'TAXAS DE CAMBIO'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': '030', 'Nome Curva': 'BOLIVIANO', 'Nome Categoria': 'TAXAS DE CAMBIO'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': '065', 'Nome Curva': 'COROA NORUE/NORUE', 'Nome Categoria': 'TAXAS DE CAMBIO'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': '865', 'Nome Curva': 'RUPIA/INDONESIA', 'Nome Categoria': 'TAXAS DE CAMBIO'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': '975', 'Nome Curva': 'ZLOTY/POLONIA', 'Nome Categoria': 'TAXAS DE CAMBIO'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': '720', 'Nome Curva': 'PESO COLOMBIANO', 'Nome Categoria': 'TAXAS DE CAMBIO'},
            {'STATUS': 'ACTIVE', 'Codigo Referencia Externa': 'C00', 'Nome Curva': 'VCP', 'Nome Categoria': 'VCP'},
        ],
    },
    'ndfc-ir-exempt': {
        'label': 'NDF Commodities IR — Exempt Clients',
        'columns': [
            {'key': 'CLIENT', 'label': 'Client'},
            {'key': 'MATCH', 'label': 'Match', 'type': 'select',
             'options': ['Exact', 'Starts with']},
            {'key': 'NOTES', 'label': 'Notes'},
        ],
        # Quem NÃO paga o IR de 0,005% do Termo de Mercadoria: instituições
        # financeiras e as pernas internas. A fórmula da planilha só trazia
        # LAWTON — as demais entraram porque foram pedidas por nome; qualquer
        # outra se registra por aqui, sem tocar em código.
        'seed': [
            {'CLIENT': 'LAWTON', 'MATCH': 'Starts with', 'NOTES': 'fórmula da planilha'},
            {'CLIENT': 'ATACAMA', 'MATCH': 'Starts with', 'NOTES': ''},
            {'CLIENT': 'BANCO', 'MATCH': 'Starts with', 'NOTES': 'instituições financeiras'},
            {'CLIENT': 'JPMORGAN', 'MATCH': 'Starts with', 'NOTES': 'perna interna'},
            {'CLIENT': 'J.P. MORGAN', 'MATCH': 'Starts with', 'NOTES': 'perna interna'},
        ],
    },
    'b3-accounts': {
        'label': 'B3 Accounts',
        'columns': [
            {'key': 'LE', 'label': 'Legal Entity'},
            {'key': 'SIMPLIFIED NAME', 'label': 'Simplified Name (B3)'},
            {'key': 'ACCOUNT', 'label': 'B3 Account'},
            # O TIPO é o que decide a regra, então é `select`: digitado à mão,
            # um "Cliente1" viraria conta PRÓPRIA em silêncio — e uma conta
            # própria tratada como guarda-chuva manda o app procurar cliente
            # pelo CNPJ onde não há cliente nenhum.
            {'key': 'ACCOUNT TYPE', 'label': 'Account Type', 'type': 'select',
             'options': list(_B3_ACCOUNT_TYPES)},
            # Como a entidade está escrita no Reference Data. O Nome
            # Simplificado ao lado é o da B3 (`INTRAGLAWTONFDO`, 20 caracteres),
            # que não é razão social nenhuma: é ele que vem nas linhas do
            # Operations B3, e sem esta coluna a mensageria de uma conta nossa
            # sairia endereçada ao apelido do arquivo.
            {'key': 'REFDATA NAME', 'label': 'Reference Data Name', 'refdata': 'name'},
            # A mensageria sai na visão desta conta? Ver `_b3_msg_view_use`.
            {'key': 'MESSAGING', 'label': 'Messaging', 'type': 'select',
             'options': list(_B3_MSG_USES)},
            {'key': 'NOTES', 'label': 'Notes'},
        ],
        # As contas da B3 de cada entidade nossa: a PRÓPRIA (posição da casa) e
        # as de CLIENTE, que são guarda-chuva. Duas perguntas saem daqui:
        #
        #  1. o Nome Simplificado do Participante no header dos arquivos TER —
        #     antes um dicionário fixo no código, com a mesma resposta escrita
        #     em dois lugares;
        #  2. a conta identifica o cliente? CLIENT 1/CLIENT 2 NÃO identificam:
        #     numa linha do Operations B3 com essa Conta Contraparte o nome que
        #     vem da B3 é o do titular do guarda-chuva, e quem é o cliente sai
        #     do CNPJ. A conta PRÓPRIA identifica — é a nossa.
        #  3. a mensageria sai na visão desta conta? A liquidação intragrupo
        #     chega pelas DUAS pontas — o mesmo pagamento, espelhado —, e sair
        #     das duas seria cobrar duas vezes. A visão que assina é a do Banco;
        #     MGT, Lawton e Atacama entram como Disregard.
        'seed': [
            {'LE': 'MGT', 'SIMPLIFIED NAME': 'MORGANBC', 'ACCOUNT': '04880.00-6',
             'ACCOUNT TYPE': 'OWN', 'REFDATA NAME': '', 'MESSAGING': 'Disregard', 'NOTES': ''},
            {'LE': 'MGT', 'SIMPLIFIED NAME': 'MORGANBC', 'ACCOUNT': '04880.10-9',
             'ACCOUNT TYPE': 'CLIENT 1', 'REFDATA NAME': '', 'MESSAGING': 'Disregard', 'NOTES': ''},
            {'LE': 'JPM', 'SIMPLIFIED NAME': 'JPMORGANBM', 'ACCOUNT': '73760.00-9',
             'ACCOUNT TYPE': 'OWN', 'REFDATA NAME': '', 'MESSAGING': 'Consider', 'NOTES': ''},
            {'LE': 'JPM', 'SIMPLIFIED NAME': 'JPMORGANBM', 'ACCOUNT': '73760.10-2',
             'ACCOUNT TYPE': 'CLIENT 1', 'REFDATA NAME': '', 'MESSAGING': 'Consider', 'NOTES': ''},
            {'LE': 'JPM', 'SIMPLIFIED NAME': 'JPMORGANBM', 'ACCOUNT': '73760.20-5',
             'ACCOUNT TYPE': 'CLIENT 2', 'REFDATA NAME': '', 'MESSAGING': 'Consider', 'NOTES': ''},
            {'LE': 'LAWTON', 'SIMPLIFIED NAME': 'INTRAGLAWTONFDO', 'ACCOUNT': '00041.00-7',
             'ACCOUNT TYPE': 'OWN', 'REFDATA NAME': '', 'MESSAGING': 'Disregard', 'NOTES': ''},
            {'LE': 'ATACAMA', 'SIMPLIFIED NAME': 'INTRAGATACAMAFDO', 'ACCOUNT': '85398.00-5',
             'ACCOUNT TYPE': 'OWN', 'REFDATA NAME': '', 'MESSAGING': 'Disregard', 'NOTES': ''},
        ],
        'upgrade': _b3_accounts_upgrade,
    },
    'ndfc-advice-split': {
        'label': 'NDF Commodities Advice — Split by Commodity',
        'columns': [
            {'key': 'CLIENT', 'label': 'Client'},
            {'key': 'MATCH', 'label': 'Match', 'type': 'select',
             'options': ['Exact', 'Starts with']},
            {'key': 'NOTES', 'label': 'Notes'},
        ],
        # Contrapartes que recebem UM AVISO POR COMMODITY. Fora desta lista, um
        # aviso pode trazer alumínio e café na mesma tabela.
        # 'MONDELEZ' com Starts with cobre as duas entidades (Brasil e Brasil
        # Norte Nordeste) — que já são contrapartes distintas no Reference Data,
        # então a quebra entre elas acontece sozinha.
        'seed': [
            {'CLIENT': 'MONDELEZ', 'MATCH': 'Starts with',
             'NOTES': 'um aviso por commodity'},
        ],
    },
    'swap-funcionalidade': {
        'label': 'Swap — Funcionalidade',
        'columns': [
            {'key': 'CODE', 'label': 'B3 Code'},
            {'key': 'LABEL', 'label': 'Text'},
        ],
        'seed': [
            {'CODE': '0', 'LABEL': 'SEM FUNCIONALIDADE'},
            {'CODE': '1', 'LABEL': 'KNOCK IN'},
            {'CODE': '2', 'LABEL': 'KNOCK OUT'},
            {'CODE': '3', 'LABEL': 'KNOCK INOUT'},
            {'CODE': '4', 'LABEL': 'SWAPTION'},
            {'CODE': '5', 'LABEL': 'COMPOUND'},
            {'CODE': '6', 'LABEL': 'OPCAO ARREPENDIMENTO'},
            {'CODE': '7', 'LABEL': 'KNOCK IN COM OPCAO'},
            {'CODE': '8', 'LABEL': 'KNOCK OUT COM OPCAO'},
            {'CODE': '9', 'LABEL': 'SWAP COM PR\u00caMIO'},
        ],
    },
    'swap-amortizacao': {
        'label': 'Swap — Tipo de Amortiza\u00e7\u00e3o',
        'columns': [
            {'key': 'CODE', 'label': 'B3 Code'},
            {'key': 'LABEL', 'label': 'Text'},
        ],
        'seed': [
            {'CODE': '0', 'LABEL': 'Sobre Valor Base Original'},
            {'CODE': '1', 'LABEL': 'Sobre Valor Base Remanescente'},
            {'CODE': '3', 'LABEL': 'Na Data de Vencimento'},
            {'CODE': '4', 'LABEL': 'Sem Troca de Amortiza\u00e7\u00e3o'},
        ],
    },
    'swap-code-labels': {
        'label': 'Swap — Sinal e Sim/N\u00e3o',
        'columns': [
            {'key': 'FIELD', 'label': 'Field', 'type': 'select',
             'options': ['Sinal Taxa', 'Sim/N\u00e3o']},
            {'key': 'CODE', 'label': 'B3 Code'},
            {'key': 'LABEL', 'label': 'Text'},
        ],
        'seed': [
            {'FIELD': 'Sinal Taxa', 'CODE': '0', 'LABEL': '+'},
            {'FIELD': 'Sinal Taxa', 'CODE': '1', 'LABEL': '-'},
            {'FIELD': 'Sim/N\u00e3o', 'CODE': '0', 'LABEL': 'Sim'},
            {'FIELD': 'Sim/N\u00e3o', 'CODE': '1', 'LABEL': 'N\u00e3o'},
        ],
    },
    # Cotações: **Código do Ativo Subjacente** (o mesmo do Index B3) → o símbolo
    # que a fonte de mercado entende. É de-para, então é cadastro (§2).
    #
    # A LISTA de instrumentos da tela NÃO sai daqui: ela é o Subjacente ao vivo
    # (`features/quotes/queries.underlyings`), para um ativo novo cadastrado
    # no Index B3
    # aparecer em Cotações no mesmo dia. Este cadastro só TRADUZ, e por isso
    # guarda apenas os códigos que já têm símbolo — as ~1.700 linhas restantes
    # em branco seriam ruído para quem edita a tabela.
    #
    # O `seed` vai VAZIO de propósito: os dois arquivos são versionados
    # (471 equities e 70 commodities, semeados do `symbol_map` do app de
    # desktop mais a regra dos tickers brasileiros — `PETR4` → `PETR4.SA`), e
    # repetir centenas de pares aqui criaria uma segunda lista para divergir da
    # primeira. Numa instância sem o pull o cadastro nasce vazio e a tela diz
    # "não está no cadastro" — visível, diferente de um símbolo errado.
    'quotes-equity': {
        'label': 'Quotes \u2014 Equities',
        'columns': [
            {'key': 'LABEL', 'label': 'Underlying Code (B3)'},
            {'key': 'SYMBOL', 'label': 'Market Symbol'},
            {'key': 'NOTES', 'label': 'Notes'},
        ],
        'seed': [],
    },
    # Commodity é o mesmo cadastro do Equities, com uma diferença que muda o
    # tamanho da tabela: contrato futuro tem VENCIMENTO, e o de-para por código
    # fechado pedia uma linha por mês de cada mercadoria — 70 linhas para 10
    # mercadorias, mais uma linha nova a cada vencimento que a B3 abre.
    #
    # As duas colunas aceitam o PADRÃO `"MY"` do cadastro Commodities × B3
    # (letra do mês + ano; `_` é espaço literal), e uma linha passa a valer para
    # todos os vencimentos daquela mercadoria:
    #
    #     BO"MY"  →  ZL"MY".CBT       BOK6  → ZLK26.CBT
    #     C_"MY"  →  ZC"MY".CBT      'C K6' → ZCK26.CBT
    #
    # Quem expande é o `quotes.symbol_lookup` — inclusive o ano de UM dígito da
    # B3 virando os DOIS que o símbolo de mercado usa. Linha sem `"MY"` continua
    # literal e vence o padrão, que é como se cadastra a exceção de um
    # vencimento só.
    'quotes-commodity': {
        'label': 'Quotes \u2014 Commodities',
        'columns': [
            {'key': 'LABEL', 'label': 'Underlying Code (B3) \u2014 code or "MY" pattern'},
            {'key': 'SYMBOL', 'label': 'Market Symbol \u2014 symbol or "MY" pattern'},
            {'key': 'NOTES', 'label': 'Notes'},
        ],
        'seed': [],
    },
    # Código de histórico do GDT (`nHistorico` do `rlctahis.csv`) → PRODUTO, para
    # a Recon Pay/Rec.
    #
    # O lado do cliente da recon é o extrato da conta interna, e o único campo
    # que diz de que produto é o lançamento é esse código. Sem o de-para, o
    # `_cli_rlctahis` classificava tudo como NDF (menos três códigos de swap
    # fixos no código): o pagamento de uma opção de commodity da Lawton entrava
    # no balde NDF, e como o `_net_client` agrupa por
    # **(contraparte, LE, PRODUTO)**, o Total Net dela somava produtos que o lado
    # do JPM tem separados. Daí as linhas "Netting não tratado pelo OTC Tracker":
    # o valor netado não batia com nada porque era a soma de coisas diferentes.
    # Netar continua sendo netar — só que DENTRO do produto.
    #
    # **PRODUCT preenchido = o código liquida aquele produto, e a linha ENTRA na
    # recon. PRODUCT em branco = o código está documentado e é IGNORADO** — é o
    # caso das duas transferências entre contas, que a mesa quer ver cadastradas
    # para saber o que são, e que não são liquidação de produto nenhum.
    #
    # `_SDCONTA_HIST_ALLOW` continua no `recon_payrec` como piso: código que
    # ainda não tem linha aqui (o `4419` e o `AA`) segue entrando com a regra
    # histórica, para o cadastro novo não apagar comportamento em silêncio.
    # Excecoes de liquidacao da Recon Pay/Rec: net type DIRECIONAL por contraparte.
    #
    # O `NET` do Reference Data e UM valor por contraparte, e algumas liquidam de
    # um jeito no que pagam e de outro no que recebem. Eram duas contrapartes
    # FIXAS no `recon_payrec` (as duas Saint-Gobain); agora e cadastro, e a ordem
    # de consulta e **excecao primeiro, Reference Data depois** - sem linha aqui,
    # nada muda para a contraparte.
    #
    # A regra vale nos DOIS lados do batimento (JPM e cliente). Aplica-la so num
    # deles faria a recon comparar coisas agrupadas de jeitos diferentes, que e a
    # origem de metade das pendencias que ninguem explica.
    #
    # A ORDEM DAS LINHAS E A PRECEDENCIA: vence a primeira que casar, e o
    # casamento e por TOKENS - todas as palavras do texto cadastrado tem de
    # aparecer no nome da contraparte (cego a acento, caixa e pontuacao). Por isso
    # a linha mais especifica vem ANTES da mais geral: `GOBAIN CANALIZACAO` antes
    # de `GOBAIN BRASIL`, e o Mondelez do Norte/Nordeste antes do `MONDELEZ
    # BRASIL`, que casaria com os dois. Era o que o codigo fixo fazia com um
    # `not 'canalizacao'`; aqui a precedencia e visivel e editavel.
    #
    # Sentido em BRANCO cai no net type do Reference Data - e como se cadastra
    # "so o pagamento e excecao".
    # Quem entra no e-mail MT300 (card do Control Panel, 19:30).
    #
    # A mensagem MT300 e confirmada por um grupo especifico de clientes, e a
    # lista e da MESA: empresa nova do grupo entra pela tela, sem release. Sem
    # linha aqui, ninguem entra — o e-mail so sai se o dia tiver operacao de
    # alguem desta lista.
    #
    # O casamento tenta TRES identificadores, nesta ordem: CNPJ (so digitos),
    # SPN e, por ultimo, o nome por tokens. O CNPJ vem primeiro porque e o
    # unico que nao muda de grafia — o mesmo cliente chega como 'NESTLE BRASIL
    # LTDA' num arquivo e 'NESTLE BRASIL LTDA.' noutro, e o SPN as vezes vem
    # vazio. Basta UM dos tres casar.
    'mt300': {
        'label': 'MT300',
        'columns': [
            # Os tres ligados ao Reference Data: preencher UM preenche os
            # outros dois (ver `wireRefdata` no mapping.html). O dominio e
            # ABERTO de proposito — cliente que ainda nao esta no RefData tem de
            # poder ser cadastrado aqui.
            {'key': 'COUNTERPARTY', 'label': 'Counterparty', 'refdata': 'name'},
            {'key': 'SPN', 'label': 'SPN', 'refdata': 'spn'},
            {'key': 'CNPJ', 'label': 'CNPJ', 'refdata': 'taxid'},
            {'key': 'NOTES', 'label': 'Notes'},
        ],
        'seed': [
            {'COUNTERPARTY': 'ABB AUTOMACAO LTDA', 'SPN': '2156027',
             'CNPJ': '33.449.965/0001-15', 'NOTES': ''},
            {'COUNTERPARTY': 'ABB ELETRIFICACAO LTDA', 'SPN': '2155922',
             'CNPJ': '33.449.988/0001-20', 'NOTES': ''},
            {'COUNTERPARTY': 'CHOCOLATES GAROTO SA', 'SPN': '8837805',
             'CNPJ': '28.053.619/0001-83', 'NOTES': ''},
            {'COUNTERPARTY': 'NESTLE BRASIL LTDA', 'SPN': '806544',
             'CNPJ': '60.409.075/0001-52', 'NOTES': ''},
            {'COUNTERPARTY': 'NESTLE NORDESTE ALIMENTOS E BEBIDAS LTDA', 'SPN': '8937851',
             'CNPJ': '08.334.818/0001-52', 'NOTES': ''},
            {'COUNTERPARTY': 'NESTLE WATERS BRASIL - BEBIDAS E ALIMENTOS LTDA',
             'SPN': '8937847', 'CNPJ': '33.062.464/0001-81', 'NOTES': ''},
        ],
    },
    'settlement-exception': {
        'label': 'Settlement Exception',
        'columns': [
            {'key': 'COUNTERPARTY', 'label': 'Counterparty (all words must match)'},
            {'key': 'PAY', 'label': 'Pay (blank = Reference Data)', 'type': 'select',
             'options': ['', 'No Net', 'Pay/Rec', 'Total Net']},
            {'key': 'RECEIVE', 'label': 'Receive (blank = Reference Data)', 'type': 'select',
             'options': ['', 'No Net', 'Pay/Rec', 'Total Net']},
            {'key': 'NOTES', 'label': 'Notes'},
        ],
        'seed': [
            {'COUNTERPARTY': 'GOBAIN CANALIZACAO', 'PAY': 'Pay/Rec', 'RECEIVE': 'No Net',
             'NOTES': 'Cada recebimento por si; os pagamentos numa perna so'},
            {'COUNTERPARTY': 'GOBAIN BRASIL', 'PAY': 'No Net', 'RECEIVE': 'No Net',
             'NOTES': 'Cada perna por si nos dois sentidos'},
            {'COUNTERPARTY': 'MONDELEZ BRASIL NORTE NORDESTE', 'PAY': 'Pay/Rec',
             'RECEIVE': 'No Net', 'NOTES': ''},
            {'COUNTERPARTY': 'MONDELEZ BRASIL', 'PAY': 'Pay/Rec', 'RECEIVE': 'No Net',
             'NOTES': ''},
        ],
    },
    'gdt-codes': {
        'label': 'GDT Codes',
        'columns': [
            {'key': 'DESCRIPTION', 'label': 'Description'},
            {'key': 'CODE', 'label': 'Code (nHistorico)'},
            # Domínio FECHADO: um produto digitado errado não dá erro — ele cria
            # um grupo que não existe do outro lado, e a linha vira uma pendência
            # que ninguém explica. A opção vazia é a primeira e aparece como "—".
            {'key': 'PRODUCT', 'label': 'Product (blank = ignore)', 'type': 'select',
             'options': ['', 'NDF', 'COMM TER', 'COMM OPT', 'SWAP', 'FXO', 'EQUITIES']},
        ],
        'seed': [
            {'DESCRIPTION': 'DEBITO NDF MANUAL', 'CODE': '9409', 'PRODUCT': 'NDF'},
            {'DESCRIPTION': 'CREDITO NDF MANUAL', 'CODE': '4407', 'PRODUCT': 'NDF'},
            {'DESCRIPTION': 'DEBITO NDF COMMODITIES MANUAL', 'CODE': '9410', 'PRODUCT': 'COMM TER'},
            {'DESCRIPTION': 'CREDITO NDF COMMODITIES MANUAL', 'CODE': '4408', 'PRODUCT': 'COMM TER'},
            {'DESCRIPTION': 'DEBITO OPCAO COMMODITIES MANUAL', 'CODE': '9411', 'PRODUCT': 'COMM OPT'},
            {'DESCRIPTION': 'CREDITO OPCAO COMMODITIES MANUAL', 'CODE': '4409', 'PRODUCT': 'COMM OPT'},
            {'DESCRIPTION': 'DEBITO SWAP AUTOMATICO', 'CODE': '4406', 'PRODUCT': 'SWAP'},
            {'DESCRIPTION': 'EDG/SWAP CEM DEBITO', 'CODE': '9385', 'PRODUCT': 'SWAP'},
            {'DESCRIPTION': 'EDG/SWAP CEM CREDITO', 'CODE': '4413', 'PRODUCT': 'SWAP'},
            {'DESCRIPTION': 'DEBITO TSS-FX', 'CODE': '9386', 'PRODUCT': 'FXO'},
            {'DESCRIPTION': 'CREDITO TSS-FX', 'CODE': '4414', 'PRODUCT': 'FXO'},
            {'DESCRIPTION': 'ESTORNO DEBITO TSS-FX', 'CODE': '9396', 'PRODUCT': 'FXO'},
            {'DESCRIPTION': 'ESTORNO CREDITO TSS-FX', 'CODE': '4424', 'PRODUCT': 'FXO'},
            # As duas transferências entre contas: documentadas, sem produto —
            # não são liquidação e não entram na recon.
            {'DESCRIPTION': 'DEB. TRANSF CTAS MM TITULARIDADE', 'CODE': '5347', 'PRODUCT': ''},
            {'DESCRIPTION': 'CRED. TRANS ENTRE CONTAS', 'CODE': '0159', 'PRODUCT': ''},
        ],
    },
    # ── DCE: os domínios do registro (Código de Referência → Domínio), um
    #    mapping por Campo da planilha "Mapping DCE". Os seeds vão VAZIOS de
    #    propósito, como os dois Quotes: os JSONs são versionados (o Underlying
    #    Asset tem ~14 mil linhas) e repeti-los aqui criaria uma segunda lista
    #    para divergir da primeira. O `lang` das colunas é a chave i18n que o
    #    mapping.html usa no cabeçalho, no filtro, no export e no modal.
    'dce-country': {
        'label': 'DCE — Country',
        'columns': [
            {'key': 'REFERENCE CODE', 'label': 'Reference Code', 'lang': 'map-col-dce-ref-code'},
            {'key': 'DOMAIN', 'label': 'Domain', 'lang': 'map-col-dce-domain'},
        ],
        'seed': [],
    },
    'dce-type-of-derivative': {
        'label': 'DCE — Type of Derivative',
        'columns': [
            {'key': 'REFERENCE CODE', 'label': 'Reference Code', 'lang': 'map-col-dce-ref-code'},
            {'key': 'DOMAIN', 'label': 'Domain', 'lang': 'map-col-dce-domain'},
        ],
        'seed': [],
    },
    'dce-type-of-swap': {
        'label': 'DCE — Type of Swap',
        'columns': [
            {'key': 'REFERENCE CODE', 'label': 'Reference Code', 'lang': 'map-col-dce-ref-code'},
            {'key': 'DOMAIN', 'label': 'Domain', 'lang': 'map-col-dce-domain'},
        ],
        'seed': [],
    },
    'dce-type-of-verification': {
        'label': 'DCE — Type of Verification',
        'columns': [
            {'key': 'REFERENCE CODE', 'label': 'Reference Code', 'lang': 'map-col-dce-ref-code'},
            {'key': 'DOMAIN', 'label': 'Domain', 'lang': 'map-col-dce-domain'},
        ],
        'seed': [],
    },
    'dce-functionality': {
        'label': 'DCE — Functionality',
        'columns': [
            {'key': 'REFERENCE CODE', 'label': 'Reference Code', 'lang': 'map-col-dce-ref-code'},
            {'key': 'DOMAIN', 'label': 'Domain', 'lang': 'map-col-dce-domain'},
        ],
        'seed': [],
    },
    'dce-underlying-asset-category': {
        'label': 'DCE — Underlying Asset Category',
        'columns': [
            {'key': 'REFERENCE CODE', 'label': 'Reference Code', 'lang': 'map-col-dce-ref-code'},
            {'key': 'DOMAIN', 'label': 'Domain', 'lang': 'map-col-dce-domain'},
        ],
        'seed': [],
    },
    'dce-underlying-asset': {
        'label': 'DCE — Underlying Asset',
        'columns': [
            {'key': 'CATEGORY', 'label': 'Category', 'lang': 'map-col-dce-category'},
            {'key': 'ATHENA INDEX NAME', 'label': 'Athena Index Name', 'lang': 'map-col-dce-athena'},
            {'key': 'REFERENCE CODE', 'label': 'Reference Code', 'lang': 'map-col-dce-ref-code'},
            {'key': 'DOMAIN', 'label': 'Domain', 'lang': 'map-col-dce-domain'},
        ],
        'seed': [],
    },
    # ── Onboarding e Reconciliação de CGD ───────────────────────────────────
    #
    # Os quatro cadastros que a esteira do CGD usa. Os três últimos eram abas do
    # `Auxiliar.xlsx` numa pasta de rede, mantidas à mão: cadastro novo só valia
    # para quem tivesse o arquivo aberto, e o batimento rodava com a lista de
    # ontem sem dizer nada.
    'cgd-stage': {
        'label': 'Onboarding — Status → Mesa',
        'columns': [
            {'key': 'STATUS', 'label': 'Status (como está na lista do SharePoint)'},
            {'key': 'STAGE', 'label': 'Mesa', 'type': 'select',
             'options': list(_cgd_mod.STAGES)},
            {'key': 'NOTES', 'label': 'Notes'},
        ],
        # Seed VAZIO de propósito: os status são texto livre de quem opera a
        # lista, e semear com os de hoje seria fixar no código o de-para que esta
        # tela existe para não fixar. Sem linha, o Overview DERIVA a mesa pelos
        # carimbos que o documento tem — e marca o item como derivado.
        'seed': [],
    },
    # As colunas seguem a ORDEM e os nomes da aba `Mapping` do Auxiliar.xlsx
    # (A = Razão Social, B = Nome Simplificado, C = CNPJ, D = Conta): é a
    # planilha de onde a lista vem, e a tela na mesma ordem é o que deixa o
    # copiar-e-colar conferível coluna a coluna. As CHAVES são as antigas — o
    # motor da recon (`recon_cgd._participantes`) e os JSONs já gravados leem
    # por elas, e renomeá-las apagaria o cadastro existente em silêncio.
    'cgd-b3-participante': {
        'label': 'CGD — B3 Participants',
        'columns': [
            {'key': 'RAZAO SOCIAL', 'label': 'Participant (Legal Name)',
             'lang': 'map-col-cgd-part-name'},
            {'key': 'NOME CONTRAPARTE', 'label': 'Participant (Short Name)',
             'lang': 'map-col-cgd-part-short'},
            {'key': 'CNPJ', 'label': 'Participant (Tax ID)',
             'lang': 'map-col-cgd-part-cnpj'},
            {'key': 'CONTA', 'label': 'Participant (Account)',
             'lang': 'map-col-cgd-part-account'},
        ],
        'seed': [],
    },
    'cgd-garantidor': {
        'label': 'CGD — Guarantors',
        'columns': [
            {'key': 'CNPJ / CPF', 'label': 'Tax ID (CNPJ / CPF)',
             'lang': 'map-col-cgd-guar-cnpj'},
            {'key': 'EMPRESA', 'label': 'Company', 'lang': 'map-col-cgd-guar-company'},
            {'key': 'CLIENTE', 'label': 'Client', 'lang': 'map-col-cgd-guar-client'},
        ],
        # O formato antigo tinha a coluna NOME: vira EMPRESA na leitura, senão o
        # cadastro existente abriria com a coluna vazia e o primeiro Save (que
        # reescreve o arquivo inteiro) apagaria o nome de todas as linhas.
        'upgrade': lambda rows: [
            dict(r, **{'EMPRESA': r.get('EMPRESA') or r.get('NOME') or ''})
            for r in rows],
        'seed': [],
    },
    'cgd-conta-encerrada': {
        'label': 'CGD — Closed Accounts',
        'columns': [
            {'key': 'CNPJ / CPF', 'label': 'Tax ID (CNPJ / CPF)',
             'lang': 'map-col-cgd-closed-cnpj'},
            {'key': 'NOME', 'label': 'Name', 'lang': 'map-col-cgd-closed-name'},
            {'key': 'NOTES', 'label': 'Notes', 'lang': 'map-col-notes'},
        ],
        'seed': [],
    },
}

_mapping_cache = {}


def _mapping_path(key):
    d = _MAPPING_DEFS[key]
    return d.get('file') or os.path.join(_MAPPINGS_DIR, key + '.json')


def _mapping_rows(key):
    """Linhas do mapping `key` (lista de dicts). Cria o arquivo com o SEED na
    primeira leitura; cacheia por mtime, então edição pela tela vale na
    requisição seguinte."""
    d = _MAPPING_DEFS.get(key)
    if not d:
        return []
    path = _mapping_path(key)
    if not os.path.isfile(path):
        # Semear sob o lock: dois requests simultâneos na primeira leitura
        # gravariam o mesmo arquivo ao mesmo tempo. O re-teste de existência
        # dentro do lock evita a segunda escrita.
        with _cache_lock:
            if not os.path.isfile(path):
                try:
                    os.makedirs(os.path.dirname(path), exist_ok=True)
                    _atomic_write_json(path, list(d.get('seed') or []))
                except Exception:
                    log.warning('[mappings] seed write failed for %s:\n%s', key, traceback.format_exc())
                    return list(d.get('seed') or [])
    try:
        mtime = os.path.getmtime(path)
        cached = _mapping_cache.get(key)
        if cached and cached[0] == mtime:
            return cached[1]
        # DB-first (fase 3): o banco DESTE cadastro (`mappings_<key>.db`, ou o
        # `<arquivo>.db` da raiz para os registros com `file`) quando o
        # manifest prova o frescor — via `_raw`, então o conteúdo é o do
        # arquivo byte a byte e o `upgrade` abaixo roda igual nas duas fontes.
        # Senão, o JSON.
        rows = None
        try:
            from apps.pages import duck_read
            rows = duck_read.dataset_records(path)
        except Exception:                                   # noqa: BLE001
            rows = None
        if rows is None:
            with open(path, encoding='utf-8') as fh:
                rows = json.load(fh) or []
        if not isinstance(rows, list):
            rows = []
        rows = [r for r in rows if isinstance(r, dict)]
        up = d.get('upgrade')
        if up:
            rows = up(rows)
        _mapping_cache[key] = (mtime, rows)
        return rows
    except Exception:
        return list(d.get('seed') or [])


def _b3_code_matches(pattern, code):
    """O Ativo Subjacente `code` corresponde a esta linha de Commodities × B3?

    A coluna B3 CODE guarda um LITERAL nas linhas FIXED ('NACX0005') e um PADRÃO
    nas PREFIX ('HO"MY"', 'C_"MY"', 'KO"MY"BNMK', §164). O padrão casa por
    prefixo + sufixo, com pelo menos um caractere de mês/ano no meio — senão
    'HO' casaria com 'HO' pelado, que não é código de contrato nenhum."""
    from apps.pages.otc_boxparse import split_b3_pattern
    pat = str(pattern or '').strip()
    cod = str(code or '').strip().upper()
    if not pat or not cod:
        return False
    if '"' not in pat and '_' not in pat:
        return pat.upper() == cod
    head, tail = split_b3_pattern(pat)
    head, tail = head.upper(), tail.upper()
    return (cod.startswith(head) and cod.endswith(tail)
            and len(cod) > len(head) + len(tail))


def _b3_quote_cfg(code):
    """Tipo de Cotação e Fonte de Informação do Ativo Subjacente, do cadastro.

    Saem das colunas do mapping Commodities × B3 (§177) — coluna vazia, ou
    subjacente sem linha, devolve o default histórico ('A' / '5' / '358').
    O flag FIXED QUOTE que escolhia F/340 foi aposentado (§252): esses valores
    estão materializados nas colunas das linhas que eram YES.

    Devolve as três de uma vez porque a linha é a mesma: procurá-la três vezes
    percorreria o cadastro inteiro a cada campo, para cada deal do arquivo."""
    row = {}
    for r in _mapping_rows('commodities-b3'):
        if _b3_code_matches(r.get('B3 CODE'), code):
            row = r
            break
    return {
        'ndf':    str(row.get('QUOTE TYPE NDF', '') or '').strip() or 'A',
        'opt':    str(row.get('QUOTE TYPE OPT', '') or '').strip() or _B3_QUOTE_OPT_DEFAULT,
        'source': str(row.get('INFO SOURCE', '') or '').strip() or '358',
    }


def _mapping_ccy_maps():
    """(athena→iso, ISO fracas, ISO→casas da inversão) da Currency Base — o ISO
    é o SIMBOLO da própria linha da moeda."""
    ath, weak, inv = {}, set(), {}
    for r in _mapping_rows('currency-base'):
        a = str(r.get('ATHENA CODE', '') or '').strip().upper()
        i = str(r.get('SIMBOLO', '') or '').strip().upper()
        if a and i and a not in ath:
            ath[a] = i
        if i and str(r.get('WEAK', '') or '').strip().upper() in ('YES', 'Y', 'SIM', 'S', 'TRUE', '1'):
            weak.add(i)
        dec = str(r.get('INV DECIMALS', '') or '').strip()
        if i and dec.isdigit():
            inv[i] = int(dec)
    return ath, weak, inv




def _refdata_by_name(refmap_spn=None):
    """Razão social normalizada (`_pc_norm`) → registro do Reference Data. Deriva
    do índice por SPN quando ele é passado, para não reler o arquivo."""
    out = {}
    for rec in (refmap_spn if refmap_spn is not None else _fxo_refdata_by_spn()).values():
        nm = _pc_norm(rec.get('COUNTERPARTY', ''))
        if nm and nm not in out:
            out[nm] = rec
    return out




# ══════════════════════════════════════════════════════════════════════════
#  COTAÇÕES (Apps › Quotes) — PTAX, Equities e Commodities
# ══════════════════════════════════════════════════════════════════════════
#
# Saiu daqui: virou a vertical `apps/pages/features/quotes/`. O MOTOR continua
# no `apps/pages/quotes.py` — a sessão Kerberos, o proxy e o de-para com padrão
# `"MY"`, com o `check_quotes.py` em cima. Ver a §10 do CLAUDE.md.


# ═════════════════════════════════════════════════════════════════════════════
#  FILE INTERPRETER — os layouts dos arquivos B3 (Batch Conecta) como templates
#  editáveis. Um JSON por template em static/data/file-interpreter/ (versionados:
#  os seeds transcrevem o manual "Transferência de Arquivos – Enviar Arquivos").
#  Nada de layout fixo no código: a página monta os blocos do que o JSON disser,
#  e template novo/atualizado da B3 entra pela tela (Create New Template).
# ═════════════════════════════════════════════════════════════════════════════
_FILE_INTERPRETER_DIR = os.path.normpath(os.path.join(
    data_dir(), 'file-interpreter'))

# ── File Interpreter: motor movido para platform/file_interpreter.py (§318) ──
# Os nomes ficam como ALIAS; o _FILE_INTERPRETER_DIR fica AQUI porque e a
# superficie de patch dos check_fi_* (a platform o le por routes.<nome>).
from apps.pages.platform import file_interpreter as _pf_fi  # noqa: E402
_FI_LEGACY_DIR = _pf_fi._FI_LEGACY_DIR
_FI_KEY_RE = _pf_fi._FI_KEY_RE
_FI_FIELD_KEYS = _pf_fi._FI_FIELD_KEYS
_FI_META_KEYS = _pf_fi._FI_META_KEYS
_fi_path = _pf_fi._fi_path
_fi_load = _pf_fi._fi_load
_fi_clean_template = _pf_fi._fi_clean_template
_fi_tpl_cache = _pf_fi._fi_tpl_cache
_fi_tpl_cached = _pf_fi._fi_tpl_cached
_FI_LE_PAIRS = _pf_fi._FI_LE_PAIRS
_fi_le_pair_norm = _pf_fi._fi_le_pair_norm
_fi_variant_key = _pf_fi._fi_variant_key
_fi_variant_file_name = _pf_fi._fi_variant_file_name
_fi_width = _pf_fi._fi_width
_fi_field_src = _pf_fi._fi_field_src
_fi_seq_key = _pf_fi._fi_seq_key
_FI_CALC_RE = _pf_fi._FI_CALC_RE
_fi_deal_get = _pf_fi._fi_deal_get
_fi_calc_value = _pf_fi._fi_calc_value
_fi_effective_seq_value = _pf_fi._fi_effective_seq_value
_fi_block_of = _pf_fi._fi_block_of
_fi_build_line = _pf_fi._fi_build_line
try:
    if os.path.isdir(_FI_LEGACY_DIR):
        os.makedirs(_FILE_INTERPRETER_DIR, exist_ok=True)
        for _fn in os.listdir(_FI_LEGACY_DIR):
            _dst = os.path.join(_FILE_INTERPRETER_DIR, _fn)
            if _fn.endswith('.json') and not os.path.exists(_dst):
                shutil.move(os.path.join(_FI_LEGACY_DIR, _fn), _dst)
except OSError:
    pass





_schedule_on_start('ndf', _ndf_api_start_scheduler)


# ==============================================================================
# API — NDF COMMODITIES CACHE (mesma lógica que opt-commodities, arquivo _ndfcomm.json)
# ==============================================================================

NDF_COMM_CACHE_DIR = os.path.normpath(os.path.join(
    data_dir(), "cache", "new deals", "NDF", "Commodities"
))

NEW_DEALS_CACHE_ROOT = os.path.normpath(os.path.join(
    data_dir(), "cache", "new deals"
))


# Network share where generated Intrag NDF .txt files are written; pende da
# raiz do share (Config.SHARED_DRIVE_ROOT), como os demais destinos.

# English month names for the "mm. Mmmm" folder (e.g. "06. June") — fixed list
# so the folder name never depends on the server locale.
_EN_MONTH_NAMES = _pf_dates._EN_MONTH_NAMES

# ── ANBIMA calendar ───────────────────────────────────────────────────────────
# O calendário mora em `apps/pages/platform/anbima.py` (a primeira fatia da
# fase platform/ — CLAUDE.md §10). Os nomes ficam aqui como ALIAS: as features
# alcançam por `routes.<nome>` (andaime declarado) e os testes que trocam a
# FUNÇÃO no `routes` seguem valendo. O ESTADO (`_ANBIMA_HOLIDAYS`,
# `_anbima_loaded`, `_anbima_hols_cache`) mora LÁ — um alias de set apontaria
# para o objeto velho quando `_load_anbima` rebinda o global.
from apps.pages.platform import anbima as _pf_anbima  # noqa: E402

_load_anbima = _pf_anbima._load_anbima
_prev_anbima_bizday = _pf_anbima._prev_anbima_bizday
_anbima_bizdays_between = _pf_anbima._anbima_bizdays_between
_weekday_bizdays_between = _pf_anbima._weekday_bizdays_between
_last_anbima_bizday_of_month = _pf_anbima._last_anbima_bizday_of_month
_pcx_is_bizday = _pf_anbima._pcx_is_bizday
_anbima_holidays = _pf_anbima._anbima_holidays
_anbima_biz_diff = _pf_anbima._anbima_biz_diff
_anbima_add_biz = _pf_anbima._anbima_add_biz

# ── Subjacente.json lookup (keyed by Codigo do Ativo Subjacente, first match) ──

def _load_subjacente_lookup():
    try:
        fp = data_path('Subjacente.json')
        with open(fp, 'r', encoding='utf-8') as fh:
            rows = json.load(fh)
        result = {}
        for row in rows:
            code = (row.get('Codigo do Ativo Subjacente') or '').strip().upper()
            if code and code not in result:
                result[code] = row
        return result
    except Exception as exc:
        log.warning('[SUBJACENTE] Failed to load: %s', exc)
        return {}


# Cache por mtime, não dict de módulo: o Subjacente.json é editado pela tela
# Index B3, e um dict carregado na subida servia o cadastro VELHO até o
# próximo restart — com o Maturity Month/Year da Intrag saindo dele, um código
# recém-cadastrado tem de valer no request seguinte (mesma regra dos mappings).
_subjacente_cache = {'mtime': None, 'data': {}}


def _subjacente_by_code():
    fp = data_path('Subjacente.json')
    try:
        mtime = os.path.getmtime(fp)
    except OSError:
        return _subjacente_cache['data']
    if _subjacente_cache['mtime'] != mtime:
        _subjacente_cache['data'] = _load_subjacente_lookup()
        _subjacente_cache['mtime'] = mtime
    return _subjacente_cache['data']

_MONTH_ABBR = {
    'JAN': '01', 'FEB': '02', 'MAR': '03', 'APR': '04',
    'MAY': '05', 'JUN': '06', 'JUL': '07', 'AUG': '08',
    'SEP': '09', 'OCT': '10', 'NOV': '11', 'DEC': '12',
}

# Intra-group accounts: 73760.00-9 = Banco J.P. Morgan, 00041.00-7 = Lawton.


# ── Intrag ID mapping (Return-folder export CSV → fill each entry's intrag_id) ──
# The Return folder holds a single Boletas CSV (Boletas.csv / Boletas(1).csv / …) with
# ALL operations and NO header — match by row content, not column names:
#   • Option: col C (idx 2) == 'OPCAO'                → col I (idx 8) = B3 ID, col A (idx 0) = Intrag ID
#   • NDF:    col B (idx 1) == 'NDF - TERMO MERCADORIA'→ col C (idx 2) = B3 ID, col A (idx 0) = Intrag ID


# ── Intrag Swap ───────────────────────────────────────────────────────────────
# Mesmo modelo da Intrag NDF: entradas persistidas em day-files
# YYYYMMDD_intrag_swap.json, ciclo New → Pending → Approved → Sent (4-eyes) e
# geração de .txt (36 colunas do layout B3 de swap, separadas por ';') na mesma
# pasta de rede da NDF. Não há feed automático — as linhas são criadas na página.




# ==============================================================================
# API — B3 JSON CRUD (Subjacente / VCP / Domínio / RefData)
# ==============================================================================

# O `/static/data/...` do NAVEGADOR também sai do `DATA_DIR`
# ─────────────────────────────────────────────────────────
# Setenta e um `fetch` espalhados por quinze telas leem JSON por URL estática —
# `RefData.json`, `Subjacente.json`, `anbima.json`, as agendas de feriado, os
# cadastros do /mapping. Como URL estática, o Flask os serve da pasta do CÓDIGO,
# e é aí que a regra do `data_paths` era furada pela ponta que ela não cobre: o
# servidor lê e grava no `DATA_DIR` (o share, na instância do JPM) e a TELA lia
# o checkout.
#
# Na dev as duas pastas são a mesma e nada aparece. Na instância do JPM não são,
# e o efeito é o pior tipo de defeito: a mesa edita o Reference Data pela tela,
# o app grava no share, a tela recarrega — e mostra a cópia versionada, de antes
# do último `git pull`. Nenhum erro, dois arquivos, e a edição que "não salvou"
# está salva no lugar certo.
#
# Esta rota é mais específica que o `/static/<path:filename>` embutido, então
# ganha dele no roteamento, e resolve pelo MESMO `data_path()` do servidor —
# `DATA_DIR` primeiro, cópia empacotada como queda. Quem não tem `DATA_DIR`
# separado (a dev) não vê diferença nenhuma: o caminho resolvido é o mesmo.
@blueprint.route('/static/data/<path:filename>')
def static_data_file(filename):
    """Serve `static/data/...` pelo `DATA_DIR`, com queda para o empacotado.

    A RAIZ e o caminho RELATIVO vão separados para o `send_from_directory` de
    propósito: quem recusa sair da pasta é o `safe_join` que ele faz por dentro,
    e ele só tem como recusar se enxergar o `..` que veio na URL. Resolver o
    caminho aqui e passar `dirname`/`basename` já resolvidos anula essa
    checagem — a pasta traversada VIRA a raiz permitida, e
    `/static/data/../../config.py` passa a servir o config.

    **Os JSONs cobertos pelos bancos saem SERVIDOS DO BANCO quando ele está
    fresco** (fase 3, HANDOFF §330): é o flip de leitura do NAVEGADOR — os
    `fetch` de RefData, CounterpartyDetails e dos arquivos de calendário
    respondem pelo DuckDB sem mudar uma linha de JS. Qualquer dúvida (banco
    frio, arquivo não coberto, subpasta) cai no arquivo, como sempre foi.
    """
    resp = _duck_static_json(filename)
    if resp is not None:
        return resp
    for raiz in (data_dir(), PACKAGED_DATA_DIR):
        try:
            return send_from_directory(raiz, filename)
        except NotFound:
            continue                      # não está no DATA_DIR: tenta a cópia do repo
    raise NotFound()


# Os arquivos de primeiro nível que NÃO são calendário de feriado — poupa a
# consulta ao registro no fetch dos pesados (o Subjacente tem 4 MB). A lista
# pode envelhecer sem quebrar nada: um nome fora dela só paga uma consulta ao
# registro que devolve "não é calendário".
_DUCK_STATIC_NAO_CALENDARIO = frozenset({
    'Subjacente.json', 'VCP.json', 'Dominio.json', 'SwapIndex.json',
    'datatables-rendering.json', 'datatables.json', 'treeview-data.json',
    'typeahead-data-2.json', 'typeahead.json', 'holiday-calendars.json',
})


def _duck_static_json(filename):
    """A resposta do banco para um `/static/data/<arquivo>` coberto — ou `None`.

    Melhor esforço de ponta a ponta: este caminho nunca pode ser a razão de um
    fetch falhar, então toda exceção vira `None` e o arquivo é servido do
    disco."""
    try:
        nome = str(filename or '').replace('\\', '/').strip('/')
        if not nome.endswith('.json'):
            return None
        from apps.pages import duck_read
        rows = None
        if '/' in nome:
            # Subpasta: só os cadastros do /mapping — o resto (translations,
            # file-interpreter) fica com o arquivo.
            if nome.startswith('mappings/') and nome.count('/') == 1:
                rows = duck_read.dataset_records(
                    os.path.join(_B3_DATA_DIR, *nome.split('/')))
            else:
                return None
        elif nome == 'RefData.json':
            rows = duck_read.refdata_rows()
        elif nome == 'CounterpartyDetails.json':
            rows = duck_read.cpd_records()
        else:
            if nome not in _DUCK_STATIC_NAO_CALENDARIO:
                from apps.pages.features.holidays.infra import persistence as _hp
                if any(str(r.get('file', '') or '').strip().lower() == nome.lower()
                       for r in _hp.calendars()):
                    rows = _hp._load_holidays_db(nome)
            if rows is None:
                # Qualquer outro JSON de raiz coberto pelos DATASETS
                # (Subjacente, Dominio, VCP, SwapIndex, …) — lista de
                # registros sai do banco; payload que não é lista fica com o
                # arquivo (dataset_records devolve None).
                rows = duck_read.dataset_records(os.path.join(_B3_DATA_DIR, nome))
        if rows is None:
            return None
        from flask import Response
        return Response(json.dumps(rows, ensure_ascii=False),
                        mimetype='application/json')
    except Exception:                                       # noqa: BLE001
        return None


_B3_DATA_DIR = data_dir()
_B3_FILE_MAP = {
    'subj':      'Subjacente.json',
    'vcp':       'VCP.json',
    'dominio':   'Dominio.json',
    'refdata':   'RefData.json',
    'swapindex': 'SwapIndex.json',
}


def _b3_load(table):
    path = os.path.join(_B3_DATA_DIR, _B3_FILE_MAP[table])
    # DB-first (fase 3): o refdata pelo reference_data.db, os demais quatro
    # (Subjacente, VCP, Dominio, SwapIndex) pelo banco de CADA UM
    # (`subjacente.db`, `vcp.db`, …) — quando o manifest prova o frescor. O
    # caminho devolvido segue sendo o do JSON: é
    # nele que o _b3_save grava, e o espelho realinha o banco em seguida.
    try:
        from apps.pages import duck_read
        rows = (duck_read.refdata_rows() if table == 'refdata'
                else duck_read.dataset_records(path))
    except Exception:                                       # noqa: BLE001
        rows = None
    if rows is not None:
        return rows, path
    with open(path, encoding='utf-8') as fh:
        return json.load(fh), path


def _b3_save(path, records):
    # Pelo FUNIL (auditoria §335): atômico e com o espelho avisado de graça —
    # o notify manual da fase 2 saiu junto com o write cru.
    _atomic_write_json(path, records)


# ── Holidays Calendar ────────────────────────────────────────────────────────
# Saiu daqui: virou a vertical `apps/pages/features/holidays/`. O
# `_anbima_holidays` NÃO foi junto — ele é o calendário de dias úteis do app
# INTEIRO (SLA da esteira, aging do CGD, schedulers, D-1 das recons), e o lugar
# dele é o `platform/`. Ver a §10 do CLAUDE.md.


# ==============================================================================
# PARSE MSG EMAIL — extrai HTML de arquivo .msg do Outlook
# ==============================================================================


# ==============================================================================
# BOX SCAN AUTOMÁTICO — varredura agendada do box (NDF Comm e Opt Comm)
#
# O botão Import continua existindo e é o caminho manual; este bloco é o mesmo
# trabalho feito sozinho a cada BOX_SCAN_POLL_MIN minutos, sem ninguém com a
# página aberta.
#
# ⚠️ O caminho manual parseia o e-mail NO NAVEGADOR (otc-fileupload.js). Aqui
# quem parseia é `otc_boxparse`, um porte da MESMA regra para Python — duas
# cópias da mesma lógica, com a armadilha que isso implica (HANDOFF §121). O que
# as mantém honestas é `scratchpad/check_boxparse.py`, que roda o JS de verdade
# no JavaScriptCore e compara campo a campo. Mexeu num lado, rode-o.
# ==============================================================================


# Maker sintético: ninguém humano importou. Mesma convenção do pull da Athena
# (que grava 'API'), e é o que mantém a trava de quatro olhos válida — qualquer
# usuário pode aprovar um deal que a máquina trouxe.


# ==============================================================================
# SEND TO CONECTA — gera arquivo TXT para B3 Batch Conecta
# ==============================================================================


# ==============================================================================
# MAPPING B3 ID — lê arquivos de retorno do Batch Conecta e atualiza B3_ID
# ==============================================================================


# ==============================================================================
# API — SETTLEMENT / CONFIRMATION E-MAILS (Premium D0 + Economic Affirmation)
# Builds the HTML drafts in apps/pages/otc_emails.py and opens them in Outlook
# for manual review (win32com — Windows/JPM only; degrades gracefully elsewhere).
# ==============================================================================
_email_drafts_response = _pf_mail._email_drafts_response


# ==============================================================================
# API — COUNTERPARTY DETAILS (Reference Data double-click editor)
# Persists CGD / Banking (PAY+RECEIVE) / Contacts to CounterpartyDetails.json,
# keyed by SPN. Replaces (or appends) the record for the edited counterparty.
# ==============================================================================




# Migra o CounterpartyDetails.json já na subida do app (e não só no primeiro
# request), para o modal do Reference Data — que lê o JSON estático direto —
# encontrar os ids estáveis desde o primeiro acesso.
try:
    _cpd_load()
except Exception:
    log.warning('[counterparty-details] startup migration failed:\n%s',
                traceback.format_exc())


# ──────────────────────────────────────────────────────────────────────────
# CGD — maker/checker (Pending → Active). Item: {id,value,status,maker,checker}
# ──────────────────────────────────────────────────────────────────────────


# ──────────────────────────────────────────────────────────────────────────
# SETTLEMENT NET TYPE — single value per counterparty, maker/checker.
# Editing proposes a change (→ Pending); a different SID approves (→ Active).
# ──────────────────────────────────────────────────────────────────────────


# ──────────────────────────────────────────────────────────────────────────
# CONTACTS — maker/checker (Pending → Active). Approval lives in `appr`;
# `status` stays the business Active/Inactive value.
# Item: {id,name,phone,email,rules,status,appr,maker,checker}
# ──────────────────────────────────────────────────────────────────────────


# ==============================================================================
# API — GENERIC NEW-DEALS CACHE (NDF FWD Start / NDF Other Publisher)
# Page + CRUD only. Same Deal+Client keyed JSON cache model as ndf-commodities.
# Import-parse / mapping-B3 / send-Conecta are product-specific and intentionally
# not wired here (handled per-product later).
# ==============================================================================

_GENERIC_ND_PRODUCTS = {
    'fwd-start': {
        'dir':    os.path.join(NEW_DEALS_CACHE_ROOT, 'NDF', 'FwdStart'),
        'suffix': '_ndffwdstart.json',
        'label':  'NDF FWD Start',
    },
    'other-publishers': {
        'dir':    os.path.join(NEW_DEALS_CACHE_ROOT, 'NDF', 'OtherPublisher'),
        'suffix': '_ndfotherpub.json',
        'label':  'NDF Other Publisher',
    },
    'vanilla': {
        'dir':    os.path.join(NEW_DEALS_CACHE_ROOT, 'NDF', 'Vanilla'),
        'suffix': '_ndfvanilla.json',
        'label':  'NDF Vanilla',
    },
}


def _generic_nd_cfg(product):
    return _GENERIC_ND_PRODUCTS.get(product)




# ==============================================================================
# NEW DEALS — CONFIRMATIONS (NDF Commodities · Termo de Mercadoria)
# Portado da macro legada (TERMO.doc): as confirmações são segregadas por
# contraparte × mercadoria × família de moeda do strike, POR trade date, sempre
# excluindo as pontas internas (Client = Banco J.P. Morgan ou Lawton — as
# pernas banco×lawton/lawton×banco não geram confirmação de cliente). Por
# enquanto só o template "strike em USD" existe na web (os demais — BRL,
# PLATTS, Palm Oil — chegam depois); grupos das outras famílias aparecem na
# listagem como "template pendente".
# ==============================================================================

# ── motor de confirmações: movido para platform/confirmations.py (§316) ──────
# Os nomes ficam como ALIAS — features e testes seguem alcançando por
# `routes.<nome>`, e todo chamador resolve o atributo em tempo de chamada.
from apps.pages.platform import confirmations as _pf_conf  # noqa: E402
_CONF_INTERNAL_RE = _pf_conf._CONF_INTERNAL_RE
_CONF_GEN_ELIGIBLE_STATUS = _pf_conf._CONF_GEN_ELIGIBLE_STATUS
_CONF_MONTHS_PT = _pf_conf._CONF_MONTHS_PT
_CONF_FUT_MONTH_CODE = _pf_conf._CONF_FUT_MONTH_CODE
_CONF_TICKER_MAP = _pf_conf._CONF_TICKER_MAP
_conf_subj_cache = _pf_conf._conf_subj_cache
_conf_subjacente_map = _pf_conf._conf_subjacente_map
_conf_fmt_date = _pf_conf._conf_fmt_date
_conf_date_extenso = _pf_conf._conf_date_extenso
_conf_fmt_cnpj = _pf_conf._conf_fmt_cnpj
_conf_to_float = _pf_conf._conf_to_float
_conf_fmt_num = _pf_conf._conf_fmt_num
_conf_prev_biz = _pf_conf._conf_prev_biz
_conf_load_ndfcomm = _pf_conf._conf_load_ndfcomm
_conf_deal_family = _pf_conf._conf_deal_family
CONF_STATE_DIR = _pf_conf.CONF_STATE_DIR
_conf_state_path = _pf_conf._conf_state_path
_conf_key = _pf_conf._conf_key
_conf_state_load = _pf_conf._conf_state_load
_conf_state_save = _pf_conf._conf_state_save
_conf_segregate = _pf_conf._conf_segregate
_conf_ndfcomm_groups = _pf_conf._conf_ndfcomm_groups
_CONF_STAGE_ORDER = _pf_conf._CONF_STAGE_ORDER
_conf_esteira_stages = _pf_conf._conf_esteira_stages
_conf_group_stage = _pf_conf._conf_group_stage
_conf_stage_counts = _pf_conf._conf_stage_counts
_conf_co12_text = _pf_conf._conf_co12_text
_conf_cgd_lookup = _pf_conf._conf_cgd_lookup
_CONF_PALMOIL_BOLSA = _pf_conf._CONF_PALMOIL_BOLSA
_CONF_PALMOIL_TAXA_CONV = _pf_conf._CONF_PALMOIL_TAXA_CONV
_CONF_FAMILY_TEMPLATES = _pf_conf._CONF_FAMILY_TEMPLATES
_conf_sort_key_venc = _pf_conf._conf_sort_key_venc
_conf_pick_eligible = _pf_conf._conf_pick_eligible
_conf_pick_ndfcomm = _pf_conf._conf_pick_ndfcomm
_conf_generation_page = _pf_conf._conf_generation_page
_CONF_CNPJ_BANCO = _pf_conf._CONF_CNPJ_BANCO
_conf_ccy_num = _pf_conf._conf_ccy_num
_conf_ccy_is_brl = _pf_conf._conf_ccy_is_brl
_conf_strike_adj = _pf_conf._conf_strike_adj
_conf_fx_legs = _pf_conf._conf_fx_legs
_conf_ndf_xml = _pf_conf._conf_ndf_xml
_conf_xml_doc = _pf_conf._conf_xml_doc
_conf_pc_set_fepweb = _pf_conf._conf_pc_set_fepweb
_conf_state_entry_or_404 = _pf_conf._conf_state_entry_or_404
_CONF_OPT_FAMILY_TEMPLATES = _pf_conf._CONF_OPT_FAMILY_TEMPLATES
_CONF_OPT_PDF_VARIANT = _pf_conf._CONF_OPT_PDF_VARIANT
_CONF_OPT_PDF_FROM_HTML = _pf_conf._CONF_OPT_PDF_FROM_HTML
_conf_load_optcomm = _pf_conf._conf_load_optcomm
_conf_opt_family = _pf_conf._conf_opt_family
_conf_optcomm_groups = _pf_conf._conf_optcomm_groups
_conf_pick_optcomm = _pf_conf._conf_pick_optcomm
_conf_opt_generation_page = _pf_conf._conf_opt_generation_page
_CONF_FXO_FAMILY_TEMPLATES = _pf_conf._CONF_FXO_FAMILY_TEMPLATES
_conf_load_optfxo = _pf_conf._conf_load_optfxo
_conf_fxo_family = _pf_conf._conf_fxo_family
_conf_optfxo_groups = _pf_conf._conf_optfxo_groups
_conf_pick_optfxo = _pf_conf._conf_pick_optfxo
_conf_fxo_strike = _pf_conf._conf_fxo_strike
_conf_fxo_conv_rate = _pf_conf._conf_fxo_conv_rate
_conf_fxo_generation_page = _pf_conf._conf_fxo_generation_page
_CONF_FWDSTART_FAMILY_TEMPLATES = _pf_conf._CONF_FWDSTART_FAMILY_TEMPLATES
_CONF_FWDSTART_PARTEA = _pf_conf._CONF_FWDSTART_PARTEA
_conf_fwdstart_partea = _pf_conf._conf_fwdstart_partea
_conf_load_ndffwdstart = _pf_conf._conf_load_ndffwdstart
_conf_fwdstart_family = _pf_conf._conf_fwdstart_family
_conf_fwdstart_moeda = _pf_conf._conf_fwdstart_moeda
_conf_fwdstart_groups = _pf_conf._conf_fwdstart_groups
_conf_pick_fwdstart = _pf_conf._conf_pick_fwdstart
_conf_fwdstart_rows = _pf_conf._conf_fwdstart_rows




# A validação do OTC na esteira NÃO acontece mais aqui. O checklist do New Deals
# fecha o ciclo do DOCUMENTO (New → Generated → Success): ele diz que o papel foi
# gerado e conferido por quem o montou. Carimbá-lo também como a validação do OTC
# fazia a confirmação nascer já na mesa seguinte, e a fila do OTC no Monitor
# ficava vazia por construção — a mesa não tinha onde conferir o que ela mesma
# acabara de emitir, com o prazo de D+3 correndo em silêncio.
#
# Gerar é gravar. A etapa do OTC é carimbada uma vez só, em
# `/manual-confirmation/validate`, aberto pelo Monitor.




# ==============================================================================
# NEW DEALS — MONITOR
# Visão por produto × status das operações importadas na reference date. Os
# caches de New Deals já são particionados por dia (YYYYMMDD_*.json), então o
# histórico "dia a dia" vem do próprio layout de arquivos — mudar a reference
# date apenas lê os arquivos daquele dia, nada precisa ser re-persistido.
# ==============================================================================

# Catálogo dos cards (ordem de exibição). 'dirs' são os caminhos relativos ao
# NEW_DEALS_CACHE_ROOT que alimentam o card (variantes de grafia incluídas).
# 'soon' = produto ainda sem desenvolvimento — o card aparece como placeholder,
# mas se um dia o diretório ganhar arquivos os números passam a contar sozinhos.

# Produtos cuja entidade intragrupo é o fundo Atacama — a perna-espelho
# (Client = Banco) conta como ATA em vez de LAW.
# NDFs genéricos (têm coluna LE e não têm perna-espelho da Lawton nos caches):
# neles LAW = operação CONTRA a Lawton, não Client = Banco.


# ──────────────────────────────────────────────────────────────────────────
# Deals Monitor — e-mail diário de pendências (19h00 e 19h30)
# ──────────────────────────────────────────────────────────────────────────
# Varre os cards do Monitor da data e manda para a mesa o que ainda NÃO está
# 100% Success. Um card entra na lista quando tem operação (total > 0) e pelo
# menos uma fora de Success; a quantidade é `total - Success`, que é exatamente
# o que sobra de ação na tela.
#
# As contagens saem do snapshot do Monitor (features/deals_monitor/engine),
# o MESMO código que alimenta a
# página — de propósito. Uma segunda contagem própria do e-mail divergiria da
# tela no primeiro ajuste de regra, e a mesa passaria a não confiar em nenhuma
# das duas.

# Tipo (zona do Monitor) × produto × detalhe de cada card. As zonas são as três
# da tela: Registration (B3), Confirmation e Intrag.


# Dois disparos por dia. O segundo é lembrete: quem não tratou às 19h00 leva o
# aviso de novo às 19h30, já com a lista atualizada.


# Desfecho do último disparo — é o que a tela mostra para responder se o aviso
# das 19h saiu, não saiu, ou não tinha o que mandar.


# ──────────────────────────────────────────────────────────────────────────
# Control Panel — Pending Confirmations Spreadsheet Metrics
# Grava a planilha "PENDING - Outstanding Confirmation OTC.xlsx" no share de
# Movimento (sobrescrevendo a anterior), todo dia útil ANBIMA às 10:45 de
# Brasília — o horário pedido foi 7:15 PM IST (Índia, UTC+5:30), que é 13:45
# UTC = 10:45 em São Paulo. As linhas são as do chip Status = Pending da
# página (situação recomputada na leitura, não o DB em que a linha mora), e as
# colunas seguem a lista que o time global pediu por extenso (§256) —
# inclusive as que a página NÃO tem (procuração, Vias, Abono…), que saem em
# branco de propósito: o cabeçalho preservado é o que deixa quem consome a
# planilha continuar achando as colunas no lugar.
# ──────────────────────────────────────────────────────────────────────────

# (cabeçalho da planilha, coluna da página, é data?). A LISTA INTEIRA — nomes
# e ORDEM — é o layout que o time global pediu por extenso (§256): coluna com
# `None` é as que a página não tem, e sai VAZIA de propósito, mantendo a
# posição, para a query do consumidor continuar achando cada nome no lugar.
# "Overall Comments" ← Comments e "JP sending documentation" / "Client return
# the document" ← Send Date / Return Date são os nomes da planilha para as
# mesmas colunas da página. EA é coluna de DATA (a data da economic
# affirmation), não texto. "Document type" continua vazia (o Signature Type
# chegou a ser usado nela e a mesa confirmou que NÃO é a mesma coisa — agora
# ele tem a própria coluna, "Signature Type").


_otc_app_url = _pf_mail._otc_app_url



def _moeda_num_code(iso):
    """ISO 3-letter → B3 3-digit currency code (mapping Currency Base, SIMBOLO →
    CODIGO DE CADASTRO). Unknown/blank → ''. Lê pelo loader do mapping, que
    cacheia por mtime: edição na tela vale na requisição seguinte."""
    m = {}
    for r in _mapping_rows('currency-base'):
        sym = str(r.get('SIMBOLO') or '').strip().upper()
        code = str(r.get('CODIGO DE CADASTRO') or '').strip()
        if sym and code and sym not in m:
            m[sym] = code
    return m.get(str(iso or '').strip().upper(), '')


# A montagem final das linhas passou a sair do cadastro do File Interface
# (template termo-multiclasses): a ordem dos campos e os literais Fixed vêm
# do JSON, com os overrides por página (page_url) resolvidos pelo motor. Os
# valores calculados continuam aqui, já na largura exata de sempre — o motor
# não reformata nada. Template/bloco ausente vira erro claro no endpoint,
# nunca arquivo montado do jeito velho em silêncio.
_TER_FI_KEY = 'termo-multiclasses'
_TER_FI_ERROR = 'File Interpreter template missing/invalid — check /file-interpreter'
# O arquivo TER sai um por VISÃO, e a visão é a entidade dona dele. O balde é o
# vocabulário do gerador ('BANCO' é como a mesa chama o Banco J.P. Morgan nos
# arquivos da CETIP); a LE é o vocabulário do cadastro (`le-accronym`,
# `b3-accounts`). Este mapa é a tradução entre os dois, e é a ÚNICA coisa que
# sobrou fixa aqui: o Nome Simplificado em si vem do cadastro.
_TER_BUCKET_LE = {
    'BANCO': 'JPM',
    'LAWTON': 'LAWTON',
    'MGT': 'MGT',
}


def _le_side_norm(s):
    """Razão social normalizada para comparação: só letra e dígito, em caixa
    alta — 'BANCO J.P. MORGAN S.A.' ≡ 'BANCO J.P MORGAN S.A'."""
    return re.sub(r'[^A-Z0-9]', '', str(s or '').upper())


def _le_spn_entity_side(name):
    """Lado do par pelo CADASTRO le-spn, por razão social normalizada EXATA.
    É o que resolve o nome que nenhuma heurística de substring saberia:
    'JPMORGAN CHASE BANK, N.A. - SAO PAULO BRANCH' é a MGT — e o regex do JPM
    casaria com o 'JPMORGAN' antes. Nome fora do cadastro devolve None e cai
    nas heurísticas de sempre."""
    alvo = _le_side_norm(name)
    if not alvo:
        return None
    for row in _mapping_rows('le-spn'):
        nm = _le_side_norm(row.get('NAME'))
        if nm and nm == alvo:
            le = str(row.get('LE', '') or '').upper()
            for tok in ('LAWTON', 'ATACAMA', 'MGT', 'JPM'):
                if tok in le:
                    return tok
    return None


def _ter_le_side(name):
    """Lado de um par de pernas a partir de um nome (LE ou contraparte):
    'JPM', 'MGT', 'LAWTON', 'ATACAMA' — ou None para cliente externo. O
    cadastro le-spn (razão social exata) vence; o resto são os mesmos testes
    de substring que o gerador sempre usou para contas
    (`_is_jpm`/`_is_mgt`/`_is_lawton`), para o par nunca discordar do bucket.
    A cópia no navegador é o `side` do fi-ter-pair.js (entidades via
    /api/mappings/le-spn)."""
    u = str(name or '').upper()
    if not u.strip():
        return None
    ent = _le_spn_entity_side(name)
    if ent:
        return ent
    if 'LAWTON' in u:
        return 'LAWTON'
    if 'ATACAMA' in u:
        return 'ATACAMA'
    if 'MGT' in u:
        return 'MGT'
    if re.search(r'J\.?P\.?\s*MORGAN', u) or u.strip() in ('JPM', 'BANCO'):
        return 'JPM'
    return None


def _ter_le_pair(our_side, client):
    """Par 'NOSSA PERNA x CONTRAPARTE' que escolhe a variante do template
    ('MGT x JPM', 'JPM x CLI', ...). Contraparte que não é entidade do grupo
    é 'CLI'. A cópia desta regra no navegador é o `static/js/fi-ter-pair.js`
    (o preview escolhe a mesma variante que o arquivo usa)."""
    return '{} x {}'.format(our_side, _ter_le_side(client) or 'CLI')


def _opc_le_pair(client):
    """Par de pernas dos geradores OPC (Opt FXO / Opt Commodities). A perna
    nossa segue os testes de SUBSTRING que o gerador sempre usou para as
    contas ('BANCO J.P MORGAN' / 'JP MORGAN' — sem regex): cliente JPM nessa
    grafia = visão Lawton. Grafia fora do padrão não casa par nenhum de grupo
    e cai no template base — que é o comportamento hardcoded de sempre. A
    cópia no navegador é o `pairOpc` do fi-ter-pair.js."""
    c = str(client or '').upper()
    is_jpm = 'BANCO J.P MORGAN' in c or 'JP MORGAN' in c
    ours = 'LAWTON' if is_jpm else 'JPM'
    if 'LAWTON' in c:
        theirs = 'LAWTON'
    elif is_jpm:
        theirs = 'JPM'
    else:
        theirs = _ter_le_side(client) or 'CLI'
    return '{} x {}'.format(ours, theirs)


def _ter_file_header(le, today, page_url, le_pair=None):
    """Header (linha tipo 0) de um arquivo TER, pelo bloco `header` do
    cadastro — só o Participante e a data são do gerador; 'TER', tipo de
    linha, código de operação e versão de layout saem do JSON.

    O Participante é o **Nome Simplificado** da entidade dona do arquivo, e ele
    sai do cadastro `b3-accounts` (era um dicionário fixo aqui). O motor
    completa com espaços até os 20 caracteres do X(20) — o valor do gerador
    nunca é truncado nem reformatado.

    LE sem Nome Simplificado cadastrado levanta ValueError dizendo qual: o
    header é obrigatório, e um Participante em branco é um arquivo que a B3
    recusa depois de a mesa já ter mandado.

    Com `le_pair`, a VARIANTE do template pode fixar o Participante (Source
    Fixed no campo 4 do header) — aí o cadastro `b3-accounts` deixa de ser
    exigido, porque o valor que vai para o arquivo é o do próprio template."""
    tpl = _fi_tpl_cached(_fi_variant_key(_TER_FI_KEY, page_url, le_pair))
    participant_fixed = False
    for f in (_fi_block_of(tpl, 'header') or {}).get('fields', []):
        if _fi_seq_key(f.get('seq')) == '4':
            participant_fixed = \
                str(_fi_field_src(f, page_url).get('source', '')) == 'Fixed'
            break
    nome = _b3_participant_name(le)
    if not nome and not participant_fixed:
        raise ValueError(
            'B3 Accounts: no Simplified Name registered for legal entity '
            '{!r} — register it at /mapping › B3 Accounts'.format(le))
    return _fi_build_line(_TER_FI_KEY, 'header',
                          {'4': nome, '5': today}, page_url=page_url,
                          le_pair=le_pair)




def _vanilla_verification_lines(deal, page_url, le_pair):
    """Linhas tipo 2 (Dados Variáveis) do DOWNLOAD do Vanilla — só ele as
    emite: o registro oficial do Vanilla é de outra ferramenta, e o arquivo
    dos demais produtos genéricos nunca as carregou (o asiático vai só na
    contagem do campo 59). Uma linha por dia útil da janela de fixing, no
    calendário do deal (FX Holiday Schedule; sem cadastro, Seg–Sex, como o
    preview). Com a Cotação para o Vencimento EFETIVA preenchida (> 0 — Fixed
    da variante ou fórmula cadastrada), cada data sai deslocada N dias úteis
    PARA FRENTE, no mesmo calendário."""
    def _s(v):
        return re.sub(r'<[^>]+>', '', str(v or '')).strip()
    a = _parse_date_any(_s(deal.get('FirstFixingDate')))
    b = _parse_date_any(_s(deal.get('LastFixingDate')))
    if not a or not b or a >= b:
        return []
    hols = set()
    sched = re.sub(r'[^A-Za-z0-9_]', '',
                   _s(deal.get('FXHolidaySchedule')).replace('-', '_').lower())
    if sched:
        try:
            with open(os.path.join(_B3_DATA_DIR, sched + '.json'), encoding='utf-8') as fh:
                hols = {(x.get('date') if isinstance(x, dict) else x)
                        for x in json.load(fh)}
        except Exception:
            hols = set()
    cotv = _fi_effective_seq_value(_TER_FI_KEY, 'registro-dados-fixos', '15', {},
                                   page_url, le_pair, deal).strip()
    shift = int(cotv) if cotv.isdigit() and int(cotv) > 0 else 0

    def _is_biz(d):
        return d.weekday() < 5 and d.strftime('%Y-%m-%d') not in hols

    def _shifted(d):
        cur, left = d, shift
        while left > 0:
            cur += timedelta(days=1)
            if _is_biz(cur):
                left -= 1
        return cur

    lines, cur = [], a
    while cur <= b:
        if _is_biz(cur):
            d8 = _shifted(cur).strftime('%Y%m%d')
            lines.append(_fi_build_line(
                _TER_FI_KEY, 'registro-dados-variaveis',
                {'4': d8.ljust(8), '6': ''.ljust(18)},
                page_url=page_url, le_pair=le_pair, deal=deal))
        cur += timedelta(days=1)
    return lines




# ==============================================================================
# SUPPORT CENTER — TICKETS
# ==============================================================================
#
# Saiu daqui: virou a vertical `apps/pages/features/support/` — entrypoint,
# commands, queries, domain e infra. Ver a §10 do CLAUDE.md.
#
# As rotas continuam registradas porque o import lá embaixo, no fim deste
# arquivo, executa os decoradores `@blueprint.route` daquele módulo.


# ==============================================================================
# NOTIFICATIONS
# ==============================================================================

# O último motivo já registrado, para o aviso do sino não repetir a cada poll.
_notif_fail_last = {'msg': ''}


def _notif_query_failed(exc):
    """Registra a falha da consulta do sino UMA vez por motivo.

    Sem o de-duplicador, oito segundos vezes o número de abas abertas enchem o
    log com o mesmo traceback e escondem tudo o que veio antes — inclusive a
    linha da migração que explica a causa.
    """
    msg = '%s: %s' % (type(exc).__name__, exc)
    if _notif_fail_last['msg'] == msg:
        return
    _notif_fail_last['msg'] = msg
    # O caminho é o do banco de NOTIFICAÇÕES: a mensagem imprimia o
    # Config.DATABASE_PATH (o banco de usuários) e mandava quem depura caçar
    # lock no arquivo errado.
    log.error("[notifications] a consulta do sino falhou (%s) — o sino fica "
              "vazio até isto ser resolvido:\n%s",
              NOTIF_DB_PATH, traceback.format_exc())


@blueprint.route('/api/notifications', methods=['GET'])
def api_get_notifications():
    if not session.get('authenticated'):
        return jsonify({"success": False}), 401
    user_role = session.get('user_role', '')
    user_sid = (session.get('user_sid') or '').strip().upper()
    # SEM LOCK: o sino é a consulta mais repetida do app (uma por aba a cada
    # poucos segundos) e é de MELHOR ESFORÇO — a que falha já devolve o sino
    # vazio logo abaixo, e o poll seguinte corrige. Dispensando o lock, ela
    # não espera nem por uma gravação de notificação em curso. É o ÚNICO
    # lugar do app autorizado a isso (`check_unlocked_reads.py`).
    # A ABERTURA também é de melhor esforço, e ela ficava de fora: o `try` de
    # baixo cobria a consulta, então um banco que não abre (não existe ainda,
    # está em uso por outro processo) virava 500 — e o sino é chamado a cada 8
    # segundos POR ABA. A tela ficava com o erro no console e o log com a mesma
    # exceção centenas de vezes por minuto, escondendo a causa no meio das
    # repetições. Sem conexão, o sino é o de sempre: vazio.
    # UMA retentativa curta antes de desistir: a falha mais comum aqui é a
    # colisão de milissegundos com uma gravação de notificação em curso — o
    # DuckDB recusa abrir read-only o arquivo que o MESMO processo tem aberto
    # em escrita ("different configuration than existing connections"). A
    # gravação é curta, então 250 ms depois a abertura quase sempre passa, e o
    # sino deixa de piscar vazio por um ciclo de poll. O ERROR do log fica
    # reservado para a falha que persiste (conexão vazada, ensure em loop),
    # que é a que pede alguém. O laço mantém UM ponto de chamada
    # `unlocked=True` — é o que o check_unlocked_reads prende por AST.
    conn = None
    for tentativa in (1, 2):
        try:
            conn = get_notif_connection(readonly=True, unlocked=True)
            break
        except Exception as exc:                            # noqa: BLE001
            if tentativa == 1:
                time.sleep(0.25)
                continue
            _notif_query_failed(exc)
            # Mesma FORMA da resposta de sucesso, `total_today` incluído: a
            # topbar lê o campo direto, e um payload pela metade trocaria o
            # sino vazio por um erro no console do navegador.
            return jsonify({"success": True, "notifications": [], "total_today": 0})
    try:
        # target_sid endereça UM usuário e é mais forte que target_role: quando
        # está preenchido, só aquele SID vê a notificação — nem o master, nem
        # quem compartilha o papel. É o que mantém as atualizações de ticket
        # restritas ao requester.
        # A consulta falhando NÃO pode derrubar o endpoint: a topbar o chama a
        # cada 8 segundos POR ABA ABERTA, então um erro aqui vira um 500 a cada 8
        # segundos por pessoa — o log enche e a causa some no meio das
        # repetições. O sino fica vazio (o pior que pode acontecer) e o motivo
        # sai UMA vez, em ERROR.
        try:
            rows = conn.execute("""
                SELECT id, actor_sid, actor_name, action, page, detail, target_role, created_at
                FROM notifications
                WHERE DATE(created_at) = CURRENT_DATE
                  AND (COALESCE(target_sid, '') = '' OR COALESCE(target_sid, '') = ?)
                  AND (COALESCE(target_role, '') = ''
                       OR list_contains(string_split(target_role, ','), ?))
                ORDER BY created_at DESC
                LIMIT 50
            """, [user_sid, user_role]).fetchall()
        except Exception as exc:
            _notif_query_failed(exc)
            rows = []
        notifs = []
        for r in rows:
            notifs.append({
                "id": r[0],
                "actor_sid": r[1] or '',
                "actor_name": r[2] or '',
                "action": r[3] or '',
                "page": r[4] or '',
                "detail": r[5] or '',
                "created_at": r[7].isoformat() if r[7] else ''
            })
    finally:
        conn.close()

    # Page-access filter: a configured user only sees notifications for the pages
    # they can reach. Master / unconfigured users see everything. A page label with
    # no known URL, or one that isn't a controllable page, is kept.
    if not _session_is_master():
        configured, allowed = _get_page_access(session.get('user_sid', ''))
        if configured:
            def _visible(n):
                url = _notif_page_url(n.get('page', ''), n.get('action', ''))
                if not url or url not in _NAV_URLS:
                    return True
                if url == '/control-panel':
                    return _cp_page_allowed(allowed)
                return url in allowed
            notifs = [n for n in notifs if _visible(n)]
    return jsonify({"success": True, "notifications": notifs, "total_today": len(notifs)})


@blueprint.route('/api/notifications', methods=['POST'])
def api_create_notification():
    if not session.get('authenticated'):
        return jsonify({"success": False}), 401
    data        = request.get_json(silent=True) or {}
    action      = data.get('action', '').strip()
    page        = data.get('page', '').strip()
    detail      = data.get('detail', '').strip()
    target_role = data.get('target_role', '').strip()
    if not action or not page:
        return jsonify({"success": False, "message": "action and page required"}), 400
    _create_notification(
        session.get('user_sid', ''), session.get('user_name', ''),
        action, page, detail, target_role
    )
    return jsonify({"success": True})


# ── Web Push (Service Worker) ─────────────────────────────────────────────────
@blueprint.route('/sw.js')
def service_worker():
    """Serve the Service Worker at the root scope so it controls the whole app.
    The Service-Worker-Allowed header lets a /static-hosted script claim '/'."""
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                        '..', 'static', 'js', 'sw-push.js')
    try:
        with open(path, 'r', encoding='utf-8') as f:
            js = f.read()
    except Exception:
        return ('', 404)
    resp = make_response(js)
    resp.headers['Content-Type'] = 'application/javascript'
    resp.headers['Service-Worker-Allowed'] = '/'
    resp.headers['Cache-Control'] = 'no-cache'
    return resp


@blueprint.route('/api/push/public-key', methods=['GET'])
def api_push_public_key():
    from apps.pages import webpush
    return jsonify({'enabled': webpush.is_enabled(), 'key': webpush.get_public_key()})


@blueprint.route('/api/push/subscribe', methods=['POST'])
def api_push_subscribe():
    if not session.get('authenticated'):
        return jsonify({'success': False}), 401
    data = request.get_json(silent=True) or {}
    sub = data.get('subscription') or data
    endpoint = str((sub or {}).get('endpoint') or '').strip()
    if not endpoint:
        return jsonify({'success': False, 'message': 'no endpoint'}), 400
    conn = get_notif_connection()
    try:
        conn.execute("DELETE FROM push_subscriptions WHERE endpoint = ?", [endpoint])
        conn.execute(
            "INSERT INTO push_subscriptions (endpoint, sid, role) VALUES (?, ?, ?)",
            [endpoint, session.get('user_sid', '') or '', session.get('user_role', '') or ''])
        conn.commit()
    finally:
        conn.close()
    return jsonify({'success': True})


@blueprint.route('/api/push/unsubscribe', methods=['POST'])
def api_push_unsubscribe():
    if not session.get('authenticated'):
        return jsonify({'success': False}), 401
    data = request.get_json(silent=True) or {}
    endpoint = str(data.get('endpoint') or '').strip()
    if endpoint:
        conn = get_notif_connection()
        try:
            conn.execute("DELETE FROM push_subscriptions WHERE endpoint = ?", [endpoint])
            conn.commit()
        finally:
            conn.close()
    return jsonify({'success': True})


# ==============================================================================
# RECONCILIAÇÃO DE COMITENTES
# ==============================================================================


# ── Reconciliation › Pay/Rec ──────────────────────────────────────────────────
#  Matches JPM-side payments/receipts against client-side amounts. Input files can
#  be dropped in the page (held until Run) or imported from the network Pay_Rec
#  folder. "End process" e-mails the final situation of the day to OTC Ops.


# ==============================================================================
#  MANUAL CONFIRMATIONS — a esteira de validação de uma confirmação gerada.
#  Duas telas: Confirmations Monitor (os cards por etapa) e Track Confirmations
#  (a tabela inteira). O motor está em apps/pages/manual_conf.py; aqui só entram
#  sessão, o SPN de quem validou e o e-mail do reject.
# ==============================================================================




# ==============================================================================
#  FXO RECONCILIATION — posição B3/CETIP (DPOSICAO .OPC) × Athena (EOD FXO).
# ==============================================================================
#
# Saiu daqui: virou a vertical `apps/pages/features/reconciliation_fxo/`. O
# MOTOR continua no `apps/pages/recon_fxo.py` — ele é o domínio desta recon e
# tem o `check_recon_fxo.py` em cima. Ver a §10 do CLAUDE.md.


# ==============================================================================
# ROTA GENÉRICA — TEMPLATES (deve ser a ÚLTIMA rota definida)
# ==============================================================================



@blueprint.route('/<template>')
def route_template(template):
    # Catch-all page renderer — require authentication so unauthenticated
    # visitors can't load internal templates and bypass the page-access gate.
    if not session.get('authenticated'):
        return redirect(url_for('pages_blueprint.sign_in_page'))
    try:
        if not template.endswith('.html'):
            template += '.html'
        segment = get_segment(request)
        log.debug("[route_template] Rendering pages/%s (segment=%s)", template, segment)
        return render_template("pages/" + template, segment=segment)
    except TemplateNotFound:
        log.warning("[route_template] Template not found: pages/%s", template)
        return render_template('pages/error-404.html'), 404
    except Exception:
        log.error("[route_template] Error rendering pages/%s:\n%s", template, traceback.format_exc())
        return render_template('pages/error-500.html'), 500


# ==============================================================================
# FUNÇÕES AUXILIARES — INTERNAS
# ==============================================================================

def _initiate_2fa(sid, email, name):
    log.info("[_initiate_2fa] Generating 2FA code for SID=%s email=%s", sid, email)
    # Throttle sends: if a code was just emailed (or too many were requested),
    # don't send another — route the user to the entry page to use the code they
    # already have. The pending session is still set so they can proceed.
    allowed, wait_msg = _code_send_allowed(sid)
    if not allowed:
        log.warning("[_initiate_2fa] Send throttled for SID=%s: %s", sid, wait_msg)
        session['pending_sid'] = sid
        session['masked_email'] = get_masked_email(email)
        session['masked_phone'] = get_masked_phone()
        flash(wait_msg, "warning")
        return redirect(url_for('pages_blueprint.two_factor_page'))
    code = generate_verification_code()
    save_verification_code(sid, code)

    email_sent = send_verification_email(email, code, name)
    log.info("[_initiate_2fa] Email sent=%s for SID=%s", email_sent, sid)
    if not email_sent:
        flash("Failed to send verification email. Please try again.", "error")
        return redirect(url_for('pages_blueprint.sign_in_page'))

    session['pending_sid'] = sid
    session['masked_email'] = get_masked_email(email)
    session['masked_phone'] = get_masked_phone()
    log.info("[_initiate_2fa] Session set for SID=%s → redirecting to 2FA page", sid)
    return redirect(url_for('pages_blueprint.two_factor_page'))




# ============================================================================
# ONBOARDING — Overview e Tracking Docs (o CGD que vem da lista do SharePoint)
# ============================================================================
#
# Saiu daqui: virou a vertical `apps/pages/features/onboarding/`. O domínio
# continua no `apps/pages/cgd_docs.py` — a Recon de CGD e o /mapping também o
# consultam, então ele é horizontal. Ver a §10 do CLAUDE.md.


# ============================================================================
# RECONCILIAÇÃO DE CGD — o batimento FEP × B3 (tradução do Alteryx `Batimento
# CGD`; a regra e o porquê de cada mudança estão no `recon_cgd`).
#
# A tela ABRE com o resultado que já foi rodado naquele dia (cache), e rodar de
# novo é decisão de quem opera: a leitura do arquivo da B3 e da lista do FEP é
# lenta e vive num share, e fazê-la a cada abertura de página deixaria a tela
# refém da rede.
# ============================================================================


def get_segment(request):
    try:
        segment = request.path.split('/')[-1]
        return segment if segment else 'index'
    except Exception:
        return None


# ==============================================================================
# AS VERTICAIS — o registro das features extraídas deste arquivo
# ==============================================================================
#
# Cada import aqui existe para EXECUTAR os `@blueprint.route` do módulo: em
# Flask, rota que ninguém importa é rota que não existe, e a página responderia
# 404 sem erro nenhum na subida.
#
# Ele vai no FIM do arquivo de propósito. As features buscam no `routes` o que
# ainda é de plataforma (`_create_notification`, `get_db_connection`, o SMTP),
# e importá-las no topo fecharia um ciclo — aqui embaixo, o módulo já está
# inteiro. A busca do lado delas é atrasada (dentro da função), o que mantém o
# ciclo impossível e preserva os monkeypatches dos testes.
from apps.pages.features.support import entrypoint as _f_support     # noqa: E402,F401
from apps.pages.features.onboarding import entrypoint as _f_onboarding  # noqa: E402,F401
from apps.pages.features.reconciliation_fxo import entrypoint as _f_recon_fxo  # noqa: E402,F401
from apps.pages.features.quotes import entrypoint as _f_quotes         # noqa: E402,F401
from apps.pages.features.holidays import entrypoint as _f_holidays     # noqa: E402,F401
from apps.pages.features.bacc import entrypoint as _f_bacc             # noqa: E402,F401
# O scheduler do BACC vive na feature; o REGISTRO é gancho de plataforma e por
# isso fica aqui, ao lado do import — chamar o _schedule_on_start de dentro do
# módulo exigiria importar o routes no corpo dele, que é o ciclo proibido.
_schedule_on_start('bacc-ea', _f_bacc.start_scheduler)
from apps.pages.features.mt300 import entrypoint as _f_mt300           # noqa: E402,F401
_schedule_on_start('mt300', _f_mt300.start_scheduler)
from apps.pages.features.mdea import entrypoint as _f_mdea             # noqa: E402,F401
_schedule_on_start('manual-deals-ea', _f_mdea.start_scheduler)
from apps.pages.features.conf_escalation import entrypoint as _f_confesc  # noqa: E402,F401
_schedule_on_start('conf-escalation', _f_confesc.start_scheduler)
from apps.pages.features.daily_metric import entrypoint as _f_daily_metric      # noqa: E402,F401
from apps.pages.features.weekly_escalation import entrypoint as _f_weekly_esc   # noqa: E402,F401
from apps.pages.features.recon_comitente import entrypoint as _f_recon_comitente  # noqa: E402,F401
from apps.pages.features.recon_payrec import entrypoint as _f_recon_payrec        # noqa: E402,F401
from apps.pages.features.recon_cgd import entrypoint as _f_recon_cgd              # noqa: E402,F401
from apps.pages.features.boxscan import entrypoint as _f_boxscan                  # noqa: E402,F401
from apps.pages.features.sigcoll import entrypoint as _f_sigcoll                  # noqa: E402,F401
from apps.pages.features.pcx import entrypoint as _f_pcx                          # noqa: E402,F401
_schedule_on_start('pending-spreadsheet', _f_pcx.start_scheduler)
from apps.pages.features.forecast import entrypoint as _f_forecast               # noqa: E402,F401
from apps.pages.features.deals_monitor import entrypoint as _f_deals_monitor     # noqa: E402,F401
_schedule_on_start('deals-monitor', _f_deals_monitor.start_scheduler)
from apps.pages.features.cetip import entrypoint as _f_cetip                     # noqa: E402,F401
from apps.pages.features.intrag import entrypoint as _f_intrag                   # noqa: E402,F401
from apps.pages.features.counterparty_details import entrypoint as _f_cpd        # noqa: E402,F401
from apps.pages.features.electronic_inventory import entrypoint as _f_ei         # noqa: E402,F401
from apps.pages.features.manual_confirmation import entrypoint as _f_mc          # noqa: E402,F401
from apps.pages.features.mtm import entrypoint as _f_mtm                         # noqa: E402,F401
from apps.pages.features.cognos import entrypoint as _f_cognos                   # noqa: E402,F401
from apps.pages.features.otm import entrypoint as _f_otm                          # noqa: E402,F401
from apps.pages.features.latam import entrypoint as _f_latam                      # noqa: E402,F401
from apps.pages.features.accrual import entrypoint as _f_accrual                 # noqa: E402,F401
_schedule_on_start('boxscan', _f_boxscan.start_scheduler)
from apps.pages.features.appver import entrypoint as _f_appver         # noqa: E402,F401
from apps.pages.features.ndf_summary import entrypoint as _f_ndf_summary          # noqa: E402,F401
from apps.pages.features.operations_b3 import entrypoint as _f_operations_b3      # noqa: E402,F401
from apps.pages.features.other_products import entrypoint as _f_other_products    # noqa: E402,F401
from apps.pages.features.file_interpreter import entrypoint as _f_file_interpreter# noqa: E402,F401
from apps.pages.features.confirmation import entrypoint as _f_confirmation        # noqa: E402,F401
from apps.pages.features.ndf_cockpit import entrypoint as _f_ndf_cockpit          # noqa: E402,F401
from apps.pages.features.ndf_other_publisher import entrypoint as _f_ndf_other_publisher# noqa: E402,F401
from apps.pages.features.pending_confirmation import entrypoint as _f_pending_confirmation# noqa: E402,F401
from apps.pages.features.live_positions import entrypoint as _f_live_positions    # noqa: E402,F401
from apps.pages.features.mapping import entrypoint as _f_mapping                  # noqa: E402,F401
from apps.pages.features.index_b3 import entrypoint as _f_index_b3                # noqa: E402,F401
from apps.pages.features.daily_settlement import entrypoint as _f_daily_settlement# noqa: E402,F401
from apps.pages.features.new_deals import entrypoint as _f_new_deals              # noqa: E402,F401