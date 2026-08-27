# -*- coding: utf-8 -*-
"""Planilha de fatores → `{cetip: (fator_parte, fator_contraparte)}`.

Três formatos convivem, um por LOB (`_ACC_FACTOR_KINDS`): o CEM, que precisa da
2ª aba para saber a VIEW de cada CETIP (e inverter os fatores quando só existe a
199), e os dois diretos (EDG e Hybrids), que diferem só nas colunas.
"""
import re

from apps.pages.features.accrual import domain


def _R():
    """Busca ATRASADA no routes — plataforma (ver features/support/infra)."""
    from apps.pages import routes
    return routes


def _acc_read_sheets(filename, raw_bytes):
    """{sheet_name: rows} for .xlsx/.xlsm; one '__main__' sheet for csv/tsv."""
    name = (filename or '').lower()
    if name.endswith(('.xlsx', '.xlsm')):
        import openpyxl
        wb = openpyxl.load_workbook(_R().io.BytesIO(raw_bytes), read_only=True, data_only=True)
        return {sn: [list(r) for r in wb[sn].iter_rows(values_only=True)] for sn in wb.sheetnames}
    return {'__main__': _R()._cc_read_rows(filename, raw_bytes)}


def _acc_parse_cem_factors(filename, raw_bytes):
    """{cetip -> (parte_factor, contra_factor)} from the CEM workbook.

    As abas são lidas por POSIÇÃO, não por nome: 1ª = summary (os fatores), 2ª =
    Kapital CETIP (col B Kapital → col E LE). Antes a 2ª era procurada pelo nome
    conter 'kapital', e o arquivo real chega com a aba nomeada de outro jeito —
    a importação morria com "CEM file is missing the 'Kapital CETIP' sheet".
    A ordem das abas é estável no arquivo que a área gera; o nome não é.

    When a CETIP ID carries view 228 it is the bank view → normal (Parte = col I,
    Contraparte = col J). When it only carries view 199 the factors are inverted
    (Parte = col J, Contraparte = col I). Duplicated 228/199 → keep the 228 row."""
    sheets = _acc_read_sheets(filename, raw_bytes)
    names = list(sheets.keys())
    if len(names) < 2:
        raise ValueError(
            'CEM file needs at least 2 sheets (1st = summary, 2nd = Kapital CETIP); '
            'found {}: {}.'.format(len(names), ', '.join(repr(n) for n in names) or 'none'))
    main_rows, kap_rows = sheets[names[0]], sheets[names[1]]
    # Registra QUAIS abas foram usadas: como a escolha agora é posicional, é o
    # log que permite descobrir uma planilha fora de ordem (ou uma aba oculta
    # que tenha entrado no meio) sem abrir o arquivo.
    _R().log.info('[accrual] CEM %s: summary=%r, Kapital CETIP=%r (de %d abas)',
             filename, names[0], names[1], len(names))

    # Kapital ID (col B) → LE digits (col E). Both sides drop the leading zeros so
    # the lookup is robust to '00123' vs '123' mismatches between the two sheets.
    def _kap_key(v):
        return str(v or '').strip().upper().lstrip('0')

    kap_le = {}
    for r in kap_rows:
        kid = _kap_key(_R()._cc_cell(r, 1))
        le  = domain._acc_le_norm(_R()._cc_cell(r, 4))
        if kid and le:
            kap_le.setdefault(kid, le)

    # Group every data row by CETIP ID (col C), tagging its LE via the Kapital map.
    groups = {}
    for r in main_rows:
        cetip = _R()._cc_cell(r, 2).strip()
        if not re.search(r'\d', cetip):                  # skip title/header/blank rows
            continue
        kid = _kap_key(_R()._cc_cell(r, 1))
        groups.setdefault(cetip, []).append(
            {'le': kap_le.get(kid, ''), 'i': _R()._cc_cell(r, 8), 'j': _R()._cc_cell(r, 9)})

    fmap = {}
    for cetip, items in groups.items():
        v228 = next((it for it in items if it['le'] == '228'), None)
        v199 = next((it for it in items if it['le'] == '199'), None)
        if v228:                                         # bank view → normal mapping
            pf, cf = domain._acc_fmt_factor(v228['i']), domain._acc_fmt_factor(v228['j'])
        elif v199:                                       # only 199 → inverted mapping
            pf, cf = domain._acc_fmt_factor(v199['j']), domain._acc_fmt_factor(v199['i'])
        else:
            continue                                     # other view (e.g. 123) → not the bank view
        domain._acc_fmap_put(fmap, cetip, pf, cf)
    return fmap


def _acc_parse_direct_factors(filename, raw_bytes, cetip_col=0, parte_col=1, contra_col=2):
    """{cetip -> (parte_factor, contra_factor)} from a direct file (no LE / no
    inversion). Column indices are 0-based:
        EDG → CETIP=A(0), Fator Parte=B(1),  Fator Contraparte=C(2)
        HYB → CETIP=B(1), Fator Parte=L(11), Fator Contraparte=M(12)"""
    sheets = _acc_read_sheets(filename, raw_bytes)
    main_rows = next(iter(sheets.values()), [])
    fmap = {}
    for r in main_rows:
        cetip = _R()._cc_cell(r, cetip_col).strip()
        if not re.search(r'\d', cetip):
            continue
        domain._acc_fmap_put(fmap, cetip,
                      domain._acc_fmt_factor(_R()._cc_cell(r, parte_col)),
                      domain._acc_fmt_factor(_R()._cc_cell(r, contra_col)))
    return fmap


_ACC_FACTOR_KINDS = {
    'cem': {'lob': 'CEM',     'parser': lambda fn, raw: _acc_parse_cem_factors(fn, raw)},
    'edg': {'lob': 'EDG',     'parser': lambda fn, raw: _acc_parse_direct_factors(fn, raw)},
    'hyb': {'lob': 'Hybrids', 'parser': lambda fn, raw: _acc_parse_direct_factors(
                                            fn, raw, cetip_col=1, parte_col=11, contra_col=12)},
}
