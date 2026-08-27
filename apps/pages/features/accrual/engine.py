# -*- coding: utf-8 -*-
"""Accrual de Swap — fatores por LOB, ciclo da linha, geração e recon.

Movido VERBATIM do routes.py. Ficaram lá: `_acc_digits` (o b3-accounts e o
NDF Commodities comparam contas com ele), `_accrual_lob` (o forecast rotula
pela mesma regra), `_accrual_parse_date` (o MtM parseia a data com ele) e a
lista `_ACC_ENDPROC_CC` (o e-mail do forecast copia as mesmas caixas).
"""
import os
import re
import traceback
from datetime import datetime



def _R():
    """Busca ATRASADA no routes — plataforma (ver features/support/infra)."""
    from apps.pages import routes
    return routes


ACCRUAL_JSON_ROOT = _R().data_write('cache', 'accrual')

ACCRUAL_SOURCE_ROOT = os.getenv('ACCRUAL_SOURCE_ROOT', os.path.join(
    _R().Config.SHARED_DRIVE_ROOT, 'Confirmation', 'Derivativos', 'OTC Tracker', 'Regulatory',
    'Accrual'))

_ACC_HEADER_ROW = 9                                # 1-based: headers on row 9

_ACC_ACCOUNT_COL = 10                              # col K — house-account filter

_ACC_ACCOUNTS   = {'73760009', '04880006'}         # col K house accounts (digits only)

_ACC_FIXED_HEADERS = [
    'Código IF', 'Data Início', 'Data Vencimento',
    'PARTE / Conta', 'PARTE / Nome Simplificado', 'PARTE / Indexador',
    'CONTRAPARTE / Conta', 'CONTRAPARTE / Nome Simplificado', 'CONTRAPARTE / Indexador',
    'Fator Parte', 'Fator Contraparte', 'Comments',
]

_ACC_DISPLAY_SRC = [0, 5, 6, 10, 11, 13, 16, 17, 19, None, None, None]

def _accrual_build_result(rows):
    """Core VCP→tables logic (no I/O). Splits the rows into the four LOB books and
    returns the result dict (without 'date'/'saved_at')."""
    records, ref_date = _R()._swap_pos_latest_records()
    lob_map = _R()._swap_pos_lob_map(records)

    buckets = {'CEM': [], 'EDG': [], 'Hybrids': [], 'Commodities': []}
    total = kept = matched = 0
    for i in range(_ACC_HEADER_ROW, len(rows)):
        row = rows[i]
        a_raw = _R()._cc_cell(row, 0)
        if not a_raw and not any(_R()._cc_cell(row, c) for c in _ACC_DISPLAY_SRC if c is not None):
            continue                                    # fully blank line
        total += 1
        contract = a_raw.replace('#', '').strip()       # col A: drop '#'
        if _R()._acc_digits(_R()._cc_cell(row, _ACC_ACCOUNT_COL)) not in _ACC_ACCOUNTS:
            continue                                    # col K: house accounts only
        kept += 1
        ident = lob_map.get(contract.upper())
        if ident is None:
            ident = lob_map.get('#' + _R()._acc_digits(contract))
        lob = _R()._accrual_lob(ident)
        if not lob:
            continue                                    # IF not found / unclassified
        matched += 1
        # Build the row aligned to _ACC_FIXED_HEADERS (None src → empty placeholder).
        cells = []
        for src in _ACC_DISPLAY_SRC:
            if src is None:      cells.append('')
            elif src == 0:       cells.append(contract)         # Código IF (# stripped)
            else:                cells.append(_R()._cc_cell(row, src))
        buckets[lob].append(cells)

    # Append, per row, the maker/checker meta and a stable id as the LAST cell.
    # Row layout: [ ...fixed data cells..., status, maker, checker, id ]
    for _lob, _rws in buckets.items():
        for _i, _rw in enumerate(_rws):
            _rw.extend(['New', '', ''])                # status, maker, checker
            _rw.append('{}-{}'.format(_lob, _i))       # stable id (last cell)

    return {
        'success': True,
        'headers': list(_ACC_FIXED_HEADERS),
        'tables': buckets,
        'counts': {k: len(v) for k, v in buckets.items()},
        'ref_date': ref_date,
        'diagnostics': {'total': total, 'kept': kept, 'matched': matched,
                        'position_records': len(records)},
    }

def _accrual_persist(result, source_file, ymd=None):
    """Persist a build result under static/data/cache/accrual/YYYY/MM/DD/. Defaults
    to today; pass ymd ('YYYYMMDD') to store under the run/reference date instead.
    Returns (path, saved_dict)."""
    now = datetime.now()
    ymd = ymd or now.strftime('%Y%m%d')
    out_dir = os.path.join(ACCRUAL_JSON_ROOT, ymd[:4], ymd[4:6], ymd[6:8])
    os.makedirs(out_dir, exist_ok=True)
    saved = dict(result)
    saved['date']        = '{}-{}-{}'.format(ymd[:4], ymd[4:6], ymd[6:8])
    saved['saved_at']    = now.strftime('%Y-%m-%d %H:%M:%S')
    saved['source_file'] = source_file
    path = os.path.join(out_dir, 'accrual_swap_{}.json'.format(ymd))
    with open(path, 'w', encoding='utf-8') as fh:
        _R().json.dump(saved, fh, ensure_ascii=False, indent=2)
    _R().log.info('[accrual] saved %s', path)
    return path, saved

def _accrual_path_for(ymd):
    return os.path.join(ACCRUAL_JSON_ROOT, ymd[:4], ymd[4:6], ymd[6:8],
                        'accrual_swap_{}.json'.format(ymd))

def _accrual_latest_ymd():
    """Newest saved accrual date as 'YYYY-MM-DD' (scans accrual_swap_*.json under
    ACCRUAL_JSON_ROOT), or None if nothing saved yet. Lets the page land on the
    most recent dataset when no explicit date is requested (e.g. from a bell
    notification), instead of an empty 'today'."""
    latest = None
    if not os.path.isdir(ACCRUAL_JSON_ROOT):
        return None
    for _root, _dirs, files in os.walk(ACCRUAL_JSON_ROOT):
        for fn in files:
            m = re.match(r'accrual_swap_(\d{8})\.json$', fn)
            if m and (latest is None or m.group(1) > latest):
                latest = m.group(1)
    return '{}-{}-{}'.format(latest[:4], latest[4:6], latest[6:8]) if latest else None

def _accrual_load(date_str):
    ymd = _R()._accrual_parse_date(date_str) or datetime.now().strftime('%Y%m%d')
    path = _accrual_path_for(ymd)
    if not os.path.isfile(path):
        return None, None
    try:
        with open(path, 'r', encoding='utf-8') as fh:
            return path, _accrual_migrate(_R().json.load(fh))
    except Exception:
        _R().log.error('[accrual] read failed %s:\n%s', path, traceback.format_exc())
        return None, None

def _accrual_migrate(data):
    """Bring rows saved under an older column layout up to the current fixed set,
    padding the data block (before the 4 meta cells status/maker/checker/id) so a
    newly-added column like 'Comments' lands in the right place for old files too."""
    nfix = len(_ACC_FIXED_HEADERS)
    for rows in (data.get('tables') or {}).values():
        for r in rows:
            ndata = len(r) - 4                       # cells before status/maker/checker/id
            while 0 <= ndata < nfix:
                r.insert(ndata, '')                  # append to the data block, push meta right
                ndata += 1
    data['headers'] = list(_ACC_FIXED_HEADERS)
    return data

def _accrual_save(path, data):
    data['counts'] = {k: len(v) for k, v in (data.get('tables') or {}).items()}
    with open(path, 'w', encoding='utf-8') as fh:
        _R().json.dump(data, fh, ensure_ascii=False, indent=2)

def _accrual_find(data, lob, rid):
    for r in (data.get('tables') or {}).get(lob, []):
        if r and str(r[-1]) == str(rid):
            return r
    return None

_ACC_FACTOR_STATUS_MISSING = 'Missing Accrual'

def _acc_parse_num(s):
    """Parse a number that may be BR (1.234,56) or US (1,234.56) formatted."""
    t = str('' if s is None else s).strip().replace('%', '').replace(' ', '')
    if not t or not re.search(r'\d', t):
        return None
    neg = t.startswith('-')
    t = t.lstrip('+-')
    has_c, has_d = ',' in t, '.' in t
    if has_c and has_d:                                  # decimal = whichever comes last
        dec = ',' if t.rfind(',') > t.rfind('.') else '.'
        t = t.replace('.' if dec == ',' else ',', '').replace(dec, '.')
    elif has_c:
        t = t.replace(',', '.')
    try:
        v = float(t)
    except ValueError:
        return None
    return -v if neg else v

def _acc_fmt_factor(s):
    """US format, 8 decimals (rounded), ALWAYS absolute (drop any '-'). '' when the
    cell is not a number."""
    v = _acc_parse_num(s)
    return '' if v is None else '{:.8f}'.format(round(abs(v), 8))

def _acc_le_norm(s):
    """Normalise an LE/view to bare digits without leading zeros (0228 → 228)."""
    return re.sub(r'\D', '', str(s or '')).lstrip('0')

def _acc_read_sheets(filename, raw_bytes):
    """{sheet_name: rows} for .xlsx/.xlsm; one '__main__' sheet for csv/tsv."""
    name = (filename or '').lower()
    if name.endswith(('.xlsx', '.xlsm')):
        import openpyxl
        wb = openpyxl.load_workbook(_R().io.BytesIO(raw_bytes), read_only=True, data_only=True)
        return {sn: [list(r) for r in wb[sn].iter_rows(values_only=True)] for sn in wb.sheetnames}
    return {'__main__': _R()._cc_read_rows(filename, raw_bytes)}

def _acc_factor_keys(code):
    """Lookup keys for a CETIP ID / Código IF: upper-cased and digits-only ('#'+d)."""
    code = str(code or '').strip()
    keys = [code.upper()]
    dg = re.sub(r'\D', '', code)
    if dg:
        keys.append('#' + dg)
    return keys

def _acc_fmap_put(fmap, cetip, parte_factor, contra_factor):
    for k in _acc_factor_keys(cetip):
        fmap.setdefault(k, (parte_factor, contra_factor))

def _acc_fmap_get(fmap, code):
    for k in _acc_factor_keys(code):
        if k in fmap:
            return fmap[k]
    return None

def _acc_parse_cem_factors(filename, raw_bytes):
    """{cetip -> (parte_factor, contra_factor)} from the CEM workbook.

    As abas são lidas por POSIÇÃO, não por nome: 1ª = summary (os fatores), 2ª =
    Kapital CETIP (col B Kapital → col E LE). Antes a 2ª era procurada pelo nome
    conter 'kapital', e o arquivo real chega com a aba nomeada de outro jeito —
    a importação morria com "CEM file is missing the 'Kapital CETIP' sheet".
    A ordem das abas é estável no arquivo que a área gera; o nome não é.

    When a CETIP ID carries view 228 it is the bank view → normal (Parte = col I,
    Contraparte = col J). When it only carries view 199 the factors are inverted
    (Parte = col J, Contraparte = col I). Duplicated 228/199 → keep the 228 row."""
    sheets = _acc_read_sheets(filename, raw_bytes)
    names = list(sheets.keys())
    if len(names) < 2:
        raise ValueError(
            'CEM file needs at least 2 sheets (1st = summary, 2nd = Kapital CETIP); '
            'found {}: {}.'.format(len(names), ', '.join(repr(n) for n in names) or 'none'))
    main_rows, kap_rows = sheets[names[0]], sheets[names[1]]
    # Registra QUAIS abas foram usadas: como a escolha agora é posicional, é o
    # log que permite descobrir uma planilha fora de ordem (ou uma aba oculta
    # que tenha entrado no meio) sem abrir o arquivo.
    _R().log.info('[accrual] CEM %s: summary=%r, Kapital CETIP=%r (de %d abas)',
             filename, names[0], names[1], len(names))

    # Kapital ID (col B) → LE digits (col E). Both sides drop the leading zeros so
    # the lookup is robust to '00123' vs '123' mismatches between the two sheets.
    def _kap_key(v):
        return str(v or '').strip().upper().lstrip('0')

    kap_le = {}
    for r in kap_rows:
        kid = _kap_key(_R()._cc_cell(r, 1))
        le  = _acc_le_norm(_R()._cc_cell(r, 4))
        if kid and le:
            kap_le.setdefault(kid, le)

    # Group every data row by CETIP ID (col C), tagging its LE via the Kapital map.
    groups = {}
    for r in main_rows:
        cetip = _R()._cc_cell(r, 2).strip()
        if not re.search(r'\d', cetip):                  # skip title/header/blank rows
            continue
        kid = _kap_key(_R()._cc_cell(r, 1))
        groups.setdefault(cetip, []).append(
            {'le': kap_le.get(kid, ''), 'i': _R()._cc_cell(r, 8), 'j': _R()._cc_cell(r, 9)})

    fmap = {}
    for cetip, items in groups.items():
        v228 = next((it for it in items if it['le'] == '228'), None)
        v199 = next((it for it in items if it['le'] == '199'), None)
        if v228:                                         # bank view → normal mapping
            pf, cf = _acc_fmt_factor(v228['i']), _acc_fmt_factor(v228['j'])
        elif v199:                                       # only 199 → inverted mapping
            pf, cf = _acc_fmt_factor(v199['j']), _acc_fmt_factor(v199['i'])
        else:
            continue                                     # other view (e.g. 123) → not the bank view
        _acc_fmap_put(fmap, cetip, pf, cf)
    return fmap

def _acc_parse_direct_factors(filename, raw_bytes, cetip_col=0, parte_col=1, contra_col=2):
    """{cetip -> (parte_factor, contra_factor)} from a direct file (no LE / no
    inversion). Column indices are 0-based:
        EDG → CETIP=A(0), Fator Parte=B(1),  Fator Contraparte=C(2)
        HYB → CETIP=B(1), Fator Parte=L(11), Fator Contraparte=M(12)"""
    sheets = _acc_read_sheets(filename, raw_bytes)
    main_rows = next(iter(sheets.values()), [])
    fmap = {}
    for r in main_rows:
        cetip = _R()._cc_cell(r, cetip_col).strip()
        if not re.search(r'\d', cetip):
            continue
        _acc_fmap_put(fmap, cetip,
                      _acc_fmt_factor(_R()._cc_cell(r, parte_col)), _acc_fmt_factor(_R()._cc_cell(r, contra_col)))
    return fmap

def _acc_apply_factors(data, lob, fmap):
    """Fill Fator Parte/Contraparte for the rows of one LOB table. A side keyed by a
    non-VCP indexer gets '-'; a VCP side with no factor flags the row 'Missing
    Accrual'. Row layout: [ ...11 data..., status, maker, checker, id ]."""
    rows = (data.get('tables') or {}).get(lob, [])
    matched = missing = 0
    for row in rows:
        if not row or len(row) < 15:
            continue
        parte_idx  = str(row[5] or '').strip().upper()
        contra_idx = str(row[8] or '').strip().upper()
        entry = _acc_fmap_get(fmap, row[0])
        if entry:
            matched += 1
        pf, cf = entry if entry else ('', '')
        row_missing = False
        if parte_idx == 'VCP':
            if pf:
                row[9] = pf
            else:
                row[9] = ''
                row_missing = True
        else:
            row[9] = '-'
        if contra_idx == 'VCP':
            if cf:
                row[10] = cf
            else:
                row[10] = ''
                row_missing = True
        else:
            row[10] = '-'
        if row_missing:
            row[-4] = _ACC_FACTOR_STATUS_MISSING
            missing += 1
    return matched, missing

_ACC_FACTOR_KINDS = {
    'cem': {'lob': 'CEM',     'parser': lambda fn, raw: _acc_parse_cem_factors(fn, raw)},
    'edg': {'lob': 'EDG',     'parser': lambda fn, raw: _acc_parse_direct_factors(fn, raw)},
    'hyb': {'lob': 'Hybrids', 'parser': lambda fn, raw: _acc_parse_direct_factors(
                                            fn, raw, cetip_col=1, parte_col=11, contra_col=12)},
}

def _accrual_source_dir(ymd):
    """ACCRUAL_SOURCE_ROOT\\YYYY\\mm. Month\\DD for a 'YYYYMMDD' run date."""
    ref = datetime.strptime(ymd, '%Y%m%d')
    month_folder = ref.strftime('%m') + '. ' + _R()._EN_MONTH_NAMES[ref.month - 1]
    return os.path.join(ACCRUAL_SOURCE_ROOT, ref.strftime('%Y'), month_folder, ref.strftime('%d'))

def _accrual_is_vcp_name(n):
    nl = n.lower()
    return ('vcp' in nl) or ('instrumentofin' in nl) or ('intrumentofin' in nl)

_ACC_FI_KEY = 'swap-atualizacao-pu-fator'            # cadastro do File Interface

_ACC_VIEW_BY_PREFIX = {'73760': 'BANCO', '04880': 'BANCO', '85398': 'ATACAMA', '00041': 'LAWTON'}

_ACC_VIEW_PART_NAME = {'BANCO': 'JPMORGANBM', 'LAWTON': 'INTRAGLAWTONFDO', 'ATACAMA': 'INTRAGATACAMAFDO'}

_ACC_LOB_TAG = {'CEM': 'CEM', 'EDG': 'EDG', 'Hybrids': 'HYB', 'Commodities': 'COMM'}

def _acc_swap_fator(f):
    """Factor → 2 integer + 8 decimal digits, no separator, absolute. 1.0 → '0100000000'."""
    try:
        n = abs(float(str(f or '').replace(',', '.')))
    except (ValueError, TypeError):
        n = 0.0
    ip, fp = '{:.8f}'.format(n).split('.')
    return ip[-2:].rjust(2, '0') + fp

def _acc_swap_header(view, today):
    """Linha de header (tipo 0) — literais e larguras do cadastro; participante
    e data entram por seq."""
    return _R()._fi_build_line(_ACC_FI_KEY, 'header',
                          {'4': _ACC_VIEW_PART_NAME.get(view, view), '5': today},
                          page_url='/accrual-swap')

def _acc_swap_records(row, today):
    """Return a list of {view, line} for one accrual row (empty when no VCP leg)."""
    codigo = str(row[0] or '').strip()
    accP, idxP = row[3], str(row[5] or '').strip().upper()
    accC, idxC = row[6], str(row[8] or '').strip().upper()
    fatP, fatC = row[9], row[10]
    digP = re.sub(r'\D', '', str(accP or '')); digC = re.sub(r'\D', '', str(accC or ''))
    numP = int(digP or '0'); numC = int(digC or '0')
    roleP = '01' if numP > numC else '00'
    roleC = '01' if numC > numP else '00'
    legs = []                                       # (curva, fator) per VCP leg
    if idxP == 'VCP': legs.append((roleP, fatP))
    if idxC == 'VCP': legs.append((roleC, fatC))
    if not legs:
        return []
    prefP, prefC = digP[:5], digC[:5]
    updaters = [(roleP, prefP)]                      # PARTE (our house entity) always updates
    if prefC in _ACC_VIEW_BY_PREFIX and prefC != prefP:
        updaters.append((roleC, prefC))             # group counterparty also submits its view
    out = []
    for papel, pref in updaters:
        view = _ACC_VIEW_BY_PREFIX.get(pref)
        if not view:
            continue
        for curva, fat in legs:
            meu = ''.join(_R().random.choice('0123456789') for _ in range(10))
            line = _R()._fi_build_line(_ACC_FI_KEY, 'registro',
                                  {'4': codigo, '5': papel, '7': curva,
                                   '8': today, '9': meu,
                                   '11': _acc_swap_fator(fat)},
                                  page_url='/accrual-swap')
            out.append({'view': view, 'line': line})
    return out

def _acc_write_batch_files(data, lob, today, evidence_dir=None):
    """Generate + write ACCRUAL_<view>-<lob>.txt for one LOB book, split by view.
    Written to the Batch Conecta folder AND (best-effort) to the evidence folder
    (Regulatory\\Accrual\\YYYY\\mm. Month\\DD). Returns [{filename, path, view, count}]."""
    by_view = {}
    for r in ((data.get('tables') or {}).get(lob) or []):
        if not r or len(r) < 15:
            continue
        if str(r[-4] or '') == _ACC_FACTOR_STATUS_MISSING:    # skip rows without a factor
            continue
        for rec in _acc_swap_records(r, today):
            by_view.setdefault(rec['view'], []).append(rec['line'])
    if not by_view:
        return []
    lob_tag = _ACC_LOB_TAG.get(lob, str(lob).upper())
    os.makedirs(_R().CONECTA_NEW_PATH, exist_ok=True)
    if evidence_dir:
        try:
            os.makedirs(evidence_dir, exist_ok=True)
        except Exception:
            _R().log.warning('[accrual] could not create evidence dir %s:\n%s', evidence_dir, traceback.format_exc())
    generated = []
    for view in ('BANCO', 'LAWTON', 'ATACAMA'):
        lines = by_view.get(view)
        if not lines:
            continue
        content = '\n'.join([_acc_swap_header(view, today)] + lines)
        fpath = _R()._unique_filepath(_R().CONECTA_NEW_PATH, 'ACCRUAL_{}-{}.txt'.format(view, lob_tag))
        with open(fpath, 'w', encoding='utf-8') as fh:
            fh.write(content)
        # Evidence copy (same base name), best-effort — never blocks the Conecta write.
        if evidence_dir and os.path.isdir(evidence_dir):
            try:
                with open(os.path.join(evidence_dir, os.path.basename(fpath)), 'w', encoding='utf-8') as fh:
                    fh.write(content)
            except Exception:
                _R().log.warning('[accrual] evidence copy failed for %s:\n%s', fpath, traceback.format_exc())
        generated.append({'filename': os.path.basename(fpath), 'path': fpath, 'view': view, 'count': len(lines)})
    return generated

def _acc_missing_accrual_rows(data, lobs):
    """Rows flagged 'Missing Accrual' (no updated factor) across the given LOB books.
    Their presence blocks file generation. Returns [{id, lob, codigo}]."""
    out = []
    for lob in lobs:
        for r in ((data.get('tables') or {}).get(lob) or []):
            if r and len(r) >= 15 and str(r[-4] or '') == _ACC_FACTOR_STATUS_MISSING:
                out.append({'id': str(r[-1]), 'lob': lob, 'codigo': str(r[0] or '')})
    return out

_ACC_RECON_ACCOUNTS = {'04880006', '73760009'}

_ACC_RECON_MARKER = 'REGISTRO DE PU/FATOR'

_ACC_RECON_HEADER_ROW = 5                  # 1-based → data starts at index 5

def _acc_run_recon(data, rows):
    """Gather registered factors per Código IF from the operacoes rows, then flag each
    VCP leg OK/Check by simple factor membership. Mutates data (recon + status)."""
    by_cif = {}                                           # cif_key -> [rounded floats]
    for i in range(_ACC_RECON_HEADER_ROW, len(rows)):     # data from row 6 (index 5)
        row = rows[i]
        if _R()._acc_digits(_R()._cc_cell(row, 1)) not in _ACC_RECON_ACCOUNTS:       # col B house account
            continue
        if _R()._cc_cell(row, 4).strip().upper() != _ACC_RECON_MARKER:          # col E marker
            continue
        cif = _R()._cc_cell(row, 7).strip()                                     # col H título
        fac = _acc_parse_num(_R()._cc_cell(row, 15))                            # col P factor (comma→dot)
        if not cif or fac is None:
            continue
        for k in _acc_factor_keys(cif):
            by_cif.setdefault(k, []).append(round(fac, 8))

    def _regs(cif):
        for k in _acc_factor_keys(cif):
            if k in by_cif:
                return by_cif[k]
        return []

    recon_out, ok_rows, check_rows = {}, 0, 0
    for table in (data.get('tables') or {}).values():
        for r in table:
            if not r or len(r) < 15:
                continue
            idxP = str(r[5] or '').strip().upper()
            idxC = str(r[8] or '').strip().upper()
            legs = []                                       # (tag, accrual_factor); p=Parte, c=Contra
            if idxP == 'VCP': legs.append(('p', r[9]))
            if idxC == 'VCP': legs.append(('c', r[10]))
            if not legs:
                continue
            regs = _regs(str(r[0] or '').strip())
            regset = set(regs)
            regdisp = ', '.join('{:.8f}'.format(x) for x in regs)
            entry, all_ok = {}, True
            for tag, acc_fac in legs:
                accv = _acc_parse_num(acc_fac)
                ok = (accv is not None and round(accv, 8) in regset)
                if not ok:
                    all_ok = False
                entry[tag] = {'ok': ok, 'reg': regdisp}
            recon_out[str(r[-1])] = entry
            r[-4] = 'Success' if all_ok else 'Check'        # status
            if all_ok: ok_rows += 1
            else:      check_rows += 1
    data['recon'] = recon_out
    return {'success_rows': ok_rows, 'check_rows': check_rows, 'map_entries': len(by_cif)}

def _acc_find_operacoes(folder):
    if not os.path.isdir(folder):
        return None
    for fn in os.listdir(folder):
        if not os.path.isfile(os.path.join(folder, fn)):
            continue
        base = os.path.splitext(fn)[0].lower()
        base = (base.replace('ç', 'c').replace('õ', 'o').replace('ã', 'a')
                    .replace('é', 'e').replace('ô', 'o'))
        if base.startswith('operac'):
            return os.path.join(folder, fn)
    return None

def _acc_check_status_rows(data):
    """Return (all_check_rows, uncommented) — rows whose status is 'Check'."""
    checks, pending = [], []
    for lob, table in (data.get('tables') or {}).items():
        for r in table:
            if not r or len(r) < 15:
                continue
            if str(r[-4] or '').strip().lower() == 'check':
                comment = str(r[-5] or '').strip()
                item = {'id': str(r[-1]), 'lob': lob, 'codigo': str(r[0] or ''), 'comment': comment}
                checks.append(item)
                if not comment:
                    pending.append(item)
    return checks, pending
