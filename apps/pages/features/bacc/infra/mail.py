# -*- coding: utf-8 -*-
"""O anexo .xlsx e o e-mail que o carrega."""
import io
import smtplib
import traceback
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

from flask import render_template

from apps.pages.features.bacc import domain


def _routes():
    """Busca ATRASADA (ver `infra/persistence.py`): o contexto de aplicação, o
    logo, o cabeçalho de e-mail e os endereços SMTP são plataforma."""
    from apps.pages import routes
    return routes


def build_xlsx(rows):
    """Workbook openpyxl com as colunas de `domain.COLUMNS`, largura ajustada ao
    conteúdo.

    O openpyxl não tem auto-fit de verdade (quem calcula a largura de um texto é
    o Excel, na hora de desenhar), então a largura sai da CONTAGEM DE CARACTERES
    da coluna inteira, cabeçalho incluído, com um piso e um teto. O teto existe
    porque o assunto do e-mail em Comments tem 120 caracteres e, sem ele, essa
    coluna sozinha empurraria as outras onze para fora da tela.

    A Trade Date sai como DATA de verdade (number_format dd/mm/yyyy) e o Aging
    como INTEIRO: escrever texto deixaria as duas colunas em General, e quem
    recebe não conseguiria ordenar nem filtrar por elas. Valor que não parseia
    sai como veio — sumir com ele seria pior.
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    R = _routes()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'EA METRICS'
    bold = Font(bold=True)
    larguras = []
    for j, (header, _src, _kind) in enumerate(domain.COLUMNS, start=1):
        c = ws.cell(row=1, column=j, value=header)
        c.font = bold
        c.alignment = Alignment(horizontal='center', vertical='center')
        larguras.append(len(header))
    for i, r in enumerate(rows, start=2):
        for j, (_header, src, kind) in enumerate(domain.COLUMNS, start=1):
            raw = str(domain.value(r, src) or '').strip()
            if not raw:
                continue
            cell = ws.cell(row=i, column=j)
            if kind == 'date':
                dt = R._parse_date_any(raw)
                if dt is not None:
                    cell.value = datetime(dt.year, dt.month, dt.day)
                    cell.number_format = 'DD/MM/YYYY'
                    raw = '00/00/0000'          # o que a célula OCUPA na tela
                else:
                    cell.value = raw            # texto livre numa coluna de data
            elif kind in ('num', 'money'):
                n = domain.num(raw)
                # Número que não parseia sai COMO VEIO: uma célula de texto no
                # meio de uma coluna de números é menos ruim do que sumir com o
                # valor que está no banco. E aí NÃO leva máscara — a máscara
                # sobre um texto não faz nada, mas prometeria um número.
                cell.value = raw if n is None else n
                if n is not None and kind == 'money':
                    cell.number_format = domain.MONEY_FMT
                    # A largura tem de medir o que se VÊ, não o que está no
                    # banco: '1500000' são 7 caracteres e a célula desenha
                    # '1.500.000,00', que são 12. Sem isto a coluna nasce
                    # estreita e o Excel mostra ####.
                    raw = '{:,.2f}'.format(n)
            else:
                cell.value = raw
            larguras[j - 1] = max(larguras[j - 1], len(raw))
    for j, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(j)].width = max(10, min(w + 3, 48))
    ws.freeze_panes = 'A2'
    return wb


def send(rows, to_list, cc_list, ref):
    """Monta e envia o e-mail com o anexo. True ou a mensagem do erro.

    O contexto de aplicação envolve a MONTAGEM INTEIRA e não só o
    `render_template`: o `_get_logo_path` lê `current_app.root_path`, e envolver
    só o render troca um "Working outside of application context" por outro três
    linhas abaixo. Dentro do request do botão Run isto é no-op — e é justamente
    por isso que o Run funcionava e o automático morria em silêncio.
    """
    from email.mime.image import MIMEImage
    from email.mime.application import MIMEApplication
    R = _routes()
    nome = domain.attach_name(ref)
    try:
        with R._app_context():
            html = render_template('pages/email-template-bacc-ea-metrics.html',
                                   ref_date_fmt=ref.strftime('%d/%m/%Y'),
                                   rows_n=len(rows), attach_name=nome,
                                   current_year=datetime.now().year)
            msg = MIMEMultipart('mixed')
            msg['Subject'] = domain.SUBJECT
            msg['From'] = R.SHARED_MAILBOX
            msg['To'] = ', '.join(to_list)
            if cc_list:
                msg['Cc'] = ', '.join(cc_list)
            corpo = MIMEMultipart('related')
            alt = MIMEMultipart('alternative')
            alt.attach(MIMEText('Please view this report in HTML.', 'plain', 'utf-8'))
            alt.attach(MIMEText(html, 'html', 'utf-8'))
            corpo.attach(alt)
            logo_path = R._get_logo_path()
            if logo_path:
                with open(logo_path, 'rb') as f:
                    limg = MIMEImage(f.read())
                limg.add_header('Content-ID', '<otc_logo>')
                limg.add_header('Content-Disposition', 'inline', filename='logo.png')
                corpo.attach(limg)
            R._attach_email_gradient(corpo)
            msg.attach(corpo)
            buf = io.BytesIO()
            build_xlsx(rows).save(buf)
            anexo = MIMEApplication(
                buf.getvalue(),
                _subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            anexo.add_header('Content-Disposition', 'attachment', filename=nome)
            msg.attach(anexo)
        with smtplib.SMTP(R.SMTP_HOST, R.SMTP_PORT, timeout=30) as server:
            server.sendmail(R.SHARED_MAILBOX, to_list + cc_list, msg.as_string())
        R.log.info('[bacc-ea] enviado — %d linha(s) · to=%s · cc=%s', len(rows), to_list, cc_list)
        return True
    except Exception as e:                                  # noqa: BLE001
        R.log.error('[bacc-ea] envio FALHOU:\n%s', traceback.format_exc())
        return '{}: {}'.format(type(e).__name__, e)
