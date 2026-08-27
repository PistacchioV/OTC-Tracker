# -*- coding: utf-8 -*-
"""Os arquivos do card: a planilha no share, o claim, o status e a foto."""
import json
import os
import traceback
from datetime import datetime

from apps.config import Config
from apps.pages.features.pcx import domain


def _routes():
    """Busca ATRASADA — ver `features/support/infra/persistence.py`."""
    from apps.pages import routes
    return routes


DIR = os.getenv('PCX_SPREADSHEET_DIR', os.path.join(
    Config.SHARED_DRIVE_ROOT, 'Confirmation', 'Derivativos', 'Movimento', 'Pending Confirmation'))
TIME_RAW = os.getenv('PCX_SPREADSHEET_TIME', '10:45')      # BRT (= 19:15 IST)


def claim_file():
    return os.path.join(_routes()._DAILY_METRIC_DIR, 'pcx_spreadsheet_sent.json')


def status_file():
    return os.path.join(_routes()._DAILY_METRIC_DIR, 'pcx_spreadsheet_status.json')


def read_status():
    try:
        with open(status_file(), encoding='utf-8') as fh:
            d = json.load(fh)
        return d if isinstance(d, dict) else {}
    except (IOError, OSError, json.JSONDecodeError):
        return {}


def snapshot_path(d):
    """A foto diária do Pending Confirmation daquele dia
    (`cache/pending-confirmation/AAAA/MM/DD/pending-confirmation_AAAAMMDD.json`),
    gravada pela manutenção das 11:30 (`_pc_snapshot_pending`)."""
    # normpath porque este caminho VAI PARA A TELA quando o snapshot não existe,
    # e `.../pages/../static/...` faz quem lê procurar a pasta errada.
    return os.path.normpath(
        os.path.join(_routes()._PC_SNAPSHOT_DIR, d.strftime('%Y'), d.strftime('%m'),
                     d.strftime('%d'),
                     'pending-confirmation_{}.json'.format(d.strftime('%Y%m%d'))))


def build_xlsx(rows):
    """Workbook openpyxl no layout combinado com o time global, linha a linha.

    O layout inteiro é contrato com o consumidor — o time global de métricas
    lê o arquivo via OLEDB (Confirmation_Latam) — e cada divergência quebrou
    de um jeito diferente (§250/§253/§256):

      * a ABA é CONFIRMATIONS (com outro nome: "CONFIRMATIONS$ is not a
        valid name");
      * os CABEÇALHOS ficam na LINHA 1 e os dados começam na 2, SEM o título
        mesclado da planilha antiga: o consumidor refez a query sobre a lista
        de colunas do §256 (o título da era legada — §253 — saiu do contrato
        junto). Cabeçalho fora da linha que o leitor espera vira "No value
        given for one or more required parameters";
      * as COLUNAS — nomes e ordem — são exatamente as de `domain.COLUMNS`,
        inclusive as vazias: tirá-las deslocaria as demais.

    As colunas de data saem como DATA de verdade com number_format dd/mm/yyyy
    (Short Date) — escrever o texto deixaria a célula General e o Excel do
    consumidor sem ordenar/filtrar por data. Valor que não parseia (texto
    livre numa coluna de data) sai como veio: sumir com ele seria pior."""
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'CONFIRMATIONS'
    bold = Font(bold=True)
    for j, (header, _src, _is_date) in enumerate(domain.COLUMNS, start=1):
        c = ws.cell(row=1, column=j, value=header)
        c.font = bold
    for i, r in enumerate(rows, start=2):
        for j, (_header, src, is_date) in enumerate(domain.COLUMNS, start=1):
            raw = str(r.get(src, '') or '').strip() if src else ''
            if not raw:
                continue
            cell = ws.cell(row=i, column=j)
            if is_date:
                dt = _routes()._parse_date_any(raw)
                if dt is not None:
                    cell.value = datetime(dt.year, dt.month, dt.day)
                    cell.number_format = 'DD/MM/YYYY'
                else:
                    cell.value = raw
            elif raw.lstrip('-').isdigit():
                cell.value = int(raw)                      # Aging ordena como número
            else:
                cell.value = raw
    for j, (header, _src, _is_date) in enumerate(domain.COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(j)].width = max(12, min(len(header) + 4, 34))
    ws.freeze_panes = 'A2'          # cabeçalho (linha 1) congelado
    return wb


def save_spreadsheet(rows):
    """Gera e grava a planilha no share. → (n linhas, caminho).

    A escrita é em arquivo temporário + `os.replace`: quem abrir a planilha no
    meio da gravação vê a versão anterior inteira, nunca um xlsx pela metade.

    **Sempre o nome canônico**, com ou sem `ref` — é assim de propósito. O time
    global de métricas lê ESSE arquivo por OLEDB (`Confirmation_Latam`), e é por
    ele que a mesa entrega uma data anterior quando pedem: grava-se a foto do dia
    pedido, o time puxa, e a próxima corrida (o Run de hoje ou a rotina das 10:45)
    devolve o arquivo à data corrente. Um arquivo datado ao lado não seria visto
    por quem consome — o consumidor tem um caminho só.

    O preço é que, entre uma coisa e outra, o arquivo do share contém a foto de
    outro dia. Quem diz isso é o **status do card** (`ref` no
    `write_status`): sem essa marca, nada na tela distinguiria o arquivo de
    hoje do de 08/08."""
    wb = build_xlsx(rows)
    os.makedirs(DIR, exist_ok=True)
    fp = os.path.join(DIR, domain.FILENAME)
    tmp = fp + '.tmp'
    wb.save(tmp)
    os.replace(tmp, fp)
    return len(rows), fp


def claim_slot(slot):
    """Reserva o disparo do dia EM DISCO (mesma razão do Deals Monitor: com
    mais de um processo, a trava em memória não impede gravação dupla — aqui o
    arquivo é idempotente, mas o claim é o que faz o catch-up saber o que já
    rodou)."""
    R = _routes()
    return R._claim_daily_slot(claim_file(), R._DAILY_METRIC_DIR, slot, 16, 'pending-spreadsheet')


def release_slot(slot):
    """Devolve o slot quando a gravação falhou (share fora, arquivo aberto com
    lock no Excel): o catch-up da próxima volta tenta de novo. Sem isto uma
    falha transitória custava a planilha do dia inteiro."""
    _routes()._release_daily_slot(claim_file(), slot, 'pending-spreadsheet')


def write_status(slot, result, when, ref=None):
    """Desfecho da última gravação. `ref` é a data da FOTO que está no arquivo
    agora — vazia quando é a situação corrente.

    Como a planilha histórica sobrescreve o nome canônico, esta é a única coisa
    que responde "o que está no share neste momento?". O arquivo é reescrito
    inteiro a cada gravação, então a marca cai sozinha na próxima corrida
    normal — que é exatamente o comportamento desejado."""
    R = _routes()
    try:
        os.makedirs(R._DAILY_METRIC_DIR, exist_ok=True)
        R._atomic_write_json(status_file(), {
            'slot': slot, 'result': result,
            'at': when.strftime('%d/%m/%Y %H:%M:%S'),
            'ref': ref.strftime('%d/%m/%Y') if ref else '',
        })
    except Exception:                                       # noqa: BLE001
        R.log.warning('[pending-spreadsheet] não consegui gravar o status:\n%s',
                    traceback.format_exc())

