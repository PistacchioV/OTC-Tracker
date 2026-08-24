# -*- coding: utf-8 -*-
"""Reconciliação de CGD — o batimento entre a lista do FEP/CMS e a posição da B3.

Tradução do workflow Alteryx `Batimento CGD` (`Alteryx CGD.yxmd`, 150 nós). A
pergunta que ele responde é uma só: **de quem temos Contrato Global de
Derivativos assinado, e de quem a B3 reconhece que temos?** As duas listas
divergem nos dois sentidos, e cada sentido é um trabalho diferente:

| Bucket | O que é | O que fazer |
|---|---|---|
| `No breaks` | o CNPJ está nos dois lados | nada |
| `Only FEP` | CGD assinado, sem contraparte na B3 | falta INCLUIR na B3 |
| `Only B3` | a B3 reconhece, e não há CGD assinado | conferir o cadastro |
| `Pending action` | está nos dois, mas o CGD não fechou | cobrar a assinatura |

E duas justificativas tiram o `Only FEP` da fila sem que ele vire pendência:
o cliente **é garantidor** de outro (o CGD vale pela ponta garantida) ou a
**conta foi encerrada**.

O que MUDOU do Alteryx, e por quê
---------------------------------

- **O calendário é o ANBIMA do app** (`static/data/anbima.json`), não a aba
  `Feriados` do `Auxiliar.xlsx`. Dois calendários no mesmo app divergem no
  primeiro feriado que só um dos dois conhece, e o D-1 daqui é o mesmo D-1 das
  outras telas.
- **As contas da B3 saem do cadastro `b3-accounts`** (`ACCOUNT TYPE = OWN` e
  `LE` em `CGD_LES`). O workflow tinha `73760.00-9` e `04880.00-6` escritos no
  filtro; são as contas PRÓPRIAS do JPM e do MGT, e elas já estão cadastradas —
  repetir os números aqui criaria uma segunda lista das mesmas contas (§6). O
  recorte por LE continua sendo do processo: os fundos do grupo também têm conta
  própria na B3 e não assinam CGD com ninguém.
- **As três tabelas do `Auxiliar.xlsx` viraram cadastro do /mapping**:
  `cgd-b3-participante` (o de-para do participante sem CNPJ), `cgd-garantidor` e
  `cgd-conta-encerrada`. Eram planilhas mantidas à mão numa pasta de rede: o
  cadastro novo só entrava por quem tivesse o arquivo aberto, e a recon rodava
  com a lista de ontem sem dizer nada.
- **CNPJ compara por DÍGITO**, dos dois lados. A B3 manda `12.345.678/0001-99`,
  a planilha manda `12345678000199`, e comparar texto casa NADA em silêncio
  (§197).
- O workflow gravava um `Dummy.xlsx` só para fechar o fluxo e um
  `CGD Reconciliation - dd.mm.yy.xlsx` ao lado do e-mail. Aqui o resultado fica
  no cache do dia (JSON) e a planilha é o Export da própria tela — um arquivo
  gravado no share por uma rotina que ninguém pediu é lixo que ninguém apaga.

O que NÃO mudou
---------------

Os filtros do FEP são os do workflow, e continuam literais porque são a regra do
processo, não um de-para: fora `TIPO OPERAÇÃO = ADITAMENTO AO CONTRATO GLOBAL DE
DERIVATIVOS`, fora `STATUS CMS = Cancelado`, e "assinado" é um dos três
(`DOC TRANSACIONAL`, `Assinatura Concluida`, `Assinatura Manualmente`). O
`DOC TRANSACIONAL` é EXIBIDO como `Docusign`, como no Alteryx.

O aging é o do workflow — **dias corridos** desde a data de criação mais RECENTE
daquele CNPJ no FEP —, e não dias úteis: é o relógio que a mesa já usa para
cobrar, e trocá-lo mudaria todo número que ela conhece. O CNPJ com mais de uma
solicitação conta pela última, que é a que está em aberto.
"""

import json
import logging
import os
import re
import tempfile
import unicodedata
from datetime import date, datetime, timedelta

from apps.config import Config

_LOG = logging.getLogger(__name__)

_MODULE_DIR = os.path.dirname(os.path.abspath(__file__))
_DATA_DIR = os.path.normpath(os.path.join(_MODULE_DIR, '..', 'static', 'data'))
_MAPPINGS_DIR = os.path.join(_DATA_DIR, 'mappings')
_CACHE_DIR = os.path.join(_DATA_DIR, 'cache', 'reconciliation', 'cgd')

# ── Entradas ────────────────────────────────────────────────────────────────
# A pasta de input do batimento (a lista do FEP). Env var como todo destino de
# rede do app; o default pende do `Config.SHARED_DRIVE_ROOT`, nunca de um `I:\`
# escrito aqui (§309).
CGD_INPUT_ROOT = os.getenv(
    'CGD_INPUT_ROOT',
    os.path.join(Config.SHARED_DRIVE_ROOT, 'Confirmation', 'Derivativos',
                 'Alteryx', 'Batimento CGD', 'Input'))
FEP_XLSX = os.getenv('CGD_FEP_XLSX', 'LISTA_CONTRATOS_CGD.xlsx')

# ── A lista do FEP vem por E-MAIL, não de uma pasta ─────────────────────────
# O relatório do FepWeb chega como ANEXO `.xlsx` numa subpasta do box
# compartilhado — a mesma caixa que a varredura de booking recap e a Recon de
# Comitentes já leem. Ninguém salva esse arquivo em lugar nenhum, então apontar
# o batimento para uma pasta era apontá-lo para um arquivo que alguém teria de
# copiar à mão todo dia (e, no dia em que esquecesse, a recon rodaria com a
# lista da semana passada sem dizer nada).
#
# A pasta é `Inbox > Automatico`, e ela recebe MAIS COISA do que este relatório:
# quem identifica o e-mail certo é o ASSUNTO. Procurar pelo anexo mais recente da
# pasta pegaria o `.xlsx` de qualquer outra rotina automática que caia ali.
#
# Lê-se o e-mail MAIS RECENTE entre os que casam: o relatório é reemitido e a
# pasta acumula. E o assunto e a data do e-mail escolhido voltam no resultado,
# porque "de que dia é esta lista" é a primeira pergunta de quem olha uma quebra.
#
# O assunto casa por PEDAÇO e normalizado (caixa, acento e espaço fora), que é o
# que sobrevive a um `RES:` na frente, a uma data no fim e à grafia de quem
# reencaminhou.
#
# Nada é apagado nem movido no box: a rotina só LÊ.
FEP_MAILBOX = os.getenv('OTC_BOX_MAILBOX', 'brazil.otc.ops@jpmorgan.com')
FEP_MAIL_FOLDER = tuple(
    x for x in os.getenv('CGD_FEP_MAIL_FOLDER', 'Automatico').split('|') if x.strip())
FEP_MAIL_SUBJECT = os.getenv(
    'CGD_FEP_MAIL_SUBJECT',
    'FEPWEB-CGD-ContratoGlobalDerivativos - SEM FILTRO DATAS')
_FEP_MAIL_EXT = ('.xlsx', '.xlsm')
# PidTagNormalizedSubjectW — o assunto como propriedade MAPI, para o Restrict.
_MAPI_SUBJECT = 'http://schemas.microsoft.com/mapi/proptag/0x0E1D001F'

# A árvore das posições CETIP é a de DESTINO da rotina Save CETIP Files
# (`CETIP_DEST_ROOT`), e não a de origem: o que a B3 despeja na pasta de download
# é o arquivo cru, com o nome do dia em que foi baixado, e quem o filtra, renomeia
# para a convenção da casa e o guarda no dia certo é a rotina. Ler a origem era
# ler antes de a rotina passar — e num dia em que ela não rodasse, a recon acharia
# um arquivo e diria que está tudo certo com a posição da véspera.
#
# É a MESMA raiz e o MESMO env var que o `recon_fxo` usa para a DPOSICAO de opção
# — duas raízes seriam duas verdades sobre onde está o arquivo do dia.
CETIP_DEST_ROOT = os.getenv(
    'CETIP_DEST_ROOT',
    os.path.join(Config.SHARED_DRIVE_ROOT, 'Confirmation', 'Derivativos',
                 'OTC Tracker', 'CETIP Files', 'Position Files'))
# O nome que a rotina GRAVA (regra `CGD (NET)`: o `.txt` entra no destino). O
# nome sem extensão é o do arquivo cru e fica como segunda tentativa, para o dia
# em que alguém puser o arquivo na pasta à mão. CSV de 10 campos, sem cabeçalho.
B3_FILE_TPL = 'CETIP21_{yymmdd}_DPOSICAO-NET.txt'
B3_FILE_ALT = 'CETIP21_{yymmdd}_DPOSICAO-NET'

_EN_MONTHS = ('January', 'February', 'March', 'April', 'May', 'June',
              'July', 'August', 'September', 'October', 'November', 'December')

# As colunas do DPOSICAO-NET que interessam, por POSIÇÃO (o arquivo não tem
# cabeçalho). Índices 0-based sobre o split por vírgula, como o TextToColumns.
B3_COLS = {1: 'Contrato B3', 2: 'Parte', 3: 'Nome Parte', 4: 'Cód. Contraparte',
           5: 'Nome Contraparte', 6: 'CNPJ', 7: 'Razão Social', 8: 'Tipo'}

# ── A regra do FEP (literais do workflow) ────────────────────────────────────
TIPO_FORA = 'ADITAMENTO AO CONTRATO GLOBAL DE DERIVATIVOS'
STATUS_FORA = 'CANCELADO'
STATUS_ASSINADO = ('DOC TRANSACIONAL', 'ASSINATURA CONCLUIDA', 'ASSINATURA MANUALMENTE')
STATUS_DOCUSIGN = 'DOC TRANSACIONAL'
# Os dois status que, convivendo no mesmo CNPJ, carimbam a observação do
# workflow: o cliente tem um documento no Docusign E um contrato no FEP.
OBS_DOCUSIGN_FEP = 'Docusign + FEP'

# Faixas do aging que o e-mail pinta. São as regras das duas tabelas do
# PortfolioComposer: a de inclusão na B3 avisa em 5 e alarma em 15; a de
# cobrança de assinatura avisa em 5 e alarma em 10.
AGING_RULES = {'pending_b3': (5, 15), 'pending_action': (5, 10)}

# ── E-mail ──────────────────────────────────────────────────────────────────
_SMTP_HOST = 'mailhost.jpmchase.net'
_SMTP_PORT = 25
_FROM = 'brazil.otc.ops@jpmorgan.com'


# ── Normalização ─────────────────────────────────────────────────────────────

def _digits(v):
    """Só os dígitos. É assim que CNPJ compara com CNPJ neste app: os dois lados
    guardam pontuação diferente e comparar texto casa nada, em silêncio (§197)."""
    return re.sub(r'\D', '', str(v or ''))


def _norm(s):
    """Sem acento, sem caixa, sem espaço duplo — para comparar status e nome."""
    s = unicodedata.normalize('NFKD', str(s or ''))
    s = ''.join(c for c in s if not unicodedata.combining(c))
    return ' '.join(s.upper().split())


def _parse_date(v):
    """Data de qualquer das grafias que chegam da planilha do FEP."""
    if v is None or v == '':
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s:
        return None
    if s.isdigit() and 20000 <= int(s) <= 80000:
        return (datetime(1899, 12, 30) + timedelta(days=int(s))).date()
    s = s.split(' ')[0].split('T')[0]
    for fmt in ('%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d/%m/%y', '%m/%d/%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _fmt_date(d):
    return d.strftime('%d/%m/%Y') if isinstance(d, date) else ''


# ── Calendário: o D-1 do arquivo da B3 ───────────────────────────────────────

_ANBIMA = {'feriados': None}


def _feriados():
    if _ANBIMA['feriados'] is None:
        try:
            with open(os.path.join(_DATA_DIR, 'anbima.json'), encoding='utf-8') as fh:
                _ANBIMA['feriados'] = {d['date'] for d in (json.load(fh) or []) if d.get('date')}
        except Exception:
            # Sem o arquivo o D-1 vira "ontem que não é fim de semana": erra por
            # feriado, mas a recon não some da tela.
            _LOG.warning('[recon-cgd] anbima.json não pôde ser lido; o D-1 vai '
                         'ignorar feriados')
            _ANBIMA['feriados'] = set()
    return _ANBIMA['feriados']


def dia_util_anterior(ref=None):
    """O último dia útil ANTES de `ref` (padrão: hoje).

    É o `GenerateRows` + `Filter` + `Join Feriados` + `Sort desc` + `Sample 1` do
    workflow, que é um jeito complicado de dizer isto. O teto de dez voltas é o
    mesmo do original: mais que isso é feriadão que não existe no Brasil, e um
    laço sem teto sobre um `anbima.json` vazio giraria para sempre.
    """
    d = (ref or date.today()) - timedelta(days=1)
    fer = _feriados()
    for _ in range(10):
        if d.weekday() < 5 and d.strftime('%Y-%m-%d') not in fer:
            return d
        d -= timedelta(days=1)
    return d


def caminho_b3(dia):
    """`{destino}/{AAAA}/{MM}. {Month}/{DD}/CETIP21_{AAMMDD}_DPOSICAO-NET.txt`.

    A árvore é a que o Save CETIP Files ESCREVE, com o mês em inglês por extenso
    — é a convenção do share, não uma escolha desta tela. Existindo só o nome sem
    extensão (arquivo posto à mão), é ele que volta: a alternativa é devolver um
    caminho que não existe e dizer que a posição do dia não chegou.
    """
    pasta = os.path.join(
        CETIP_DEST_ROOT, dia.strftime('%Y'),
        '{}. {}'.format(dia.strftime('%m'), _EN_MONTHS[dia.month - 1]),
        dia.strftime('%d'))
    principal = os.path.join(pasta, B3_FILE_TPL.format(yymmdd=dia.strftime('%y%m%d')))
    if os.path.isfile(principal):
        return principal
    alt = os.path.join(pasta, B3_FILE_ALT.format(yymmdd=dia.strftime('%y%m%d')))
    return alt if os.path.isfile(alt) else principal


# ── Cadastros (/mapping) ─────────────────────────────────────────────────────

_MAP_CACHE = {}


def _mapping_rows(key):
    """As linhas de um cadastro do /mapping, cacheadas por mtime.

    Lido DIRETO do JSON: importar o `routes` daqui seria circular, e é o mesmo
    caminho que o `recon_fxo` usa para os cadastros dele. Arquivo ausente devolve
    lista vazia — a instância em que ninguém abriu a tela de /mapping roda a
    recon do mesmo jeito, só sem as regras que ninguém cadastrou.
    """
    path = os.path.join(_MAPPINGS_DIR, '{}.json'.format(key))
    try:
        mt = os.path.getmtime(path)
    except OSError:
        return []
    ent = _MAP_CACHE.get(key)
    if ent and ent[0] == mt:
        return ent[1]
    try:
        with open(path, encoding='utf-8') as fh:
            linhas = json.load(fh) or []
    except Exception:
        linhas = []
    linhas = linhas if isinstance(linhas, list) else []
    _MAP_CACHE[key] = (mt, linhas)
    return linhas


# As entidades que assinam CGD com CLIENTE. O workflow filtrava as contas
# `73760.00-9` e `04880.00-6`, que são as contas PRÓPRIAS do Banco J.P. Morgan e
# do JPMorgan Chase — os números saem do cadastro `b3-accounts` (§6: um de-para
# das mesmas contas escrito aqui envelheceria sozinho), mas o RECORTE continua
# sendo o do processo: as demais entidades do grupo (os fundos) têm conta própria
# na B3 e não assinam CGD com ninguém. Incluí-las encheria o `Only B3` de
# contraparte que não é cliente.
CGD_LES = ('JPM', 'MGT')


def contas_proprias():
    """As contas PRÓPRIAS das entidades que assinam CGD (`b3-accounts`).

    `ACCOUNT TYPE = OWN` e `LE` em `CGD_LES`. Cadastro vazio devolve vazio, e a
    recon AVISA em vez de silenciosamente não filtrar nada — sem filtro, o
    arquivo inteiro da B3 entraria no batimento.
    """
    out = []
    for r in _mapping_rows('b3-accounts'):
        if _norm(r.get('ACCOUNT TYPE')) != 'OWN':
            continue
        if _norm(r.get('LE')) not in CGD_LES:
            continue
        c = str(r.get('ACCOUNT') or '').strip()
        if c:
            out.append(c)
    return out


def _participantes():
    """`{nome simplificado normalizado: (CNPJ, Razão Social)}` do cadastro
    `cgd-b3-participante`.

    A linha da B3 que vem SEM CNPJ é a do participante (o omnibus), e o que a
    identifica é o nome simplificado. Sem cadastro ela é descartada — como no
    workflow, onde o `Filter CNPJ IsNull` corta o que não casou."""
    out = {}
    for r in _mapping_rows('cgd-b3-participante'):
        nome = _norm(r.get('NOME CONTRAPARTE'))
        if nome:
            out[nome] = (str(r.get('CNPJ') or '').strip(),
                         str(r.get('RAZAO SOCIAL') or '').strip())
    return out


def _cnpjs_do_cadastro(key, coluna='CNPJ / CPF'):
    """O conjunto de documentos de um cadastro, só dígitos."""
    out = set()
    for r in _mapping_rows(key):
        d = _digits(r.get(coluna) or r.get('CNPJ') or r.get('CPF'))
        if d:
            out.add(d)
    return out


# ── Lado B3 ──────────────────────────────────────────────────────────────────

def ler_b3(dia, avisos):
    """`{cnpj_digits: {'cnpj', 'nome'}}` das contrapartes da posição do dia.

    Três passos do workflow, na ordem: filtra as linhas das NOSSAS contas
    próprias, resolve pelo cadastro de participante as que vêm sem CNPJ, e
    descarta as que continuaram sem.
    """
    path = caminho_b3(dia)
    if not os.path.isfile(path):
        avisos.append('Arquivo da B3 não encontrado: {}'.format(path))
        return {}, path

    contas = {c.strip() for c in contas_proprias()}
    if not contas:
        avisos.append('Cadastro `b3-accounts` sem conta PRÓPRIA (ACCOUNT TYPE = OWN): '
                      'sem elas não há como saber quais linhas da B3 são nossas.')
        return {}, path

    part = _participantes()
    achou = {}
    lidas = sem_cnpj = resolvidas = 0
    # `latin-1`: é o code page 28591 do workflow, e o arquivo da CETIP vem nele.
    with open(path, encoding='latin-1', errors='replace') as fh:
        for linha in fh:
            campos = linha.rstrip('\n').split(',')
            if len(campos) < 9:
                continue
            reg = {nome: campos[i].strip() for i, nome in B3_COLS.items()}
            if reg['Parte'] not in contas:
                continue
            lidas += 1
            cnpj, razao = reg['CNPJ'], reg['Razão Social']
            if not _digits(cnpj):
                sem_cnpj += 1
                p = part.get(_norm(reg['Nome Contraparte']))
                if not p or not _digits(p[0]):
                    continue                       # sem cadastro, sai — como no fluxo
                cnpj, razao = p[0], (p[1] or reg['Nome Contraparte'])
                resolvidas += 1
            achou.setdefault(_digits(cnpj), {'cnpj': cnpj, 'nome': razao or reg['Nome Contraparte']})
    if sem_cnpj and sem_cnpj != resolvidas:
        # Linha que some sem dizer nada vira "sumiu um cliente da recon".
        avisos.append('{} linha(s) da B3 vieram sem CNPJ e {} foram resolvidas pelo '
                      'cadastro `cgd-b3-participante`; as demais ficaram de fora.'
                      .format(sem_cnpj, resolvidas))
    return achou, path


# ── Lado FEP ─────────────────────────────────────────────────────────────────

def _col_idx(cabecalho):
    """`{campo conhecido: índice}` do cabeçalho da planilha do FEP, casando por
    nome normalizado — a coluna muda de lugar quando alguém mexe na exportação,
    e por posição a leitura erraria calada."""
    alvo = {
        'CPF/CNPJ CLIENTE CMS': 'cnpj',
        'NOME CLIENTE CMS': 'cliente',
        'STATUS CMS': 'status',
        'TIPO OPERACAO': 'tipo',
        'ID DA OPERACAO': 'op',
        'CRIACAO': 'criacao',
    }
    out = {}
    for i, c in enumerate(cabecalho):
        k = alvo.get(_norm(c))
        if k and k not in out:
            out[k] = i
    return out


def _subpasta(folder, nome):
    """A subpasta pelo nome, cega a caixa e a espaço nas pontas, ou `None`.

    Pelo índice não dá: o nome da pasta é o que o time vê e digita, e a ordem
    delas no Outlook muda quando alguém cria outra.
    """
    subs = folder.Folders
    alvo = str(nome or '').strip().lower()
    for i in range(1, subs.Count + 1):
        f = subs.Item(i)
        if str(f.Name).strip().lower() == alvo:
            return f
    return None


def baixar_fep_do_box(avisos, destino=None):
    """Salva o anexo `.xlsx` do e-mail MAIS RECENTE cujo assunto é o do FepWeb.

    Devolve `(caminho, descricao)` — a descrição é o assunto e a data do e-mail
    lido, para o painel dizer DE QUE dia é a lista. `(None, '')` quando não deu
    para ler, sempre com o motivo em `avisos`: uma recon que rodou sem um dos
    lados parece uma recon limpa.

    Windows-only (COM/MAPI). Fora do Windows levanta `EnvironmentError`, e quem
    chama cai para o arquivo em pasta.
    """
    try:
        import win32com.client as _w
        import pythoncom
    except ImportError:
        raise EnvironmentError(
            'win32com não disponível: ler o anexo do FepWeb requer Windows com '
            'Outlook instalado.')

    caminho_txt = 'Inbox > ' + ' > '.join(FEP_MAIL_FOLDER)
    pythoncom.CoInitialize()
    try:
        ns = _w.Dispatch('Outlook.Application').GetNamespace('MAPI')
        pasta = ns.Folders[FEP_MAILBOX].Folders['Inbox']
        for nome in FEP_MAIL_FOLDER:
            sub = _subpasta(pasta, nome)
            if sub is None:
                avisos.append('Pasta do FepWeb não encontrada no box {}: {}'
                              .format(FEP_MAILBOX, caminho_txt))
                return None, ''
            pasta = sub

        # O pré-filtro do MAPI evita percorrer uma pasta que recebe todo dia.
        # Recusado (acontece em algumas caixas), varre-se tudo: o teste de assunto
        # abaixo é o que decide, e ele roda nos dois caminhos.
        try:
            itens = pasta.Items.Restrict(
                '@SQL="%s" LIKE \'%%%s%%\'' % (_MAPI_SUBJECT, FEP_MAIL_SUBJECT))
        except Exception:
            itens = pasta.Items
        # Mais recente primeiro. O relatório é reemitido e a pasta acumula:
        # pegar o primeiro da ordem natural devolveria o mais ANTIGO, e o
        # batimento rodaria com a lista de semanas atrás sem erro nenhum.
        try:
            itens.Sort('[ReceivedTime]', True)
        except Exception:
            _LOG.warning('[recon-cgd] não consegui ordenar a pasta por data')
        alvo_assunto = _norm(FEP_MAIL_SUBJECT)
        for i in range(1, itens.Count + 1):
            try:
                msg = itens.Item(i)
                if alvo_assunto not in _norm(getattr(msg, 'Subject', '')):
                    continue
                anexos = msg.Attachments
                for a in range(1, anexos.Count + 1):
                    at = anexos.Item(a)
                    nome = str(at.FileName or '')
                    if not nome.lower().endswith(_FEP_MAIL_EXT):
                        continue          # assinatura, imagem embutida, .msg
                    alvo = destino or os.path.join(
                        tempfile.gettempdir(), 'fepweb-cgd-' + os.path.basename(nome))
                    at.SaveAsFile(alvo)
                    recebido = ''
                    try:
                        recebido = msg.ReceivedTime.strftime('%d/%m/%Y %H:%M')
                    except Exception:
                        pass
                    desc = '{} — {} ({})'.format(nome, str(msg.Subject or '').strip(),
                                                 recebido or 's/ data')
                    _LOG.info('[recon-cgd] lista do FEP: %s', desc)
                    return alvo, desc
            except Exception:
                _LOG.exception('[recon-cgd] falha lendo um item da pasta do FepWeb')
        avisos.append('Nenhum e-mail com assunto "{}" e anexo .xlsx em {}.'
                      .format(FEP_MAIL_SUBJECT, caminho_txt))
        return None, ''
    finally:
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass


def ler_fep(avisos, path=None):
    """As linhas da lista do FepWeb, já filtradas pela regra do fluxo.

    A fonte é o ANEXO do e-mail (`baixar_fep_do_box`). `path` explícito vence —
    é o upload manual e o caminho dos testes. Sem Outlook (o servidor que não é
    Windows, a máquina de desenvolvimento), cai para o arquivo em
    `CGD_INPUT_ROOT`, avisando qual das duas fontes acabou valendo: rodar com a
    lista errada e não saber é a única falha aqui que não aparece.

    Devolve `{cnpj_digits: {...}}` com UMA entrada por CNPJ: o status é o do
    documento mais recente e a data de criação também. O workflow chegava no
    mesmo lugar por `Summarize` + `MultiRowFormula` ordenando por data.
    """
    origem = ''
    if not path:
        try:
            path, origem = baixar_fep_do_box(avisos)
        except EnvironmentError as e:
            avisos.append(str(e))
        if not path:
            path = os.path.join(CGD_INPUT_ROOT, FEP_XLSX)
            avisos.append('Usei a lista em pasta ({}) em vez do anexo do e-mail.'
                          .format(path))
    if not os.path.isfile(path):
        avisos.append('Lista do FEP não encontrada: {}'.format(path))
        return {}, origem or path
    # `path` é o ARQUIVO, que o openpyxl abre; `rotulo` é o que o painel e o
    # e-mail MOSTRAM. Anexo salvo em temporário tem nome que não diz nada — o
    # que a mesa precisa ler é o assunto e a data do e-mail de onde ele saiu.
    rotulo = origem or path
    try:
        from openpyxl import load_workbook
    except Exception:
        avisos.append('openpyxl não está instalado: sem ele não dá para ler a lista do FEP.')
        return {}, rotulo

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    linhas = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    if not linhas:
        avisos.append('Lista do FEP está vazia.')
        return {}, rotulo

    idx = _col_idx(linhas[0])
    faltando = [k for k in ('cnpj', 'status') if k not in idx]
    if faltando:
        avisos.append('A lista do FEP não tem as colunas {} — nada foi lido dela.'
                      .format(', '.join(faltando)))
        return {}, rotulo

    por_cnpj = {}
    for l in linhas[1:]:
        def cel(k):
            i = idx.get(k)
            return l[i] if (i is not None and i < len(l)) else ''
        status = _norm(cel('status'))
        tipo = _norm(cel('tipo'))
        if not status or status == STATUS_FORA:
            continue                                  # [41] STATUS CMS != Cancelado
        if tipo == _norm(TIPO_FORA):
            continue                                  # [216] fora o aditamento
        d = _digits(cel('cnpj'))
        if not d:
            continue
        criado = _parse_date(cel('criacao'))
        rec = por_cnpj.get(d)
        atual = {
            'cnpj': str(cel('cnpj') or '').strip(),
            'cliente': str(cel('cliente') or '').strip(),
            'status': str(cel('status') or '').strip(),
            'status_norm': status,
            'criacao': criado,
            'qtd': 1,
            'statuses': {status},
        }
        if not rec:
            por_cnpj[d] = atual
            continue
        rec['qtd'] += 1
        rec['statuses'].add(status)
        # A linha MAIS RECENTE manda no status e na data: é a solicitação que
        # está em aberto, e é dela que o aging conta.
        if criado and (not rec['criacao'] or criado > rec['criacao']):
            rec['criacao'] = criado
            rec['status'] = atual['status']
            rec['status_norm'] = status
            rec['cliente'] = atual['cliente'] or rec['cliente']
        elif not rec['cliente']:
            rec['cliente'] = atual['cliente']

    # A observação do workflow: o CNPJ que tem documento no Docusign E contrato
    # concluído no FEP ao mesmo tempo.
    for rec in por_cnpj.values():
        rec['obs'] = (OBS_DOCUSIGN_FEP
                      if {'DOC TRANSACIONAL', 'ASSINATURA CONCLUIDA'} <= rec['statuses']
                      else '')
    return por_cnpj, rotulo


# ── O batimento ──────────────────────────────────────────────────────────────

def _status_exibido(status):
    """`DOC TRANSACIONAL` se lê `Docusign` no relatório — é como a mesa chama."""
    return 'Docusign' if _norm(status) == STATUS_DOCUSIGN else status


def _assinado(rec):
    return rec.get('status_norm') in STATUS_ASSINADO


def executar(ref=None, fep_path=None):
    """Roda o batimento e devolve o resultado do dia.

    `ref` é a data da POSIÇÃO (o D-1 do arquivo da B3); o padrão é o último dia
    útil. Rodar de novo o mesmo dia dá o mesmo resultado: nada aqui depende de
    quando a rotina rodou, só do arquivo e dos cadastros.
    """
    avisos = []
    dia = ref or dia_util_anterior()
    b3, b3_path = ler_b3(dia, avisos)
    fep, fep_file = ler_fep(avisos, fep_path)

    garantidores = _cnpjs_do_cadastro('cgd-garantidor')
    encerradas = _cnpjs_do_cadastro('cgd-conta-encerrada')

    hoje = date.today()
    linhas = []

    for d, rec in fep.items():
        aging = (hoje - rec['criacao']).days if rec['criacao'] else ''
        base = {
            'cnpj': rec['cnpj'], 'client': rec['cliente'],
            'status': _status_exibido(rec['status']), 'obs': rec.get('obs', ''),
            'creation_date': _fmt_date(rec['criacao']), 'aging': aging,
        }
        if d in b3:
            # Está nos dois lados. Assinado é `No breaks`; não assinado é a
            # cobrança — o cliente opera e o contrato não fechou.
            base['check'] = 'No breaks'
            base['bucket'] = 'matched' if _assinado(rec) else 'pending_action'
        else:
            base['check'] = 'Only FEP'
            if not _assinado(rec):
                # Só no FEP e ainda não assinado: não é pendência de inclusão na
                # B3 — não há o que incluir. Fica na cobrança de assinatura.
                base['bucket'] = 'pending_action'
            elif d in garantidores:
                base['bucket'] = 'justified'
                base['obs'] = 'Guarantee'
            elif d in encerradas:
                base['bucket'] = 'justified'
                base['obs'] = 'Closed Account'
            else:
                base['bucket'] = 'pending_b3'
        linhas.append(base)

    for d, rec in b3.items():
        if d in fep:
            continue
        linhas.append({'check': 'Only B3', 'bucket': 'only_b3', 'cnpj': rec['cnpj'],
                       'client': rec['nome'], 'status': '', 'obs': '',
                       'creation_date': '', 'aging': ''})

    def _idade(r):
        return r['aging'] if isinstance(r['aging'], int) else -1
    linhas.sort(key=lambda r: (r['check'], -_idade(r), _norm(r['client'])))

    buckets = {}
    for r in linhas:
        buckets.setdefault(r['bucket'], []).append(r)

    return {
        'ref': dia.strftime('%Y-%m-%d'), 'ref_fmt': _fmt_date(dia),
        'generated_at': datetime.now().strftime('%d/%m/%Y %H:%M'),
        'b3_file': b3_path, 'fep_file': fep_file,
        'b3_count': len(b3), 'fep_count': len(fep),
        'rows': linhas,
        'counts': {k: len(buckets.get(k, [])) for k in
                   ('matched', 'pending_b3', 'pending_action', 'only_b3', 'justified')},
        'warnings': avisos,
    }


# ── Cache do dia ─────────────────────────────────────────────────────────────

def _cache_path(ref):
    # Sem data, o dia é o MESMO default do `executar` — o último dia útil, não
    # hoje. Com `date.today()` aqui, a tela gravava em D-1 e lia em D0: o
    # batimento rodava, e o GET seguinte voltava vazio dizendo que ninguém
    # rodou.
    d = _parse_date(ref) or dia_util_anterior()
    return os.path.join(_CACHE_DIR, d.strftime('%Y'), d.strftime('%m'),
                        'cgd-recon_{}.json'.format(d.strftime('%Y%m%d')))


def salvar(res):
    path = _cache_path(res.get('ref'))
    try:
        os.makedirs(os.path.dirname(path), exist_ok=True)
        tmp = path + '.tmp'
        with open(tmp, 'w', encoding='utf-8') as fh:
            json.dump(res, fh, ensure_ascii=False)
        os.replace(tmp, path)                    # escrita atômica: nunca meio JSON
    except Exception:
        _LOG.exception('[recon-cgd] não consegui gravar o cache do dia')
    return path


def carregar(ref=None):
    """O resultado já rodado daquele dia, ou `None`. A tela abre com ele: rodar
    de novo é uma decisão de quem opera, não do carregamento da página."""
    path = _cache_path(ref)
    try:
        with open(path, encoding='utf-8') as fh:
            return json.load(fh)
    except Exception:
        return None


# ── E-mail ───────────────────────────────────────────────────────────────────

def _linha_faixa(rows, regra):
    """As linhas de uma tabela do e-mail, já com a cor do aging.

    As duas faixas são as do PortfolioComposer do workflow: aviso na primeira,
    alarme na segunda.
    """
    aviso, alarme = AGING_RULES.get(regra, (5, 15))
    out = []
    for r in rows:
        a = r.get('aging')
        cor = ''
        if isinstance(a, int):
            cor = 'late' if a >= alarme else ('warn' if a >= aviso else '')
        out.append(dict(r, aging_class=cor))
    return out


def montar_email(res):
    """`(assunto, html)` do relatório — as mesmas seções do layout do Alteryx."""
    from flask import render_template
    rows = res.get('rows', [])
    def b(nome):
        return [r for r in rows if r.get('bucket') == nome]
    html = render_template(
        'pages/email-template-recon-cgd.html',
        ref_fmt=res.get('ref_fmt', ''),
        counts=res.get('counts', {}),
        pending_b3=_linha_faixa(b('pending_b3'), 'pending_b3'),
        pending_action=_linha_faixa(b('pending_action'), 'pending_action'),
        justified=b('justified'),
        only_b3=b('only_b3'),
        warnings=res.get('warnings', []),
        current_year=datetime.now().year,
    )
    return 'CGD Matching - {}'.format(res.get('ref_fmt', '')), html


def enviar_email(res, to, cc=None):
    """Manda o relatório. `to`/`cc` vêm de quem chama (o Control Panel guarda os
    destinatários do app) — lista vazia é desfecho, não erro: relatório que não
    saiu de casa tem de aparecer como tal em vez de "enviado"."""
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText

    to = [e for e in (to or []) if str(e).strip()]
    if not to:
        return False, 'no_recipient'
    cc = [e for e in (cc or []) if str(e).strip()]
    assunto, html = montar_email(res)

    msg = MIMEMultipart('alternative')
    msg['Subject'] = assunto
    msg['From'] = _FROM
    msg['To'] = ', '.join(to)
    if cc:
        msg['Cc'] = ', '.join(cc)
    msg.attach(MIMEText('CGD reconciliation for {}. View in an HTML client.'
                        .format(res.get('ref_fmt', '')), 'plain'))
    msg.attach(MIMEText(html, 'html'))
    try:
        with smtplib.SMTP(_SMTP_HOST, _SMTP_PORT) as server:
            server.sendmail(_FROM, to + cc, msg.as_string())
        return True, ''
    except Exception as exc:
        _LOG.warning('[recon-cgd] envio falhou: %s', exc)
        return False, str(exc)
